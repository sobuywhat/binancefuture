# 암호화폐 자동매매 스크립트 (로테이션 모드)
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
[5개 시트 Excel 생성]
- 1분봉: 12400개 수집 → 12000개 남김
- 5분봉: 2601개 수집 → 미완성 1개 제거 → 200개 제거 → 2400개 남김
- 15분봉: 1601개(801+800) 수집 → 미완성 1개 제거 → 800개 제거 → 800개 남김
- 1시간봉(1600개), 일봉(200개), 주봉(200개)
- {TICKER}USDT5M, {TICKER}USDT15M, {TICKER}USDT1H, {TICKER}USDT1D, {TICKER}USDTW

[설정]
- 로테이션: BTC → ETH → XRP → SOL → BNB (순환)
- 거래단위: BTC, ETH, XRP, SOL 모두 16 USDT, BNB는 5.5 USDT
- 실행시간: 매 15분1초, 30분1초, 45분1초, 0분1초 (15분 간격)
- 저장위치: ./cryptodaily15min/{TICKER폴더}/

[필요 패키지]
pip install pandas requests openpyxl python-dateutil PyJWT
"""
import os
import time
import datetime as dt
from typing import Optional, List, Union, Any
import gc  # 가비지 컬렉션 추가
import json
import threading
import sys
import csv
import concurrent.futures  # 병렬 수집을 위한 모듈

import pandas as pd
import requests
from dateutil import tz
import numpy as np
import openpyxl
import subprocess

# ---------- 공통 데이터 정제 함수 ----------
def clean_df_display_format(df, sheet_type=None):
    """섞인 날짜 타입을 하나로 통일하고 숫자 쉼표 제거 (시트별 포맷 개별화)
    
    Args:
        df: DataFrame
        sheet_type: 시트 타입 ('1m', '5m', '15m', '1h', '1h4x', '1d', 'w' 또는 None)
                   None이면 자동 감지 시도
    """
    if df is None or df.empty:
        return df
    
    # 0. 필수: 원본 보호 및 슬라이스 경고 방지
    df = df.copy()
    
    # 1. 날짜 정규화: 이미 datetime64 타입이면 변환 스킵, 문자열만 처리
    if 'Date(UTC)' in df.columns:
        # 이미 datetime64 타입이면 변환 스킵 (데이터 파괴 방지)
        if pd.api.types.is_datetime64_any_dtype(df['Date(UTC)']):
            # 이미 Timestamp 객체이면 그대로 사용
            pass
        else:
            # 문자열인 경우에만 처리
            s = df['Date(UTC)'].astype(str).str.strip()
            
            # 시트별 포맷 정의
            if sheet_type is None:
                # 자동 감지: 샘플 데이터로 포맷 추정
                sample = s.iloc[0] if len(s) > 0 else ''
                if ',' in sample and ':' in sample.split(',')[1] and len(sample.split(',')[1].split(':')) == 2:
                    # 분까지 있는 경우 (1M, 5M, 15M)
                    sheet_type = 'minute'
                elif ',' in sample and ':00' in sample:
                    # 시간만 있는 경우 (1H, 1H4x, 1D, W)
                    sheet_type = 'hour'
                else:
                    sheet_type = 'minute'  # 기본값
            
            # 시트별 포맷 적용
            if sheet_type in ['1m', '5m', '15m', 'minute']:
                # 1M, 5M, 15M: %y/%m/%d,%H:%M (쉼표 포함, 분까지)
                mask_comma = s.str.contains(',', na=False)
                if mask_comma.any():
                    s_comma = s[mask_comma].str.replace(',', ' ', regex=False)
                    df.loc[mask_comma, 'Date(UTC)'] = pd.to_datetime(s_comma, format='%y/%m/%d %H:%M', errors='coerce')
                # 쉼표 없는 형식 fallback
                mask_no_comma = ~mask_comma
                if mask_no_comma.any():
                    df.loc[mask_no_comma, 'Date(UTC)'] = pd.to_datetime(s[mask_no_comma], format='%y/%m/%d %H:%M', errors='coerce')
            elif sheet_type in ['1h', '1h4x', '1d', 'w', 'hour']:
                # 1H, 1H4x, 1D, W: %y/%m/%d,%H:00 (쉼표 포함, 시간만)
                mask_comma = s.str.contains(',', na=False)
                if mask_comma.any():
                    s_comma = s[mask_comma].str.replace(',', ' ', regex=False)
                    df.loc[mask_comma, 'Date(UTC)'] = pd.to_datetime(s_comma, format='%y/%m/%d %H:%M', errors='coerce')
                # 쉼표 없는 형식 fallback
                mask_no_comma = ~mask_comma
                if mask_no_comma.any():
                    df.loc[mask_no_comma, 'Date(UTC)'] = pd.to_datetime(s[mask_no_comma], format='%y/%m/%d %H:%M', errors='coerce')
            else:
                # 알 수 없는 타입: 기본 처리
                mask_comma = s.str.contains(',', na=False)
                if mask_comma.any():
                    s_comma = s[mask_comma].str.replace(',', ' ', regex=False)
                    df.loc[mask_comma, 'Date(UTC)'] = pd.to_datetime(s_comma, format='%y/%m/%d %H:%M', errors='coerce')
                mask_no_comma = ~mask_comma
                if mask_no_comma.any():
                    df.loc[mask_no_comma, 'Date(UTC)'] = pd.to_datetime(s[mask_no_comma], format='%y/%m/%d %H:%M', errors='coerce')
            
            # 파싱 실패한 경우 경고 억제하고 자동 인식
            mask_failed = df['Date(UTC)'].isna()
            if mask_failed.any():
                import warnings
                with warnings.catch_warnings():
                    warnings.simplefilter("ignore", UserWarning)
                    df.loc[mask_failed, 'Date(UTC)'] = pd.to_datetime(s[mask_failed], errors='coerce')
        
        # NaT 제거 안 하면 정렬 시 또 터짐 (데이터 유실 방지)
        df = df.dropna(subset=['Date(UTC)'])
    
    # 2. 숫자 정규화 (모든 지표 열 포함)
    numeric_cols = ['종', '시', '고', '저', 'Vol.', 'SMA3', 'SMA5', 'SMA7', 'SMA10', 'SMA12', 'SMAF', 'SMA20', 'SMA25', 'SMA28', 'SMA40', 'SMA15', 'SMA35', 'SMA50', 'SMA80', 'SMA100', 'SMA200', 'SMA400', 'SMA800', 
                    'SFast', 'Fast', 'Base', '1HMSFast', '1HMSF', 'SPRD', 'SPRD2', 'Max200', 'Min200', 'Max100', 'Min100', 'Max70', 'Min70', 'Max400', 'Min400', 'Max15', 'Min15', 'Max25', 'Min25',
                    '하단', '상단', 'buyside', 'sellside', '1HCL', '-1HCL', 'p', 'p1H', 'TP', 'StoSP', 'StoSU', 'TPC', 'TPCS', 'NBS']
    for col in numeric_cols:
        if col in df.columns:
            if df[col].dtype == 'object':
                df[col] = df[col].astype(str).str.replace(',', '', regex=False)
            df[col] = pd.to_numeric(df[col], errors='coerce')
    return df

# ---------- Binance auth & order ----------
import hashlib

# -------------------- 공통 --------------------
KST = tz.gettz("Asia/Seoul")

# ---------- 캔들 개수 설정 (Y = 15분봉 최종 개수) ----------
Y = 400  # 15분봉 최종 개수 (기본값)

import math

def calculate_candle_count(y: int = Y) -> dict:
    """
    최종 캔들 개수 (CANDLE_COUNT) 계산
    용도: 엑셀 저장 전 데이터 제한 시 사용
    
    Args:
        y: 15분봉 최종 개수 (기본값: 400)
    
    Returns:
        dict: {'1m': int, '5m': int, '15m': int, '1h': int, '1d': int, '1w': int}
    """
    return {
        '1m': 12000,  # 1분봉: 12400개 수집 → 400개 제거 → 12000개
        '5m': 2400,   # 5분봉: 2601개 수집 → 미완성 1개 제거 → 200개 제거 → 2400개
        '15m': 800,   # 15분봉: 1601개 수집 → 미완성 1개 제거 → 800개 제거 → 800개
        '1h': 1600,  # 1시간봉: Source 기준 1600개 (2400개 수집 후 과거 800개 제거)
        '1d': math.ceil(y / 4 / 24 / 200) * 200,  # 1일봉: roundup(Y/4/24/200) × 200
        '1w': math.ceil(y / 4 / 24 / 7 / 200) * 200  # 7일봉: roundup(Y/4/24/7/200) × 200
    }

def calculate_collection_count(y: int = Y) -> dict:
    """
    수집 캔들 개수 (COLLECTION_COUNT) 계산
    용도: API에서 데이터 수집 시 사용
    
    Args:
        y: 15분봉 최종 개수 (기본값: 400)
    
    Returns:
        dict: {'1m': int, '5m': int, '15m': int, '1h': int, '1d': int, '1w': int}
    """
    return {
        '1m': 12000 + 400,  # 1분봉: 12400개 수집 → 12000개 남김 (400개 제거)
        '5m': 2401 + 200,   # 5분봉: 2601개 수집 → 미완성 1개 제거 → 200개 제거 → 2400개 남김
        '15m': 801 + 800,   # 15분봉: 1601개 수집 → 미완성 1개 제거 → 800개 제거 → 800개 남김
        '1h': 2400,  # 1시간봉: Source 기준 2400개 수집 (SMA800 계산을 위해 최소 800개 이상 필요, 최종 1600개 저장)
        '1d': math.ceil((y + 200) / 4 / 24 / 200) * 200,  # 1일봉: roundup((Y+200)/4/24/200) × 200
        '1w': math.ceil((y + 200) / 4 / 24 / 7 / 200) * 200  # 7일봉: roundup((Y+200)/4/24/7/200) × 200
    }

# ---------- Binance API 설정 ----------
BINANCE_API_BASE = "https://api.binance.com"
BINANCE_FUTURES_BASE = "https://fapi.binance.com"

# ---------- 선물 전략 설정 (LS 시그널, 종가 기준 1만 달러 단위 식) ----------
FUTURES_BASE_PRICE = 60_000
FUTURES_BASE_TOTAL_USDT = 210
FUTURES_BASE_TP_USDT = 70
FUTURES_STEP_TOTAL_PER_10K = 30
FUTURES_STEP_TP_PER_10K = 10
FUTURES_MIN_QTY_BTC = 0.001
FUTURES_MIN_NOTIONAL_USDT = 5    # 거래소 최소 주문 금액(참고), 진입은 아래 금액 사용
FUTURES_POSITION_USDT = 725      # BTC, ETH, XRP, SOL 진입 725 USDT
FUTURES_TP_PART_USDT = 100       # 25%×3 TP: 각 100 USDT (나머지 25% = 100 USDT 추세전환까지 유지)
FUTURES_BNB_TOTAL_USDT = 100     # BNB만 진입 100 USDT
FUTURES_BNB_TP_PART_USDT = 25    # BNB 4분할: 25%×3 = 각 25 USDT TP, 나머지 25% 유지 (25 USDT ≥ 거래소 최소)
FUTURES_SL_PERCENT = 0.012      # SL 1.2% 3분할 (진입가 대비, 롱/숏 모두)
# BE(Break-Even) 이동: 1차 익절(TP 1/3) 체결 시 남은 물량 SL을 '진입가 + 왕복 수수료'로 이동.
# 수익 났던 포지션이 손실로 전환되는 것을 막기 위함. 단순 진입가(0%)면 수수료로 미세 손실 발생.
# 바이낸스 선물 BNB 할인 기준 왕복 수수료 상당 → 진입가보다 약 +0.06% 위 지점에 SL 배치.
FUTURES_BE_OFFSET_PERCENT = 0.0006  # 롱: 진입가+0.06%에서 매도 트리거 / 숏: 진입가-0.06%에서 매수 트리거 (수수료 상쇄)
ENABLE_FUTURES_LS_STRATEGY = True   # True 시 LS=1/-1일 때 선물 전략 실행 (ENABLE_TRADING=True여야 함)
# 선물 주문 허용: ROTATION_TICKERS 5개(BTC,ETH,XRP,SOL,BNB) 모두 주문 가능

# 심볼 메타데이터 캐시 (precision, filters 등)
_symbol_info_cache = {}
QUOTE_PRECISION_MAP = {}

def binance_get_symbol_info(symbol: str) -> dict:
    """Binance exchangeInfo에서 심볼 정보를 조회하고 캐시합니다."""
    try:
        if symbol in _symbol_info_cache and (time.time() - _symbol_info_cache[symbol]['_ts'] < 3600):
            return _symbol_info_cache[symbol]['data']
        r = requests.get(f"{BINANCE_API_BASE}/api/v3/exchangeInfo", params={"symbol": symbol}, timeout=10)
        r.raise_for_status()
        data = r.json()
        if 'symbols' in data and data['symbols']:
            info = data['symbols'][0]
            _symbol_info_cache[symbol] = { 'data': info, '_ts': time.time() }
            return info
        raise RuntimeError(f"exchangeInfo empty for {symbol}")
    except Exception as e:
        raise RuntimeError(f"exchangeInfo fetch failed for {symbol}: {e}")

def init_symbol_quote_precisions(symbols: list[str]) -> None:
    """지정 심볼들의 quotePrecision(또는 quoteAssetPrecision)을 한 번 조회하여 캐시에 고정합니다."""
    global QUOTE_PRECISION_MAP
    for sym in symbols:
        try:
            info = binance_get_symbol_info(sym)
            qp = info.get('quotePrecision') if 'quotePrecision' in info else info.get('quoteAssetPrecision', 5)
            try:
                QUOTE_PRECISION_MAP[sym] = int(qp)
            except:
                QUOTE_PRECISION_MAP[sym] = 5
        except Exception:
            # 실패 시 보수적 기본값
            QUOTE_PRECISION_MAP[sym] = 5


# ---------- 폴링 및 주문 설정 ----------
ENABLE_POLLING = True   # 폴링(스케줄러) 활성화/비활성화
ENABLE_TRADING = True   # 주문 전송 활성화 (선물·스팟 분리는 아래 플래그)
ENABLE_SPOT_TRADING = False  # 스팟 주문 전송 (False=스팟 미전송, 선물만 주문)
ENABLE_ASSET_RECORDING = False  # 자산기록(잔고 스냅샷·PNLcal 호출) 🔒 비활성화 (로그는 LOG_DIR에 별도 기록)
USE_SERVER_TIME = True  # 서버 시간 동기화 사용 (시간 오차 대응)

# ---------- 거래 단위 설정 (USDT) ----------
TRADING_UNIT = 16  # 1unit = 16 USDT (기본 거래 단위)

# ---------- 티커 설정 ----------
TICKER = "ETH"  # 거래할 암호화폐 티커 (기본값)

# ---------- 로테이션 설정 ----------
ROTATION_TICKERS = ["BTC", "ETH", "XRP", "SOL", "BNB"]  # 로테이션 순서
ROTATION_TRADING_UNITS = {
    "BTC": 30,    # USDT
    "ETH": 30,    # USDT
    "XRP": 30,    # USDT
    "SOL": 30,    # USDT
    "BNB": 5.5    # USDT
}

# ---------- 티커별 LOT_SIZE 설정 (stepSize) ----------
SYMBOL_STEP_SIZE = {
    "BTC": 0.00001,  # BTC의 stepSize
    "ETH": 0.0001,   # ETH의 stepSize (0.0001 단위)
    "XRP": 0.1,      # XRP의 stepSize
    "SOL": 0.001,    # SOL의 stepSize
    "BNB": 0.001     # BNB의 stepSize
}

# ---------- 바이낸스 심볼별 수량 정밀도 (stepSize) ----------
SYMBOL_QTY_PRECISION = {
    "BTCUSDT": 5,  # 0.00001 BTC
    "ETHUSDT": 4,  # 0.0001 ETH
    "SOLUSDT": 3,  # 0.001 SOL
    "XRPUSDT": 1,  # 0.1 XRP
    "BNBUSDT": 3   # 0.001 BNB
}

# ---------- 바이낸스 심볼별 USDT 금액 정밀도 (실제 사이트 기준) ----------
SYMBOL_USDT_PRECISION = {
    "BTCUSDT": 7,  # 5.5386805 USDT (7자리)
    "ETHUSDT": 6,  # 5.114265 USDT (6자리)
    "SOLUSDT": 5,  # 5.00500 USDT (5자리)
    "XRPUSDT": 5,  # 5.12211 USDT (5자리)
    "BNBUSDT": 5   # 5.12345 USDT (5자리)
}

# ---------- 바이낸스 거래 수수료 설정 ----------
BINANCE_TRADING_FEE = 0.00075  # 바이낸스 스팟 거래 수수료 0.075% (0.00075)
BINANCE_FUTURES_FEE = 0.0004   # 바이낸스 선물 Taker 수수료 0.04% (시장가)
BINANCE_FUTURES_MAKER_FEE = 0.0002  # 바이낸스 선물 Maker 수수료 0.02% (지정가 진입·TP 전부 지정가 사용)
BINANCE_BNB_FEE_DISCOUNT = 0.9  # BNB로 수수료 결제 시 10% 할인

# ==========================================
# 로그 설정: 로그 보관 기간 (일 단위)
# ==========================================
LOG_DIR = 'logs'
DAYS_TO_KEEP = 30

# ==========================================
# 터미널 로그 기록 시스템
# ==========================================
# 스크립트 디렉토리 경로
script_dir = os.path.dirname(os.path.abspath(__file__))
LOG_DIR_ABS = os.path.join(script_dir, LOG_DIR)

# 로그 디렉토리 생성
if not os.path.exists(LOG_DIR_ABS):
    os.makedirs(LOG_DIR_ABS)

# UTC 기준 로그 파일명 생성 함수
def get_current_log_filename(exchange="BINANCE"):
    """UTC 기준 현재 날짜로 로그 파일명을 생성합니다.
    
    Args:
        exchange: 거래소 이름 (기본값: "BINANCE")
    
    Returns:
        str: 로그 파일의 전체 경로
    """
    now_utc = dt.datetime.now(tz.UTC)
    current_date_utc = now_utc.strftime("%Y%m%d")
    return os.path.join(LOG_DIR_ABS, f"{exchange}_log_{current_date_utc}.txt")

def get_utc_date_str(offset_days=0):
    """UTC 기준 날짜 문자열을 반환합니다 (오프셋 지원).
    
    Args:
        offset_days: 날짜 오프셋 (기본값: 0, -1이면 어제, 1이면 내일)
    
    Returns:
        str: YYYYMMDD 형식의 날짜 문자열
    """
    now_utc = dt.datetime.now(tz.UTC)
    target_date = now_utc + dt.timedelta(days=offset_days)
    return target_date.strftime("%Y%m%d")

# DualLogger 클래스: 터미널과 파일에 동시 출력 (UTC 0시 기준 자동 파일 분리)
class DualLogger:
    def __init__(self, log_dir, exchange="BINANCE"):
        self.terminal = sys.stdout  # 원본 터미널 출력
        self.log_dir = log_dir
        self.exchange = exchange
        self.current_date_utc = None  # 현재 사용 중인 UTC 날짜
        self.log = None
        self._open_log_file()  # 초기 파일 열기
    
    def _get_current_date_utc(self):
        """현재 UTC 날짜를 YYYYMMDD 형식으로 반환합니다."""
        now_utc = dt.datetime.now(tz.UTC)
        return now_utc.strftime("%Y%m%d")
    
    def _open_log_file(self):
        """현재 UTC 날짜에 맞는 로그 파일을 엽니다."""
        current_date = self._get_current_date_utc()
        filename = get_current_log_filename(self.exchange)
        
        # 기존 파일이 열려있으면 닫기
        if self.log is not None:
            try:
                self.log.close()
            except:
                pass
        
        # 새 파일 열기
        self.log = open(filename, "a", encoding='utf-8')
        self.current_date_utc = current_date
    
    def write(self, message):
        # UTC 날짜가 바뀌었는지 체크 (UTC 0시 기준)
        current_date = self._get_current_date_utc()
        if current_date != self.current_date_utc:
            # 날짜가 바뀌었으면 새 파일 열기
            self._open_log_file()
        
        self.terminal.write(message)  # 터미널에 출력
        self.log.write(message)      # 파일에도 기록
        self.flush()
    
    def flush(self):
        self.terminal.flush()  # 터미널 버퍼 플러시
        if self.log is not None:
            self.log.flush()       # 파일 버퍼 플러시

# 로거 연결: 선물 전용 로그 파일 (BINANCE_FUTURES_log_YYYYMMDD.txt)
logger = DualLogger(LOG_DIR_ABS, exchange="BINANCE_FUTURES")
sys.stdout = logger  # 표준 출력을 DualLogger로 리다이렉트
sys.stderr = logger  # 표준 에러도 DualLogger로 리다이렉트

# ==========================================
# [설정] 디스코드 웹후크 (파일에서 읽기)
# ==========================================
WEBHOOK_FILE_PATH = r"C:\Users\upharm\Desktop\UPBIT PY\discordwebhook\binancefwebhook.txt"

def get_timestamp(include_ms: bool = True):
    """현재 시간을 [KST HH:MM:SS.fff](UTC HH:MM:SS.fff) 형식으로 반환합니다.
    
    Args:
        include_ms: True면 밀리초 포함, False면 HH:MM:SS만 (디스코드 전송용 등)
    
    Returns:
        str: [KST 시간](UTC 시간) 형식의 타임스탬프 문자열
    
    Note:
        - KST는 참고용으로만 표시 (계산 로직에는 사용하지 않음)
        - UTC 시간이 실제 기준 시간 (모든 로직은 UTC 기준)
    """
    now_kst = dt.datetime.now(KST)  # 참고용
    now_utc = dt.datetime.now(tz.UTC)  # 실제 기준 시간
    if include_ms:
        return f"[{now_kst.strftime('%H:%M:%S.%f')[:-3]}](UTC {now_utc.strftime('%H:%M:%S.%f')[:-3]})"
    return f"[{now_kst.strftime('%H:%M:%S')}](UTC {now_utc.strftime('%H:%M:%S')})"

def wait_for_file_ready(file_path: str, max_wait_seconds: int = 5, check_interval: float = 0.1) -> bool:
    """
    파일이 완전히 저장되고 읽을 수 있을 때까지 대기합니다.
    
    Args:
        file_path: 확인할 파일 경로
        max_wait_seconds: 최대 대기 시간 (초)
        check_interval: 확인 간격 (초)
    
    Returns:
        파일이 준비되었으면 True, 그렇지 않으면 False
    """
    import time
    start_time = time.time()
    last_size = -1
    
    while time.time() - start_time < max_wait_seconds:
        if not os.path.exists(file_path):
            time.sleep(check_interval)
            continue
        
        try:
            current_size = os.path.getsize(file_path)
            # 파일 크기가 안정화되었는지 확인 (연속 3번 같은 크기)
            if current_size == last_size and current_size > 0:
                # 파일이 잠겨있지 않은지 확인 (읽기 모드로 열어보기)
                try:
                    with open(file_path, 'rb') as f:
                        f.read(1)
                    return True
                except (IOError, PermissionError):
                    time.sleep(check_interval)
                    continue
            last_size = current_size
            time.sleep(check_interval)
        except (OSError, IOError):
            time.sleep(check_interval)
            continue
    
    # 최대 대기 시간 초과
    return False

# 서버 시간 캐시 (중복 호출 방지)
_server_time_cache = None
_server_time_cache_time = 0

def get_binance_server_time():
    """바이낸스 서버 시간 조회 (시간 동기화용) - 캐시 적용"""
    global _server_time_cache, _server_time_cache_time
    
    current_time = time.time()
    # 5초 이내 캐시된 시간이 있으면 재사용
    if _server_time_cache and (current_time - _server_time_cache_time) < 5:
        return _server_time_cache
    
    try:
        r = requests.get(f"{BINANCE_API_BASE}/api/v3/time", timeout=5)
        if r.status_code == 200:
            server_time = r.json()['serverTime']
            _server_time_cache = server_time
            _server_time_cache_time = current_time
            # print(f"{get_timestamp()} 🕐 서버 시간 동기화: {server_time}")
            return server_time
        else:
            print(f"{get_timestamp()} ⚠️ 서버 시간 조회 실패, 로컬 시간 사용")
            return int(time.time() * 1000)
    except Exception as e:
        print(f"{get_timestamp()} ⚠️ 서버 시간 조회 오류: {e}, 로컬 시간 사용")
        return int(time.time() * 1000)

def get_futures_server_time() -> int:
    """선물 서버 시간(ms). 서명 -1022 방지용으로 fapi 기준 사용."""
    try:
        r = requests.get(f"{BINANCE_FUTURES_BASE}/fapi/v1/time", timeout=5)
        if r.status_code == 200:
            return int(r.json()["serverTime"])
    except Exception:
        pass
    return int(time.time() * 1000)

def _binance_headers(query_string: str = ""):
    """바이낸스 API 인증 헤더 생성
    
    Args:
        query_string: 쿼리 문자열 (파라미터)
    
    Returns:
        tuple: (headers, signature, timestamp, recv_window)
    
    Note:
        - timestamp는 UTC 기준 밀리초 타임스탬프 사용
        - 서버 시간 동기화 옵션(USE_SERVER_TIME) 사용 시 바이낸스 서버 시간 사용
        - 그 외에는 로컬 시간을 UTC 기준으로 변환하여 사용
    """
    # 스크립트 폴더의 binanceaccountinfo 폴더에서 키 파일 읽기
    script_dir = os.path.dirname(os.path.abspath(__file__))
    api_key_path = os.path.join(script_dir, "binanceaccountinfo", "binanceapikey.txt")
    api_secret_path = os.path.join(script_dir, "binanceaccountinfo", "binanceapisecret.txt")
    
    try:
        with open(api_key_path, 'r') as f:
            api_key = f.read().strip()
        with open(api_secret_path, 'r') as f:
            api_secret = f.read().strip()
    except FileNotFoundError as e:
        raise RuntimeError(f"키 파일을 찾을 수 없습니다: {e}")
    except Exception as e:
        raise RuntimeError(f"키 파일 읽기 실패: {e}")
    
    if not api_key or not api_secret:
        raise RuntimeError("binanceapikey.txt 또는 binanceapisecret.txt 파일이 비어있습니다.")

    # timestamp와 recvWindow 추가 (서버 시간 동기화 옵션)
    if USE_SERVER_TIME:
        timestamp = get_binance_server_time()
    else:
        timestamp = int(time.time() * 1000)
    recv_window = 10000  # 10초 (네트워크 지연 대응)
    
    # 간단한 방식으로 query_string 구성 (디버깅용)
    if query_string:
        query_string = f"{query_string}&timestamp={timestamp}&recvWindow={recv_window}"
    else:
        query_string = f"timestamp={timestamp}&recvWindow={recv_window}"
    
    # HMAC SHA256 서명 생성
    import hmac as hmac_lib
    signature = hmac_lib.new(
        api_secret.encode('utf-8'),
        query_string.encode('utf-8'),
        hashlib.sha256
    ).hexdigest()
    
    # DEBUG 출력 제거
    
    return {
        "X-MBX-APIKEY": api_key,
        "Content-Type": "application/json"
    }, signature, timestamp, recv_window

def _binance_fapi_headers(query_string: str = ""):
    """바이낸스 선물(fapi) API 인증 헤더 생성. binancefapikey.txt / binancefapisecret.txt 사용."""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    api_key_path = os.path.join(script_dir, "binanceaccountinfo", "binancefapikey.txt")
    api_secret_path = os.path.join(script_dir, "binanceaccountinfo", "binancefapisecret.txt")
    try:
        with open(api_key_path, 'r', encoding='utf-8-sig') as f:
            api_key = f.read().strip()
        with open(api_secret_path, 'r', encoding='utf-8-sig') as f:
            api_secret = f.read().strip()
    except FileNotFoundError as e:
        raise RuntimeError(f"선물 키 파일을 찾을 수 없습니다: {e}")
    except Exception as e:
        raise RuntimeError(f"선물 키 파일 읽기 실패: {e}")
    if not api_key or not api_secret:
        raise RuntimeError("binancefapikey.txt 또는 binancefapisecret.txt 파일이 비어있습니다.")
    if USE_SERVER_TIME:
        timestamp = get_futures_server_time()
    else:
        timestamp = int(time.time() * 1000)
    recv_window = 10000
    if query_string:
        query_string = f"{query_string}&timestamp={timestamp}&recvWindow={recv_window}"
    else:
        query_string = f"timestamp={timestamp}&recvWindow={recv_window}"
    import hmac as hmac_lib
    signature = hmac_lib.new(
        api_secret.encode('utf-8'),
        query_string.encode('utf-8'),
        hashlib.sha256
    ).hexdigest()
    return {
        "X-MBX-APIKEY": api_key,
        "Content-Type": "application/json"
    }, signature, timestamp, recv_window

def binance_test_order(symbol: str, side: str, order_type: str, **kwargs) -> bool:
    """바이낸스 드라이런 테스트 (POST /api/v3/order/test)"""
    try:
        # query_string 생성 (파라미터 정렬)
        test_params = {
            "symbol": symbol,
            "side": side,
            "type": order_type,
            **kwargs
        }
        
        # 파라미터를 정렬하여 query_string 생성
        sorted_params = sorted(test_params.items())
        query_string = '&'.join([f"{key}={value}" for key, value in sorted_params])
        
        headers, signature, timestamp, recv_window = _binance_headers(query_string)
        
        # 최종 테스트 파라미터 구성
        final_params = {
            "symbol": symbol,
            "side": side,
            "type": order_type,
            "timestamp": timestamp,
            "recvWindow": recv_window,
            "signature": signature,
            **kwargs
        }
        
        # 드라이런 테스트 실행
        r = requests.post(f"{BINANCE_API_BASE}/api/v3/order/test", params=final_params, headers=headers, timeout=10)
        
        if r.status_code == 200:
            print(f"{get_timestamp()} ✅ 드라이런 테스트 성공: {r.json()}")
            return True
        else:
            print(f"{get_timestamp()} ❌ 드라이런 테스트 실패: {r.status_code} {r.text}")
            return False
            
    except Exception as e:
        print(f"{get_timestamp()} ❌ 드라이런 테스트 오류: {e}")
        return False

def binance_ticker_price(symbol: str = None) -> float:
    """바이낸스 현재가 조회 (USDT)"""
    if symbol is None:
        symbol = f"{TICKER}USDT"
    r = requests.get(f"{BINANCE_API_BASE}/api/v3/ticker/price", params={"symbol": symbol}, timeout=10)
    r.raise_for_status()
    data = r.json()
    return float(data["price"])

def binance_fapi_ticker_price(symbol: str = None) -> float:
    """바이낸스 선물 현재가 조회 (fapi)"""
    if symbol is None:
        symbol = f"{TICKER}USDT"
    r = requests.get(f"{BINANCE_FUTURES_BASE}/fapi/v1/ticker/price", params={"symbol": symbol}, timeout=10)
    r.raise_for_status()
    data = r.json()
    return float(data["price"])

def get_futures_orderbook_snapshot(symbol: str):
    """선물 호가창(fapi/v1/depth)에서 ask, bid, ask_q, bid_q 조회. 스마트 주문 엔진용."""
    try:
        r = requests.get(f"{BINANCE_FUTURES_BASE}/fapi/v1/depth", params={"symbol": symbol, "limit": 20}, timeout=10)
        r.raise_for_status()
        data = r.json()
        if data and "asks" in data and "bids" in data and len(data["asks"]) > 0 and len(data["bids"]) > 0:
            ask = float(data["asks"][0][0])
            ask_q = float(data["asks"][0][1])
            bid = float(data["bids"][0][0])
            bid_q = float(data["bids"][0][1])
            return ask, bid, ask_q, bid_q
        return None, None, None, None
    except Exception as e:
        print(f"{get_timestamp()} ❌ 선물 호가창 조회 실패: {e}")
        return None, None, None, None

# 선물 exchangeInfo 캐시 (LOT_SIZE, MIN_NOTIONAL 등 — 동적 규칙 반영)
_futures_exchange_info_cache = {}
_futures_exchange_info_ts = 0
FUTURES_EXCHANGE_INFO_CACHE_TTL = 0  # 0 = 캐시 없음, 시행마다 exchangeInfo 조회

def get_futures_exchange_info(symbol: str = "BTCUSDT", use_cache: bool = False) -> dict:
    """
    GET /fapi/v1/exchangeInfo 로 해당 심볼의 LOT_SIZE(minQty, stepSize), MIN_NOTIONAL(notional) 등을 조회.
    시행마다 최신 규칙 확인 (캐시 미사용).
    Returns:
        dict: minQty, stepSize, notional, quantityPrecision, pricePrecision (없으면 기본값)
    """
    global _futures_exchange_info_cache, _futures_exchange_info_ts
    now = time.time()
    if use_cache and symbol in _futures_exchange_info_cache and (now - _futures_exchange_info_ts) < FUTURES_EXCHANGE_INFO_CACHE_TTL:
        return _futures_exchange_info_cache[symbol]
    try:
        r = requests.get(f"{BINANCE_FUTURES_BASE}/fapi/v1/exchangeInfo", timeout=10)
        r.raise_for_status()
        data = r.json()
    except Exception as e:
        print(f"{get_timestamp()} ⚠️ 선물 exchangeInfo 조회 실패: {e}")
        return {"minQty": 0.001, "stepSize": 0.001, "notional": 5.0, "quantityPrecision": 3, "pricePrecision": 1, "tickSize": 0.01}
    out = {"minQty": 0.001, "stepSize": 0.001, "notional": 5.0, "quantityPrecision": 3, "pricePrecision": 1, "tickSize": 0.01}
    for s in data.get("symbols", []):
        if s.get("symbol") != symbol:
            continue
        out["quantityPrecision"] = int(s.get("quantityPrecision", 3))
        out["pricePrecision"] = int(s.get("pricePrecision", 1))
        for f in s.get("filters", []):
            ft = f.get("filterType", "")
            if ft == "LOT_SIZE":
                out["minQty"] = float(f.get("minQty", "0.001"))
                out["stepSize"] = float(f.get("stepSize", "0.001"))
            elif ft == "PRICE_FILTER":
                raw = f.get("tickSize", "0.01")
                out["tickSize"] = float(raw) if not isinstance(raw, (int, float)) else raw
            elif ft == "MIN_NOTIONAL":
                out["notional"] = float(f.get("notional", "5"))
        break
    # notional은 API 값 그대로 사용 (티커별 최소 주문 금액)
    _futures_exchange_info_cache[symbol] = out
    _futures_exchange_info_ts = now
    return out

def adjust_price_to_tick_futures(symbol: str, price: float) -> float:
    """선물 가격을 해당 심볼의 틱 사이즈(fapi exchangeInfo PRICE_FILTER)에 맞춤. -4014 방지."""
    info = get_futures_exchange_info(symbol, use_cache=True)
    tick_size = info.get("tickSize", 0.01)
    if tick_size <= 0:
        return price
    precision = _precision_from_tick_or_step(tick_size)
    adjusted = round(round(price / tick_size) * tick_size, precision)
    return adjusted

def _format_qty_for_futures(symbol: str, quantity: float) -> str:
    """선물 수량을 stepSize 기준으로 내림·포맷 (부동소수점 잔여 제거)."""
    info = get_futures_exchange_info(symbol, use_cache=True)
    step_size = info.get("stepSize", 0.001)
    qty_rounded = _round_down_to_step(quantity, step_size)
    decimals = min(_precision_from_tick_or_step(step_size), 8)
    qty_rounded = round(qty_rounded, decimals)  # n*step_size 부동소수점 잔여 제거
    s = f"{qty_rounded:.8f}".rstrip("0").rstrip(".")
    return s

def _format_price_for_futures(symbol: str, price: float) -> str:
    """선물 가격을 tickSize 기준으로 포맷 문자열 (API 전송용)."""
    info = get_futures_exchange_info(symbol, use_cache=True)
    tick_size = info.get("tickSize", 0.01)
    prec = _precision_from_tick_or_step(tick_size)
    p = round(round(price / tick_size) * tick_size, prec) if tick_size > 0 else price
    return f"{p:.8f}".rstrip("0").rstrip(".")

def _round_down_to_step(qty: float, step_size: float) -> float:
    """수량을 stepSize 단위로 내림. step_size 예: 0.001, 1e-5 등 (티커별 소수점 오류 방지)"""
    if step_size <= 0:
        return qty
    n = int(qty / step_size)
    result = n * step_size
    decimals = _precision_from_tick_or_step(step_size)
    return round(result, min(decimals, 8))

def _round_up_to_step(qty: float, step_size: float) -> float:
    """수량을 stepSize 단위로 올림 (최소 notional 등 충족용)"""
    if step_size <= 0:
        return qty
    n = int(np.ceil(qty / step_size)) if qty > 0 else 0
    result = n * step_size
    decimals = _precision_from_tick_or_step(step_size)
    return round(result, min(decimals, 8))

def print_futures_exchange_info_summary():
    """5개 티커(BTC,ETH,XRP,SOL,BNB) 선물 exchangeInfo 호출값 정리 → 터미널 출력 + 디스코드 전송."""
    separator = "=" * 80
    lines = []
    header = f"{get_timestamp()} 📋 선물 exchangeInfo (5개 티커) minQty | stepSize | notional(USDT) | qtyPrec | pricePrec"
    lines.append(header)
    for ticker in ROTATION_TICKERS:
        symbol = f"{ticker}USDT"
        try:
            info = get_futures_exchange_info(symbol, use_cache=False)
            line = f"[{ticker}] minQty={info['minQty']} stepSize={info['stepSize']} notional={info['notional']} qtyPrec={info['quantityPrecision']} pricePrec={info['pricePrecision']}"
            lines.append(line)
        except Exception as e:
            line = f"[{ticker}] 조회 실패: {e}"
            lines.append(line)
    full_msg = "\n".join(lines)
    print(separator)
    print(full_msg)
    print(separator)
    send_discord_message(full_msg)

def get_futures_account() -> dict:
    """선물 계정 요약 조회 (fapi/v2/account). 잔고·미실현손익·마진 등."""
    try:
        headers, signature, timestamp, recv_window = _binance_fapi_headers("")
        query_signed = f"timestamp={timestamp}&recvWindow={recv_window}&signature={signature}"
        r = requests.get(f"{BINANCE_FUTURES_BASE}/fapi/v2/account?{query_signed}", headers=headers, timeout=10)
        if r.status_code == 200:
            return r.json()
        return {}
    except Exception as e:
        print(f"{get_timestamp()} ⚠️ 선물 계정 조회 실패: {e}")
        return {}

def get_futures_position_risk(symbol: str) -> list:
    """선물 포지션 조회 (fapi/v2/positionRisk). 해당 심볼 포지션 목록 반환."""
    headers, signature, timestamp, recv_window = _binance_fapi_headers(f"symbol={symbol}")
    query_signed = f"symbol={symbol}&timestamp={timestamp}&recvWindow={recv_window}&signature={signature}"
    r = requests.get(f"{BINANCE_FUTURES_BASE}/fapi/v2/positionRisk?{query_signed}", headers=headers, timeout=10)
    r.raise_for_status()
    data = r.json()
    return data if isinstance(data, list) else []

def cancel_all_futures_orders(symbol: str) -> bool:
    """선물 미체결 주문 전량 취소 (DELETE fapi/v1/allOpenOrders)"""
    try:
        headers, signature, timestamp, recv_window = _binance_fapi_headers(f"symbol={symbol}")
        query_signed = f"symbol={symbol}&timestamp={timestamp}&recvWindow={recv_window}&signature={signature}"
        r = requests.delete(f"{BINANCE_FUTURES_BASE}/fapi/v1/allOpenOrders?{query_signed}", headers=headers, timeout=10)
        if r.status_code == 200:
            print(f"{get_timestamp()} ✅ 선물 미체결 주문 전량 취소 완료: {symbol}")
            return True
        print(f"{get_timestamp()} ⚠️ 선물 주문 취소 응답: {r.status_code} {r.text}")
        return False
    except Exception as e:
        print(f"{get_timestamp()} ❌ 선물 주문 취소 실패: {e}")
        return False

def get_futures_open_orders(symbol: str) -> list:
    """선물 미체결 주문 조회 (GET fapi/v1/openOrders). reduce_only TP/SL 구분용. (일반 LIMIT 등만 포함, Algo STOP_MARKET 제외)"""
    try:
        headers, signature, timestamp, recv_window = _binance_fapi_headers(f"symbol={symbol}")
        query_signed = f"symbol={symbol}&timestamp={timestamp}&recvWindow={recv_window}&signature={signature}"
        r = requests.get(f"{BINANCE_FUTURES_BASE}/fapi/v1/openOrders?{query_signed}", headers=headers, timeout=10)
        if r.status_code == 200:
            data = r.json()
            return data if isinstance(data, list) else []
        return []
    except Exception as e:
        print(f"{get_timestamp()} ⚠️ 선물 미체결 주문 조회 실패: {e}")
        return []

def get_futures_open_algo_orders(symbol: str) -> list:
    """선물 Algo 미체결 조회 (GET fapi/v1/openAlgoOrders). STOP_MARKET 등 조건부 SL/TP는 여기만 있음."""
    try:
        headers, signature, timestamp, recv_window = _binance_fapi_headers(f"symbol={symbol}")
        query_signed = f"symbol={symbol}&timestamp={timestamp}&recvWindow={recv_window}&signature={signature}"
        r = requests.get(f"{BINANCE_FUTURES_BASE}/fapi/v1/openAlgoOrders?{query_signed}", headers=headers, timeout=10)
        if r.status_code == 200:
            data = r.json()
            return data if isinstance(data, list) else []
        return []
    except Exception as e:
        print(f"{get_timestamp()} ⚠️ 선물 Algo 미체결 조회 실패: {e}")
        return []

def cancel_futures_algo_order(algo_id: int) -> bool:
    """선물 Algo 주문 취소 (DELETE fapi/v1/algoOrder). algoId만 필요."""
    try:
        from urllib.parse import urlencode
        params = {"algoId": algo_id}
        query_string = urlencode(sorted(params.items()))
        headers, signature, timestamp, recv_window = _binance_fapi_headers(query_string)
        full_query = f"{query_string}&timestamp={timestamp}&recvWindow={recv_window}&signature={signature}"
        r = requests.delete(f"{BINANCE_FUTURES_BASE}/fapi/v1/algoOrder?{full_query}", headers=headers, timeout=10)
        if r.status_code == 200:
            return True
        return False
    except Exception:
        return False

def cancel_futures_order(symbol: str, order_id: int) -> bool:
    """선물 단일 주문 취소 (DELETE fapi/v1/order)."""
    try:
        headers, signature, timestamp, recv_window = _binance_fapi_headers(f"symbol={symbol}&orderId={order_id}")
        query_signed = f"symbol={symbol}&orderId={order_id}&timestamp={timestamp}&recvWindow={recv_window}&signature={signature}"
        r = requests.delete(f"{BINANCE_FUTURES_BASE}/fapi/v1/order?{query_signed}", headers=headers, timeout=10)
        if r.status_code == 200:
            return True
        return False
    except Exception:
        return False

def set_futures_leverage(symbol: str, leverage: int = 1) -> bool:
    """선물 레버리지 설정 (POST fapi/v1/leverage). 롱/숏 전부 1배만 사용하도록 주문 전 호출."""
    try:
        from urllib.parse import urlencode
        params = [("leverage", leverage), ("symbol", symbol)]
        query_string = urlencode(params)
        headers, signature, timestamp, recv_window = _binance_fapi_headers(query_string)
        full_query = f"{query_string}&timestamp={timestamp}&recvWindow={recv_window}&signature={signature}"
        url = f"{BINANCE_FUTURES_BASE}/fapi/v1/leverage?{full_query}"
        r = requests.post(url, headers=headers, timeout=10)
        if r.status_code == 200:
            print(f"{get_timestamp()} ✅ 선물 레버리지 설정: {symbol} {leverage}배")
            return True
        print(f"{get_timestamp()} ⚠️ 선물 레버리지 설정 응답: {r.status_code} {r.text}")
        return False
    except Exception as e:
        print(f"{get_timestamp()} ❌ 선물 레버리지 설정 실패: {e}")
        return False

def close_current_position(symbol: str) -> bool:
    """선물 포지션 전량 시장가 청산 (포지션 조회 후 반대 방향 reduceOnly 시장가)"""
    try:
        positions = get_futures_position_risk(symbol)
        for pos in positions:
            amt = float(pos.get("positionAmt", 0))
            if amt == 0:
                continue
            side = "SELL" if amt > 0 else "BUY"
            qty = abs(amt)
            res = binance_fapi_order(symbol, side, qty, order_type="MARKET", reduce_only=True)
            if res:
                print(f"{get_timestamp()} ✅ 선물 포지션 청산: {side} {qty} @ {symbol}")
            else:
                print(f"{get_timestamp()} ⚠️ 선물 포지션 청산 주문 실패")
            return bool(res)
        print(f"{get_timestamp()} ℹ️ 청산할 선물 포지션 없음: {symbol}")
        return True
    except Exception as e:
        print(f"{get_timestamp()} ❌ 선물 포지션 청산 실패: {e}")
        return False

def get_futures_position_params(close_price: float) -> tuple:
    """종가(close) 기준 1만 달러 단위 계단식으로 진입·익절 금액 계산. (total_usdt, tp_per_stage_usdt) 반환."""
    if close_price <= 0:
        return (FUTURES_BASE_TOTAL_USDT, FUTURES_BASE_TP_USDT)
    band = max(0, math.floor((close_price - FUTURES_BASE_PRICE) / 10_000))
    total_usdt = FUTURES_BASE_TOTAL_USDT + band * FUTURES_STEP_TOTAL_PER_10K
    tp_usdt = FUTURES_BASE_TP_USDT + band * FUTURES_STEP_TP_PER_10K
    return (total_usdt, tp_usdt)

def binance_fapi_order(symbol: str, side: str, quantity: float, price: Optional[float] = None, order_type: str = "MARKET", reduce_only: bool = False) -> Optional[dict]:
    """선물 주문 (fapi/v1/order). 서명 오류(-1022) 방지를 위해 파라미터를 URL 쿼리 스트링으로 전송. 수량/가격은 stepSize·tickSize 기준 포맷."""
    try:
        from urllib.parse import urlencode
        qty_str = _format_qty_for_futures(symbol, quantity)
        params = {
            "symbol": symbol,
            "side": side,
            "type": order_type,
            "quantity": qty_str,
        }
        if order_type == "LIMIT":
            params["price"] = _format_price_for_futures(symbol, price)
            params["timeInForce"] = "GTC"
        if reduce_only:
            params["reduceOnly"] = "true"
        # 1. 파라미터 정렬 및 쿼리 스트링 생성
        sorted_pairs = sorted(params.items())
        query_string = urlencode(sorted_pairs)
        # 2. 헤더 및 서명 생성 (timestamp, recvWindow 포함됨)
        headers, signature, timestamp, recv_window = _binance_fapi_headers(query_string)
        # 3. POST 요청 시에도 파라미터를 URL에 포함 (서명 검증과 동일한 문자열로 전달)
        full_query = f"{query_string}&timestamp={timestamp}&recvWindow={recv_window}&signature={signature}"
        url = f"{BINANCE_FUTURES_BASE}/fapi/v1/order?{full_query}"
        r = requests.post(url, headers=headers, timeout=10)
        if r.status_code == 200:
            return r.json()
        print(f"{get_timestamp()} ❌ 선물 주문 실패: {r.status_code} {r.text}")
        return None
    except Exception as e:
        print(f"{get_timestamp()} ❌ 선물 주문 예외: {e}")
        return None

def binance_fapi_stop_market(symbol: str, side: str, quantity: float, stop_price: float, reduce_only: bool = True) -> Optional[dict]:
    """
    선물 스탑마켓 주문 (STOP_MARKET) – SL용.
    ⚠️ Binance 변경 사항(2025-12 이후): STOP_MARKET/TAKE_PROFIT_MARKET 등 조건부 주문은
    기존 /fapi/v1/order 가 아닌 새 Algo Order 엔드포인트(/fapi/v1/algoOrder)를 사용해야 함.
    - algoType=CONDITIONAL
    - type=STOP_MARKET
    - triggerPrice=stop_price
    - reduceOnly=true (부분 청산용)
    """
    try:
        from urllib.parse import urlencode

        qty_str = _format_qty_for_futures(symbol, quantity)
        trigger_str = _format_price_for_futures(symbol, stop_price)

        # Algo Order API 파라미터 (USDT-M Futures)
        params = {
            "algoType": "CONDITIONAL",
            "symbol": symbol,
            "side": side,
            "type": "STOP_MARKET",
            "quantity": qty_str,
            "triggerPrice": trigger_str,
            "workingType": "CONTRACT_PRICE",  # 마크가격 대신 선물이름(계약가) 기준
            "reduceOnly": "true" if reduce_only else "false",
        }

        sorted_pairs = sorted(params.items())
        query_string = urlencode(sorted_pairs)
        headers, signature, timestamp, recv_window = _binance_fapi_headers(query_string)
        full_query = f"{query_string}&timestamp={timestamp}&recvWindow={recv_window}&signature={signature}"
        url = f"{BINANCE_FUTURES_BASE}/fapi/v1/algoOrder?{full_query}"
        r = requests.post(url, headers=headers, timeout=10)
        if r.status_code == 200:
            return r.json()
        print(f"{get_timestamp()} ❌ 선물 스탑마켓(SL) 주문 실패(Algo): {r.status_code} {r.text}")
        return None
    except Exception as e:
        print(f"{get_timestamp()} ❌ 선물 스탑마켓(SL) 주문 예외(Algo): {e}")
        return None

def execute_futures_strategy(ls_signal: int, symbol: str = None, stage_prefix: str = "", K: float = None):
    """LS 시그널(1 또는 -1)에 따른 선물 진입 및 3단계 분할 익절. K=LS 판정된 종가 → 스마트 주문 엔진으로 주문가 결정."""
    if symbol is None:
        symbol = f"{TICKER}USDT"
    sym_upper = symbol.replace("USDT", "").upper()
    if sym_upper not in ROTATION_TICKERS:
        print(f"{get_timestamp()} [{stage_prefix}] 🔒 선물 잠금: {sym_upper} (선물 주문은 {ROTATION_TICKERS} 만 가능)")
        return
    if ls_signal not in (1, -1):
        print(f"{get_timestamp()} [{stage_prefix}] ⚠️ execute_futures_strategy: ls_signal은 1 또는 -1이어야 함 (현재: {ls_signal})")
        return
    try:
        current_price = float(binance_fapi_ticker_price(symbol))
        if current_price <= 0:
            print(f"{get_timestamp()} [{stage_prefix}] ❌ 선물 현재가 조회 실패")
            return
        # K = LS 판정된 종가 (미전달 시 현재가 사용)
        ref_K = float(K) if K is not None and K > 0 else current_price
        # 동적 규칙: exchangeInfo 에서 LOT_SIZE, MIN_NOTIONAL 로드
        info = get_futures_exchange_info(symbol, use_cache=False)
        min_qty = max(info["minQty"], FUTURES_MIN_QTY_BTC)
        step_size = info["stepSize"]
        notional = info["notional"]
        price_prec = int(info.get("pricePrecision", 2))
        position_usdt, tp_stage_usdt = get_futures_position_params(ref_K)
        position_usdt = round(float(position_usdt), 2)
        tp_stage_usdt = round(float(tp_stage_usdt), 2)
        # BNB: 진입 100 USDT, 4분할(25%×3 TP + 25% 추세전환까지 유지). 그 외: 725 USDT, 25%×3 TP + 25% 유지
        if symbol == "BNBUSDT":
            position_usdt = float(FUTURES_BNB_TOTAL_USDT)
            tp_stage_usdt = round(float(FUTURES_BNB_TP_PART_USDT), 2)
        else:
            position_usdt = float(FUTURES_POSITION_USDT)
            tp_stage_usdt = round(float(FUTURES_TP_PART_USDT), 2)
        if position_usdt < notional:
            print(f"{get_timestamp()} [{stage_prefix}] ❌ 진입 금액 부족: {position_usdt} USDT < 거래소 최소 주문금액 {notional} USDT")
            return
        cancel_all_futures_orders(symbol)
        close_current_position(symbol)
        # 롱/숏 전부 1배만 사용
        set_futures_leverage(symbol, leverage=1)
        # 스마트 주문 엔진: 선물 호가 + K(종가) → 진입 주문가 결정
        ask, bid, ask_q, bid_q = get_futures_orderbook_snapshot(symbol)
        if ask is None or bid is None:
            print(f"{get_timestamp()} [{stage_prefix}] ⚠️ 선물 호가 실패, 현재가로 진입가 사용")
            entry_price = round(current_price, price_prec)
            smart_log_lines = []
        else:
            is_buy = ls_signal == 1
            entry_price, smart_log_lines = execute_smart_order(is_buy=is_buy, K=ref_K, ask=ask, bid=bid, ask_q=ask_q, bid_q=bid_q, symbol=symbol)
            if entry_price is None:
                entry_price = round(current_price, price_prec)
                smart_log_lines = ["⚠️ 스마트 주문 실패 → 현재가 사용"]
            else:
                entry_price = round(entry_price, price_prec)
            for line in smart_log_lines:
                print(f"{get_timestamp()} [{stage_prefix}] 📐 선물 스마트주문: {line}")
        # 선물 틱 사이즈에 맞춤 (-4014 Price not increased by tick size 방지)
        entry_price = adjust_price_to_tick_futures(symbol, entry_price)
        total_qty = _round_down_to_step(position_usdt / entry_price, step_size)
        if total_qty < min_qty:
            print(f"{get_timestamp()} [{stage_prefix}] ❌ 잔고 또는 금액 부족으로 주문 불가 (필요 수량 {total_qty} < 최소 {min_qty})")
            return
        notional_min = float(info.get("notional", 5))
        # 100% → TP 분할. TP 1.5% 3분할 (0.5% / 1.0% / 1.5%), 롱/숏 모두
        tp_total_qty = _round_down_to_step(total_qty * 1.0, step_size)
        held_qty = 0
        min_qty_per_tp = max(min_qty, _round_up_to_step(notional_min / entry_price, step_size))
        n_tps = int(tp_total_qty / min_qty_per_tp) if min_qty_per_tp > 0 else 0
        n_max_tps, step_pct, tp_range_str = 3, 0.015 / 3, "1.5% 3분할"   # step_pct=0.5% → 0.5%, 1.0%, 1.5%
        if n_tps > n_max_tps:
            n_tps = n_max_tps
        side = "BUY" if ls_signal == 1 else "SELL"
        print(f"{get_timestamp()} [{stage_prefix}] 🚀 신규 진입: {side} {total_qty} @ {entry_price} USDT (K={ref_K:.2f}, 진입 {position_usdt:.2f} USDT, 100% TP {tp_range_str} {n_tps}개)")
        entry_res = binance_fapi_order(symbol, side, total_qty, price=entry_price, order_type="LIMIT")
        if not entry_res:
            return
        # ReduceOnly TP/SL은 포지션 존재 시에만 가능 → 진입 체결 대기
        wait_sec = 60
        poll_interval = 2
        position_seen = False
        for _ in range(max(1, wait_sec // poll_interval)):
            time.sleep(poll_interval)
            positions = get_futures_position_risk(symbol)
            for pos in positions:
                amt = float(pos.get("positionAmt", 0) or 0)
                if amt == 0:
                    continue
                if (side == "BUY" and amt > 0) or (side == "SELL" and amt < 0):
                    position_seen = True
                    break
            if position_seen:
                break
        if not position_seen:
            print(f"{get_timestamp()} [{stage_prefix}] ⚠️ 진입 체결 대기 {wait_sec}초 초과, TP/SL 건너뜀 (수동 예약 필요)")
            return
        tp_side = "SELL" if side == "BUY" else "BUY"
        if n_tps >= 1:
            if n_tps == 1:
                tp_unit_qty = tp_total_qty
            else:
                tp_unit_qty = _round_down_to_step((tp_total_qty - min_qty_per_tp) / (n_tps - 1), step_size)
                if tp_unit_qty < min_qty_per_tp:
                    tp_unit_qty = min_qty_per_tp
            for i in range(n_tps):
                pct = step_pct * (i + 1)
                tp_price = entry_price * (1.0 + pct if side == "BUY" else 1.0 - pct)
                tp_price = adjust_price_to_tick_futures(symbol, tp_price)
                qty_this = (tp_total_qty - tp_unit_qty * (n_tps - 1)) if i == n_tps - 1 else tp_unit_qty
                if qty_this >= min_qty and qty_this * tp_price >= notional_min:
                    binance_fapi_order(symbol, tp_side, qty_this, price=tp_price, order_type="LIMIT", reduce_only=True)
        else:
            print(f"{get_timestamp()} [{stage_prefix}] ⚠️ TP 0개 (min notional {notional_min} USDT·minQty {min_qty} 충족 불가), 전량 유지")
        # SL 1.2% 3분할, 롱/숏 모두. 숏=진입가 상승 1.2% 트리거, 롱=진입가 하락 1.2% 트리거
        sl_res = None
        sl_price = entry_price * (1.0 + FUTURES_SL_PERCENT) if side == "SELL" else entry_price * (1.0 - FUTURES_SL_PERCENT)
        sl_price = adjust_price_to_tick_futures(symbol, sl_price)
        min_sl_qty = max(min_qty, _round_up_to_step(notional_min / sl_price, step_size))
        n_sls = 3
        if total_qty >= n_sls * min_sl_qty:
            sl_unit_qty = _round_down_to_step((total_qty - min_sl_qty) / (n_sls - 1), step_size)
            if sl_unit_qty < min_sl_qty:
                sl_unit_qty = min_sl_qty
            for i in range(n_sls):
                sl_qty = (total_qty - sl_unit_qty * (n_sls - 1)) if i == n_sls - 1 else sl_unit_qty
                if sl_qty >= min_sl_qty:
                    sl_res = binance_fapi_stop_market(symbol, tp_side, sl_qty, sl_price, reduce_only=True)
                    if sl_res:
                        print(f"{get_timestamp()} [{stage_prefix}] 🛑 손절(SL {FUTURES_SL_PERCENT*100:.1f}% 3분할 {i+1}/{n_sls}) 예약: {tp_side} {sl_qty} @ 트리거 {sl_price}")
        else:
            sl_res = binance_fapi_stop_market(symbol, tp_side, total_qty, sl_price, reduce_only=True)
            if sl_res:
                print(f"{get_timestamp()} [{stage_prefix}] 🛑 손절(SL {FUTURES_SL_PERCENT*100:.1f}%) 예약: {tp_side} {total_qty} @ 트리거 {sl_price}")
        tp_desc = f"TP {tp_range_str} {n_max_tps}분할"
        sl_desc = f"SL {FUTURES_SL_PERCENT*100:.1f}% 3분할"
        print(f"{get_timestamp()} [{stage_prefix}] ✅ {tp_desc} + {sl_desc} 예약 완료")
    except Exception as e:
        print(f"{get_timestamp()} [{stage_prefix}] ❌ 전략 실행 중 오류: {e}")

def check_and_move_sl_to_be(symbol: str, stage_prefix: str = ""):
    """BE(Break-Even) 이동: 1차 목표가(1/3 물량 익절) 도달 시, 남은 물량의 SL을 진입가(0%)가 아닌
    [진입가 + 왕복 수수료] 지점으로 끌어올려, 수익→손실 전환을 막고 기하평균 수익률을 방어.
    - 변경 전: 익절 +1.5% / 손절 -1.2%
    - 변경 후(1차 익절 시): 익절 +1.5% / 손절 0% (실제로는 진입가 ±0.06% = 수수료 상쇄)
    바이낸스 선물 BNB 할인 기준: 롱=진입가+0.06%에서 매도, 숏=진입가-0.06%에서 매수 트리거."""
    try:
        positions = get_futures_position_risk(symbol)
        entry_price, position_amt, side = None, 0.0, None
        for pos in positions:
            amt = float(pos.get("positionAmt", 0) or 0)
            if amt == 0:
                continue
            entry_price = float(pos.get("entryPrice", 0) or 0)
            position_amt = abs(amt)
            side = "BUY" if amt > 0 else "SELL"
            break
        if entry_price is None or entry_price <= 0 or position_amt <= 0:
            return
        open_orders = get_futures_open_orders(symbol)
        tp_orders = [o for o in open_orders if (o.get("type") or "").upper() == "LIMIT" and (o.get("reduceOnly") in (True, "true", "TRUE") or str(o.get("reduceOnly", "")).lower() == "true")]
        # SL은 Algo Order API로만 등록되므로 openAlgoOrders에서 조회 (triggerPrice 사용)
        open_algo = get_futures_open_algo_orders(symbol)
        sl_orders = [o for o in open_algo if (o.get("orderType") or "").upper() == "STOP_MARKET" and (o.get("reduceOnly") in (True, "true", "TRUE") or str(o.get("reduceOnly", "")).lower() == "true")]
        if len(sl_orders) == 0:
            return
        if len(tp_orders) >= 3:
            return
        first_sl_stop = float(sl_orders[0].get("triggerPrice", 0) or 0)
        if first_sl_stop <= 0:
            return
        # 실제 본절가 = 진입가 + 왕복 수수료 (0.06%). 롱=아래로 떨어질 때 진입+0.06%에서 매도, 숏=올라갈 때 진입-0.06%에서 매수
        be_long = entry_price * (1.0 + FUTURES_BE_OFFSET_PERCENT)
        be_short = entry_price * (1.0 - FUTURES_BE_OFFSET_PERCENT)
        tol_pct = 0.0005
        if side == "BUY":
            already_be = abs(first_sl_stop - be_long) / entry_price <= tol_pct
            original_sl = entry_price * (1.0 - FUTURES_SL_PERCENT)
            is_original = abs(first_sl_stop - original_sl) / entry_price <= tol_pct
        else:
            already_be = abs(first_sl_stop - be_short) / entry_price <= tol_pct
            original_sl = entry_price * (1.0 + FUTURES_SL_PERCENT)
            is_original = abs(first_sl_stop - original_sl) / entry_price <= tol_pct
        if already_be or not is_original:
            return
        for o in sl_orders:
            aid = o.get("algoId")
            if aid is not None:
                cancel_futures_algo_order(int(aid))
        info = get_futures_exchange_info(symbol, use_cache=False)
        step_size = info["stepSize"]
        min_qty = info["minQty"]
        notional_min = float(info.get("notional", 5))
        be_price = be_long if side == "BUY" else be_short
        be_price = adjust_price_to_tick_futures(symbol, be_price)
        min_sl_qty = max(min_qty, _round_up_to_step(notional_min / be_price, step_size))
        n_sls = 3
        tp_side = "SELL" if side == "BUY" else "BUY"
        disp_symbol = symbol.replace("USDT", "")
        if position_amt >= n_sls * min_sl_qty:
            sl_unit_qty = _round_down_to_step((position_amt - min_sl_qty) / (n_sls - 1), step_size)
            if sl_unit_qty < min_sl_qty:
                sl_unit_qty = min_sl_qty
            for i in range(n_sls):
                sl_qty = (position_amt - sl_unit_qty * (n_sls - 1)) if i == n_sls - 1 else sl_unit_qty
                if sl_qty >= min_sl_qty:
                    binance_fapi_stop_market(symbol, tp_side, sl_qty, be_price, reduce_only=True)
            print(f"{get_timestamp()} [{stage_prefix}] 📌 (BEP) 이동: {disp_symbol} ({be_price:.4f})")
        else:
            binance_fapi_stop_market(symbol, tp_side, position_amt, be_price, reduce_only=True)
            print(f"{get_timestamp()} [{stage_prefix}] 📌 (BEP) 이동: {disp_symbol} ({be_price:.4f})")
    except Exception as e:
        print(f"{get_timestamp()} [{stage_prefix}] ⚠️ BE 이동 확인 중 오류: {e}")

def binance_orderbook_bid(symbol: str) -> float:
    """바이낸스 매수 1호가(bid) 조회 - 즉시 체결 가능한 가격"""
    try:
        r = requests.get(f"{BINANCE_API_BASE}/api/v3/depth", params={"symbol": symbol, "limit": 5}, timeout=10)
        r.raise_for_status()
        data = r.json()
        if data and "bids" in data and len(data["bids"]) > 0:
            # 매수 1호가는 bids의 첫 번째 가격
            bid_price = float(data["bids"][0][0])
            return bid_price
        return 0.0
    except Exception as e:
        print(f"{get_timestamp()} ❌ 매수 1호가(bid) 조회 실패: {e}")
        return 0.0

def get_binance_orderbook_snapshot(symbol: str):
    """
    바이낸스 호가창에서 매도1호가(ask), 매수1호가(bid), 매도잔량(ask_q), 매수잔량(bid_q)을 한 번에 조회
    
    Returns:
        tuple: (ask, bid, ask_q, bid_q) 또는 실패 시 (None, None, None, None)
    """
    try:
        r = requests.get(f"{BINANCE_API_BASE}/api/v3/depth", params={"symbol": symbol, "limit": 20}, timeout=10)
        r.raise_for_status()
        data = r.json()
        if data and "asks" in data and "bids" in data:
            if len(data["asks"]) > 0 and len(data["bids"]) > 0:
                # 매도 1호가 (ask)
                ask = float(data["asks"][0][0])
                ask_q = float(data["asks"][0][1])
                # 매수 1호가 (bid)
                bid = float(data["bids"][0][0])
                bid_q = float(data["bids"][0][1])
                return ask, bid, ask_q, bid_q
        return None, None, None, None
    except Exception as e:
        print(f"{get_timestamp()} ❌호가창 조회 실패: {e}")
        return None, None, None, None

def _precision_from_tick_or_step(value) -> int:
    """
    틱/스텝 크기에서 소수 자릿수 계산. 과학적 표기(1e-5 등) 포함 처리.
    스마트주문가격·수량 반올림 시 티커별 소수점 오류 방지용.
    """
    if value is None or value <= 0:
        return 8
    s = str(value).strip().lower()
    if 'e' in s:
        parts = s.split('e')
        if len(parts) == 2:
            try:
                exp = int(parts[1])
                return max(0, -exp)
            except ValueError:
                pass
        return 8
    if '.' in s:
        return len(s.split('.')[-1].rstrip('0'))
    return 0


def get_binance_tick_size(symbol: str, price: float) -> float:
    """
    바이낸스 심볼의 틱 사이즈 조회 (exchangeInfo에서 priceFilter의 tickSize 사용)
    
    Args:
        symbol: 심볼 (예: "BTCUSDT")
        price: 가격 (참고용, 실제로는 심볼 정보에서 틱 사이즈 조회)
    
    Returns:
        틱 사이즈 (float)
    """
    try:
        info = binance_get_symbol_info(symbol)
        filters = info.get('filters', [])
        for f in filters:
            if f.get('filterType') == 'PRICE_FILTER':
                raw = f.get('tickSize', '0.01')
                tick_size = float(raw) if not isinstance(raw, (int, float)) else raw
                return tick_size
        # 기본값: 가격에 따라 추정
        if price >= 1000: return 0.01
        elif price >= 100: return 0.001
        elif price >= 10: return 0.0001
        elif price >= 1: return 0.00001
        else: return 0.000001
    except Exception as e:
        # 기본값 반환
        if price >= 1000: return 0.01
        elif price >= 100: return 0.001
        elif price >= 10: return 0.0001
        elif price >= 1: return 0.00001
        else: return 0.000001

def adjust_price_to_tick_binance(symbol: str, price: float) -> float:
    """
    바이낸스 틱 사이즈에 맞춰 가격 조정
    
    Args:
        symbol: 심볼 (예: "BTCUSDT")
        price: 가격
    
    Returns:
        조정된 가격
    """
    tick_size = get_binance_tick_size(symbol, price)
    if tick_size <= 0:
        return price
    # 틱 사이즈에 맞춰 반올림 (티커별 소수점 정밀도: 1e-5 등 과학적 표기 포함)
    precision = _precision_from_tick_or_step(tick_size)
    adjusted = round(round(price / tick_size) * tick_size, precision)
    return adjusted

def round_price_to_tick_size_binance(symbol: str, price: float) -> float:
    """
    바이낸스 틱 사이즈에 맞춰 가격 반올림 (execute_smart_order에서 사용)
    
    Args:
        symbol: 심볼 (예: "BTCUSDT")
        price: 가격
    
    Returns:
        반올림된 가격
    """
    return adjust_price_to_tick_binance(symbol, price)

def ceil_price_to_tick_size_binance(symbol: str, price: float) -> float:
    """
    바이낸스 틱 사이즈에 맞춰 가격 올림 처리 (최소매도가 계산용)
    
    Args:
        symbol: 심볼 (예: "BTCUSDT")
        price: 가격
    
    Returns:
        올림 처리된 가격 (반올림이 아닌 올림)
    """
    tick_size = get_binance_tick_size(symbol, price)
    if tick_size <= 0:
        return price
    # 틱 사이즈로 나눈 후 올림하고 다시 곱함 (반올림이 아닌 올림)
    adjusted = math.ceil(price / tick_size) * tick_size
    # 부동소수점 오차 방지를 위해 정밀도 조정 (티커별: 1e-5 등 과학적 표기 포함)
    precision = _precision_from_tick_or_step(tick_size)
    return round(adjusted, precision)

def binance_orderbook_cumulative(symbol: str, target_amount: float) -> tuple:
    """
    바이낸스 호가창에서 목표 금액(USDT)을 채우기 위한 누적 금액 확인 후,
    매수 1호가로 정확한 수량을 계산합니다.
    
    Returns:
        tuple: (계산된 수량, 예상 총액, 사용된 호가 수, 매수1호가)
    """
    try:
        r = requests.get(f"{BINANCE_API_BASE}/api/v3/depth", params={"symbol": symbol, "limit": 20}, timeout=10)
        r.raise_for_status()
        data = r.json()
        
        if not data or "bids" not in data or len(data["bids"]) == 0:
            return 0.0, 0.0, 0, 0.0
        
        orderbook = data["bids"]
        
        # 매수 1호가 가격 (가장 높은 매수 호가)
        bid_price_1st = float(orderbook[0][0])
        
        cumulative_amount = 0.0
        used_levels = 0
        
        # 매수 1호가부터 아래로 누적하여 목표 금액을 넘는지 확인
        for level in orderbook:
            bid_price = float(level[0])
            bid_size = float(level[1])
            
            # 이 호가에서 가능한 최대 거래량
            level_amount = bid_size * bid_price
            
            # 누적 계산
            cumulative_amount += level_amount
            used_levels += 1
            
            # 목표 금액을 넘었는지 확인
            if cumulative_amount >= target_amount:
                break
        
        # 누적 금액이 목표 금액을 넘었는지 확인
        if cumulative_amount < target_amount:
            print(f"{get_timestamp()} ⚠️ 호가창 누적 금액 부족: {cumulative_amount:.2f} USDT < {target_amount:.2f} USDT")
            return 0.0, 0.0, 0, bid_price_1st
        
        # 매수 1호가로 정확한 수량 계산
        target_volume = target_amount / bid_price_1st
        
        # 바이낸스 최소 단위는 0.00000001까지 가능
        volume = float(f"{target_volume:.8f}")
        
        # 최종 예상 금액 계산 (매수 1호가 기준)
        final_amount = volume * bid_price_1st
        
        # 안전 마진: 소수점 반올림과 호가 변동을 고려하여 0.1% 여유분 추가
        safety_margin = 1.001
        final_volume = volume * safety_margin
        final_amount_with_margin = final_volume * bid_price_1st
        
        return final_volume, final_amount_with_margin, used_levels, bid_price_1st
        
    except Exception as e:
        print(f"{get_timestamp()} ❌ 호가창 누적 계산 실패: {e}")
        return 0.0, 0.0, 0, 0.0

def binance_get_order_status(symbol: str, order_id: int) -> dict:
    """바이낸스 주문 상태 조회 및 체결 정보 반환
    
    Args:
        symbol: 거래 심볼 (예: "BTCUSDT")
        order_id: 주문 ID
    
    Returns:
        dict: 주문 상태 정보 (executed_price, executed_volume, remaining_volume, state, order_data)
    
    Note:
        - 모든 시간 처리는 UTC 기준으로 수행
        - 로깅 시간은 get_timestamp() 함수 사용 (UTC 표시 포함)
    """
    try:
        query_string = f"symbol={symbol}&orderId={order_id}"
        headers, signature, timestamp, recv_window = _binance_headers(query_string)
        
        r = requests.get(
            f"{BINANCE_API_BASE}/api/v3/order",
            params={"symbol": symbol, "orderId": order_id, "timestamp": timestamp, "recvWindow": recv_window, "signature": signature},
            headers=headers,
            timeout=10
        )
        
        if r.status_code == 200:
            order_data = r.json()
            
            # 체결 정보 추출
            executed_price = float(order_data.get('price', 0))  # 체결 가격
            executed_qty = float(order_data.get('executedQty', 0))  # 체결 수량
            orig_qty = float(order_data.get('origQty', 0))  # 원래 주문 수량
            status = order_data.get('status', '')  # 주문 상태 (NEW, PARTIALLY_FILLED, FILLED, CANCELED, etc.)
            
            return {
                'executed_price': executed_price,
                'executed_volume': executed_qty,
                'remaining_volume': orig_qty - executed_qty,
                'state': status,
                'order_data': order_data
            }
        else:
            print(f"{get_timestamp()} ❌ 주문 상태 조회 실패: {r.status_code} {r.text}")
            return None
            
    except Exception as e:
        print(f"{get_timestamp()} ❌ 주문 상태 조회 중 오류: {e}")
        return None

def binance_get_account_balance(asset: str = None, subtract_bnb_fee: bool = True) -> dict:
    """바이낸스 계좌 잔고 조회 (특정 자산)
    
    Args:
        asset: 조회할 자산 (기본값: TICKER)
        subtract_bnb_fee: BNB에서 10 USDT 상당을 차감할지 여부 (기본값: True, 자산기록 시 False)
    """
    if asset is None:
        asset = TICKER
    try:
        # 빈 query_string으로 시작 (account 조회는 추가 파라미터 없음)
        query_string = ""
        headers, signature, timestamp, recv_window = _binance_headers(query_string)
        
        r = requests.get(
            f"{BINANCE_API_BASE}/api/v3/account",
            params={"timestamp": timestamp, "recvWindow": recv_window, "signature": signature},
            headers=headers,
            timeout=10
        )
        
        if r.status_code == 200:
            account_data = r.json()
            balances = account_data.get('balances', [])
            
            # 특정 자산 잔고 찾기
            for balance_info in balances:
                if balance_info.get('asset') == asset:
                    free = float(balance_info.get('free', 0))  # 사용 가능한 수량
                    locked = float(balance_info.get('locked', 0))  # 주문 중인 수량
                    total = free + locked
                    
                    # BNB의 경우: Available에서 10 USDT 상당의 BNB를 빼서 보유량으로 계산 (subtract_bnb_fee가 True일 때만)
                    if asset == "BNB" and subtract_bnb_fee:
                        try:
                            bnb_price = binance_ticker_price("BNBUSDT")
                            bnb_amount_to_subtract = 10.0 / bnb_price if bnb_price > 0 else 0  # 10 USDT 상당의 BNB 수량
                            # available에서 10 USDT 상당을 빼되, 0보다 작아지지 않도록 처리
                            adjusted_free = max(0.0, free - bnb_amount_to_subtract)
                            adjusted_total = adjusted_free + locked
                        except Exception as e:
                            print(f"{get_timestamp()} ⚠️ BNB 가격 조회 실패, 원본 잔고 사용: {e}")
                            adjusted_free = free
                            adjusted_total = total
                    else:
                        adjusted_free = free
                        adjusted_total = total
                    
                    return {
                        'currency': asset,
                        'balance': adjusted_total,
                        'locked': locked,
                        'avg_buy_price': 0.0,  # 바이낸스 API는 평균 매수가 제공 안함
                        'available': adjusted_free,  # 사용 가능한 수량 (BNB는 subtract_bnb_fee=True일 때만 10 USDT 상당 차감)
                        'free_precise': adjusted_free,  # 정밀한 사용 가능 수량 (BNB는 subtract_bnb_fee=True일 때만 10 USDT 상당 차감)
                        'total_precise': adjusted_total,  # 정밀한 총 수량 (BNB는 subtract_bnb_fee=True일 때만 10 USDT 상당 차감)
                        'free_raw': balance_info.get('free', '0'),  # 원시 문자열 값 (정밀도 보존)
                        'total_raw': str(float(balance_info.get('free', 0)) + float(balance_info.get('locked', 0)))  # 원시 문자열 총 수량
                    }
            
            # 해당 자산이 없으면 0으로 반환
            return {
                'currency': asset,
                'balance': 0.0,
                'locked': 0.0,
                'avg_buy_price': 0.0,
                'available': 0.0,
                'free_precise': 0.0,
                'total_precise': 0.0
            }
        else:
            print(f"{get_timestamp()} ❌ 계좌 잔고 조회 실패: {r.status_code} {r.text}")
            return None
            
    except Exception as e:
        print(f"{get_timestamp()} ❌ 계좌 잔고 조회 중 오류: {e}")
        return None

def binance_market_buy(symbol: str, usdt_amount: float, stage_prefix: str = ""):
    """바이낸스 시장가 매수: quoteOrderQty (USDT 금액 기준)
    
    Args:
        symbol: 거래 심볼 (예: "BTCUSDT")
        usdt_amount: 매수할 USDT 금액
    
    Returns:
        dict: 주문 결과 (orderId 포함)
    
    Note:
        - 모든 시간 처리는 UTC 기준으로 수행
        - API 타임스탬프는 UTC 기준 밀리초 타임스탬프 사용
        - 로깅 시간은 get_timestamp() 함수 사용 (UTC 표시 포함)
    """
    try:
        # 심볼 precision을 exchangeInfo에서 조회하여 quoteOrderQty 정밀도 결정
        # 미리 초기화된 정밀도 사용(없으면 1회 조회 후 캐시)
        usdt_precision = QUOTE_PRECISION_MAP.get(symbol)
        if usdt_precision is None:
            info = binance_get_symbol_info(symbol)
            usdt_precision = info.get('quotePrecision') if 'quotePrecision' in info else info.get('quoteAssetPrecision', 5)
            try:
                usdt_precision = int(usdt_precision)
            except:
                usdt_precision = 5
            QUOTE_PRECISION_MAP[symbol] = usdt_precision
        # Decimal로 정밀 반올림 및 문자열 보전
        from decimal import Decimal, ROUND_DOWN, getcontext
        getcontext().prec = 28
        quant = Decimal('1') if usdt_precision == 0 else Decimal('1.' + ('0'*usdt_precision))
        usdt_amount_dec = Decimal(str(usdt_amount))
        usdt_amount_rounded_dec = usdt_amount_dec.quantize(quant, rounding=ROUND_DOWN)
        usdt_amount_rounded_str = format(usdt_amount_rounded_dec, 'f')
        print(f"{get_timestamp()} [{stage_prefix}] 🔍 시장가 매수 주문 전송 중... 목표: {usdt_amount_rounded_str} USDT (quotePrecision {usdt_precision})")
        
        # 드라이런 테스트 제거 - 바로 실제 주문 실행
        print(f"{get_timestamp()} [{stage_prefix}] 🚀 실제 주문 실행")
        
        # quoteOrderQty를 사용하여 USDT 금액 기준으로 매수
        query_string = f"symbol={symbol}&side=BUY&type=MARKET&quoteOrderQty={usdt_amount_rounded_str}"
        headers, signature, timestamp, recv_window = _binance_headers(query_string)
        
        params = {
            "symbol": symbol,
            "side": "BUY",
            "type": "MARKET",
            "quoteOrderQty": usdt_amount_rounded_str,
            "timestamp": timestamp,
            "recvWindow": recv_window,
            "signature": signature
        }
        
        r = requests.post(f"{BINANCE_API_BASE}/api/v3/order", params=params, headers=headers, timeout=10)
        
        if r.status_code != 200:
            error_data = r.json() if r.text else {}
            error_code = error_data.get('code', 'UNKNOWN')
            error_msg = error_data.get('msg', r.text)
            print(f"{get_timestamp()} [{stage_prefix}] ❌ 매수 주문 실패: {r.status_code} - {error_code}: {error_msg}")
            
            # 바이낸스 특정 에러 코드 처리
            if error_code == -1022:
                print(f"{get_timestamp()} [{stage_prefix}] 🔍 서명 오류: API 키 또는 서명이 잘못됨")
            elif error_code == -1021:
                print(f"{get_timestamp()} [{stage_prefix}] 🔍 시간 오류: 타임스탬프가 서버 시간과 맞지 않음")
            elif error_code == -2010:
                print(f"{get_timestamp()} [{stage_prefix}] 🔍 잔고 부족: 계좌 잔고가 부족함")
            elif error_code == -1013:
                print(f"{get_timestamp()} [{stage_prefix}] 🔍 수량 오류: 주문 수량이 최소/최대 범위를 벗어남")
            
            raise RuntimeError(f"매수 실패: {error_code} - {error_msg}")
        
        result = r.json()
        order_id = result.get('orderId')
        
        print(f"{get_timestamp()} [{stage_prefix}] ✅ 매수 주문 접수 성공: OrderID={order_id}")
        
        # 주문 체결 상태 확인 및 실제 체결 가격 조회
        if order_id:
            print(f"{get_timestamp()} [{stage_prefix}] 🔍 매수 주문 체결 확인 중... OrderID: {order_id}")
            
            # 잠시 대기 후 체결 상태 확인
            time.sleep(2)
            
            # 주문 상태 조회
            order_status = binance_get_order_status(symbol, order_id)
            if order_status:
                executed_qty = order_status.get('executed_volume', 0)
                
                if executed_qty > 0:
                    ticker_balance = binance_get_account_balance(TICKER)
                    current_price = binance_ticker_price(symbol)
                    print(f"{get_timestamp()} [{stage_prefix}] 🎯 매수 체결 완료: {executed_qty:.8f} {TICKER} @ {current_price:.2f} USDT | 잔고: {ticker_balance['total_precise']:.8f} {TICKER}")
                else:
                    print(f"{get_timestamp()} [{stage_prefix}] ⚠️ 매수 주문 체결 정보를 가져올 수 없습니다")
            else:
                print(f"{get_timestamp()} [{stage_prefix}] ⚠️ 주문 상태 확인 실패")
        
        print(f"{get_timestamp()} [{stage_prefix}] 🎉 매수 주문 성공: {usdt_amount_rounded_str} USDT")
        return result
    except Exception as e:
        print(f"{get_timestamp()} [{stage_prefix}] ❌ 매수 주문 중 오류 발생: {e}")
        raise

def binance_market_sell(symbol: str, usdt_amount: float, price_hint: float | None = None, use_safety_margin: bool = True, exact_volume: float = None, stage_prefix: str = "", decision_price: float | None = None, min_sell_price: float | None = None):
    """바이낸스 지정가 매도: 매수 1호가 가격으로 수량 계산하여 매도
    
    Args:
        symbol: 거래 심볼 (예: "BTCUSDT")
        usdt_amount: 매도할 USDT 금액
        price_hint: prev_tp (TP = 목표가격) - 예상수익률 계산에 사용 (평균단가 역산용)
        use_safety_margin: 안전 마진 적용 여부
        exact_volume: 정확한 수량 (현물 부족/짜투리 처리용)
        decision_price: 결정가격 = 주문 발생 가격 = trigger = 종가 - 스마트 주문의 K값으로 사용
    
    Returns:
        dict: 주문 결과 (orderId 포함)
    
    Note:
        - 모든 시간 처리는 UTC 기준으로 수행
        - API 타임스탬프는 UTC 기준 밀리초 타임스탬프 사용
        - 로깅 시간은 get_timestamp() 함수 사용 (UTC 표시 포함)
        - 예상수익률 계산: price_hint(prev_tp)로 평균단가 역산 → smart_price 기준 수익률 계산 → 수수료 차감
    """
    try:
        # 1. 호가창 전체 스냅샷 조회 (선물 호가·잔량 — 스마트 주문용)
        ask, bid, ask_q, bid_q = get_futures_orderbook_snapshot(symbol)
        
        # 호가 조회 실패 시 안전장치 (기존 로직 유지)
        if ask is None or bid is None:
            print(f"{get_timestamp()} [{stage_prefix}] ❌선물 호가 조회 실패. 매도 취소")
            return None

        # 2. 스마트 주문 가격 결정 (전략 트리 실행!)
        # K = 결정가격(decision_price) = 주문 발생 가격 = trigger = 종가
        # 스마트 주문의 K는 항상 결정가격이어야 함
        K = decision_price if decision_price is not None and decision_price > 0 else ask 

        # 매도이므로 is_buy=False
        smart_price, smart_log = execute_smart_order(is_buy=False, K=K, ask=ask, bid=bid, ask_q=ask_q, bid_q=bid_q, symbol=symbol)
        
        # 안전장치: 가격이 없거나 0이면 매수호가로 폴백
        if smart_price is None or smart_price <= 0:
            smart_price = bid
            print(f"{get_timestamp()} [{stage_prefix}] ⚠️스마트 가격 계산 실패. 매수 1호가({bid:.6f} USDT)로 설정")
            smart_log = []
        else:
            # 가격을 틱 사이즈 규칙에 맞게 보정
            smart_price = round_price_to_tick_size_binance(symbol, smart_price)
            # 스마트 주문 로그 출력
            for log_msg in smart_log:
                print(f"{get_timestamp()} [{stage_prefix}] 📊{log_msg}")
                send_discord_message(f"{get_timestamp()} [{stage_prefix}] 📊{log_msg}")
        
        # 최소매도가 체크: 스마트 가격이 최소매도가보다 낮으면 최소매도가로 설정
        if min_sell_price is not None and min_sell_price > 0:
            # USDT 정밀도 가져오기
            usdt_precision = QUOTE_PRECISION_MAP.get(symbol, SYMBOL_USDT_PRECISION.get(symbol, 5))
            if smart_price < min_sell_price:
                min_price_msg = f"{get_timestamp()} [{stage_prefix}] ⚠️ 스마트 가격({smart_price:.{usdt_precision}f} USDT)이 최소매도가({min_sell_price:.{usdt_precision}f} USDT)보다 낮아 최소매도가로 조정"
                print(min_price_msg)
                send_discord_message(min_price_msg)
                smart_price = min_sell_price
            else:
                # 스마트 가격이 최소매도가보다 높거나 같으면 통과 메시지 출력
                pass_msg = f"{get_timestamp()} [{stage_prefix}] ✅ 스마트 가격({smart_price:.{usdt_precision}f} USDT)이 최소매도가({min_sell_price:.{usdt_precision}f} USDT) 이상으로 통과"
                print(pass_msg)
                send_discord_message(pass_msg)
        
        # bid_price를 smart_price로 대체
        bid_price = smart_price
        
        # 심볼별 수량 정밀도 가져오기
        qty_precision = SYMBOL_QTY_PRECISION.get(symbol, 8)  # 기본값 8
        
        # 잔고 부족 사전 체크 및 처리
        try:
            ticker_balance = binance_get_account_balance(TICKER)
            current_balance = ticker_balance['total_precise']  # 정밀한 총 수량 사용
            current_balance_usdt = current_balance * bid_price
            # 유닛 계산
            current_balance_unit = current_balance_usdt / TRADING_UNIT if TRADING_UNIT > 0 else 0
            
            print(f"{get_timestamp()} [{stage_prefix}] ✅{TICKER} 충분 Available: {current_balance:.8f} {TICKER} ({current_balance_unit:.2f} U {current_balance_usdt:.2f} USDT)")
            
            # 정확한 수량이 제공된 경우 해당 수량 사용 (현물 부족/짜투리 처리)
            if exact_volume is not None:
                # 티커별 stepSize 가져오기
                step_size = SYMBOL_STEP_SIZE.get(TICKER, 0.001)  # 기본값 0.001
                adjusted_volume = int(exact_volume / step_size) * step_size  # floor 사용
                final_volume = adjusted_volume
                exact_vol_msg = f"{get_timestamp()} [{stage_prefix}] 🔍 정확한 수량 사용: {final_volume:.8f} {TICKER} (stepSize 조정: {exact_volume:.8f} → {final_volume:.8f})"
                print(exact_vol_msg)
                send_discord_message(exact_vol_msg)
            else:
                # 간단한 수량 계산: 목표 금액 / 매수 1호가
                target_volume = usdt_amount / bid_price
                
                # 심볼별 정밀도에 맞춰 반올림
                volume = round(target_volume, qty_precision)
                
                # 안전 마진 적용 여부 결정 (짜투리 처리 시에는 제거)
                if use_safety_margin:
                    safety_margin = 1.001
                    final_volume = round(volume * safety_margin, qty_precision)
                else:
                    final_volume = volume
            
            # 최종 예상 금액 계산 (스마트 가격 사용)
            estimated_amount = final_volume * smart_price
            
            # 수량 계산 완료 로그 제거 (SOURCE 스타일 - 주문 전송 메시지에 포함됨)
            
            # Case 1: 매도 계산값 수량 > 보유잔고 이면서 보유잔고 >= 5 USDT (전량 매도)
            if final_volume > current_balance and current_balance_usdt >= 5:
                print(f"{get_timestamp()} [{stage_prefix}] 🔄 Case 1: 전량 매도 실행 (계산값: {final_volume:.{qty_precision}f}, 보유: {current_balance:.8f})")
                # stepSize에 맞춰 보유 잔고 조정 (LOT_SIZE 필터 준수)
                step_size = SYMBOL_STEP_SIZE.get(TICKER, 0.001)
                final_volume = int(current_balance / step_size) * step_size  # floor 사용하여 stepSize 배수로 조정
                stepsize_msg = f"{get_timestamp()} [{stage_prefix}] 🔍 stepSize 조정: {current_balance:.8f} → {final_volume:.8f} {TICKER}"
                print(stepsize_msg)
                send_discord_message(stepsize_msg)
                estimated_amount = final_volume * smart_price
                adjust_msg = f"{get_timestamp()} [{stage_prefix}] 🔍 수량 조정 완료: {final_volume:.8f} {TICKER} @ {smart_price:.6f} USDT (예상 {estimated_amount:.2f} USDT)"
                print(adjust_msg)
                send_discord_message(adjust_msg)
            
            # Case 2: 짜투리 처리 - 매도 계산값 수량 <= 보유잔고 이면서 (보유잔고 - 매도계산값수량) <= 5 USDT
            elif final_volume <= current_balance and (current_balance_usdt - estimated_amount) <= 5:
                case2_msg = f"{get_timestamp()} [{stage_prefix}] 🔄 Case 2: 짜투리 처리 - 보유 잔고만큼 매도 실행 (남은 잔고: {current_balance_usdt - estimated_amount:.2f} USDT)"
                print(case2_msg)
                send_discord_message(case2_msg)
                # stepSize에 맞춰 보유 잔고 조정
                step_size = SYMBOL_STEP_SIZE.get(TICKER, 0.001)
                final_volume = int(current_balance / step_size) * step_size
                stepsize_msg = f"{get_timestamp()} [{stage_prefix}] 🔍 stepSize 조정: {current_balance:.8f} → {final_volume:.8f} {TICKER}"
                print(stepsize_msg)
                send_discord_message(stepsize_msg)
                estimated_amount = final_volume * smart_price
                adjust_msg2 = f"{get_timestamp()} [{stage_prefix}] 🔍 수량 조정 완료: {final_volume:.8f} {TICKER} @ {smart_price:.6f} USDT (예상 {estimated_amount:.2f} USDT)"
                print(adjust_msg2)
                send_discord_message(adjust_msg2)
            
            # Case 3: 보유잔고 < 5 USDT (최소 주문 금액 미달)
            elif current_balance_usdt < 5:
                print(f"{get_timestamp()} [{stage_prefix}] ❌ 매도 주문 실패: 보유 잔고가 5 USDT 미만입니다 ({current_balance_usdt:.2f} USDT)")
                raise RuntimeError(f"보유 잔고가 5 USDT 미만입니다: {current_balance_usdt:.2f} USDT")
            
            # 최소 주문 금액 검증 (바이낸스 최소 5 USDT)
            if estimated_amount < 5:
                print(f"{get_timestamp()} [{stage_prefix}] ❌ 매도 주문 실패: 예상 금액이 5 USDT 미만입니다 ({estimated_amount:.2f} USDT)")
                raise RuntimeError(f"예상 금액이 5 USDT 미만입니다: {estimated_amount:.2f} USDT")
                
        except Exception as balance_error:
            print(f"{get_timestamp()} [{stage_prefix}] ❌ 잔고 조회 실패: {balance_error}")
            raise RuntimeError(f"잔고 조회 실패: {balance_error}")
        
        # 지정가 매도 주문 전송 (스마트 가격 사용)
        # stepSize에 맞춰 수량 조정 (BTC 등 소수 자릿수 제한으로 주문 실패 방지)
        step_size = SYMBOL_STEP_SIZE.get(TICKER, 0.001)
        final_volume = round(final_volume / step_size) * step_size
        final_volume = round(final_volume, qty_precision)
        qty_str = f"{final_volume:.{qty_precision}f}"
        price_str = f"{smart_price:.8f}"
        query_string = f"symbol={symbol}&side=SELL&type=LIMIT&timeInForce=GTC&quantity={qty_str}&price={price_str}"
        headers, signature, timestamp, recv_window = _binance_headers(query_string)
        
        params = {
            "symbol": symbol,
            "side": "SELL",
            "type": "LIMIT",
            "timeInForce": "GTC",
            "quantity": qty_str,
            "price": price_str,
            "timestamp": timestamp,
            "recvWindow": recv_window,
            "signature": signature
        }
        
        # 예상 수익률 계산 (Maker 주문은 즉시 체결 안 되므로 주문 전송 시점에 예상 수익률 표시)
        # price_hint = prev_tp (TP = 목표가격): TP 기준으로 평단가 역산하여 수익률 계산
        tp_percent = 0.01
        tp_ref = price_hint if price_hint and price_hint > 0 else smart_price  # TP 우선, 없으면 smart_price 사용
        assumed_avg = tp_ref / (1 + tp_percent)  # TP에서 평단가 역산 (TP = 평단가 × 1.01)
        gross_pct = ((smart_price - assumed_avg) / assumed_avg) * 100 if assumed_avg > 0 else 0  # 예상 체결가격(smart_price) 기준
        fee_rate = BINANCE_TRADING_FEE  # 0.075%
        expected_profit_rate = gross_pct - (fee_rate * 2 * 100)  # 매수+매도 수수료 차감
        expected_trade_usdt = smart_price * final_volume
        expected_pnl_usdt = expected_trade_usdt * (expected_profit_rate / 100.0)
        # 유닛 계산
        sell_unit = expected_trade_usdt / TRADING_UNIT if TRADING_UNIT > 0 else 0
        order_msg = f"{get_timestamp()} [{stage_prefix}] 📤매도 주문 전송 중 가격: {smart_price:.6f} USDT, 수량: {final_volume:.{qty_precision}f} {TICKER} ({sell_unit:.2f}U, {expected_trade_usdt:.2f} USDT) 예상수익률: {expected_profit_rate:+.2f}% ({expected_pnl_usdt:+.2f} USDT)"
        print(order_msg)
        send_discord_message(order_msg)
        r = requests.post(f"{BINANCE_API_BASE}/api/v3/order", params=params, headers=headers, timeout=10)
        
        if r.status_code != 200:
            error_data = r.json() if r.text else {}
            error_code = error_data.get('code', 'UNKNOWN')
            error_msg = error_data.get('msg', r.text)
            print(f"{get_timestamp()} [{stage_prefix}] ❌ 매도 주문 실패: {r.status_code} - {error_code}: {error_msg}")
            print(f"{get_timestamp()} [{stage_prefix}] 🔍 실패한 계산값: 수량={final_volume:.{qty_precision}f} {TICKER} (정밀도: {qty_precision}), 예상금액={estimated_amount:.2f} USDT, 스마트가격={smart_price:.6f} USDT")
            
            # 바이낸스 특정 에러 코드 처리
            if error_code == -1022:
                print(f"{get_timestamp()} [{stage_prefix}] 🔍 서명 오류: API 키 또는 서명이 잘못됨")
            elif error_code == -1021:
                print(f"{get_timestamp()} [{stage_prefix}] 🔍 시간 오류: 타임스탬프가 서버 시간과 맞지 않음")
            elif error_code == -2010:
                print(f"{get_timestamp()} [{stage_prefix}] 🔍 잔고 부족: 계좌 잔고가 부족함")
            elif error_code == -1013:
                print(f"{get_timestamp()} [{stage_prefix}] 🔍 수량 오류: 주문 수량이 최소/최대 범위를 벗어남")
            elif error_code == -1016:
                print(f"{get_timestamp()} [{stage_prefix}] 🔍 가격 오류: 주문 가격이 허용 범위를 벗어남")
            
            raise RuntimeError(f"매도 실패: {error_code} - {error_msg}")
        
        result = r.json()
        order_id = result.get('orderId')
        
        accept_msg = f"{get_timestamp()} [{stage_prefix}] ✅매도 주문 접수 성공: OrderID={order_id}"
        print(accept_msg)
        send_discord_message(accept_msg)
        
        # 주문 체결 상태 확인 및 실제 체결 가격 조회
        if order_id:
            print(f"{get_timestamp()} [{stage_prefix}] 🔍매도 주문 체결 확인 중 UUID: {order_id}")
            
            # 잠시 대기 후 체결 상태 확인
            time.sleep(2)
            
            # 주문 상태 조회
            order_status = binance_get_order_status(symbol, order_id)
            if order_status:
                executed_qty = order_status.get('executed_volume', 0)
                executed_price = order_status.get('executed_price', 0)
                remaining_qty = order_status.get('remaining_volume', 0)
                order_state = order_status.get('state', '')
                
                # 부분 체결 또는 대기 중인 경우 메시지 출력
                if order_state == 'PARTIALLY_FILLED' or (executed_qty >= 0 and remaining_qty > 0):
                    # USDT 정밀도 가져오기
                    usdt_precision = QUOTE_PRECISION_MAP.get(symbol, 2)
                    if executed_price > 0:
                        executed_price_str = f"{executed_price:.{usdt_precision}f} USDT"
                    else:
                        # 주문 가격 사용 (체결 가격이 없을 경우)
                        executed_price_str = f"{price_hint:.{usdt_precision}f} USDT" if price_hint and price_hint > 0 else "0.00 USDT"
                    partial_msg = f"{get_timestamp()} [{stage_prefix}] ⏳매도 주문 부분 체결: 체결 {executed_qty:.8f} @ {executed_price_str}, 미체결 {remaining_qty:.8f} {TICKER} (상태: {order_state.lower() if order_state else 'wait'})"
                    print(partial_msg)
                    send_discord_message(partial_msg)
                
                if executed_qty > 0 and executed_price > 0:
                    # 수익률 계산 (SOURCE 스타일)
                    tp_percent = 0.01
                    tp_ref = price_hint if price_hint and price_hint > 0 else executed_price
                    assumed_avg = tp_ref / (1 + tp_percent)  # TP에서 평단가 역산
                    gross_pct = ((executed_price - assumed_avg) / assumed_avg) * 100 if assumed_avg > 0 else 0
                    fee_rate = BINANCE_TRADING_FEE  # 0.075%
                    profit_rate = gross_pct - (fee_rate * 2 * 100)  # 매수+매도 수수료 차감
                    trade_usdt = executed_price * executed_qty
                    pnl_usdt = trade_usdt * (profit_rate / 100.0)
                    # 유닛 계산
                    sell_unit = trade_usdt / TRADING_UNIT if TRADING_UNIT > 0 else 0
                    # 통합된 체결 완료 메시지 (SOURCE 스타일)
                    complete_msg = f"{get_timestamp()} [{stage_prefix}] 🎯매도 체결 완료: {executed_qty:.8f} {TICKER} ({sell_unit:.2f}U, {trade_usdt:.2f} USDT) | 체결가격: {executed_price:.6f} USDT | 수익률: {profit_rate:+.2f}% ({pnl_usdt:+.2f} USDT)"
                    print(complete_msg)
                    send_discord_message(complete_msg)
            else:
                print(f"{get_timestamp()} [{stage_prefix}] ⚠️ 주문 상태 확인 실패")
        return result
        
    except Exception as e:
        print(f"{get_timestamp()} [{stage_prefix}] ❌ 매도 주문 중 오류 발생: {e}")
        raise

def binance_limit_buy(symbol: str, usdt_amount: float, price: float, stage_prefix: str = ""):
    """
    바이낸스 지정가 매수: 가격과 USDT 금액을 지정하여 매수
    
    Args:
        symbol: 거래 심볼 (예: "BTCUSDT")
        usdt_amount: 매수할 USDT 금액
        price: 지정가 가격
        stage_prefix: 로그 프리픽스
    
    Returns:
        dict: 주문 결과 (orderId 포함)
    """
    try:
        # 틱 사이즈에 맞춰 가격 조정
        price = adjust_price_to_tick_binance(symbol, price)
        
        # 심볼 precision 조회
        usdt_precision = QUOTE_PRECISION_MAP.get(symbol)
        if usdt_precision is None:
            info = binance_get_symbol_info(symbol)
            usdt_precision = info.get('quotePrecision') if 'quotePrecision' in info else info.get('quoteAssetPrecision', 5)
            try:
                usdt_precision = int(usdt_precision)
            except:
                usdt_precision = 5
            QUOTE_PRECISION_MAP[symbol] = usdt_precision
        
        # 수량 계산: 목표 금액 / 지정가
        target_volume = usdt_amount / price
        qty_precision = SYMBOL_QTY_PRECISION.get(symbol, 8)
        step_size = SYMBOL_STEP_SIZE.get(symbol.replace("USDT", ""), 0.001)
        volume = round(target_volume / step_size) * step_size
        volume = round(volume, qty_precision)
        qty_str = f"{volume:.{qty_precision}f}"
        
        # 최소 주문 금액 체크 (5 USDT)
        estimated_amount = volume * price
        if estimated_amount < 5:
            print(f"{get_timestamp()} [{stage_prefix}] ❌지정가 매수 주문 실패: 예상 금액이 5 USDT 미만입니다 ({estimated_amount:.2f} USDT)")
            return None
        
        # 지정가 매수 주문 전송
        query_string = f"symbol={symbol}&side=BUY&type=LIMIT&timeInForce=GTC&quantity={qty_str}&price={price}"
        headers, signature, timestamp, recv_window = _binance_headers(query_string)
        
        params = {
            "symbol": symbol,
            "side": "BUY",
            "type": "LIMIT",
            "timeInForce": "GTC",
            "quantity": qty_str,
            "price": str(price),
            "timestamp": timestamp,
            "recvWindow": recv_window,
            "signature": signature
        }
        
        buy_unit = usdt_amount / TRADING_UNIT if TRADING_UNIT > 0 else 0
        order_msg = f"{get_timestamp()} [{stage_prefix}] 📤매수 주문 전송 중 가격: {price:.6f} USDT, 수량: {volume:.8f} {TICKER} ({buy_unit:.2f}U, {usdt_amount:.2f} USDT)"
        print(order_msg)
        send_discord_message(order_msg)
        
        r = requests.post(f"{BINANCE_API_BASE}/api/v3/order", params=params, headers=headers, timeout=10)
        
        if r.status_code != 200:
            error_data = r.json() if r.text else {}
            error_code = error_data.get('code', 'UNKNOWN')
            error_msg = error_data.get('msg', r.text)
            print(f"{get_timestamp()} [{stage_prefix}] ❌지정가 매수 주문 실패: {r.status_code} - {error_code}: {error_msg}")
            return None
        
        result = r.json()
        order_id = result.get('orderId')
        accept_msg = f"{get_timestamp()} [{stage_prefix}] ✅매수 주문 접수 성공: OrderID={order_id}"
        print(accept_msg)
        send_discord_message(accept_msg)
        
        # 주문 체결 상태 확인 및 실제 체결 가격 조회
        if order_id:
            print(f"{get_timestamp()} [{stage_prefix}] 🔍매수 주문 체결 확인 중... OrderID: {order_id}")
            
            # 잠시 대기 후 체결 상태 확인
            time.sleep(2)
            
            # 주문 상태 조회
            order_status = binance_get_order_status(symbol, order_id)
            if order_status:
                executed_qty = order_status.get('executed_volume', 0)
                executed_price = order_status.get('executed_price', 0)
                remaining_qty = order_status.get('remaining_volume', 0)
                order_state = order_status.get('state', '')
                
                # 부분 체결 또는 대기 중인 경우 메시지 출력
                if order_state == 'PARTIALLY_FILLED' or (executed_qty >= 0 and remaining_qty > 0):
                    # USDT 정밀도 가져오기
                    usdt_precision = QUOTE_PRECISION_MAP.get(symbol, 2)
                    if executed_price > 0:
                        executed_price_str = f"{executed_price:.{usdt_precision}f} USDT"
                    else:
                        # 주문 가격 사용 (체결 가격이 없을 경우)
                        executed_price_str = f"{price:.{usdt_precision}f} USDT"
                    partial_msg = f"{get_timestamp()} [{stage_prefix}] ⏳매수 주문 부분 체결: 체결 {executed_qty:.8f} @ {executed_price_str}, 미체결 {remaining_qty:.8f} {TICKER} (상태: {order_state.lower() if order_state else 'wait'})"
                    print(partial_msg)
                    send_discord_message(partial_msg)
                
                if executed_qty > 0 and executed_price > 0:
                    trade_usdt = executed_price * executed_qty
                    buy_unit = trade_usdt / TRADING_UNIT if TRADING_UNIT > 0 else 0
                    complete_msg = f"{get_timestamp()} [{stage_prefix}] 🎯매수 체결 완료: {executed_qty:.8f} {TICKER} ({buy_unit:.2f}U, {trade_usdt:.2f} USDT) | 체결가격: {executed_price:.6f} USDT"
                    print(complete_msg)
                    send_discord_message(complete_msg)
                elif executed_qty == 0:
                    waiting_msg = f"{get_timestamp()} [{stage_prefix}] ⏳매수 주문 대기 중: 아직 체결되지 않음 (상태: {order_state.lower() if order_state else 'NEW'})"
                    print(waiting_msg)
                    send_discord_message(waiting_msg)
            else:
                print(f"{get_timestamp()} [{stage_prefix}] ⚠️ 주문 상태 확인 실패")
        
        return result
        
    except Exception as e:
        print(f"{get_timestamp()} [{stage_prefix}] ❌지정가 매수 주문 중 오류 발생: {e}")
        import traceback
        traceback.print_exc()
        return None

def binance_limit_sell(symbol: str, price: float, volume: float, stage_prefix: str = ""):
    """
    바이낸스 지정가 매도: 가격과 수량을 지정하여 매도 (GTC 동작)
    
    Args:
        symbol: 거래 심볼 (예: "BTCUSDT")
        price: 지정가 가격
        volume: 매도할 수량
        stage_prefix: 로그 프리픽스
    
    Returns:
        dict: 주문 결과 (orderId 포함)
    """
    try:
        # 틱 사이즈에 맞춰 가격 조정
        price = adjust_price_to_tick_binance(symbol, price)
        
        # 수량 정밀도·stepSize 조정 (BTC 등 소수 자릿수 제한으로 주문 실패 방지)
        qty_precision = SYMBOL_QTY_PRECISION.get(symbol, 8)
        step_size = SYMBOL_STEP_SIZE.get(symbol.replace("USDT", ""), 0.001)
        volume = round(volume / step_size) * step_size
        volume = round(volume, qty_precision)
        qty_str = f"{volume:.{qty_precision}f}"
        
        # 최소 주문 금액 체크 (5 USDT)
        estimated_amount = volume * price
        if estimated_amount < 5:
            skip_msg = f"{get_timestamp()} [{stage_prefix}] ⚠️지정가 매도 스킵: 예상 금액이 5 USDT 미만입니다 ({estimated_amount:.2f} USDT)"
            print(skip_msg)
            send_discord_message(skip_msg)
            return None
        
        send_msg = f"{get_timestamp()} [{stage_prefix}] 📤[주문 전송] {TICKER} 자동 지정가 매도\n가격: {price:.6f} USDT\n수량: {volume:.{qty_precision}f}\n예상금액: {estimated_amount:.2f} USDT"
        print(send_msg)
        send_discord_message(send_msg)
        
        # 지정가 매도 주문 전송
        query_string = f"symbol={symbol}&side=SELL&type=LIMIT&timeInForce=GTC&quantity={qty_str}&price={price}"
        headers, signature, timestamp, recv_window = _binance_headers(query_string)
        
        params = {
            "symbol": symbol,
            "side": "SELL",
            "type": "LIMIT",
            "timeInForce": "GTC",
            "quantity": qty_str,
            "price": str(price),
            "timestamp": timestamp,
            "recvWindow": recv_window,
            "signature": signature
        }
        
        r = requests.post(f"{BINANCE_API_BASE}/api/v3/order", params=params, headers=headers, timeout=10)
        
        if r.status_code != 200:
            error_data = r.json() if r.text else {}
            error_code = error_data.get('code', 'UNKNOWN')
            error_msg = error_data.get('msg', r.text)
            print(f"{get_timestamp()} [{stage_prefix}] ❌지정가 매도 주문 실패: {r.status_code} - {error_code}: {error_msg}")
            return None
        
        result = r.json()
        order_id = result.get('orderId')
        accept_msg = f"{get_timestamp()} [{stage_prefix}] ✅매도 주문 접수 성공: OrderID={order_id}"
        print(accept_msg)
        send_discord_message(accept_msg)
        
        # 주문 체결 상태 확인 및 실제 체결 가격 조회
        if order_id:
            print(f"{get_timestamp()} [{stage_prefix}] 🔍매도 주문 체결 확인 중... OrderID: {order_id}")
            
            # 잠시 대기 후 체결 상태 확인
            time.sleep(2)
            
            # 주문 상태 조회
            order_status = binance_get_order_status(symbol, order_id)
            if order_status:
                executed_qty = order_status.get('executed_volume', 0)
                executed_price = order_status.get('executed_price', 0)
                remaining_qty = order_status.get('remaining_volume', 0)
                order_state = order_status.get('state', '')
                
                # 부분 체결 또는 대기 중인 경우 메시지 출력
                if order_state == 'PARTIALLY_FILLED' or (executed_qty >= 0 and remaining_qty > 0):
                    # USDT 정밀도 가져오기
                    usdt_precision = QUOTE_PRECISION_MAP.get(symbol, 2)
                    if executed_price > 0:
                        executed_price_str = f"{executed_price:.{usdt_precision}f} USDT"
                    else:
                        # 주문 가격 사용 (체결 가격이 없을 경우)
                        executed_price_str = f"{price:.{usdt_precision}f} USDT"
                    partial_msg = f"{get_timestamp()} [{stage_prefix}] ⏳매도 주문 부분 체결: 체결 {executed_qty:.8f} @ {executed_price_str}, 미체결 {remaining_qty:.8f} {TICKER} (상태: {order_state.lower() if order_state else 'wait'})"
                    print(partial_msg)
                    send_discord_message(partial_msg)
                
                if executed_qty > 0 and executed_price > 0:
                    trade_usdt = executed_price * executed_qty
                    sell_unit = trade_usdt / TRADING_UNIT if TRADING_UNIT > 0 else 0
                    complete_msg = f"{get_timestamp()} [{stage_prefix}] 🎯매도 체결 완료: {executed_qty:.8f} {TICKER} ({sell_unit:.2f}U, {trade_usdt:.2f} USDT) | 체결가격: {executed_price:.6f} USDT"
                    print(complete_msg)
                    send_discord_message(complete_msg)
                elif executed_qty == 0:
                    waiting_msg = f"{get_timestamp()} [{stage_prefix}] ⏳매도 주문 대기 중: 아직 체결되지 않음 (상태: {order_state.lower() if order_state else 'NEW'})"
                    print(waiting_msg)
                    send_discord_message(waiting_msg)
            else:
                print(f"{get_timestamp()} [{stage_prefix}] ⚠️ 주문 상태 확인 실패")
        
        return result
        
    except Exception as e:
        print(f"{get_timestamp()} [{stage_prefix}] ❌지정가 매도 주문 중 오류 발생: {e}")
        import traceback
        traceback.print_exc()
        return None

def calculate_ksc_multiplier(ksc_value, ksc_stack: int = 0, p_value: int = 3) -> int:
    """
    KSC 스택에 따른 multiplier를 계산합니다.
    
    Args:
        ksc_value: KSC 값 (숫자 스택)
        ksc_stack: 사용하지 않음 (호환성을 위해 유지)
        p_value: p 값 (기본값 3, p = 3 + p1H, 15M 열 p 사용)
    
    Returns:
        int: multiplier 값 (0 또는 p)
    
    수열 규칙:
        - KSC가 p의 배수일 때만 p, 나머지는 0
        - KSC = 1, 2, ..., p-1 → multiplier = 0
        - KSC = p, 2p, 3p, ... → multiplier = p
    
    수식: multiplier = (KSC % p == 0) ? p : 0
    
    예시 (p=3):
        KSC = 1  → multiplier = 0
        KSC = 2  → multiplier = 0
        KSC = 3  → multiplier = 3
        KSC = 4  → multiplier = 0
        KSC = 5  → multiplier = 0
        KSC = 6  → multiplier = 3
        KSC = 9  → multiplier = 3
        KSC = 12 → multiplier = 3
        KSC = 15 → multiplier = 3
        KSC = 13 → multiplier = 0 (3의 배수 아님)
    
    Note:
        - p_value는 동적으로 계산될 수 있음 (p = 3 + p1H, 15M 열 p 사용)
        - multiplier와 B값은 독립적으로 계산됨
    """
    # KSC를 숫자로 변환
    ksc_numeric = int(ksc_value) if isinstance(ksc_value, (int, float)) else 0
    
    if ksc_numeric <= 0 or p_value <= 0:
        return 0
    
    # p의 배수일 때만 p 반환
    if ksc_numeric % p_value == 0:
        return p_value
    else:
        return 0

def calculate_bomb_b_value(multiplier: int, ksc_stack: int, p_value: int = 3) -> int:
    """
    Bomb 발생 시 B 값을 계산합니다.
    
    Args:
        multiplier: KSC 스택에 따른 multiplier 값
        ksc_stack: BombCount 값 (Bomb 발생 시점의 스택 카운트, 사용하지 않음)
        p_value: p 값 (기본값 3, p = 3 + p1H, 15M 열 p 사용)
    
    Returns:
        int: B 값
        - multiplier == 0이면: 1 (고정값)
        - multiplier != 0이면: 0
    
    Note:
        - SOURCE 버전: multiplier == 0일 때 항상 1 반환 (고정값)
        - ROOT 기존 버전과 다름: 기존에는 ksc_stack 기반으로 1,2,3 반복했지만 SOURCE는 1 고정
    """
    # multiplier가 0이 아니면 B값도 0
    if multiplier != 0:
        return 0
    
    # multiplier == 0일 때: B값 = 1 (고정값)
    return 1

def execute_smart_order(is_buy: bool, K: float, ask: float, bid: float, ask_q: float, bid_q: float, symbol: str = None):
    """
    [Binance 최종] Tight(K 기반) + Wide(잔량 기반, K 무시) 통합 스마트 주문 엔진.
    티커별 소수점 정확도: symbol을 반드시 전달해야 틱 사이즈가 심볼별로 적용됩니다.

    롱/숏 대칭 (숏 = 롱의 거울 적용):
    - 롱(매수): 낮은 가격(Bid)에 사고 싶음. 매도벽(ask_q) 두꺼우면 Bid 대기 [Maker], 아니면 Ask 긁기 [Taker].
    - 숏(매도): 높은 가격(Ask)에 팔고 싶음. 매수벽(bid_q) 두꺼우면 Ask 대기 [Maker], 아니면 Bid 던지기 [Taker].
    - 비교: 롱은 K>=Ask/K<=Bid, 숏은 K>=Bid/K<Bid. 벽 확인은 롱=ask_q(매도벽), 숏=bid_q(매수벽).
    """
    log = []
    def logprint(msg): log.append(msg)

    # 1. 입력값 기본 검증
    if ask is None or bid is None or ask <= 0 or bid <= 0:
        logprint("⚠️호가 데이터 오류")
        return None, log
    
    ask_q = ask_q if ask_q is not None else 0
    bid_q = bid_q if bid_q is not None else 0
    # 틱 사이즈 및 가격 보정 (symbol 있으면 티커별 tickSize 적용, 없으면 기본값 사용)
    if symbol:
        K = adjust_price_to_tick_binance(symbol, K)
        t_ask = get_binance_tick_size(symbol, ask)
        t_bid = get_binance_tick_size(symbol, bid)
    else:
        K = round(K, 8)
        t_ask = t_bid = 0.01 if ask >= 1 else 0.0001
    spread = ask - bid
    epsilon = max(t_ask, t_bid) * 0.0001
    is_tight = spread <= t_bid + epsilon
    
    def is_eq(a, b): return abs(a - b) < epsilon
    def is_gt(a, b): return a > b + epsilon
    def is_lt(a, b): return a < b - epsilon
    base = f"[{'매수' if is_buy else '매도'}] K={K:.6f}, Ask={ask:.6f}({ask_q:.4f}), Bid={bid:.6f}({bid_q:.4f})"
    # ─────────────── [매수 로직 (is_buy = True)] ────────────────
    if is_buy:
        # --- 1. Tight Spread (호가가 붙어있음: K 위치 고려) ---
        if is_tight:
            # 상황 A: K >= ask (높은 가격 매수 의사)
            if is_gt(K, ask) or is_eq(K, ask):
                if is_gt(ask_q, bid_q * 3):
                    logprint(f"{base} | [Tight] K>=Ask & 3배매도벽↑ -> Bid 대기 [Maker]")
                    return bid, log
                else:
                    logprint(f"{base} | [Tight] K>=Ask & 벽 3배미만 -> Ask 긁기 [Taker]")
                    return ask, log
            
            # 상황 B: K <= bid (낮은 가격 매수 의사)
            else:
                if is_gt(ask_q, bid_q * 3):
                    logprint(f"{base} | [Tight] K<=Bid & 3배매도벽↑ -> Bid 대기 [Maker]")
                    return bid, log
                else:
                    logprint(f"{base} | [Tight] K<=Bid & 벽 3배미만 -> Ask 긁기 [Taker]")
                    return ask, log
        # --- 2. Wide Spread (호가가 벌어짐: K 무시, 잔량 기반) ---
        else:
            if is_gt(ask_q, bid_q * 3):
                target = adjust_price_to_tick_binance(symbol, bid + t_bid)
                if target >= ask: target = bid
                logprint(f"{base} | [Wide] 3배매도벽↑ -> Bid+1틱 대기 [Maker]")
                return target, log
            elif is_gt(ask_q, bid_q):
                target = adjust_price_to_tick_binance(symbol, ask - t_ask)
                if target <= bid: target = bid
                logprint(f"{base} | [Wide] 일반매도벽 -> Ask-1틱 대기 [Maker]")
                return target, log
            else:
                logprint(f"{base} | [Wide] 매수우세 -> Ask 긁기 [Taker]")
                return ask, log

    # ─────────────── [매도 로직 (is_buy = False)] ────────────────
    else:
        # --- 1. Tight Spread (호가가 붙어있음: K 위치 고려) ---
        if is_tight:
            # 상황 A: K >= bid (높거나 현재가 매도 의사)
            if is_gt(K, bid) or is_eq(K, bid):
                if is_gt(bid_q, ask_q * 3):
                    logprint(f"{base} | [Tight] K>=Bid & 3배매수벽↑ -> Ask 대기 [Maker]")
                    return ask, log
                else:
                    logprint(f"{base} | [Tight] K>=Bid & 벽 3배미만 -> Bid 던지기 [Taker]")
                    return bid, log

            # 상황 B: K < bid (낮은 가격 매도 의사)
            else:
                logprint(f"{base} | [Tight] K<Bid(신호하향) -> Bid 던지기 [Taker]")
                return bid, log

        # --- 2. Wide Spread (호가가 벌어짐: K 무시, 잔량 기반) ---
        else:
            # 상황 A: 매도잔량 < 매수잔량 (매수세가 더 강함)
            if is_gt(bid_q, ask_q):
                if is_gt(bid_q, ask_q * 3):
                    target = adjust_price_to_tick_binance(symbol, ask - t_ask)
                    if target <= bid: target = ask
                    logprint(f"{base} | [Wide] 3배매수벽↑ -> Ask-1틱 대기 [Maker]")
                    return target, log
                else:
                    target = adjust_price_to_tick_binance(symbol, bid + t_bid)
                    if target >= ask: target = bid
                    logprint(f"{base} | [Wide] 일반매수벽 -> Bid+1틱 대기 [Maker]")
                    return target, log
            
            # 상황 B: 매수잔량 < 매도잔량 (매도세가 더 강함)
            else:
                logprint(f"{base} | [Wide] 매도우세 -> Bid 던지기 [Taker]")
                return bid, log

    return bid, log

def _calc_h_factor(hcl: Optional[float]) -> float:
    """
    H 팩터를 계산합니다.
    
    Args:
        hcl: 1HCL + (-1HCL) 값
    
    Returns:
        float: H 팩터 (0.1 ~ 2.0 범위로 클램핑)
        - 계산식: H = 1 - 0.3 * hcl
        - round(H, 4) 후 [0.1, 2.0] 범위로 클램핑
    """
    if hcl is None:
        return 1.0
    try:
        hcl_val = float(hcl)
    except (TypeError, ValueError):
        return 1.0
    
    # H = 1 - 0.3 * hcl
    H = 1.0 - 0.3 * hcl_val
    H = round(H, 4)
    
    # [0.1, 2.0] 범위로 클램핑
    H = max(0.1, min(2.0, H))
    
    return H

def get_buy_risk_factor(order_value: str, hmsfast: Optional[float] = None, hcl: Optional[float] = None) -> float:
    """
    Buy 계열 주문에 H 팩터를 적용합니다.
    
    Args:
        order_value: 주문 신호 ("Buy5", "Buy10")
        hmsfast: 1HMSFast 값 (사용하지 않음, 호환성 유지)
        hcl: 1HCL + (-1HCL) 값
    
    Returns:
        float: H 팩터 (Buy5/Buy10에만 적용, 그 외는 1.0)
    """
    if order_value not in ("Buy5", "Buy10"):
        return 1.0
    
    return _calc_h_factor(hcl)

# ==========================================
# 유틸리티 함수들
# ==========================================
def _safe_float(val: Any, default: float = 0.0) -> float:
    """안전하게 float로 변환합니다."""
    try:
        return float(val) if pd.notna(val) else default
    except:
        return default

def _safe_float_opt(val: Any) -> Optional[float]:
    """안전하게 float로 변환합니다 (None 허용)."""
    try:
        return float(val) if pd.notna(val) else None
    except:
        return None

# ==========================================
# 파일 관리 함수들
# ==========================================
def check_file_cleanup_done_today():
    """
    UTC 0시 기준으로 오늘 날짜에 파일 정리를 이미 실행했는지 확인합니다.
    Returns: True(이미 실행함), False(아직 실행 안 함)
    """
    try:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        log_dir = os.path.join(script_dir, LOG_DIR)
        marker_file = os.path.join(log_dir, "binance_file_cleanup_last_date.txt")
        
        if not os.path.exists(marker_file):
            return False  # 마커 파일이 없으면 아직 실행 안 함
        
        # 마커 파일에서 마지막 정리 날짜 읽기
        with open(marker_file, 'r', encoding='utf-8') as f:
            last_date_str = f.read().strip()
        
        # 오늘 날짜 (UTC 0시 기준, YYYY-MM-DD)
        today_str = dt.datetime.now(tz.UTC).strftime('%Y-%m-%d')
        
        # 마지막 정리 날짜가 오늘과 같으면 이미 실행함
        return last_date_str == today_str
        
    except Exception as e:
        # 에러 발생 시 안전하게 False 반환 (정리 실행하도록)
        return False

def mark_file_cleanup_done_today():
    """
    UTC 0시 기준으로 오늘 날짜에 파일 정리를 완료했다고 표시합니다.
    """
    try:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        log_dir = os.path.join(script_dir, LOG_DIR)
        if not os.path.exists(log_dir):
            os.makedirs(log_dir)
        
        marker_file = os.path.join(log_dir, "binance_file_cleanup_last_date.txt")
        
        # 오늘 날짜 (UTC 0시 기준, YYYY-MM-DD)
        today_str = dt.datetime.now(tz.UTC).strftime('%Y-%m-%d')
        
        # 마커 파일에 오늘 날짜 기록
        with open(marker_file, 'w', encoding='utf-8') as f:
            f.write(today_str)
            
    except Exception as e:
        # 마커 파일 기록 실패는 무시 (다음 실행 시 다시 시도)
        pass

def delete_old_excel_files(days_to_keep=7):
    """
    cryptodaily15min 폴더의 티커별 엑셀 생성 폴더에서 생성된 지 days_to_keep일이 지난 
    .xlsx 파일을 삭제합니다. (임시파일 ~$ 제외)
    """
    script_dir = os.path.dirname(os.path.abspath(__file__))
    base_dir = os.path.join(script_dir, "cryptodaily15min")
    
    if not os.path.exists(base_dir):
        return
    
    # 티커별 폴더 매핑 (Binance용, 공백 포함)
    ticker_folder_mapping = {
        "BTC": "F BINANCE 1BTC",
        "ETH": "F BINANCE 2ETH",
        "XRP": "F BINANCE 3XRP",
        "SOL": "F BINANCE 4SOL",
        "BNB": "F BINANCE 5BNB"
    }
    
    # 현재 시간에서 보관 기간을 뺀 기준 시간 계산 (초 단위)
    cutoff_time = time.time() - (days_to_keep * 24 * 60 * 60)
    deleted_count = 0
    
    # 티커별 폴더만 스캔
    for ticker in ROTATION_TICKERS:
        folder_name = ticker_folder_mapping.get(ticker, f"CRYPTO_{ticker}")
        ticker_folder = os.path.join(base_dir, folder_name)
        
        if not os.path.exists(ticker_folder):
            continue
        
        # 해당 티커 폴더의 파일만 스캔
        try:
            for filename in os.listdir(ticker_folder):
                # 엑셀 파일이면서 엑셀 임시 파일(~$...)이 아닌 경우
                if filename.endswith(".xlsx") and not filename.startswith("~$"):
                    file_path = os.path.join(ticker_folder, filename)
                    try:
                        # 파일 수정 시간 확인
                        file_mod_time = os.path.getmtime(file_path)
                        
                        # 기준 시간보다 오래된 파일이면 삭제
                        if file_mod_time < cutoff_time:
                            os.remove(file_path)
                            deleted_count += 1
                    except Exception as e:
                        pass
        except Exception as e:
            pass
    
    if deleted_count > 0:
        print(f"🧹[자동 정리] {days_to_keep}일 지난 엑셀 파일 {deleted_count}개 정리 완료")
    
    # 정리 완료 후 오늘 날짜 마커 기록 (UTC 0시 기준)
    mark_file_cleanup_done_today()

def delete_old_logs_abs():
    """오래된 로그 파일을 삭제합니다."""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    log_dir_abs = os.path.join(script_dir, LOG_DIR)
    
    if not os.path.exists(log_dir_abs):
        return
    
    cutoff_time = time.time() - (DAYS_TO_KEEP * 24 * 60 * 60)
    for filename in os.listdir(log_dir_abs):
        file_path = os.path.join(log_dir_abs, filename)
        if os.path.isfile(file_path) and filename.endswith(".txt"):
            file_mod_time = os.path.getmtime(file_path)
            if file_mod_time < cutoff_time:
                try:
                    os.remove(file_path)
                    print(f"🧹[자동 정리] {DAYS_TO_KEEP}일 지난 로그 삭제됨: {filename}")
                except Exception as e:
                    print(f"⚠️삭제 실패 ({filename}): {e}")

# ==========================================
# Discord 관련 함수들
# ==========================================
def get_discord_webhook_url():
    """파일에서 디스코드 웹후크 URL을 읽어옵니다."""
    try:
        if os.path.exists(WEBHOOK_FILE_PATH):
            try:
                with open(WEBHOOK_FILE_PATH, 'r', encoding='utf-8') as f:
                    url = f.read().strip()
            except UnicodeDecodeError:
                with open(WEBHOOK_FILE_PATH, 'r', encoding='cp949') as f:
                    url = f.read().strip()
            
            if url and url.startswith("http"):
                print(f"✅ 디스코드 웹후크 URL 로드 성공")
                return url
            else:
                print(f"⚠️ 디스코드 웹후크 URL 형식이 올바르지 않습니다. (http로 시작해야 함, 현재: {url[:50] if url else '빈 파일'})")
        else:
            print(f"⚠️ 디스코드 웹후크 파일이 없습니다: {WEBHOOK_FILE_PATH}")
    except Exception as e:
        print(f"⚠️ 디스코드 웹후크 파일 로드 실패: {e}")
        import traceback
        traceback.print_exc()
    return None

# 전역 변수 초기화
DISCORD_WEBHOOK_URL = get_discord_webhook_url()

def _send_discord_actual_with_delay(msg):
    """디스코드 웹후크로 메시지를 실제로 전송하는 함수 (내부 함수, 딜레이 포함)"""
    try:
        # 메시지 간 딜레이: 50밀리초 (API 차단 방지)
        time.sleep(0.05)
        
        global DISCORD_WEBHOOK_URL
        if not DISCORD_WEBHOOK_URL:
            DISCORD_WEBHOOK_URL = get_discord_webhook_url()
            if not DISCORD_WEBHOOK_URL:
                return
        
        # 메시지 길이 제한 처리 (2000자)
        if len(msg) > 1900:
            msg = msg[:1900] + "\n...(내용 잘림)..."
        
        payload = {"content": f"```{msg}```"}
        headers = {"Content-Type": "application/json"}
        
        # 타임아웃 설정 (5초: 네트워크 지연 대응하면서도 매매 로직 방해 최소화)
        requests.post(DISCORD_WEBHOOK_URL, data=json.dumps(payload), headers=headers, timeout=5)
    except Exception as e:
        print(f"⚠️디스코드 전송 실패: {e}")

def send_discord_message(msg):
    """디스코드 웹후크로 메시지를 비동기(스레드)로 전송합니다. (메인 봇은 멈추지 않음)"""
    if not msg or not msg.strip():
        return
    
    # 웹후크 URL 확인
    global DISCORD_WEBHOOK_URL
    if not DISCORD_WEBHOOK_URL:
        DISCORD_WEBHOOK_URL = get_discord_webhook_url()
    
    if not DISCORD_WEBHOOK_URL:
        print(f"⚠️ 디스코드 메시지 전송 실패: 웹후크 URL이 없습니다. 파일 확인: {WEBHOOK_FILE_PATH}")
        return
    
    # 별도 스레드에서 전송 (메인 봇은 멈추지 않음, 5ms 딜레이 적용)
    threading.Thread(target=_send_discord_actual_with_delay, args=(msg,), daemon=True).start()

# ==========================================
# 파일 처리 함수들
# ==========================================
def wait_for_file_ready(file_path: str, max_wait_seconds: int = 5, check_interval: float = 0.1) -> bool:
    """
    파일이 완전히 저장되고 읽을 수 있을 때까지 대기합니다.
    
    Args:
        file_path: 확인할 파일 경로
        max_wait_seconds: 최대 대기 시간 (초)
        check_interval: 확인 간격 (초)
    
    Returns:
        파일이 준비되었으면 True, 그렇지 않으면 False
    """
    start_time = time.time()
    last_size = -1
    
    while time.time() - start_time < max_wait_seconds:
        if not os.path.exists(file_path):
            time.sleep(check_interval)
            continue
        
        try:
            current_size = os.path.getsize(file_path)
            # 파일 크기가 안정화되었는지 확인 (연속 3번 같은 크기)
            if current_size == last_size and current_size > 0:
                # 파일이 잠겨있지 않은지 확인 (읽기 모드로 열어보기)
                try:
                    with open(file_path, 'rb') as f:
                        f.read(1)
                    return True
                except (IOError, PermissionError):
                    time.sleep(check_interval)
                    continue
            last_size = current_size
            time.sleep(check_interval)
        except (OSError, IOError):
            time.sleep(check_interval)
            continue
    
    # 최대 대기 시간 초과
    return False

# ==========================================
# 1H4X 관련 함수들
# ==========================================
def calculate_buy_1h4x(fore_or_one, sellside):
    """
    Buy 지표를 계산합니다. (1H4x 시트용)
    4or1과 sellside를 사용하여 계산합니다.
    판정 기준: sellside <= 0.1
    """
    # NaN 체크
    if pd.isna(fore_or_one) or pd.isna(sellside):
        return ""
    
    # 조건: 4or1 < 4이고 sellside <= 0.1일 때 "buy" 반환
    if fore_or_one < 4 and sellside <= 0.1:
        return "buy"
    else:
        return ""

def calculate_sell_short_1h4x(buyside):
    """
    Sell 지표를 계산합니다. (1H4x 시트용)
    4or1 없이 buyside만 사용하여 계산합니다.
    판정 기준: buyside <= 0.1150
    """
    # NaN 체크
    if pd.isna(buyside):
        return ""
    
    # 조건: buyside <= 0.1150일 때 "sell" 반환
    if buyside <= 0.1150:
        return "sell"
    else:
        return ""

def calculate_1hmsfast(close, sma25, sma100):
    """
    1HMSFast 지표를 계산합니다. (1시간봉용)
    Source 기준: Fast 계산 함수와 동일한 방식으로 종가, SMA25, SMA100을 사용합니다.
    """
    # NaN 체크
    if pd.isna(close) or pd.isna(sma25) or pd.isna(sma100):
        return np.nan
    
    close_val, sma25_val, sma100_val = float(close), float(sma25), float(sma100)
    eps = 0.0  # 정확한 비교를 위한 epsilon 값

    def gt(a, b):
        return a > b + eps

    def eq(a, b):
        return abs(a - b) <= eps

    # 1) Strict phase
    phase_strict = 0
    if gt(close_val, sma25_val) and gt(sma25_val, sma100_val):
        phase_strict = 1
    elif gt(sma25_val, close_val) and gt(close_val, sma100_val):
        phase_strict = 2
    elif gt(sma25_val, sma100_val) and gt(sma100_val, close_val):
        phase_strict = 3
    elif gt(sma100_val, sma25_val) and gt(sma25_val, close_val):
        phase_strict = 4
    elif gt(sma100_val, close_val) and gt(close_val, sma25_val):
        phase_strict = 5
    elif gt(close_val, sma100_val) and gt(sma100_val, sma25_val):
        phase_strict = 6

    # 2) Beta
    beta = 0.0
    if gt(close_val, sma25_val) and gt(sma25_val, sma100_val):
        beta = (sma25_val - sma100_val) / (close_val - sma100_val) if not eq(close_val, sma100_val) else 0.0
    elif gt(sma25_val, close_val) and gt(close_val, sma100_val):
        beta = 1.0 - (close_val - sma100_val) / (sma25_val - sma100_val) if not eq(sma25_val, sma100_val) else 0.0
    elif gt(sma25_val, sma100_val) and gt(sma100_val, close_val):
        beta = (sma100_val - close_val) / (sma25_val - close_val) if not eq(sma25_val, close_val) else 0.0
    elif gt(sma100_val, sma25_val) and gt(sma25_val, close_val):
        beta = 1.0 - (sma25_val - close_val) / (sma100_val - close_val) if not eq(sma100_val, close_val) else 0.0
    elif gt(sma100_val, close_val) and gt(close_val, sma25_val):
        beta = (close_val - sma25_val) / (sma100_val - sma25_val) if not eq(sma100_val, sma25_val) else 0.0
    elif gt(close_val, sma100_val) and gt(sma100_val, sma25_val):
        beta = 1.0 - (sma100_val - sma25_val) / (close_val - sma25_val) if not eq(close_val, sma25_val) else 0.0

    # 3) add6 조건
    phase_plus_beta = phase_strict + beta
    add6 = 6 if (phase_plus_beta > 0 and phase_plus_beta < 1.5) else 0

    # 4) Equal phase
    equal_phase = 0
    if eq(close_val, sma25_val) and gt(close_val, sma100_val):
        equal_phase = 2
    elif eq(close_val, sma100_val) and gt(sma25_val, close_val):
        equal_phase = 3
    elif eq(sma25_val, sma100_val) and gt(sma25_val, close_val):
        equal_phase = 4
    elif eq(sma25_val, close_val) and gt(sma100_val, sma25_val):
        equal_phase = 5
    elif eq(sma100_val, close_val) and gt(sma100_val, sma25_val):
        equal_phase = 6
    elif eq(sma100_val, sma25_val) and gt(close_val, sma100_val):
        equal_phase = 7

    # 5) 최종값
    final_value = phase_plus_beta + add6 + equal_phase

    # 반올림 처리 제거 - 원본 값 그대로 반환
    return final_value

def calculate_1hmsfast_15m(smaf, sma100, sma200):
    """
    1HMSFast 지표를 계산합니다. (15분봉용)
    Fast 계산 함수와 동일한 방식으로 SMAF, SMA100, SMA200을 사용합니다. (종가 없음)
    """
    # NaN 체크
    if pd.isna(smaf) or pd.isna(sma100) or pd.isna(sma200):
        return np.nan
    
    smaf_val, sma100_val, sma200_val = float(smaf), float(sma100), float(sma200)
    eps = 0.0  # 정확한 비교를 위한 epsilon 값

    def gt(a, b):
        return a > b + eps

    def eq(a, b):
        return abs(a - b) <= eps

    # 1) Strict phase
    phase_strict = 0
    if gt(smaf_val, sma100_val) and gt(sma100_val, sma200_val):
        phase_strict = 1
    elif gt(sma100_val, smaf_val) and gt(smaf_val, sma200_val):
        phase_strict = 2
    elif gt(sma100_val, sma200_val) and gt(sma200_val, smaf_val):
        phase_strict = 3
    elif gt(sma200_val, sma100_val) and gt(sma100_val, smaf_val):
        phase_strict = 4
    elif gt(sma200_val, smaf_val) and gt(smaf_val, sma100_val):
        phase_strict = 5
    elif gt(smaf_val, sma200_val) and gt(sma200_val, sma100_val):
        phase_strict = 6

    # 2) Beta
    beta = 0.0
    if gt(smaf_val, sma100_val) and gt(sma100_val, sma200_val):
        beta = (sma100_val - sma200_val) / (smaf_val - sma200_val) if not eq(smaf_val, sma200_val) else 0.0
    elif gt(sma100_val, smaf_val) and gt(smaf_val, sma200_val):
        beta = 1.0 - (smaf_val - sma200_val) / (sma100_val - sma200_val) if not eq(sma100_val, sma200_val) else 0.0
    elif gt(sma100_val, sma200_val) and gt(sma200_val, smaf_val):
        beta = (sma200_val - smaf_val) / (sma100_val - smaf_val) if not eq(sma100_val, smaf_val) else 0.0
    elif gt(sma200_val, sma100_val) and gt(sma100_val, smaf_val):
        beta = 1.0 - (sma100_val - smaf_val) / (sma200_val - smaf_val) if not eq(sma200_val, smaf_val) else 0.0
    elif gt(sma200_val, smaf_val) and gt(smaf_val, sma100_val):
        beta = (smaf_val - sma100_val) / (sma200_val - sma100_val) if not eq(sma200_val, sma100_val) else 0.0
    elif gt(smaf_val, sma200_val) and gt(sma200_val, sma100_val):
        beta = 1.0 - (sma200_val - sma100_val) / (smaf_val - sma100_val) if not eq(smaf_val, sma100_val) else 0.0

    # 3) add6 조건
    phase_plus_beta = phase_strict + beta
    add6 = 6 if (phase_plus_beta > 0 and phase_plus_beta < 1.5) else 0

    # 4) Equal phase
    equal_phase = 0
    if eq(smaf_val, sma100_val) and gt(smaf_val, sma200_val):
        equal_phase = 2
    elif eq(smaf_val, sma200_val) and gt(sma100_val, smaf_val):
        equal_phase = 3
    elif eq(sma100_val, sma200_val) and gt(sma100_val, smaf_val):
        equal_phase = 4
    elif eq(sma100_val, smaf_val) and gt(sma200_val, sma100_val):
        equal_phase = 5
    elif eq(sma200_val, smaf_val) and gt(sma200_val, sma100_val):
        equal_phase = 6
    elif eq(sma200_val, sma100_val) and gt(smaf_val, sma200_val):
        equal_phase = 7

    # 5) 최종값
    final_value = phase_plus_beta + add6 + equal_phase

    # 반올림 처리 제거 - 원본 값 그대로 반환
    return final_value

def calculate_1hmsfast_1h4x(close, sma100, sma200):
    """
    1HMSFast 지표를 계산합니다. (1H4x 시트용)
    Fast 계산 함수와 동일한 방식으로 종가, SMA100, SMA200을 사용합니다.
    """
    # NaN 체크
    if pd.isna(close) or pd.isna(sma100) or pd.isna(sma200):
        return np.nan
    
    close_val, sma100_val, sma200_val = float(close), float(sma100), float(sma200)
    eps = 0.0  # 정확한 비교를 위한 epsilon 값

    def gt(a, b):
        return a > b + eps

    def eq(a, b):
        return abs(a - b) <= eps

    # 1) Strict phase
    phase_strict = 0
    if gt(close_val, sma100_val) and gt(sma100_val, sma200_val):
        phase_strict = 1
    elif gt(sma100_val, close_val) and gt(close_val, sma200_val):
        phase_strict = 2
    elif gt(sma100_val, sma200_val) and gt(sma200_val, close_val):
        phase_strict = 3
    elif gt(sma200_val, sma100_val) and gt(sma100_val, close_val):
        phase_strict = 4
    elif gt(sma200_val, close_val) and gt(close_val, sma100_val):
        phase_strict = 5
    elif gt(close_val, sma200_val) and gt(sma200_val, sma100_val):
        phase_strict = 6

    # 2) Beta
    beta = 0.0
    if gt(close_val, sma100_val) and gt(sma100_val, sma200_val):
        beta = (sma100_val - sma200_val) / (close_val - sma200_val) if not eq(close_val, sma200_val) else 0.0
    elif gt(sma100_val, close_val) and gt(close_val, sma200_val):
        beta = 1.0 - (close_val - sma200_val) / (sma100_val - sma200_val) if not eq(sma100_val, sma200_val) else 0.0
    elif gt(sma100_val, sma200_val) and gt(sma200_val, close_val):
        beta = (sma200_val - close_val) / (sma100_val - close_val) if not eq(sma100_val, close_val) else 0.0
    elif gt(sma200_val, sma100_val) and gt(sma100_val, close_val):
        beta = 1.0 - (sma100_val - close_val) / (sma200_val - close_val) if not eq(sma200_val, close_val) else 0.0
    elif gt(sma200_val, close_val) and gt(close_val, sma100_val):
        beta = (close_val - sma100_val) / (sma200_val - sma100_val) if not eq(sma200_val, sma100_val) else 0.0
    elif gt(close_val, sma200_val) and gt(sma200_val, sma100_val):
        beta = 1.0 - (sma200_val - sma100_val) / (close_val - sma100_val) if not eq(close_val, sma100_val) else 0.0

    # 3) add6 조건
    phase_plus_beta = phase_strict + beta
    add6 = 6 if (phase_plus_beta > 0 and phase_plus_beta < 1.5) else 0

    # 4) Equal phase
    equal_phase = 0
    if eq(close_val, sma100_val) and gt(close_val, sma200_val):
        equal_phase = 2
    elif eq(close_val, sma200_val) and gt(sma100_val, close_val):
        equal_phase = 3
    elif eq(sma100_val, sma200_val) and gt(sma100_val, close_val):
        equal_phase = 4
    elif eq(sma100_val, close_val) and gt(sma200_val, sma100_val):
        equal_phase = 5
    elif eq(sma200_val, close_val) and gt(sma200_val, sma100_val):
        equal_phase = 6
    elif eq(sma200_val, sma100_val) and gt(close_val, sma200_val):
        equal_phase = 7

    # 5) 최종값
    final_value = phase_plus_beta + add6 + equal_phase

    # 반올림 처리 제거 - 원본 값 그대로 반환
    return final_value

def calculate_all_indicators_1h4x(df, market_type):
    """
    1H4x 시트용 모든 지표를 한 번에 계산합니다. (정밀도 강화 버전)
    
    [로직 개선]
    - 기존의 문자열 파싱 및 단순 루프 방식을 제거하고, datetime 객체 기반의 딕셔너리 매핑을 사용합니다.
    - SMA100 등 계산 시: 현재 캔들 + 과거 (N/4 - 1)개의 '매 시간 45분' 캔들을 핀포인트로 찾아 계산합니다.
    """
    if df.empty:
        return df
    
    # Date(UTC) 컬럼 정규화 (Timestamp와 문자열 혼합 방지)
    if 'Date(UTC)' in df.columns:
        if df['Date(UTC)'].dtype == 'object':
            df['Date(UTC)'] = df['Date(UTC)'].astype(str).str.strip().str.replace(',', ' ', regex=False)
            df['Date(UTC)'] = pd.to_datetime(df['Date(UTC)'], format='%y/%m/%d %H:%M', errors='coerce')
        else:
            df['Date(UTC)'] = pd.to_datetime(df['Date(UTC)'], format='%y/%m/%d,%H:%M', errors='coerce')
    
    # 1. 데이터 정렬 보장 (최신 -> 과거)
    df = df.sort_values("Date(UTC)", ascending=False).reset_index(drop=True)
    
    # 2. 날짜 파싱 및 인덱스 매핑 (속도 최적화: O(1) 조회)
    # 모든 날짜를 datetime 객체로 변환하여 딕셔너리에 저장
    date_map = {}
    
    # 날짜 파싱 헬퍼 (Timestamp 객체도 처리 가능하도록 수정)
    def parse_dt_safe(date_val):
        try:
            # 이미 Timestamp 객체인 경우 그대로 반환
            if isinstance(date_val, pd.Timestamp):
                return date_val
            elif pd.api.types.is_datetime64_any_dtype(pd.Series([date_val])):
                return pd.to_datetime(date_val)
            # 문자열인 경우 파싱
            clean_str = str(date_val).replace(',', ' ').strip()
            return pd.to_datetime(clean_str, format="%y/%m/%d %H:%M", errors='coerce')
        except:
            return None
    
    # 전체 행에 대해 날짜 파싱
    # Date(UTC)가 이미 datetime64 타입이면 그대로 사용, 아니면 파싱
    if pd.api.types.is_datetime64_any_dtype(df["Date(UTC)"]):
        dates_series = df["Date(UTC)"]
    else:
        dates_series = df["Date(UTC)"].apply(parse_dt_safe)
    
    for idx, dt_val in enumerate(dates_series):
        if pd.notna(dt_val):
            # 초 단위 제거 (정확한 매칭을 위해)
            dt_key = dt_val.replace(second=0, microsecond=0)
            date_map[dt_key] = idx
    
    # 3. SMA 설정
    sma_counts = {
        "SMA12": 3,   # 12/4
        "SMA20": 5,   # 20/4
        "SMA28": 7,   # 28/4
        "SMA40": 10,  # 40/4
        "SMA80": 20,  # 80/4
        "SMA100": 25, # 100/4
        "SMA200": 50  # 200/4
    }
    
    # 결과 저장용 리스트 초기화
    sma_results = {k: [np.nan] * len(df) for k in sma_counts}
    max200_results = [np.nan] * len(df)
    min200_results = [np.nan] * len(df)
    
    # 데이터 미리 가져오기 (NumPy 배열로 변환하여 속도 향상)
    closes = df["종"].values
    opens = df["시"].values
    highs = df["고"].values
    lows = df["저"].values
    
    # 4. 전체 행 순회하며 계산
    for idx in range(len(df)):
        current_dt = dates_series[idx]
        
        if pd.isna(current_dt):
            continue
            
        current_dt = current_dt.replace(second=0, microsecond=0)
        
        # --- 과거 데이터 수집 ---
        # SMA200 기준 최대 50개(현재 포함)가 필요
        # 현재 행 데이터 담기
        collected_indices = [idx] 
        
        target_dt = current_dt
        
        # 최대 49개의 과거 데이터 탐색 (SMA200용)
        for _ in range(49):
            # 1시간 전으로 이동
            target_dt = target_dt - pd.Timedelta(hours=1)
            # 분을 무조건 45분으로 고정
            target_45 = target_dt.replace(minute=45)
            
            # 딕셔너리에서 인덱스 찾기
            if target_45 in date_map:
                found_idx = date_map[target_45]
                # 현재보다 과거 데이터여야 함 (데이터 정렬이 꼬였을 경우 대비)
                if found_idx > idx:
                    collected_indices.append(found_idx)
            else:
                # 데이터가 없으면 건너뜀 (결측)
                pass
                
        # 수집된 인덱스로 값 가져오기
        valid_closes = closes[collected_indices]
        
        # --- SMA 계산 ---
        for sma_name, count in sma_counts.items():
            if len(valid_closes) >= count:
                # 필요한 개수만큼 슬라이싱하여 평균
                sma_results[sma_name][idx] = np.mean(valid_closes[:count])
        
        # --- Max200, Min200 계산 ---
        # 수집된 캔들이 4개 이상일 때만 계산
        if len(collected_indices) >= 4:
            # 수집된 모든 인덱스의 OHLC 값을 모음
            # (해당 시간대의 고가, 저가 등을 모두 포함해야 함)
            relevant_indices = collected_indices # 이미 수집된 인덱스들
            
            # 벡터화된 연산으로 Max/Min 찾기
            batch_highs = highs[relevant_indices]
            batch_lows = lows[relevant_indices]
            batch_opens = opens[relevant_indices]
            batch_closes = closes[relevant_indices]
            
            curr_max = max(batch_highs.max(), batch_opens.max(), batch_closes.max()) # 보통 고가가 Max
            curr_min = min(batch_lows.min(), batch_opens.min(), batch_closes.min()) # 보통 저가가 Min
            
            max200_results[idx] = curr_max
            min200_results[idx] = curr_min
    
    # 5. 결과 DataFrame에 할당
    for sma_name in sma_counts:
        df[sma_name] = sma_results[sma_name]
        
    df["Max200"] = max200_results
    df["Min200"] = min200_results
    
    # 하단, 상단 계산 (Max200/Min200이 NaN이면 NaN)
    df["하단"] = df.apply(lambda row: abs((row["종"] - row["Min200"]) / row["Min200"]) if not pd.isna(row["Min200"]) else np.nan, axis=1)
    df["상단"] = df.apply(lambda row: abs((row["종"] - row["Max200"]) / row["Max200"]) if not pd.isna(row["Max200"]) else np.nan, axis=1)
    
    # SFast 계산 (SMA12, SMA20, SMA28 사용)
    df["SFast"] = df.apply(lambda row: calculate_superfast(row["SMA12"], row["SMA20"], row["SMA28"]), axis=1)
    
    # Fast 계산 (SMA20, SMA28, SMA40 사용)
    df["Fast"] = df.apply(lambda row: calculate_fast(row["SMA20"], row["SMA28"], row["SMA40"]), axis=1)
    
    # Base 계산 (SMA28, SMA40, SMA80 사용)
    df["Base"] = df.apply(lambda row: calculate_base(row["SMA28"], row["SMA40"], row["SMA80"]), axis=1)
    
    # 4or1 계산
    df["4or1"] = df.apply(lambda row: calculate_4or1(row["하단"], row["상단"]), axis=1)
    
    # buyside 계산
    df["buyside"] = df.apply(lambda row: calculate_buyside(row["SFast"], row["Fast"], row["Base"]), axis=1)
    
    # sellside 계산
    df["sellside"] = df.apply(lambda row: calculate_sellside(row["SFast"], row["Fast"], row["Base"]), axis=1)
    
    # Buy 계산 (1H4x 시트용: sellside <= 0.1)
    df["Buy"] = df.apply(lambda row: calculate_buy_1h4x(row["4or1"], row["sellside"]), axis=1)
    
    # Sell 계산 (1H4x 시트용: buyside <= 0.1)
    df["Sell"] = df.apply(lambda row: calculate_sell_short_1h4x(row["buyside"]), axis=1)
    
    # 1HMSFast 계산 (종가, SMA100, SMA200 사용)
    df["1HMSFast"] = df.apply(lambda row: calculate_1hmsfast_1h4x(row["종"], row["SMA100"], row["SMA200"]), axis=1)
    
    # 숫자 컬럼 정리
    num_cols = ["종", "시", "고", "저", "Vol.", "SMA12", "SMA20", "SMA28", "SMA40", "SMA80", "SMA100", "SMA200", "Max200", "Min200", "하단", "상단", "SFast", "Fast", "Base", "4or1", "buyside", "sellside", "1HMSFast"]
    df[num_cols] = df[num_cols].apply(pd.to_numeric, errors="coerce")
    
    return df

def calculate_latest_row_only_1h4x(df, market_type):
    """
    1H4x 시트 After 단계 최적화: 최신 1개 행만 지표 계산 (previous 지표 유지)
    
    입력: [새 데이터(idx=0), Previous(idx=1~)] (최신→과거 순서)
    출력: [새 데이터(지표 계산됨), Previous(그대로)] (최신→과거 순서 유지)
    
    계산 방식:
    - SMA 계산: idx=0만 계산, 과거 xx:45 캔들 사용
    - 모든 지표는 idx=0 + previous 데이터를 사용
    """
    if df.empty:
        return df
    
    # [Cursor 패치] 엑셀에서 읽은 데이터의 쉼표 제거 및 숫자/날짜 강제 변환
    # 이 과정이 없으면 데이터가 제대로 읽히지 않아 len(df)가 줄어들고 fallback으로 튕김
    df = clean_df_display_format(df)
    
    # Date(UTC) 컬럼 정규화 (Timestamp와 문자열 혼합 방지)
    if 'Date(UTC)' in df.columns:
        if df['Date(UTC)'].dtype == 'object':
            df['Date(UTC)'] = df['Date(UTC)'].astype(str).str.strip().str.replace(',', ' ', regex=False)
            df['Date(UTC)'] = pd.to_datetime(df['Date(UTC)'], format='%y/%m/%d %H:%M', errors='coerce')
        else:
            df['Date(UTC)'] = pd.to_datetime(df['Date(UTC)'], format='%y/%m/%d,%H:%M', errors='coerce')
    
    # ⚠️중요: 입력 데이터를 확실하게 현재→과거 순서로 정렬 (SMA 계산 일관성 보장)
    df = df.sort_values("Date(UTC)", ascending=False).reset_index(drop=True)
    
    # 데이터가 조금이라도 있으면 최대한 latest_row_only 모드 유지
    # 50개 미만이어도 새 데이터만 계산 시도 (previous 데이터가 있으면 활용)
    if len(df) < 12:
        # 최소한 SMA12 계산을 위해 12개는 필요
        return calculate_all_indicators_1h4x(df, market_type)
    
    # Buy/Sell 컬럼이 없으면 생성
    if "Buy" not in df.columns:
        df["Buy"] = ""
    if "Sell" not in df.columns:
        df["Sell"] = ""
    
    # Previous 데이터의 Buy/Sell이 NaN이면 빈 문자열로 변환
    df["Buy"] = df["Buy"].fillna("")
    df["Sell"] = df["Sell"].fillna("")
    
    # 날짜와 시간 파싱 헬퍼 함수
    def parse_datetime(date_str):
        """Date(UTC) 문자열에서 날짜와 시간 추출 (예: "25/01/15,11:15" -> (date_obj, 11, 15))
        ⚠️중요: Timestamp 객체를 최우선으로 인식하여 데이터 파괴 방지
        """
        try:
            # 1. Timestamp 객체인 경우 직접 처리 (최우선)
            if isinstance(date_str, pd.Timestamp):
                return date_str, date_str.hour, date_str.minute
            
            # 2. datetime 객체인 경우 (datetime.datetime, numpy.datetime64 등)
            if hasattr(date_str, 'hour') and hasattr(date_str, 'minute'):
                try:
                    return date_str, date_str.hour, date_str.minute
                except:
                    pass
            
            # 3. datetime64 타입인 경우
            if pd.api.types.is_datetime64_any_dtype(pd.Series([date_str])):
                dt_obj = pd.to_datetime(date_str)
                return dt_obj, dt_obj.hour, dt_obj.minute
            
            # 4. 문자열인 경우에만 파싱 시도
            date_str = str(date_str).strip()
            if not date_str or date_str.lower() in ['nan', 'nat', 'none', '']:
                return None, None, None
            
            # 콤마가 있는 경우 (YY/MM/DD,HH:MM 형식)
            if ',' in date_str:
                date_part, time_part = date_str.split(',', 1)
                if ':' in time_part:
                    try:
                        hour, minute = map(int, time_part.split(':'))
                        # 날짜 파싱 (YY/MM/DD 형식)
                        date_obj = pd.to_datetime(date_part, format="%y/%m/%d", errors='coerce')
                        if pd.notna(date_obj):
                            return date_obj, hour, minute
                    except:
                        pass
            
            # 콤마가 없는 경우 또는 위에서 실패한 경우
            # format 없이 자동 인식으로 파싱 (데이터 삭제 방지)
            try:
                date_obj = pd.to_datetime(date_str, errors='coerce')
                if pd.notna(date_obj):
                    return date_obj, date_obj.hour, date_obj.minute
            except Exception as e:
                pass
        except Exception as e:
            pass
        return None, None, None
    
    # idx=0만 계산
    idx = 0
    
    # SMA 계산: idx=0만 계산, 과거 xx:45 캔들 사용
    current_time_str = df.iloc[idx]["Date(UTC)"]
    current_date, current_hour, current_minute = parse_datetime(current_time_str)
    
    if pd.notna(current_date) and current_hour is not None and current_minute is not None:
        # SMA별 필요한 캔들 개수 (SMA값/4)
        sma_counts = {
            "SMA12": 3,   # 12/4
            "SMA20": 5,   # 20/4
            "SMA28": 7,   # 28/4
            "SMA40": 10,  # 40/4
            "SMA80": 20,  # 80/4
            "SMA100": 25, # 100/4
            "SMA200": 50  # 200/4
        }
        
        # 각 SMA 계산
        for sma_name, count in sma_counts.items():
            selected_closes = []
            
            # 1. 현재 행의 종가 추가 (마지막 시간)
            selected_closes.append(df.iloc[idx]["종"])
            
            # 2. 과거로 가면서 xx:45 캔들만 선택
            search_idx = idx + 1  # 다음 행부터 검색
            target_date = current_date
            target_hour = current_hour
            
            # 시간을 과거로 이동 (1시간씩)
            while len(selected_closes) < count and search_idx < len(df):
                # 목표 시간 계산 (1시간 전)
                target_hour -= 1
                if target_hour < 0:
                    target_hour = 23
                    # 날짜도 하루 전으로 이동
                    target_date = target_date - pd.Timedelta(days=1)
                
                # 해당 시간대의 45분 캔들 찾기
                found = False
                temp_idx = search_idx
                while temp_idx < len(df) and not found:
                    time_str = df.iloc[temp_idx]["Date(UTC)"]
                    row_date, hour, minute = parse_datetime(time_str)
                    
                    # NaT 체크 추가 (pd.notna 사용)
                    if pd.notna(row_date) and hour is not None and minute is not None:
                        # 날짜와 시간 모두 일치하는지 확인
                        if row_date.date() == target_date.date() and hour == target_hour and minute == 45:
                            selected_closes.append(df.iloc[temp_idx]["종"])
                            found = True
                            search_idx = temp_idx + 1  # 다음 검색 시작 위치
                        elif row_date.date() < target_date.date() or (row_date.date() == target_date.date() and (hour < target_hour or (hour == target_hour and minute < 45))):
                            # 시간이 지나갔으면 다음 시간대로
                            break
                    temp_idx += 1
                
                if not found:
                    # 해당 시간대의 45분 캔들을 찾지 못하면 종료
                    break
            
            # 평균 계산
            if len(selected_closes) == count:
                df.loc[idx, sma_name] = np.mean(selected_closes)
            else:
                df.loc[idx, sma_name] = np.nan
        
        # Max200, Min200 계산: 200/4 = 50개 캔들 사용 (마지막 시간 + 과거 49개의 xx:45 캔들)
        selected_values = []
        
        # 1. 현재 행의 시고저종 추가 (마지막 시간)
        selected_values.extend([
            df.iloc[idx]["시"],
            df.iloc[idx]["고"],
            df.iloc[idx]["저"],
            df.iloc[idx]["종"]
        ])
        
        # 2. 과거로 가면서 xx:45 캔들의 시고저종 선택
        search_idx = idx + 1
        target_date = current_date
        target_hour = current_hour
        
        while len(selected_values) < 50 * 4 and search_idx < len(df):  # 50개 캔들 * 4개 값(시고저종)
            # 목표 시간 계산 (1시간 전)
            target_hour -= 1
            if target_hour < 0:
                target_hour = 23
                # 날짜도 하루 전으로 이동
                target_date = target_date - pd.Timedelta(days=1)
            
            # 해당 시간대의 45분 캔들 찾기
            found = False
            temp_idx = search_idx
            while temp_idx < len(df) and not found:
                time_str = df.iloc[temp_idx]["Date(UTC)"]
                row_date, hour, minute = parse_datetime(time_str)
                
                if pd.notna(row_date) and hour is not None and minute is not None:
                    # 날짜와 시간 모두 일치하는지 확인
                    if row_date.date() == target_date.date() and hour == target_hour and minute == 45:
                        selected_values.extend([
                            df.iloc[temp_idx]["시"],
                            df.iloc[temp_idx]["고"],
                            df.iloc[temp_idx]["저"],
                            df.iloc[temp_idx]["종"]
                        ])
                        found = True
                        search_idx = temp_idx + 1
                    elif row_date.date() < target_date.date() or (row_date.date() == target_date.date() and (hour < target_hour or (hour == target_hour and minute < 45))):
                        break
                temp_idx += 1
            
            if not found:
                break
        
        # Max200, Min200 계산
        if len(selected_values) >= 4:  # 최소 1개 캔들 이상
            df.loc[idx, "Max200"] = max(selected_values)
            df.loc[idx, "Min200"] = min(selected_values)
        else:
            df.loc[idx, "Max200"] = np.nan
            df.loc[idx, "Min200"] = np.nan
    else:
        # 시간 파싱 실패 시 NaN
        df.loc[idx, "SMA12"] = np.nan
        df.loc[idx, "SMA20"] = np.nan
        df.loc[idx, "SMA28"] = np.nan
        df.loc[idx, "SMA40"] = np.nan
        df.loc[idx, "SMA80"] = np.nan
        df.loc[idx, "SMA100"] = np.nan
        df.loc[idx, "SMA200"] = np.nan
        df.loc[idx, "Max200"] = np.nan
        df.loc[idx, "Min200"] = np.nan
    
    # 하단, 상단 계산 (Max200/Min200이 NaN이면 NaN)
    if not pd.isna(df.loc[idx, "Min200"]):
        df.loc[idx, "하단"] = abs((df.loc[idx, "종"] - df.loc[idx, "Min200"]) / df.loc[idx, "Min200"])
    else:
        df.loc[idx, "하단"] = np.nan
    
    if not pd.isna(df.loc[idx, "Max200"]):
        df.loc[idx, "상단"] = abs((df.loc[idx, "종"] - df.loc[idx, "Max200"]) / df.loc[idx, "Max200"])
    else:
        df.loc[idx, "상단"] = np.nan
    
    # SFast 계산 (SMA12, SMA20, SMA28 사용)
    df.loc[idx, "SFast"] = calculate_superfast(df.loc[idx, "SMA12"], df.loc[idx, "SMA20"], df.loc[idx, "SMA28"])
    
    # Fast 계산 (SMA20, SMA28, SMA40 사용)
    df.loc[idx, "Fast"] = calculate_fast(df.loc[idx, "SMA20"], df.loc[idx, "SMA28"], df.loc[idx, "SMA40"])
    
    # Base 계산 (SMA28, SMA40, SMA80 사용)
    df.loc[idx, "Base"] = calculate_base(df.loc[idx, "SMA28"], df.loc[idx, "SMA40"], df.loc[idx, "SMA80"])
    
    # 4or1 계산
    df.loc[idx, "4or1"] = calculate_4or1(df.loc[idx, "하단"], df.loc[idx, "상단"])
    
    # buyside 계산
    df.loc[idx, "buyside"] = calculate_buyside(df.loc[idx, "SFast"], df.loc[idx, "Fast"], df.loc[idx, "Base"])
    
    # sellside 계산
    df.loc[idx, "sellside"] = calculate_sellside(df.loc[idx, "SFast"], df.loc[idx, "Fast"], df.loc[idx, "Base"])
    
    # Buy 계산 (1H4x 시트용: sellside <= 0.1)
    df.loc[idx, "Buy"] = calculate_buy_1h4x(df.loc[idx, "4or1"], df.loc[idx, "sellside"])
    
    # Sell 계산 (1H4x 시트용: buyside <= 0.1)
    df.loc[idx, "Sell"] = calculate_sell_short_1h4x(df.loc[idx, "buyside"])
    
    # 1HMSFast 계산 (종가, SMA100, SMA200 사용)
    df.loc[idx, "1HMSFast"] = calculate_1hmsfast_1h4x(df.loc[idx, "종"], df.loc[idx, "SMA100"], df.loc[idx, "SMA200"])
    
    # 숫자 컬럼 정리
    num_cols = ["종", "시", "고", "저", "Vol.", "SMA12", "SMA20", "SMA28", "SMA40", "SMA80", "SMA100", "SMA200", "Max200", "Min200", "하단", "상단", "SFast", "Fast", "Base", "4or1", "buyside", "sellside", "1HMSFast"]
    for col in num_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    
    return df

def calculate_all_indicators_1m(df, market_type):
    """
    1분봉용 모든 지표를 계산합니다. (Max400/Min400 사용)
    SMA: 15, 25, 35, 50, 100
    sfast: 15, 25, 35
    fast: 25, 35, 50
    base: 35, 50, 100
    """
    if df.empty:
        return df
    
    # ⚠️중요: 엑셀에서 읽은 데이터의 숫자 컬럼을 강제로 숫자로 변환 (쉼표 제거 후 변환)
    # 엑셀에서 불러온 숫자가 문자열("88,123.45")로 저장되어 있을 수 있음
    numeric_cols = ['종', '시', '고', '저', 'Vol.']
    for col in numeric_cols:
        if col in df.columns:
            # 문자열인 경우에만 쉼표 제거 시도
            if df[col].dtype == 'object':
                df[col] = df[col].astype(str).str.replace(',', '')
            # 강제 숫자 변환
            df[col] = pd.to_numeric(df[col], errors='coerce')
    
    # Date(UTC) 컬럼 정규화 (Timestamp와 문자열 혼합 방지)
    if 'Date(UTC)' in df.columns:
        if df['Date(UTC)'].dtype == 'object':
            # format 명시하여 파싱 시도 (연도/일 혼동 방지)
            try:
                df['Date(UTC)'] = pd.to_datetime(df['Date(UTC)'], format='%y/%m/%d,%H:%M', errors='coerce')
            except:
                # 쉼표 제거 후 형식 시도
                try:
                    df['Date(UTC)'] = df['Date(UTC)'].astype(str).str.replace(',', ' ', regex=False).str.strip()
                    df['Date(UTC)'] = pd.to_datetime(df['Date(UTC)'], format='%y/%m/%d %H:%M', errors='coerce')
                except:
                    # fallback: format 없이 파싱
                    df['Date(UTC)'] = pd.to_datetime(df['Date(UTC)'], errors='coerce')
        else:
            # 이미 datetime 타입이면 그대로 사용
            pass
    
    # ⚠️중요: 입력 데이터를 확실하게 현재→과거 순서로 정렬 (SMA 계산 일관성 보장)
    # 데이터 수집 과정에서 정렬이 여러 번 섞일 수 있으므로, 계산 직전에 확실하게 정렬
    df = df.sort_values("Date(UTC)", ascending=False).reset_index(drop=True)
    
    # SMA 계산: 각 행(idx)에서 그 행부터 앞으로(과거로) window개까지의 평균
    for idx in range(len(df)):
        df.loc[idx, "SMA15"] = df.iloc[idx:idx+15]["종"].mean() if idx + 15 <= len(df) else np.nan
        df.loc[idx, "SMA25"] = df.iloc[idx:idx+25]["종"].mean() if idx + 25 <= len(df) else np.nan
        df.loc[idx, "SMA35"] = df.iloc[idx:idx+35]["종"].mean() if idx + 35 <= len(df) else np.nan
        df.loc[idx, "SMA50"] = df.iloc[idx:idx+50]["종"].mean() if idx + 50 <= len(df) else np.nan
        df.loc[idx, "SMA100"] = df.iloc[idx:idx+100]["종"].mean() if idx + 100 <= len(df) else np.nan
    
    # Max400, Min400 계산: 각 행(idx)에서 그 행부터 앞으로(과거로) 400개까지의 최고가/최저가
    for idx in range(len(df)):
        if idx + 400 <= len(df):
            window_data = df.iloc[idx:idx+400][["시", "고", "저", "종"]]
            df.loc[idx, "Max400"] = window_data.values.max()
            df.loc[idx, "Min400"] = window_data.values.min()
        else:
            df.loc[idx, "Max400"] = np.nan
            df.loc[idx, "Min400"] = np.nan
    
    # 하단, 상단 계산 (Max400/Min400이 NaN이면 NaN)
    df["하단"] = df.apply(lambda row: abs((row["종"] - row["Min400"]) / row["Min400"]) if not pd.isna(row["Min400"]) else np.nan, axis=1)
    df["상단"] = df.apply(lambda row: abs((row["종"] - row["Max400"]) / row["Max400"]) if not pd.isna(row["Max400"]) else np.nan, axis=1)
    
    # SFast 계산 (SMA15, SMA25, SMA35 사용)
    df["SFast"] = df.apply(lambda row: calculate_superfast(row["SMA15"], row["SMA25"], row["SMA35"]), axis=1)
    
    # Fast 계산 (SMA25, SMA35, SMA50 사용)
    df["Fast"] = df.apply(lambda row: calculate_fast(row["SMA25"], row["SMA35"], row["SMA50"]), axis=1)
    
    # Base 계산 (SMA35, SMA50, SMA100 사용)
    df["Base"] = df.apply(lambda row: calculate_base(row["SMA35"], row["SMA50"], row["SMA100"]), axis=1)
    
    # 4or1 계산
    df["4or1"] = df.apply(lambda row: calculate_4or1(row["하단"], row["상단"]), axis=1)
    
    # buyside 계산
    df["buyside"] = df.apply(lambda row: calculate_buyside(row["SFast"], row["Fast"], row["Base"]), axis=1)
    
    # sellside 계산
    df["sellside"] = df.apply(lambda row: calculate_sellside(row["SFast"], row["Fast"], row["Base"]), axis=1)
    
    # Buy 계산
    df["Buy"] = df.apply(lambda row: calculate_buy(row["4or1"], row["sellside"]), axis=1)
    
    # Sell 계산 (4or1 없이 buyside만 사용)
    df["Sell"] = df.apply(lambda row: calculate_sell_short(row["buyside"]), axis=1)
    
    # 최신→과거 순서로 다시 정렬
    # 이미 최신→과거 순서이므로 재정렬 불필요 (정렬은 위에서 이미 완료)
    
    return df

def calculate_latest_row_only_1m(df, market_type):
    """
    1분봉 After 단계 최적화: 최신 행만 지표 계산 (previous 지표 유지)
    
    입력: [새 데이터(idx=0), Previous(idx=1~)] (최신→과거 순서)
    출력: [새 데이터(지표 계산됨), Previous(그대로)] (최신→과거 순서 유지)
    
    계산 방식:
    - SMA: 15, 25, 35, 50, 100
    - Max400, Min400
    - SFast, Fast, Base
    - 모든 지표는 새 데이터(idx=0) + previous 데이터를 사용
    """
    if df.empty:
        return df
    
    # [Cursor 패치] 엑셀에서 읽은 데이터의 쉼표 제거 및 숫자/날짜 강제 변환
    # 이 과정이 없으면 데이터가 제대로 읽히지 않아 len(df)가 줄어들고 fallback으로 튕김
    df = clean_df_display_format(df)
    
    # Date(UTC) 컬럼 정규화 (Timestamp와 문자열 혼합 방지)
    if 'Date(UTC)' in df.columns:
        if df['Date(UTC)'].dtype == 'object':
            df['Date(UTC)'] = df['Date(UTC)'].astype(str).str.strip().str.replace(',', ' ', regex=False)
            df['Date(UTC)'] = pd.to_datetime(df['Date(UTC)'], format='%y/%m/%d %H:%M', errors='coerce')
        else:
            df['Date(UTC)'] = pd.to_datetime(df['Date(UTC)'], format='%y/%m/%d,%H:%M', errors='coerce')
    
    # ⚠️중요: 입력 데이터를 확실하게 현재→과거 순서로 정렬 (SMA 계산 일관성 보장)
    df = df.sort_values("Date(UTC)", ascending=False).reset_index(drop=True)
    
    # SMA100 계산을 위해 최소 100개 필요
    # 하지만 데이터가 조금이라도 있으면 최대한 latest_row_only 모드 유지
    # 100개 미만이어도 새 데이터만 계산 시도 (previous 데이터가 있으면 활용)
    if len(df) < 15:
        # 최소한 SMA15 계산을 위해 15개는 필요
        return calculate_all_indicators_1m(df, market_type)
    
    # Buy/Sell 컬럼이 없으면 생성
    if "Buy" not in df.columns:
        df["Buy"] = ""
    if "Sell" not in df.columns:
        df["Sell"] = ""
    
    # Previous 데이터의 Buy/Sell이 NaN이면 빈 문자열로 변환
    df["Buy"] = df["Buy"].fillna("")
    df["Sell"] = df["Sell"].fillna("")
    
    # ⚠️중요: 새 데이터(지표가 NaN인 행)를 모두 계산
    # previous 데이터는 이미 지표가 계산되어 있으므로 건드리지 않음
    new_data_indices = []
    if "SMA100" in df.columns:
        new_data_indices = df[df["SMA100"].isna()].index.tolist()
    elif "SMA15" in df.columns:
        # SMA100이 없으면 SMA15로 판단
        new_data_indices = df[df["SMA15"].isna()].index.tolist()
    else:
        # 지표 컬럼이 없으면 최신 1개만 계산 (하위 호환성)
        new_data_indices = [0]
    
    # 새 데이터가 없으면 그대로 반환
    if not new_data_indices:
        return df
    
    # 새 데이터의 각 행에 대해 지표 계산
    for idx in new_data_indices:
        # SMA 계산: idx 포함하여 계산
        df.loc[idx, "SMA15"] = df.iloc[idx:idx+15]["종"].mean() if idx + 15 <= len(df) else np.nan
        df.loc[idx, "SMA25"] = df.iloc[idx:idx+25]["종"].mean() if idx + 25 <= len(df) else np.nan
        df.loc[idx, "SMA35"] = df.iloc[idx:idx+35]["종"].mean() if idx + 35 <= len(df) else np.nan
        df.loc[idx, "SMA50"] = df.iloc[idx:idx+50]["종"].mean() if idx + 50 <= len(df) else np.nan
        df.loc[idx, "SMA100"] = df.iloc[idx:idx+100]["종"].mean() if idx + 100 <= len(df) else np.nan
        
        # Max400, Min400 계산: idx 포함 400개 캔들
        if idx + 400 <= len(df):
            window_data = df.iloc[idx:idx+400][["시", "고", "저", "종"]]
            df.loc[idx, "Max400"] = window_data.values.max()
            df.loc[idx, "Min400"] = window_data.values.min()
        else:
            df.loc[idx, "Max400"] = np.nan
            df.loc[idx, "Min400"] = np.nan
        
        # 하단, 상단 계산
        current_price = df.loc[idx, "종"]
        min400 = df.loc[idx, "Min400"]
        max400 = df.loc[idx, "Max400"]
        
        if pd.notna(min400) and min400 != 0:
            df.loc[idx, "하단"] = abs((current_price - min400) / min400)
        else:
            df.loc[idx, "하단"] = np.nan
            
        if pd.notna(max400) and max400 != 0:
            df.loc[idx, "상단"] = abs((current_price - max400) / max400)
        else:
            df.loc[idx, "상단"] = np.nan
        
        # SFast, Fast, Base 계산
        df.loc[idx, "SFast"] = calculate_superfast(df.loc[idx, "SMA15"], df.loc[idx, "SMA25"], df.loc[idx, "SMA35"])
        df.loc[idx, "Fast"] = calculate_fast(df.loc[idx, "SMA25"], df.loc[idx, "SMA35"], df.loc[idx, "SMA50"])
        df.loc[idx, "Base"] = calculate_base(df.loc[idx, "SMA35"], df.loc[idx, "SMA50"], df.loc[idx, "SMA100"])
        
        # 4or1 계산
        df.loc[idx, "4or1"] = calculate_4or1(df.loc[idx, "하단"], df.loc[idx, "상단"])
        
        # buyside, sellside 계산
        df.loc[idx, "buyside"] = calculate_buyside(df.loc[idx, "SFast"], df.loc[idx, "Fast"], df.loc[idx, "Base"])
        df.loc[idx, "sellside"] = calculate_sellside(df.loc[idx, "SFast"], df.loc[idx, "Fast"], df.loc[idx, "Base"])
        
        # Buy, Sell 계산
        df.loc[idx, "Buy"] = calculate_buy(df.loc[idx, "4or1"], df.loc[idx, "sellside"])
        df.loc[idx, "Sell"] = calculate_sell_short(df.loc[idx, "buyside"])
    
    return df

def calculate_sb1m_for_15m(df_15m, df_1m):
    """
    1분봉 데이터를 기반으로 15분봉에 SB1M 신호를 추가합니다.
    1분봉 15개씩 묶어서 buy1~buy15, sell1~sell15까지 계산합니다.
    """
    if df_1m.empty or df_15m.empty:
        return df_15m
    
    # 원본 보존
    df_1m = df_1m.copy()
    df_15m = df_15m.copy()
    
    if 'Date(UTC)' not in df_1m.columns or 'Date(UTC)' not in df_15m.columns:
        if 'SB1M' not in df_15m.columns:
            df_15m['SB1M'] = ''
        return df_15m
    
    # 날짜 파싱 (이미 datetime64면 그대로 사용, 문자열만 파싱)
    if pd.api.types.is_datetime64_any_dtype(df_1m['Date(UTC)']):
        df_1m['Date(UTC)_dt'] = df_1m['Date(UTC)']
    else:
        # 문자열인 경우에만 파싱 (UserWarning 억제)
        s = df_1m['Date(UTC)'].astype(str).str.strip().str.replace(',', ' ', regex=False)
        import warnings
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", UserWarning)
            df_1m['Date(UTC)_dt'] = pd.to_datetime(s, errors='coerce')
    df_1m = df_1m[df_1m['Date(UTC)_dt'].notna()].copy()
    
    if df_1m.empty:
        if 'SB1M' not in df_15m.columns:
            df_15m['SB1M'] = ''
        return df_15m
    
    # 시간 정렬
    df_1m = df_1m.sort_values('Date(UTC)_dt', ascending=True).reset_index(drop=True)
    df_1m['15min_group'] = df_1m['Date(UTC)_dt'].dt.floor('15min')
    
    # 15분봉 날짜 파싱 (이미 datetime64면 그대로 사용, 문자열만 파싱)
    if pd.api.types.is_datetime64_any_dtype(df_15m['Date(UTC)']):
        df_15m['Date(UTC)_dt'] = df_15m['Date(UTC)']
    else:
        # 문자열인 경우에만 파싱 (UserWarning 억제)
        s = df_15m['Date(UTC)'].astype(str).str.strip().str.replace(',', ' ', regex=False)
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", UserWarning)
            df_15m['Date(UTC)_dt'] = pd.to_datetime(s, errors='coerce')
    
    if '15min_group' not in df_15m.columns:
        df_15m['15min_group'] = df_15m['Date(UTC)_dt'].dt.floor('15min')
    else:
        mask_na = df_15m['15min_group'].isna()
        if mask_na.any():
            df_15m.loc[mask_na, '15min_group'] = df_15m.loc[mask_na, 'Date(UTC)_dt'].dt.floor('15min')
    
    # --- 15분 그룹별 SB1M 신호 계산 ---
    sb1m_data = []
    for group_time, group_df in df_1m.groupby('15min_group', sort=True):
        buy_count = 0
        sell_count = 0
        
        for _, row in group_df.iterrows():
            buy_val = str(row.get('Buy', '')).strip().lower()
            sell_val = str(row.get('Sell', '')).strip().lower()
            
            if buy_val == 'buy':
                buy_count += 1
            elif sell_val == 'sell':
                sell_count += 1
        
        # 우선순위 로직: buy15 > buy14 > ... > buy01 > sell15 > sell14 > ... > sell01
        if buy_count == 15: sb1m_signal = 'buy15'
        elif buy_count == 14: sb1m_signal = 'buy14'
        elif buy_count == 13: sb1m_signal = 'buy13'
        elif buy_count == 12: sb1m_signal = 'buy12'
        elif buy_count == 11: sb1m_signal = 'buy11'
        elif buy_count == 10: sb1m_signal = 'buy10'
        elif buy_count == 9: sb1m_signal = 'buy09'
        elif buy_count == 8: sb1m_signal = 'buy08'
        elif buy_count == 7: sb1m_signal = 'buy07'
        elif buy_count == 6: sb1m_signal = 'buy06'
        elif buy_count == 5: sb1m_signal = 'buy05'
        elif buy_count == 4: sb1m_signal = 'buy04'
        elif buy_count == 3: sb1m_signal = 'buy03'
        elif buy_count == 2: sb1m_signal = 'buy02'
        elif buy_count == 1: sb1m_signal = 'buy01'
        elif sell_count == 15: sb1m_signal = 'sell15'
        elif sell_count == 14: sb1m_signal = 'sell14'
        elif sell_count == 13: sb1m_signal = 'sell13'
        elif sell_count == 12: sb1m_signal = 'sell12'
        elif sell_count == 11: sb1m_signal = 'sell11'
        elif sell_count == 10: sb1m_signal = 'sell10'
        elif sell_count == 9: sb1m_signal = 'sell09'
        elif sell_count == 8: sb1m_signal = 'sell08'
        elif sell_count == 7: sb1m_signal = 'sell07'
        elif sell_count == 6: sb1m_signal = 'sell06'
        elif sell_count == 5: sb1m_signal = 'sell05'
        elif sell_count == 4: sb1m_signal = 'sell04'
        elif sell_count == 3: sb1m_signal = 'sell03'
        elif sell_count == 2: sb1m_signal = 'sell02'
        elif sell_count == 1: sb1m_signal = 'sell01'
        else: sb1m_signal = ''
        
        sb1m_data.append({
            '15min_group': group_time,
            'SB1M': sb1m_signal
        })
    
    if sb1m_data:
        sb1m_df = pd.DataFrame(sb1m_data)
        sb1m_map = dict(zip(sb1m_df['15min_group'], sb1m_df['SB1M']))
        if 'SB1M' not in df_15m.columns:
            df_15m['SB1M'] = ''
        df_15m['SB1M'] = df_15m['15min_group'].map(sb1m_map).fillna('')
    else:
        if 'SB1M' not in df_15m.columns:
            df_15m['SB1M'] = ''
    
    # 임시 컬럼 정리
    cols_to_drop = ['Date(UTC)_dt', '15min_group']
    df_15m.drop(columns=[c for c in cols_to_drop if c in df_15m.columns], inplace=True)
    
    return df_15m

def calculate_sb1m_for_5m(df_5m, df_1m):
    """
    1분봉 데이터를 기반으로 5분봉에 SB1M 신호를 추가합니다.
    1분봉 5개씩 그룹화해서 Buy/Sell 개수를 세어 판정합니다.
    Buy1~Buy5, Sell1~Sell5까지 사용합니다.
    ⚠️중요: Date(UTC) 기준으로만 수행 (KST는 기록용일 뿐, 계산에는 사용하지 않음)
    """
    if df_1m.empty or df_5m.empty:
        return df_5m
    
    # 원본 보존
    df_1m = df_1m.copy()
    df_5m = df_5m.copy()
    
    # ⚠️중요: 모든 그룹화와 정렬은 Date(UTC) 기준으로만 수행 (KST는 기록용일 뿐, 정렬/그룹화에는 사용하지 않음)
    # Date(UTC) 기준 그룹화
    if 'Date(UTC)' not in df_1m.columns or 'Date(UTC)' not in df_5m.columns:
        if 'SB1M' not in df_5m.columns:
            df_5m['SB1M'] = ''
        return df_5m
    
    # 날짜 파싱 (이미 datetime64면 그대로 사용, 문자열만 파싱)
    if pd.api.types.is_datetime64_any_dtype(df_1m['Date(UTC)']):
        df_1m['Date(UTC)_dt'] = df_1m['Date(UTC)']
    else:
        # 문자열인 경우에만 파싱 (UserWarning 억제)
        s = df_1m['Date(UTC)'].astype(str).str.strip().str.replace(',', ' ', regex=False)
        import warnings
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", UserWarning)
            df_1m['Date(UTC)_dt'] = pd.to_datetime(s, errors='coerce')
    df_1m = df_1m[df_1m['Date(UTC)_dt'].notna()].copy()
    
    if df_1m.empty:
        if 'SB1M' not in df_5m.columns:
            df_5m['SB1M'] = ''
        if 'Date(UTC)_dt' in df_1m.columns:
            df_1m = df_1m.drop('Date(UTC)_dt', axis=1)
        return df_5m
    
    # 시간 정렬 (과거 → 현재)
    df_1m = df_1m.sort_values('Date(UTC)_dt', ascending=True).reset_index(drop=True)
    
    # 1분봉: 5분 그룹 생성
    df_1m['5min_group'] = df_1m['Date(UTC)_dt'].dt.floor('5min')
    
    # 5분봉: 5min_group 생성 (이미 datetime64면 그대로 사용, 문자열만 파싱)
    if pd.api.types.is_datetime64_any_dtype(df_5m['Date(UTC)']):
        df_5m['Date(UTC)_dt'] = df_5m['Date(UTC)']
    else:
        # 문자열인 경우에만 파싱 (UserWarning 억제)
        s = df_5m['Date(UTC)'].astype(str).str.strip().str.replace(',', ' ', regex=False)
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", UserWarning)
            df_5m['Date(UTC)_dt'] = pd.to_datetime(s, errors='coerce')
    
    if '5min_group' not in df_5m.columns:
        df_5m['5min_group'] = df_5m['Date(UTC)_dt'].dt.floor('5min')
    else:
        mask_na = df_5m['5min_group'].isna()
        if mask_na.any():
            df_5m.loc[mask_na, '5min_group'] = df_5m.loc[mask_na, 'Date(UTC)_dt'].dt.floor('5min')
    
    # --- 5분 그룹별 SB1M 신호 계산 ---
    sb1m_data = []
    for group_time, group_df in df_1m.groupby('5min_group', sort=True):
        buy_count = 0
        sell_count = 0
        
        for _, row in group_df.iterrows():
            buy_val = row.get('Buy', '')
            sell_val = row.get('Sell', '')
            
            if pd.notna(buy_val) and isinstance(buy_val, str) and buy_val.strip().lower() == 'buy':
                buy_count += 1
            elif pd.notna(sell_val) and isinstance(sell_val, str) and sell_val.strip().lower() == 'sell':
                sell_count += 1
        
        if buy_count == 5:
            sb1m_signal = 'buy5'
        elif buy_count == 4:
            sb1m_signal = 'buy4'
        elif buy_count == 3:
            sb1m_signal = 'buy3'
        elif buy_count == 2:
            sb1m_signal = 'buy2'
        elif buy_count == 1:
            sb1m_signal = 'buy1'
        elif sell_count == 5:
            sb1m_signal = 'sell5'
        elif sell_count == 4:
            sb1m_signal = 'sell4'
        elif sell_count == 3:
            sb1m_signal = 'sell3'
        elif sell_count == 2:
            sb1m_signal = 'sell2'
        elif sell_count == 1:
            sb1m_signal = 'sell1'
        else:
            sb1m_signal = ''
        
        sb1m_data.append({
            '5min_group': group_time,
            'SB1M': sb1m_signal
        })
    
    # --- merge 제거, map 사용 ---
    if sb1m_data:
        sb1m_df = pd.DataFrame(sb1m_data)
        
        # 타입 통일 (datetime64[ns])
        try:
            if sb1m_df['5min_group'].dtype != 'datetime64[ns]':
                sb1m_df['5min_group'] = pd.to_datetime(sb1m_df['5min_group'], errors='coerce')
            if df_5m['5min_group'].dtype != 'datetime64[ns]':
                df_5m['5min_group'] = pd.to_datetime(df_5m['5min_group'], errors='coerce')
        except Exception:
            pass
        
        # 5min_group → SB1M 매핑 딕셔너리
        sb1m_map = dict(zip(sb1m_df['5min_group'], sb1m_df['SB1M']))
        
        # SB1M 열 추가 (없으면 생성)
        if 'SB1M' not in df_5m.columns:
            df_5m['SB1M'] = ''
        
        # map을 사용하여 SB1M 값 할당
        df_5m['SB1M'] = df_5m['5min_group'].map(sb1m_map).fillna('')
    else:
        if 'SB1M' not in df_5m.columns:
            df_5m['SB1M'] = ''
    
    # 임시 컬럼 제거
    if 'Date(UTC)_dt' in df_5m.columns:
        df_5m = df_5m.drop('Date(UTC)_dt', axis=1)
    if '5min_group' in df_5m.columns:
        df_5m = df_5m.drop('5min_group', axis=1)
    if 'Date(UTC)_dt' in df_1m.columns:
        df_1m = df_1m.drop('Date(UTC)_dt', axis=1)
    if '5min_group' in df_1m.columns:
        df_1m = df_1m.drop('5min_group', axis=1)
    
    return df_5m

def trade_on_order_signal(order_value: str, symbol: str = None, samount: float = 0, bamount: float = 0, bomb_multiplier: int = 1, prft_value: Optional[Union[str, int, float]] = None, ksc_numeric: int = 0, stage_prefix: str = "", stosu: float = 0.0, hmsfast: Optional[float] = None, buyside: Optional[float] = None, tpc_value: float = 0.0, minus_1hcl: Optional[float] = None, decision_price: Optional[float] = None, h1cl: Optional[int] = None, prev_tp: Optional[float] = None, prft_multiplier: float = 1.0, p_value: Optional[Union[int, float]] = None):
    """
    ORDER 신호에 따른 자동매매 실행 (Binance 버전)
    """
    if symbol is None:
        symbol = f"{TICKER}USDT"
    
    # order_value 정규화
    order_value = (order_value or "").strip()
    
    try:
        # 주문 전 계좌 현황 조회
        ticker_balance_before = binance_get_account_balance(TICKER)
        usdt_balance = binance_get_account_balance("USDT")
        
        if not ticker_balance_before or not usdt_balance:
            print(f"{get_timestamp()} [{stage_prefix}] ❌ 계좌 잔고 조회 실패")
            return None
        
        # 티커별 USDT 정밀도 가져오기
        usdt_precision = SYMBOL_USDT_PRECISION.get(symbol, 5)  # 기본값 5자리
        qty_precision = SYMBOL_QTY_PRECISION.get(symbol, 8)  # 심볼별 수량 정밀도
        
        if order_value == "Buy5":
            # [알림] 신호 감지
            detect_msg = f"{get_timestamp()} [{stage_prefix}] 🚨ORDER 신호 감지: {TICKER} Buy5"
            send_discord_message(detect_msg)
            
            # 미체결 잔량 확인 및 표시 (Available / Locked, UNIT, USDT 표시)
            if ticker_balance_before['locked'] > 0:
                current_price = binance_ticker_price(symbol)
                avail_usdt = ticker_balance_before['free_precise'] * current_price
                locked_usdt = ticker_balance_before['locked'] * current_price
                avail_unit = avail_usdt / TRADING_UNIT if TRADING_UNIT > 0 else 0
                locked_unit = locked_usdt / TRADING_UNIT if TRADING_UNIT > 0 else 0
                locked_msg = (
                    f"{get_timestamp()} [{stage_prefix}] ⚠️{TICKER} Available: "
                    f"{ticker_balance_before['free_precise']:.8f} {TICKER} ({avail_unit:.2f} U {avail_usdt:.2f} USDT) | "
                    f"Locked: {ticker_balance_before['locked']:.8f} {TICKER} ({locked_unit:.2f} U {locked_usdt:.2f} USDT)"
                )
                print(locked_msg)
                send_discord_message(locked_msg)
            
            # Z = multiplier(수열) + if(bomb발생, B값, 0)
            # 주문량 계산:
            # - KSC 스택이 쌓이는 상황 (KSC >= 1): (1유닛 + bamount) × Z
            # - KSC 스택이 쌓이지 않는 상황 (KSC = 0): 1유닛 + bamount (Z값 무시)
            base_amount = TRADING_UNIT + bamount
            # bomb_multiplier 파라미터는 Z 값을 전달받음
            if ksc_numeric == 0:
                # KSC 스택이 쌓이지 않는 상황: Z값 무시하고 기본 주문량 사용
                buy_amount = base_amount
            else:
                # KSC 스택이 쌓이는 상황: Z값으로 주문량 컨트롤
                if bomb_multiplier == 0:
                    # Z가 0이면 주문 차단 (KSC >= 1인데 Z = 0인 경우) — 메시지 없이 차단만
                    return None
                buy_amount = base_amount * bomb_multiplier
            
            price_info = f" | 결정가 {decision_price:.{usdt_precision}f} USDT" if decision_price is not None else ""
            
            # H 팩터 계산: hcl = 1HCL + (-1HCL)
            hcl_val = 0.0
            if h1cl is not None:
                try:
                    hcl_val += float(h1cl)
                except (TypeError, ValueError):
                    pass
            if minus_1hcl is not None:
                try:
                    hcl_val += float(minus_1hcl)
                except (TypeError, ValueError):
                    pass
            
            # Buy H 팩터 적용
            risk_factor = get_buy_risk_factor(order_value, hmsfast, hcl_val)
            risk_multiplier_text = ""
            if risk_factor != 1.0:
                risk_msg = f"{get_timestamp()} [{stage_prefix}] ⚠️H 팩터 적용: 주문량 {buy_amount:.{usdt_precision}f} USDT → {buy_amount * risk_factor:.{usdt_precision}f} USDT (H={risk_factor:.2f}, hcl={hcl_val:.1f})"
                print(risk_msg)
                send_discord_message(risk_msg)
                risk_multiplier_text = f" × H={risk_factor:.2f}"
            buy_amount *= risk_factor
            buy_unit = buy_amount / TRADING_UNIT
            bamount_unit = bamount / TRADING_UNIT
            
            if ksc_numeric == 0:
                buy_msg = f"{get_timestamp()} [{stage_prefix}] 💰 매수 예정: {buy_amount:.{usdt_precision}f} USDT (1unit {TRADING_UNIT:.2f} USDT + Bamount {bamount:.{usdt_precision}f} USDT({bamount_unit:.2f} UNIT), KSC=0이므로 기본 주문량){risk_multiplier_text} = {buy_unit:.2f} UNIT{price_info}"
                print(buy_msg)
                send_discord_message(buy_msg)
            elif bomb_multiplier > 1:
                buy_msg = f"{get_timestamp()} [{stage_prefix}] 💰 매수 예정: {buy_amount:.{usdt_precision}f} USDT ((1unit {TRADING_UNIT:.2f} USDT + Bamount {bamount:.{usdt_precision}f} USDT({bamount_unit:.2f} UNIT)) × Z({bomb_multiplier})){risk_multiplier_text} = {buy_unit:.2f} UNIT{price_info}"
                print(buy_msg)
                send_discord_message(buy_msg)
            else:
                buy_msg = f"{get_timestamp()} [{stage_prefix}] 💰 매수 예정: {buy_amount:.{usdt_precision}f} USDT (1unit {TRADING_UNIT:.2f} USDT + Bamount {bamount:.{usdt_precision}f} USDT({bamount_unit:.2f} UNIT)){risk_multiplier_text} = {buy_unit:.2f} UNIT{price_info}"
                print(buy_msg)
                send_discord_message(buy_msg)
            
            # USDT 잔고 확인
            MIN_ORDER_AMOUNT = 5.5  # 바이낸스 최소 주문액 (USDT, H 적용 후 최소 5.5 USDT 보장)
            SAFETY_MARGIN_RATIO = 0.01  # 수수료 및 여유분 고려 (1%)
            if usdt_balance['free_precise'] < buy_amount:
                # 잔고 부족 시: 정밀도에 맞춰 내림 처리 + 수수료 여유분 차감
                available_adjusted = usdt_balance['free_precise'] * (1 - SAFETY_MARGIN_RATIO)
                # USDT 정밀도에 맞춰 반올림
                adjusted_buy_amount = round(available_adjusted, usdt_precision)
                
                if adjusted_buy_amount >= MIN_ORDER_AMOUNT:
                    print(f"{get_timestamp()} [{stage_prefix}] ⚠️USDT 잔고 부족: 보유 {usdt_balance['free_precise']:.{usdt_precision}f} USDT < 필요 {buy_amount:.{usdt_precision}f} USDT, 잔고만큼 주문량 조정: {adjusted_buy_amount:.{usdt_precision}f} USDT (수수료 여유분 {SAFETY_MARGIN_RATIO*100:.0f}% 차감)")
                    buy_amount = adjusted_buy_amount
                else:
                    cancel_msg = f"{get_timestamp()} [{stage_prefix}] ❌ 매수 주문 취소: USDT 잔고 부족 (보유: {usdt_balance['free_precise']:.{usdt_precision}f} USDT, 필요: {buy_amount:.{usdt_precision}f} USDT, 조정 후: {adjusted_buy_amount:.{usdt_precision}f} USDT < 최소 주문액: {MIN_ORDER_AMOUNT} USDT)"
                    print(cancel_msg)
                    send_discord_message(cancel_msg)
                return None
            else:
                # 잔고 충분 시에도 정밀도에 맞춰 반올림
                buy_amount = round(buy_amount, usdt_precision)
                # H 적용 등으로 최소 주문액보다 작아진 경우 5.5 USDT로 보정 (잔고가 충분할 때만)
                if buy_amount < MIN_ORDER_AMOUNT:
                    if usdt_balance['free_precise'] >= MIN_ORDER_AMOUNT:
                        buy_amount = MIN_ORDER_AMOUNT
                    else:
                        cancel_msg = (
                            f"{get_timestamp()} [{stage_prefix}] ❌ 매수 주문 취소: "
                            f"주문액 {buy_amount:.{usdt_precision}f} USDT < 최소 주문액 {MIN_ORDER_AMOUNT} USDT "
                            f"(Available: {usdt_balance['free_precise']:.{usdt_precision}f} USDT)"
                        )
                        print(cancel_msg)
                        send_discord_message(cancel_msg)
                        return None
                balance_msg = f"{get_timestamp()} [{stage_prefix}] ✅ USDT 충분 Available: {usdt_balance['free_precise']:.{usdt_precision}f} USDT, 주문액: {buy_amount:.{usdt_precision}f} USDT"
                print(balance_msg)
                send_discord_message(balance_msg)
            
            # 스마트 주문 로직 적용 (선물 호가·잔량 사용)
            # 결정가격(decision_price) = 종가(trigger): 매수 시 스마트 주문의 K값으로 사용
            K = decision_price if decision_price is not None and decision_price > 0 else binance_ticker_price(symbol)
            ask, bid, ask_q, bid_q = get_futures_orderbook_snapshot(symbol)
            
            if ask is None or bid is None:
                # 호가 조회 실패 시 주문 취소
                error_msg = f"{get_timestamp()} [{stage_prefix}] ❌선물 호가 조회 실패, 매수 주문 취소"
                print(error_msg)
                send_discord_message(error_msg)
                res = None
            else:
                # 스마트 주문으로 최적 가격 계산
                smart_price, smart_log = execute_smart_order(is_buy=True, K=K, ask=ask, bid=bid, ask_q=ask_q, bid_q=bid_q, symbol=symbol)
                
                if smart_price is None or smart_price <= 0:
                    # 스마트 주문 계산 실패 시 주문 취소
                    error_msg = f"{get_timestamp()} [{stage_prefix}] ❌스마트 주문 계산 실패, 매수 주문 취소"
                    print(error_msg)
                    send_discord_message(error_msg)
                    res = None
                else:
                    # 스마트 주문 로그 출력
                    for log_msg in smart_log:
                        print(f"{get_timestamp()} [{stage_prefix}] 📊{log_msg}")
                        send_discord_message(f"{get_timestamp()} [{stage_prefix}] 📊{log_msg}")
                    
                    # 가격을 틱 사이즈 규칙에 맞게 보정
                    smart_price = round_price_to_tick_size_binance(symbol, smart_price)
                    # 지정가 매수 주문 실행
                    res = binance_limit_buy(symbol, buy_amount, smart_price, stage_prefix=stage_prefix)
            
        elif order_value == "Buy10":
            # [알림] 신호 감지
            detect_msg = f"{get_timestamp()} [{stage_prefix}] 🚨ORDER 신호 감지: {TICKER} Buy10"
            send_discord_message(detect_msg)
            
            # 미체결 잔량 확인 및 표시 (Available / Locked, UNIT, USDT 표시)
            if ticker_balance_before['locked'] > 0:
                current_price = binance_ticker_price(symbol)
                avail_usdt = ticker_balance_before['free_precise'] * current_price
                locked_usdt = ticker_balance_before['locked'] * current_price
                avail_unit = avail_usdt / TRADING_UNIT if TRADING_UNIT > 0 else 0
                locked_unit = locked_usdt / TRADING_UNIT if TRADING_UNIT > 0 else 0
                locked_msg = (
                    f"{get_timestamp()} [{stage_prefix}] ⚠️{TICKER} Available: "
                    f"{ticker_balance_before['free_precise']:.8f} {TICKER} ({avail_unit:.2f} U {avail_usdt:.2f} USDT) | "
                    f"Locked: {ticker_balance_before['locked']:.8f} {TICKER} ({locked_unit:.2f} U {locked_usdt:.2f} USDT)"
                )
                print(locked_msg)
                send_discord_message(locked_msg)
            
            # Buy5와 동일한 로직
            base_amount = TRADING_UNIT + bamount
            if ksc_numeric == 0:
                buy_amount = base_amount
            else:
                if bomb_multiplier == 0:
                    return None
                buy_amount = base_amount * bomb_multiplier
            
            price_info = f" | 결정가 {decision_price:.{usdt_precision}f} USDT" if decision_price is not None else ""
            
            # H 팩터 계산: hcl = 1HCL + (-1HCL)
            hcl_val = 0.0
            if h1cl is not None:
                try:
                    hcl_val += float(h1cl)
                except (TypeError, ValueError):
                    pass
            if minus_1hcl is not None:
                try:
                    hcl_val += float(minus_1hcl)
                except (TypeError, ValueError):
                    pass
            
            # Buy H 팩터 적용
            risk_factor = get_buy_risk_factor(order_value, hmsfast, hcl_val)
            risk_multiplier_text = ""
            if risk_factor != 1.0:
                risk_msg = f"{get_timestamp()} [{stage_prefix}] ⚠️H 팩터 적용: 주문량 {buy_amount:.{usdt_precision}f} USDT → {buy_amount * risk_factor:.{usdt_precision}f} USDT (H={risk_factor:.2f}, hcl={hcl_val:.1f})"
                print(risk_msg)
                send_discord_message(risk_msg)
                risk_multiplier_text = f" × H={risk_factor:.2f}"
            buy_amount *= risk_factor
            buy_unit = buy_amount / TRADING_UNIT
            bamount_unit = bamount / TRADING_UNIT
            
            if ksc_numeric == 0:
                buy_msg = f"{get_timestamp()} [{stage_prefix}] 💰 매수 예정: {buy_amount:.{usdt_precision}f} USDT (1unit {TRADING_UNIT:.2f} USDT + Bamount {bamount:.{usdt_precision}f} USDT({bamount_unit:.2f} UNIT), KSC=0이므로 기본 주문량){risk_multiplier_text} = {buy_unit:.2f} UNIT{price_info}"
                print(buy_msg)
                send_discord_message(buy_msg)
            elif bomb_multiplier > 1:
                buy_msg = f"{get_timestamp()} [{stage_prefix}] 💰 매수 예정: {buy_amount:.{usdt_precision}f} USDT ((1unit {TRADING_UNIT:.2f} USDT + Bamount {bamount:.{usdt_precision}f} USDT({bamount_unit:.2f} UNIT)) × Z({bomb_multiplier})){risk_multiplier_text} = {buy_unit:.2f} UNIT{price_info}"
                print(buy_msg)
                send_discord_message(buy_msg)
            else:
                buy_msg = f"{get_timestamp()} [{stage_prefix}] 💰 매수 예정: {buy_amount:.{usdt_precision}f} USDT (1unit {TRADING_UNIT:.2f} USDT + Bamount {bamount:.{usdt_precision}f} USDT({bamount_unit:.2f} UNIT)){risk_multiplier_text} = {buy_unit:.2f} UNIT{price_info}"
                print(buy_msg)
                send_discord_message(buy_msg)
            
            MIN_ORDER_AMOUNT = 5.5  # 바이낸스 최소 주문액 (USDT, H 적용 후 최소 5.5 USDT 보장)
            SAFETY_MARGIN_RATIO = 0.01
            if usdt_balance['free_precise'] < buy_amount:
                available_adjusted = usdt_balance['free_precise'] * (1 - SAFETY_MARGIN_RATIO)
                adjusted_buy_amount = round(available_adjusted, usdt_precision)
                
                if adjusted_buy_amount >= MIN_ORDER_AMOUNT:
                    print(f"{get_timestamp()} [{stage_prefix}] ⚠️USDT 잔고 부족: 보유 {usdt_balance['free_precise']:.{usdt_precision}f} USDT < 필요 {buy_amount:.{usdt_precision}f} USDT, 잔고만큼 주문량 조정: {adjusted_buy_amount:.{usdt_precision}f} USDT (수수료 여유분 {SAFETY_MARGIN_RATIO*100:.0f}% 차감)")
                    buy_amount = adjusted_buy_amount
                else:
                    cancel_msg = f"{get_timestamp()} [{stage_prefix}] ❌ 매수 주문 취소: USDT 잔고 부족 (보유: {usdt_balance['free_precise']:.{usdt_precision}f} USDT, 필요: {buy_amount:.{usdt_precision}f} USDT, 조정 후: {adjusted_buy_amount:.{usdt_precision}f} USDT < 최소 주문액: {MIN_ORDER_AMOUNT} USDT)"
                    print(cancel_msg)
                    send_discord_message(cancel_msg)
                return None
            else:
                buy_amount = round(buy_amount, usdt_precision)
                # H 적용 등으로 최소 주문액보다 작아진 경우 5.5 USDT로 보정 (잔고가 충분할 때만)
                if buy_amount < MIN_ORDER_AMOUNT:
                    if usdt_balance['free_precise'] >= MIN_ORDER_AMOUNT:
                        buy_amount = MIN_ORDER_AMOUNT
                    else:
                        cancel_msg = (
                            f"{get_timestamp()} [{stage_prefix}] ❌ 매수 주문 취소: "
                            f"주문액 {buy_amount:.{usdt_precision}f} USDT < 최소 주문액 {MIN_ORDER_AMOUNT} USDT "
                            f"(Available: {usdt_balance['free_precise']:.{usdt_precision}f} USDT)"
                        )
                        print(cancel_msg)
                        send_discord_message(cancel_msg)
                        return None
                balance_msg = f"{get_timestamp()} [{stage_prefix}] ✅ USDT 충분 Available: {usdt_balance['free_precise']:.{usdt_precision}f} USDT, 주문액: {buy_amount:.{usdt_precision}f} USDT"
                print(balance_msg)
                send_discord_message(balance_msg)
            
            # 스마트 주문 로직 적용 (선물 호가·잔량 사용)
            # 결정가격(decision_price) = 종가(trigger): 매수 시 스마트 주문의 K값으로 사용
            K = decision_price if decision_price is not None and decision_price > 0 else binance_ticker_price(symbol)
            ask, bid, ask_q, bid_q = get_futures_orderbook_snapshot(symbol)
            
            if ask is None or bid is None:
                # 호가 조회 실패 시 주문 취소
                error_msg = f"{get_timestamp()} [{stage_prefix}] ❌선물 호가 조회 실패, 매수 주문 취소"
                print(error_msg)
                send_discord_message(error_msg)
                res = None
            else:
                # 스마트 주문으로 최적 가격 계산
                smart_price, smart_log = execute_smart_order(is_buy=True, K=K, ask=ask, bid=bid, ask_q=ask_q, bid_q=bid_q, symbol=symbol)
                
                if smart_price is None or smart_price <= 0:
                    # 스마트 주문 계산 실패 시 주문 취소
                    error_msg = f"{get_timestamp()} [{stage_prefix}] ❌스마트 주문 계산 실패, 매수 주문 취소"
                    print(error_msg)
                    send_discord_message(error_msg)
                    res = None
                else:
                    # 스마트 주문 로그 출력
                    for log_msg in smart_log:
                        print(f"{get_timestamp()} [{stage_prefix}] 📊{log_msg}")
                        send_discord_message(f"{get_timestamp()} [{stage_prefix}] 📊{log_msg}")
                    
                    # 가격을 틱 사이즈 규칙에 맞게 보정
                    smart_price = round_price_to_tick_size_binance(symbol, smart_price)
                    # 지정가 매수 주문 실행
                    res = binance_limit_buy(symbol, buy_amount, smart_price, stage_prefix=stage_prefix)
            
        elif order_value == "Sell5":
            # [알림] 신호 감지
            detect_msg = f"{get_timestamp()} [{stage_prefix}] 🚨ORDER 신호 감지: {TICKER} Sell5"
            send_discord_message(detect_msg)
            
            # TPOVER 체크 (TPOVER일 때는 수수료 조건 무시)
            is_tpover = prft_value is not None and isinstance(prft_value, str) and str(prft_value).strip() == 'TPOVER'
            
            # 모든 티커 수수료 조건 체크 (TPOVER가 아닐 때만)
            if not is_tpover:
                # 현재 가격 조회
                current_price = binance_ticker_price(symbol)
                
                # 티커별 tp_percent 설정 (모든 티커 1%로 통일, Binance는 USDT 특별 처리 없음)
                tp_percent = 0.01
                
                # 평균단가 역산: 직전행TP / (1 + tp_percent)
                if prev_tp is not None and prev_tp > 0:
                    avg_buy_price = prev_tp / (1 + tp_percent)
                    
                    # 수수료 조건: 매도가격 > 평균단가 * (1+2f)
                    # 바이낸스 수수료는 0.075% (0.00075)로 동일
                    # 티커별 틱단위 고려하여 올림 처리
                    min_sell_price_raw = avg_buy_price * (1 + 2 * BINANCE_TRADING_FEE)
                    min_sell_price = ceil_price_to_tick_size_binance(symbol, min_sell_price_raw)
                    if current_price <= min_sell_price:
                        skip_msg = f"{get_timestamp()} [{stage_prefix}] ⏸️{TICKER} Sell5 스킵 (수수료 조건 미충족: 현재가 {current_price:.{usdt_precision}f} USDT <= 최소매도가 {min_sell_price:.{usdt_precision}f} USDT, 역산평균단가 {avg_buy_price:.{usdt_precision}f} USDT(TP={prev_tp:.{usdt_precision}f} USDT), f={BINANCE_TRADING_FEE*100}%)"
                        print(skip_msg)
                        send_discord_message(skip_msg)
                        return None
                    # 수수료 조건 충족 시 메시지 출력
                    fee_msg = f"{get_timestamp()} [{stage_prefix}] ✅{TICKER} Sell5 수수료 조건 충족: 현재가 {current_price:.{usdt_precision}f} USDT > 최소매도가 {min_sell_price:.{usdt_precision}f} USDT (역산평균단가 {avg_buy_price:.{usdt_precision}f} USDT(TP={prev_tp:.{usdt_precision}f} USDT), f={BINANCE_TRADING_FEE*100}%)"
                    print(fee_msg)
                    send_discord_message(fee_msg)
                else:
                    # 직전행TP가 없으면 조건 체크 스킵
                    print(f"{get_timestamp()} [{stage_prefix}] ⚠️{TICKER} Sell5 직전행TP 없음, 수수료 조건 체크 스킵")
                    current_price = binance_ticker_price(symbol)
                    min_sell_price = None
            elif is_tpover:
                # TPOVER일 때는 현재 가격만 조회 (수수료 조건 체크 스킵)
                current_price = binance_ticker_price(symbol)
                min_sell_price = None
                
                # 평균단가 역산: 직전행TP / (1 + tp_percent) (로그 표시용)
                tp_percent = 0.01
                if prev_tp is not None and prev_tp > 0:
                    avg_buy_price = prev_tp / (1 + tp_percent)
                    tpover_msg = f"{get_timestamp()} [{stage_prefix}] 🔥 {TICKER} Sell5 TPOVER 모드: 수수료 조건 무시하고 StoSU만큼 매도 진행 (현재가 {current_price:.{usdt_precision}f} USDT, 역산평균단가 {avg_buy_price:.{usdt_precision}f} USDT(TP={prev_tp:.{usdt_precision}f} USDT))"
                else:
                    tpover_msg = f"{get_timestamp()} [{stage_prefix}] 🔥 {TICKER} Sell5 TPOVER 모드: 수수료 조건 무시하고 StoSU만큼 매도 진행 (현재가 {current_price:.{usdt_precision}f} USDT)"
                print(tpover_msg)
            
            # PRFT multiplier: 파라미터로 전달받은 값 사용 (호출부에서 계산됨)
            # 만약 파라미터가 전달되지 않았거나 기본값(1.0)이면 함수 내부에서 계산
            if prft_multiplier == 1.0 and hmsfast is not None:
                hmsfast_val = float(hmsfast)
                # PRFT 작동 조건: or(1HMSFast>=7, 1HMSFast<2)
                if hmsfast_val >= 7.0 or hmsfast_val < 2.0:
                    # PRFT 활성화: prft_multiplier = 1 + (1 - buyside) = 2 - buyside
                    if buyside is not None:
                        buyside_val = float(buyside)
                        prft_multiplier = 1 + (1 - buyside_val)  # = 2 - buyside
                    # buyside가 없으면 기본값 1 사용
                # or(1HMSFast>=7, 1HMSFast<2)가 아니면 prft_multiplier = 1 (기본값 유지)
            
            # 기본 공식: sell_amount = (1unit + samount) * prft_multiplier
            base_amount = TRADING_UNIT + samount
            sell_amount = base_amount * prft_multiplier
            price_info = f" | 결정가 {decision_price:.{usdt_precision}f} USDT" if decision_price is not None else ""
            
            # TPOVER일 때는 StoSU 비율만큼 추가 주문
            if is_tpover:
                stosu_multiplier = 0.0
                if stosu > 0:
                    if tpc_value == 1:
                        stosu_multiplier = 0.75
                    elif tpc_value == 2:
                        stosu_multiplier = 0.2
                    elif tpc_value == 3:
                        stosu_multiplier = 0.05
                sell_amount = base_amount if stosu_multiplier == 0 else base_amount + (stosu * TRADING_UNIT * stosu_multiplier)
                sell_unit = sell_amount / TRADING_UNIT
                samount_unit = samount / TRADING_UNIT
                if stosu_multiplier > 0:
                    sell_msg = f"{get_timestamp()} [{stage_prefix}] 💰 TPOVER 매도 예정: {sell_amount:.{usdt_precision}f} USDT ((1unit {TRADING_UNIT:.2f} USDT + Samount {samount:.{usdt_precision}f} USDT({samount_unit:.2f} UNIT)) + StoSU {stosu:.2f} UNIT × {stosu_multiplier:.2f} × {TRADING_UNIT:.2f} USDT) = {sell_unit:.2f} UNIT (TPC={tpc_value}){price_info}"
                    print(sell_msg)
                    send_discord_message(sell_msg)
                else:
                    sell_msg = f"{get_timestamp()} [{stage_prefix}] 💰 TPOVER 매도 예정: {sell_amount:.{usdt_precision}f} USDT (TPC {tpc_value} → StoSU 추가 미적용) = {sell_unit:.2f} UNIT{price_info}"
                    print(sell_msg)
                    send_discord_message(sell_msg)
            else:
                # 일반 매도: H 팩터 적용
                # hcl = 1HCL + (-1HCL)
                hcl_val = 0.0
                if h1cl is not None:
                    try:
                        hcl_val += float(h1cl)
                    except (TypeError, ValueError):
                        pass
                if minus_1hcl is not None:
                    try:
                        hcl_val += float(minus_1hcl)
                    except (TypeError, ValueError):
                        pass
                
                h_factor = _calc_h_factor(hcl_val)
                sell_amount *= h_factor
                
                # H 적용 후 최소 주문액 체크 (5.5 USDT)
                MIN_ORDER_AMOUNT = 5.5  # 바이낸스 최소 주문액 (USDT, H 적용 후 최소 5.5 USDT 보장)
                sell_amount = round(sell_amount, usdt_precision)
                if sell_amount < MIN_ORDER_AMOUNT:
                    # 현재 가격으로 보유 현물 가치 확인 (current_price는 이미 위에서 조회됨)
                    if 'current_price' not in locals() or current_price is None:
                        current_price = binance_ticker_price(symbol)
                    보유수량_체크 = ticker_balance_before['free_precise']
                    보유수량_usdt_체크 = 보유수량_체크 * current_price if current_price > 0 else 0
                    
                    if 보유수량_usdt_체크 >= MIN_ORDER_AMOUNT:
                        sell_amount = MIN_ORDER_AMOUNT
                        min_adjust_msg = f"{get_timestamp()} [{stage_prefix}] ⚠️H 적용 후 주문액 {sell_amount:.{usdt_precision}f} USDT < 최소 주문액 {MIN_ORDER_AMOUNT} USDT, {MIN_ORDER_AMOUNT} USDT로 조정 (보유 현물 가치: {보유수량_usdt_체크:.2f} USDT)"
                        print(min_adjust_msg)
                        send_discord_message(min_adjust_msg)
                    else:
                        cancel_msg = (
                            f"{get_timestamp()} [{stage_prefix}] ❌ 매도 주문 취소: "
                            f"주문액 {sell_amount:.{usdt_precision}f} USDT < 최소 주문액 {MIN_ORDER_AMOUNT} USDT "
                            f"(보유 현물 가치: {보유수량_usdt_체크:.2f} USDT)"
                        )
                        print(cancel_msg)
                        send_discord_message(cancel_msg)
                        return None
                
                sell_unit = sell_amount / TRADING_UNIT
                samount_unit = samount / TRADING_UNIT
                
                if prft_multiplier > 1:
                    sell_msg = f"{get_timestamp()} [{stage_prefix}] 💰 매도 예정: {sell_amount:.{usdt_precision}f} USDT (1unit {TRADING_UNIT:.2f} USDT + Samount {samount:.{usdt_precision}f} USDT({samount_unit:.2f} UNIT)) × {prft_multiplier} × H={h_factor:.2f} = {sell_unit:.2f} UNIT{price_info}"
                    print(sell_msg)
                    send_discord_message(sell_msg)
                else:
                    sell_msg = f"{get_timestamp()} [{stage_prefix}] 💰 매도 예정: {sell_amount:.{usdt_precision}f} USDT (1unit {TRADING_UNIT:.2f} USDT + Samount {samount:.{usdt_precision}f} USDT({samount_unit:.2f} UNIT)) × H={h_factor:.2f} = {sell_unit:.2f} UNIT{price_info}"
                    print(sell_msg)
                    send_discord_message(sell_msg)
            
            # {TICKER} 잔고 확인을 위해 필요한 {TICKER} 수량 계산
            # current_price는 이미 위에서 조회되었거나, H 적용 후 최소 주문액 체크에서 조회됨
            if 'current_price' not in locals() or current_price is None:
                current_price = binance_ticker_price(symbol)
            required_ticker = round(sell_amount / current_price, qty_precision)  # 정밀도 적용
            보유수량 = ticker_balance_before['free_precise']
            계산수량 = required_ticker
            
            # 현물이 부족할 때 처리
            if 계산수량 > 보유수량:
                보유수량_usdt = 보유수량 * current_price  # USDT 가치 계산
                
                if 보유수량_usdt >= 5:  # USDT 가치가 5 USDT 이상
                    # Available 전체 매도 (현물 부족하지만 5 USDT 이상, 정확한 보유수량 사용)
                    shortage_msg = f"{get_timestamp()} [{stage_prefix}] 🔄 현물 부족 상황: 계산수량({계산수량:.{qty_precision}f}) > 보유수량({보유수량:.{qty_precision}f}), 보유수량_usdt >= 5 USDT"
                    print(shortage_msg)
                    send_discord_message(shortage_msg)
                    available_msg = f"{get_timestamp()} [{stage_prefix}] 💰 Available 전체 매도: {보유수량:.{qty_precision}f} {TICKER} ({보유수량_usdt:.2f} USDT)"
                    print(available_msg)
                    send_discord_message(available_msg)
                    # 정확한 잔고 전달 (원시 문자열 값 사용)
                    정확한_보유수량 = float(ticker_balance_before['free_raw'])
                    # price_hint = prev_tp (TP = 목표가격): TP 우선, 없으면 decision_price (결정가격 = 종가 = trigger) 사용
                    price_hint_value = prev_tp if prev_tp is not None and prev_tp > 0 else decision_price
                    res = binance_market_sell(symbol, 보유수량_usdt, exact_volume=정확한_보유수량, price_hint=price_hint_value, decision_price=decision_price, stage_prefix=stage_prefix, min_sell_price=min_sell_price)
                else:  # USDT 가치가 5 USDT 미만
                    # 매도주문취소 (현물 부족하고 5 USDT 미만)
                    cancel_msg = f"{get_timestamp()} [{stage_prefix}] ❌ 매도 주문 취소: 현물 부족 상황 (계산수량: {계산수량:.{qty_precision}f}, 보유수량: {보유수량:.{qty_precision}f}) + 보유수량_usdt < 5 USDT ({보유수량_usdt:.2f} USDT)"
                    print(cancel_msg)
                    send_discord_message(cancel_msg)
                    return None
            elif ticker_balance_before['free_precise'] < required_ticker:
                cancel_msg = f"{get_timestamp()} [{stage_prefix}] ❌ 매도 주문 취소: {TICKER} 잔고 부족 (보유: {ticker_balance_before['free_precise']:.{qty_precision}f} {TICKER}, 필요: {required_ticker:.{qty_precision}f} {TICKER})"
                print(cancel_msg)
                send_discord_message(cancel_msg)
                return None
            else:
                # 정상 매도 전 짜투리 처리 체크
                step_size = SYMBOL_STEP_SIZE.get(TICKER, 0.001)  # 티커별 stepSize
                잔여수량 = 보유수량 - 계산수량
                
                # LOT_SIZE에 맞춰 잔여수량 조정
                adjusted_잔여수량 = int(잔여수량 / step_size) * step_size  # floor 사용
                잔여수량_usdt = adjusted_잔여수량 * current_price
                
                if 잔여수량_usdt < 5:
                    # 짜투리 처리: Available 전부 매도 (정확한 보유수량 사용)
                    보유수량_usdt = 보유수량 * current_price
                    leftover_msg = f"{get_timestamp()} [{stage_prefix}] 🔄 짜투리 처리: 잔여수량({adjusted_잔여수량:.{qty_precision}f} {TICKER}, {잔여수량_usdt:.2f} USDT) < 5 USDT"
                    print(leftover_msg)
                    send_discord_message(leftover_msg)
                    available_msg2 = f"{get_timestamp()} [{stage_prefix}] 💰 Available 전체 매도: {보유수량:.{qty_precision}f} {TICKER} ({보유수량_usdt:.2f} USDT)"
                    print(available_msg2)
                    send_discord_message(available_msg2)
                    # 정확한 잔고 전달
                    정확한_보유수량 = ticker_balance_before['free_precise']
                    # price_hint = prev_tp (TP = 목표가격): TP 우선, 없으면 decision_price (결정가격 = 종가 = trigger) 사용
                    price_hint_value = prev_tp if prev_tp is not None and prev_tp > 0 else decision_price
                    res = binance_market_sell(symbol, 보유수량_usdt, exact_volume=정확한_보유수량, price_hint=price_hint_value, decision_price=decision_price, stage_prefix=stage_prefix, min_sell_price=min_sell_price)
                else:
                    # 정상 매도
                    # price_hint = prev_tp (TP = 목표가격): TP 우선, 없으면 decision_price (결정가격 = 종가 = trigger) 사용
                    price_hint_value = prev_tp if prev_tp is not None and prev_tp > 0 else decision_price
                    res = binance_market_sell(symbol, sell_amount, price_hint=price_hint_value, decision_price=decision_price, stage_prefix=stage_prefix, min_sell_price=min_sell_price)
            
        elif order_value == "Sell10":
            # [알림] 신호 감지
            detect_msg = f"{get_timestamp()} [{stage_prefix}] 🚨ORDER 신호 감지: {TICKER} Sell10"
            send_discord_message(detect_msg)
            
            # Sell5와 동일한 로직
            is_tpover = prft_value is not None and isinstance(prft_value, str) and str(prft_value).strip() == 'TPOVER'
            
            if not is_tpover:
                current_price = binance_ticker_price(symbol)
                tp_percent = 0.01
                
                if prev_tp is not None and prev_tp > 0:
                    avg_buy_price = prev_tp / (1 + tp_percent)
                    # 수수료 조건: 매도가격 > 평균단가 * (1+2f)
                    # 바이낸스 수수료는 0.075% (0.00075)로 동일
                    # 티커별 틱단위 고려하여 올림 처리
                    min_sell_price_raw = avg_buy_price * (1 + 2 * BINANCE_TRADING_FEE)
                    min_sell_price = ceil_price_to_tick_size_binance(symbol, min_sell_price_raw)
                    if current_price <= min_sell_price:
                        skip_msg = f"{get_timestamp()} [{stage_prefix}] ⏸️{TICKER} Sell10 스킵 (수수료 조건 미충족: 현재가 {current_price:.{usdt_precision}f} USDT <= 최소매도가 {min_sell_price:.{usdt_precision}f} USDT, 역산평균단가 {avg_buy_price:.{usdt_precision}f} USDT(TP={prev_tp:.{usdt_precision}f} USDT), f={BINANCE_TRADING_FEE*100}%)"
                        print(skip_msg)
                        send_discord_message(skip_msg)
                        return None
                    # 수수료 조건 충족 시 메시지 출력
                    fee_msg = f"{get_timestamp()} [{stage_prefix}] ✅{TICKER} Sell10 수수료 조건 충족: 현재가 {current_price:.{usdt_precision}f} USDT > 최소매도가 {min_sell_price:.{usdt_precision}f} USDT (역산평균단가 {avg_buy_price:.{usdt_precision}f} USDT(TP={prev_tp:.{usdt_precision}f} USDT), f={BINANCE_TRADING_FEE*100}%)"
                    print(fee_msg)
                    send_discord_message(fee_msg)
                else:
                    print(f"{get_timestamp()} [{stage_prefix}] ⚠️{TICKER} Sell10 직전행TP 없음, 수수료 조건 체크 스킵")
                    current_price = binance_ticker_price(symbol)
                    min_sell_price = None
            elif is_tpover:
                current_price = binance_ticker_price(symbol)
                min_sell_price = None
                tp_percent = 0.01
                if prev_tp is not None and prev_tp > 0:
                    avg_buy_price = prev_tp / (1 + tp_percent)
                    tpover_msg = f"{get_timestamp()} [{stage_prefix}] 🔥 {TICKER} Sell10 TPOVER 모드: 수수료 조건 무시하고 StoSU만큼 매도 진행 (현재가 {current_price:.{usdt_precision}f} USDT, 역산평균단가 {avg_buy_price:.{usdt_precision}f} USDT(TP={prev_tp:.{usdt_precision}f} USDT))"
                else:
                    tpover_msg = f"{get_timestamp()} [{stage_prefix}] 🔥 {TICKER} Sell10 TPOVER 모드: 수수료 조건 무시하고 StoSU만큼 매도 진행 (현재가 {current_price:.{usdt_precision}f} USDT)"
                print(tpover_msg)
            
            # PRFT multiplier: 파라미터로 전달받은 값 사용 (호출부에서 계산됨)
            # 만약 파라미터가 전달되지 않았거나 기본값(1.0)이면 함수 내부에서 계산
            if prft_multiplier == 1.0 and hmsfast is not None:
                hmsfast_val = float(hmsfast)
                if hmsfast_val >= 7.0 or hmsfast_val < 2.0:
                    if buyside is not None:
                        buyside_val = float(buyside)
                        prft_multiplier = 1 + (1 - buyside_val)
            
            base_amount = TRADING_UNIT + samount
            sell_amount = base_amount * prft_multiplier
            price_info = f" | 결정가 {decision_price:.{usdt_precision}f} USDT" if decision_price is not None else ""
            
            if is_tpover:
                stosu_multiplier = 0.0
                if stosu > 0:
                    if tpc_value == 1:
                        stosu_multiplier = 0.75
                    elif tpc_value == 2:
                        stosu_multiplier = 0.2
                    elif tpc_value == 3:
                        stosu_multiplier = 0.05
                sell_amount = base_amount if stosu_multiplier == 0 else base_amount + (stosu * TRADING_UNIT * stosu_multiplier)
                sell_unit = sell_amount / TRADING_UNIT
                samount_unit = samount / TRADING_UNIT
                if stosu_multiplier > 0:
                    sell_msg = f"{get_timestamp()} [{stage_prefix}] 💰 TPOVER 매도 예정: {sell_amount:.{usdt_precision}f} USDT ((1unit {TRADING_UNIT:.2f} USDT + Samount {samount:.{usdt_precision}f} USDT({samount_unit:.2f} UNIT)) + StoSU {stosu:.2f} UNIT × {stosu_multiplier:.2f} × {TRADING_UNIT:.2f} USDT) = {sell_unit:.2f} UNIT (TPC={tpc_value}){price_info}"
                    print(sell_msg)
                else:
                    sell_msg = f"{get_timestamp()} [{stage_prefix}] 💰 TPOVER 매도 예정: {sell_amount:.{usdt_precision}f} USDT (TPC {tpc_value} → StoSU 추가 미적용) = {sell_unit:.2f} UNIT{price_info}"
                    print(sell_msg)
            else:
                # 일반 매도: H 팩터 적용
                # hcl = 1HCL + (-1HCL)
                hcl_val = 0.0
                if h1cl is not None:
                    try:
                        hcl_val += float(h1cl)
                    except (TypeError, ValueError):
                        pass
                if minus_1hcl is not None:
                    try:
                        hcl_val += float(minus_1hcl)
                    except (TypeError, ValueError):
                        pass
                
                h_factor = _calc_h_factor(hcl_val)
                sell_amount *= h_factor
                
                # H 적용 후 최소 주문액 체크 (5.5 USDT)
                MIN_ORDER_AMOUNT = 5.5  # 바이낸스 최소 주문액 (USDT, H 적용 후 최소 5.5 USDT 보장)
                sell_amount = round(sell_amount, usdt_precision)
                if sell_amount < MIN_ORDER_AMOUNT:
                    # 현재 가격으로 보유 현물 가치 확인 (current_price는 이미 위에서 조회됨)
                    if 'current_price' not in locals() or current_price is None:
                        current_price = binance_ticker_price(symbol)
                    보유수량_체크 = ticker_balance_before['free_precise']
                    보유수량_usdt_체크 = 보유수량_체크 * current_price if current_price > 0 else 0
                    
                    if 보유수량_usdt_체크 >= MIN_ORDER_AMOUNT:
                        sell_amount = MIN_ORDER_AMOUNT
                        min_adjust_msg = f"{get_timestamp()} [{stage_prefix}] ⚠️H 적용 후 주문액 {sell_amount:.{usdt_precision}f} USDT < 최소 주문액 {MIN_ORDER_AMOUNT} USDT, {MIN_ORDER_AMOUNT} USDT로 조정 (보유 현물 가치: {보유수량_usdt_체크:.2f} USDT)"
                        print(min_adjust_msg)
                        send_discord_message(min_adjust_msg)
                    else:
                        cancel_msg = (
                            f"{get_timestamp()} [{stage_prefix}] ❌ 매도 주문 취소: "
                            f"주문액 {sell_amount:.{usdt_precision}f} USDT < 최소 주문액 {MIN_ORDER_AMOUNT} USDT "
                            f"(보유 현물 가치: {보유수량_usdt_체크:.2f} USDT)"
                        )
                        print(cancel_msg)
                        send_discord_message(cancel_msg)
                        return None
                
                sell_unit = sell_amount / TRADING_UNIT
                samount_unit = samount / TRADING_UNIT
                
                if prft_multiplier > 1:
                    sell_msg = f"{get_timestamp()} [{stage_prefix}] 💰 매도 예정: {sell_amount:.{usdt_precision}f} USDT (1unit {TRADING_UNIT:.2f} USDT + Samount {samount:.{usdt_precision}f} USDT({samount_unit:.2f} UNIT)) × {prft_multiplier} × H={h_factor:.2f} = {sell_unit:.2f} UNIT{price_info}"
                    print(sell_msg)
                else:
                    sell_msg = f"{get_timestamp()} [{stage_prefix}] 💰 매도 예정: {sell_amount:.{usdt_precision}f} USDT (1unit {TRADING_UNIT:.2f} USDT + Samount {samount:.{usdt_precision}f} USDT({samount_unit:.2f} UNIT)) × H={h_factor:.2f} = {sell_unit:.2f} UNIT{price_info}"
                    print(sell_msg)
            
            # {TICKER} 잔고 확인을 위해 필요한 {TICKER} 수량 계산
            # current_price는 이미 위에서 조회되었거나, H 적용 후 최소 주문액 체크에서 조회됨
            if 'current_price' not in locals() or current_price is None:
                current_price = binance_ticker_price(symbol)
            required_ticker = round(sell_amount / current_price, qty_precision)
            보유수량 = ticker_balance_before['free_precise']
            계산수량 = required_ticker
            
            if 계산수량 > 보유수량:
                보유수량_usdt = 보유수량 * current_price
                
                if 보유수량_usdt >= 5:
                    shortage_msg = f"{get_timestamp()} [{stage_prefix}] 🔄 현물 부족 상황: 계산수량({계산수량:.{qty_precision}f}) > 보유수량({보유수량:.{qty_precision}f}), 보유수량_usdt >= 5 USDT"
                    print(shortage_msg)
                    send_discord_message(shortage_msg)
                    available_msg = f"{get_timestamp()} [{stage_prefix}] 💰 Available 전체 매도: {보유수량:.{qty_precision}f} {TICKER} ({보유수량_usdt:.2f} USDT)"
                    print(available_msg)
                    send_discord_message(available_msg)
                    정확한_보유수량 = float(ticker_balance_before['free_raw'])
                    # price_hint = prev_tp (TP = 목표가격): TP 우선, 없으면 decision_price (결정가격 = 종가 = trigger) 사용
                    price_hint_value = prev_tp if prev_tp is not None and prev_tp > 0 else decision_price
                    res = binance_market_sell(symbol, 보유수량_usdt, exact_volume=정확한_보유수량, price_hint=price_hint_value, decision_price=decision_price, stage_prefix=stage_prefix, min_sell_price=min_sell_price)
                else:
                    cancel_msg = f"{get_timestamp()} [{stage_prefix}] ❌ 매도 주문 취소: 현물 부족 상황 (계산수량: {계산수량:.{qty_precision}f}, 보유수량: {보유수량:.{qty_precision}f}) + 보유수량_usdt < 5 USDT ({보유수량_usdt:.2f} USDT)"
                    print(cancel_msg)
                    send_discord_message(cancel_msg)
                    return None
            elif ticker_balance_before['free_precise'] < required_ticker:
                cancel_msg = f"{get_timestamp()} [{stage_prefix}] ❌ 매도 주문 취소: {TICKER} 잔고 부족 (보유: {ticker_balance_before['free_precise']:.{qty_precision}f} {TICKER}, 필요: {required_ticker:.{qty_precision}f} {TICKER})"
                print(cancel_msg)
                send_discord_message(cancel_msg)
                return None
            else:
                step_size = SYMBOL_STEP_SIZE.get(TICKER, 0.001)
                잔여수량 = 보유수량 - 계산수량
                adjusted_잔여수량 = int(잔여수량 / step_size) * step_size
                잔여수량_usdt = adjusted_잔여수량 * current_price
                
                if 잔여수량_usdt < 5:
                    보유수량_usdt = 보유수량 * current_price
                    leftover_msg = f"{get_timestamp()} [{stage_prefix}] 🔄 짜투리 처리: 잔여수량({adjusted_잔여수량:.{qty_precision}f} {TICKER}, {잔여수량_usdt:.2f} USDT) < 5 USDT"
                    print(leftover_msg)
                    send_discord_message(leftover_msg)
                    available_msg = f"{get_timestamp()} [{stage_prefix}] 💰 Available 전체 매도: {보유수량:.{qty_precision}f} {TICKER} ({보유수량_usdt:.2f} USDT)"
                    print(available_msg)
                    send_discord_message(available_msg)
                    정확한_보유수량 = ticker_balance_before['free_precise']
                    # price_hint = prev_tp (TP = 목표가격): TP 우선, 없으면 decision_price (결정가격 = 종가 = trigger) 사용
                    price_hint_value = prev_tp if prev_tp is not None and prev_tp > 0 else decision_price
                    res = binance_market_sell(symbol, 보유수량_usdt, exact_volume=정확한_보유수량, price_hint=price_hint_value, decision_price=decision_price, stage_prefix=stage_prefix, min_sell_price=min_sell_price)
                else:
                    # price_hint = prev_tp (TP = 목표가격): TP 우선, 없으면 decision_price (결정가격 = 종가 = trigger) 사용
                    price_hint_value = prev_tp if prev_tp is not None and prev_tp > 0 else decision_price
                    res = binance_market_sell(symbol, sell_amount, price_hint=price_hint_value, decision_price=decision_price, stage_prefix=stage_prefix, min_sell_price=min_sell_price)
            
        else:
            print(f"{get_timestamp()} [{stage_prefix}] ORDER: '{order_value}' → 주문 없음")
            return None
        
        if res and isinstance(res, dict) and res.get('orderId'):
            # 주문 성공 메시지 제거 (SOURCE 스타일 - 체결 완료 메시지로 대체)
            return res
        else:
            print(f"{get_timestamp()} [{stage_prefix}] ⚠️ 주문 응답 이상")
            return None
            
    except Exception as e:
        print(f"{get_timestamp()} [{stage_prefix}] ❌ 주문 실패: {e}")
        import traceback
        traceback.print_exc()
        return None

def copy_1hclass_to_15m(df_15m: pd.DataFrame, df_1h: pd.DataFrame) -> pd.DataFrame:
    """
    15분봉 시트에 1시간봉 시트의 1HCLASS 값을 시간 매칭하여 복사합니다 (1HCL로 저장).
    1시간 구간을 기준으로 1시간봉의 1HCLASS 값을 매칭합니다.
    ⚠️중요: Date(UTC) 기준으로만 수행 (KST는 기록용일 뿐, 계산에는 사용하지 않음)
    """
    if df_15m.empty or df_1h.empty:
        return df_15m
    
    df_15m_copy = df_15m.copy()
    
    # ⚠️중요: 1HCLASS 복사는 Date(UTC) 기준으로만 수행 (KST는 기록용일 뿐, 계산에는 사용하지 않음)
    # Date(UTC) 컬럼이 있으면 우선 사용, 없으면 KST 사용 (하위 호환성)
    time_col_15m = 'Date(UTC)' if 'Date(UTC)' in df_15m_copy.columns else ('KST' if 'KST' in df_15m_copy.columns else None)
    time_col_1h = 'Date(UTC)' if 'Date(UTC)' in df_1h.columns else ('KST' if 'KST' in df_1h.columns else None)
    
    if time_col_15m is None or time_col_1h is None:
        return df_15m_copy
    
    # 1시간봉 데이터를 딕셔너리로 변환
    hourly_dict = {}
    
    for _, row in df_1h.iterrows():
        time_val = row.get(time_col_1h, '')
        # Timestamp 객체인 경우 문자열로 변환
        if isinstance(time_val, pd.Timestamp) or hasattr(time_val, 'strftime'):
            time_str = time_val.strftime("%y/%m/%d,%H:%M") if hasattr(time_val, 'strftime') else str(time_val)
        else:
            time_str = str(time_val)
        
        if time_str and ',' in time_str:
            # YY/MM/DD,HH:MM 형식 직접 파싱
            hour_dt = pd.to_datetime(time_str, format="%y/%m/%d,%H:%M", errors='coerce')
        else:
            # 콤마가 없는 경우 일반 파싱 시도
            hour_dt = pd.to_datetime(time_str, errors='coerce')
        
        if pd.notna(hour_dt):
            # 1시간 구간의 시작 시간을 키로 사용
            hour_start = hour_dt.replace(minute=0, second=0, microsecond=0)
            hour_key = hour_start.strftime("%y/%m/%d,%H:%M")
            hourly_dict[hour_key] = row.get('1HCLASS', np.nan)
    
    def get_1hclass_value(time_str):
        # Timestamp 객체인 경우 문자열로 변환
        if hasattr(time_str, 'strftime'):
            time_str = time_str.strftime("%y/%m/%d,%H:%M")
        
        time_str = str(time_str)
        if time_str and ',' in time_str:
            # YY/MM/DD,HH:MM 형식 직접 파싱
            time_dt = pd.to_datetime(time_str, format="%y/%m/%d,%H:%M", errors='coerce')
            if pd.notna(time_dt):
                # 해당 1시간 구간의 시작 시간 계산
                hour_start = time_dt.replace(minute=0, second=0, microsecond=0)
                hour_key = hour_start.strftime("%y/%m/%d,%H:%M")
                
                if hour_key in hourly_dict:
                    return hourly_dict[hour_key]
        return np.nan
    
    # 1HCL 열에 값 복사 (1HCLASS를 1HCL로 저장)
    df_15m_copy['1HCL'] = df_15m_copy[time_col_15m].apply(get_1hclass_value)
    
    return df_15m_copy

def copy_minus_1hclass_to_15m(df_15m: pd.DataFrame, df_1h: pd.DataFrame) -> pd.DataFrame:
    """
    15분봉 시트에 1시간봉 시트의 -1HCLASS 값을 시간 매칭하여 복사합니다 (-1HCL로 저장).
    1시간 구간을 기준으로 1시간봉의 -1HCLASS 값을 매칭합니다.
    ⚠️중요: Date(UTC) 기준으로만 수행 (KST는 기록용일 뿐, 계산에는 사용하지 않음)
    """
    if df_15m.empty or df_1h.empty:
        return df_15m
    
    df_15m_copy = df_15m.copy()
    
    # ⚠️중요: -1HCLASS 복사는 Date(UTC) 기준으로만 수행 (KST는 기록용일 뿐, 계산에는 사용하지 않음)
    # Date(UTC) 컬럼이 있으면 우선 사용, 없으면 KST 사용 (하위 호환성)
    time_col_15m = 'Date(UTC)' if 'Date(UTC)' in df_15m_copy.columns else ('KST' if 'KST' in df_15m_copy.columns else None)
    time_col_1h = 'Date(UTC)' if 'Date(UTC)' in df_1h.columns else ('KST' if 'KST' in df_1h.columns else None)
    
    if time_col_15m is None or time_col_1h is None:
        return df_15m_copy
    
    # 1시간봉 데이터를 딕셔너리로 변환
    hourly_dict = {}
    
    for _, row in df_1h.iterrows():
        time_val = row.get(time_col_1h, '')
        # Timestamp 객체인 경우 문자열로 변환
        if isinstance(time_val, pd.Timestamp) or hasattr(time_val, 'strftime'):
            time_str = time_val.strftime("%y/%m/%d,%H:%M") if hasattr(time_val, 'strftime') else str(time_val)
        else:
            time_str = str(time_val)
        
        if time_str and ',' in time_str:
            # YY/MM/DD,HH:MM 형식 직접 파싱
            hour_dt = pd.to_datetime(time_str, format="%y/%m/%d,%H:%M", errors='coerce')
        else:
            # 콤마가 없는 경우 일반 파싱 시도
            hour_dt = pd.to_datetime(time_str, errors='coerce')
        
        if pd.notna(hour_dt):
            # 1시간 구간의 시작 시간을 키로 사용
            hour_start = hour_dt.replace(minute=0, second=0, microsecond=0)
            hour_key = hour_start.strftime("%y/%m/%d,%H:%M")
            hourly_dict[hour_key] = row.get('-1HCLASS', np.nan)
    
    def get_minus_1hclass_value(time_str):
        # Timestamp 객체인 경우 문자열로 변환
        if hasattr(time_str, 'strftime'):
            time_str = time_str.strftime("%y/%m/%d,%H:%M")
        
        time_str = str(time_str)
        if time_str and ',' in time_str:
            # YY/MM/DD,HH:MM 형식 직접 파싱
            time_dt = pd.to_datetime(time_str, format="%y/%m/%d,%H:%M", errors='coerce')
            if pd.notna(time_dt):
                # 해당 1시간 구간의 시작 시간 계산
                hour_start = time_dt.replace(minute=0, second=0, microsecond=0)
                hour_key = hour_start.strftime("%y/%m/%d,%H:%M")
                
                if hour_key in hourly_dict:
                    return hourly_dict[hour_key]
        return np.nan
    
    # -1HCL 열에 값 복사 (-1HCLASS를 -1HCL로 저장)
    df_15m_copy['-1HCL'] = df_15m_copy[time_col_15m].apply(get_minus_1hclass_value)
    
    return df_15m_copy

def copy_p1h_to_15m_and_set_p(df_15m: pd.DataFrame, df_1h: pd.DataFrame) -> pd.DataFrame:
    """
    15분봉 시트에 1시간봉 시트의 p1H 값을 시간 매칭하여 가져온 뒤, p = 3 + p1H 로 계산합니다.
    ⚠️중요: Date(UTC) 기준으로만 수행 (KST는 기록용일 뿐, 계산에는 사용하지 않음)
    """
    if df_15m.empty or df_1h.empty:
        return df_15m
    
    df_15m_copy = df_15m.copy()
    
    time_col_15m = 'Date(UTC)' if 'Date(UTC)' in df_15m_copy.columns else ('KST' if 'KST' in df_15m_copy.columns else None)
    time_col_1h = 'Date(UTC)' if 'Date(UTC)' in df_1h.columns else ('KST' if 'KST' in df_1h.columns else None)
    
    if time_col_15m is None or time_col_1h is None:
        return df_15m_copy
    
    hourly_dict = {}
    for _, row in df_1h.iterrows():
        time_val = row.get(time_col_1h, '')
        if isinstance(time_val, pd.Timestamp) or hasattr(time_val, 'strftime'):
            time_str = time_val.strftime("%y/%m/%d,%H:%M") if hasattr(time_val, 'strftime') else str(time_val)
        else:
            time_str = str(time_val)
        if time_str and ',' in time_str:
            hour_dt = pd.to_datetime(time_str, format="%y/%m/%d,%H:%M", errors='coerce')
        else:
            hour_dt = pd.to_datetime(time_str, errors='coerce')
        if pd.notna(hour_dt):
            hour_start = hour_dt.replace(minute=0, second=0, microsecond=0)
            hour_key = hour_start.strftime("%y/%m/%d,%H:%M")
            p1h_val = row.get('p1H', np.nan)
            try:
                p1h_val = float(p1h_val) if pd.notna(p1h_val) else np.nan
            except (TypeError, ValueError):
                p1h_val = np.nan
            hourly_dict[hour_key] = p1h_val
    
    def get_p_value(time_str):
        if hasattr(time_str, 'strftime'):
            time_str = time_str.strftime("%y/%m/%d,%H:%M")
        time_str = str(time_str)
        if time_str and ',' in time_str:
            time_dt = pd.to_datetime(time_str, format="%y/%m/%d,%H:%M", errors='coerce')
            if pd.notna(time_dt):
                hour_start = time_dt.replace(minute=0, second=0, microsecond=0)
                hour_key = hour_start.strftime("%y/%m/%d,%H:%M")
                if hour_key in hourly_dict:
                    p1h = hourly_dict[hour_key]
                    if pd.notna(p1h):
                        return 3 + p1h
        return np.nan
    
    df_15m_copy['p'] = df_15m_copy[time_col_15m].apply(get_p_value)
    
    return df_15m_copy

def calculate_stosp_stosu(df_15m: pd.DataFrame) -> pd.DataFrame:
    """
    15분봉 시트에 StoSP, StoSU 열을 계산합니다.
    
    StoSP 로직:
    - 과거 캔들부터 누적으로 계산
    - 과거로부터 Sell과 Sell 사이에 누적
    - KSC와 bomb 고려해서
    - 실제로 주문이 전송된 행들의 종가의 누적 주문량 UNIT의 가중평균을 구함
    - Sell5, Sell10 발생 시 초기화
    - PRFT 열에 'TPOVER' 발생 시 초기화
    
    TP 로직:
    - StoSP × 1.01 (또는 USDT는 1.006) 값을 계산하여 표시
    - StoSP가 Sell로 초기화되어 0/NaN이 되어도 직전 TP 값을 유지
    - StoSP가 다시 유효한 값이 되면 해당 시점의 StoSP × 배수(1.01 또는 1.006)로 업데이트
    
    StoSU 로직:
    - 과거로부터 Sell과 Sell 사이에 누적
    - KSC와 bomb 고려해서
    - 실제로 주문이 전송된 행들의 종가의 누적 주문량 UNIT 값을 표시함
    - Sell5, Sell10 발생 시 초기화
    - PRFT 열에 'TPOVER' 발생 시 초기화
    
    TPC 로직:
    - 과거 캔들부터 순회하며 TPOVER 발생 횟수를 누적
    - StoSP가 0/NaN 상태에서 다시 유효한 값으로 시작하면 0으로 초기화
    """
    if df_15m.empty:
        return df_15m
    
    df_15m_copy = df_15m.copy()
    
    # ⚠️중요: 정렬은 Date(UTC) 기준으로만 수행
    sort_col = 'Date(UTC)'
    
    # 과거→현재 순서로 정렬 (맨 아래부터 계산)
    df_15m_copy = df_15m_copy.sort_values(sort_col, ascending=True).reset_index(drop=True)
    
    # StoSP, StoSU, TPC, TPCS 열 초기화
    stossp_values = []
    stosu_values = []
    tpc_values = []
    tpcs_values = []
    tpc_count = 0
    
    # 누적 변수 (Sell 발생 시 초기화)
    cumulative_price_unit = 0.0  # sum(종가 * 주문량 UNIT)
    cumulative_unit = 0.0  # sum(주문량 UNIT)
    
    
    for idx, row in df_15m_copy.iterrows():
        order = str(row.get('ORDER', '')).strip()
        ksc = row.get('KSC', 0)
        bomb = row.get('Bomb', '')
        close = row.get('종', np.nan)
        bamount = row.get('Bamount', 0)
        prft = row.get('PRFT', 0)
        prev_active = cumulative_unit > 0
        
        # PRFT 열에 'TPOVER' 발생 시 처리 (Sell5/Sell10 우선)
        if isinstance(prft, str) and prft.strip() == 'TPOVER':
            tpc_count += 1
            cumulative_price_unit = 0.0
            cumulative_unit = 0.0
            stossp_values.append(np.nan)
            stosu_values.append(0.0)
            tpc_values.append(tpc_count)
            # TPCS는 StoSU 최종값(sticky rule 적용) 계산 후에 계산됨 (아래에서 처리)
            tpcs_values.append(0)  # 임시값, 나중에 재계산
            continue
        
        # Sell5, Sell10 발생 시 초기화
        if order in ['Sell5', 'Sell10']:
            cumulative_price_unit = 0.0
            cumulative_unit = 0.0
            stossp_values.append(np.nan)
            stosu_values.append(0.0)
            tpc_values.append(tpc_count)
            # TPCS는 StoSU 최종값(sticky rule 적용) 계산 후에 계산됨 (아래에서 처리)
            tpcs_values.append(0)  # 임시값, 나중에 재계산
            continue
        
        # KSC 값 확인 (초기화는 하지 않음, 주문 전송 여부 판단용으로만 사용)
        ksc_numeric = 0
        if isinstance(ksc, (int, float)):
            ksc_numeric = int(ksc)
        elif isinstance(ksc, str):
            try:
                ksc_numeric = int(float(ksc))
            except (ValueError, TypeError):
                ksc_numeric = 0
        
        # Buy5 또는 Buy10인 경우만 처리
        if order in ['Buy5', 'Buy10']:
            # KSC 값은 이미 위에서 확인했으므로 재사용
            
            # Bomb 확인
            is_bomb = (isinstance(bomb, str) and bomb.strip() == 'Bomb')
            
            # Z 값 계산 (주문 전송 여부 판단용)
            z = 0
            if ksc_numeric == 0:
                # KSC = 0이면 기본 주문량 (Z = 1로 간주)
                z = 1
            else:
                # p값: 15M 열 p(= 3+p1H) 우선, 없으면 3 + 1HCL
                p_val = row.get('p', np.nan)
                if pd.notna(p_val):
                    try:
                        p_value = int(float(p_val))
                    except (TypeError, ValueError):
                        p_value = 3 + int(_safe_float(row.get('1HCL', 0), 0.0))
                else:
                    p_value = 3 + int(_safe_float(row.get('1HCL', 0), 0.0))
                
                # KSC >= 1인 경우 multiplier와 B 값 계산
                ksc_multiplier = calculate_ksc_multiplier(ksc_numeric, 0, p_value)
                
                # BombCount 확인 - 안전한 타입 변환
                bomb_count = int(_safe_float(row.get('BombCount', 0), 0.0))
                
                # Bomb 발생 시 B 값 계산
                bomb_b_value = 0
                if is_bomb:
                    bomb_b_value = calculate_bomb_b_value(ksc_multiplier, bomb_count, p_value)
                
                # Z = multiplier + B값 (제한 없음)
                z = ksc_multiplier + bomb_b_value
                
                # KSC = 1인 경우: p의 배수가 아니어도 주문 전송 (Z = 0이면 Z = 1로 강제 설정)
                # 실제 주문 전송 로직과 동일하게 처리하여 StoSP 계산 일치
                if ksc_numeric == 1 and z == 0:
                    z = 1
            
            # 주문 전송 여부 판단
            # KSC = 0이면 주문 전송, KSC >= 1이고 Z > 0이면 주문 전송
            order_sent = False
            if ksc_numeric == 0:
                order_sent = True
            elif z > 0:
                order_sent = True
            
            # 주문이 전송된 경우만 누적
            if order_sent and not pd.isna(close):
                # 주문량 계산 (bamount는 이미 UNIT으로 환산된 값)
                bamount_unit = bamount / TRADING_UNIT if not pd.isna(bamount) else 0
                base_unit = 1.0 + bamount_unit  # 1 UNIT + bamount UNIT
                
                if ksc_numeric == 0:
                    buy_unit = base_unit
                else:
                    buy_unit = base_unit * z
                
                # [수정] row.get은 object를 반환하므로 float로 안전하게 변환
                _hms_val = row.get('1HMSFast', np.nan)
                _hms_float: Optional[float] = None
                if pd.notna(_hms_val):
                    try:
                        _hms_float = float(_hms_val)
                    except (TypeError, ValueError):
                        _hms_float = None
                
                # hcl = 1HCL + (-1HCL)
                _hcl_1 = row.get('1HCL', np.nan)
                _hcl_minus1 = row.get('-1HCL', np.nan)
                hcl_val = 0.0
                if pd.notna(_hcl_1):
                    try:
                        hcl_val += float(_hcl_1)
                    except (TypeError, ValueError):
                        pass
                if pd.notna(_hcl_minus1):
                    try:
                        hcl_val += float(_hcl_minus1)
                    except (TypeError, ValueError):
                        pass
                
                risk_factor = get_buy_risk_factor(order, _hms_float, hcl_val)
                buy_unit *= risk_factor
                
                # 누적 계산
                cumulative_price_unit += float(close) * buy_unit
                cumulative_unit += buy_unit
        
        # StoSP 계산 (가중평균)
        current_stosp = np.nan
        if cumulative_unit > 0:
            current_stosp = cumulative_price_unit / cumulative_unit
            stossp_values.append(current_stosp)
        else:
            stossp_values.append(np.nan)
        
        # NBS 신호는 StoSP 값들을 모두 계산한 후에 처리 (아래에서 처리)
        
        # StoSU 계산 (누적 UNIT)
        stosu_values.append(cumulative_unit)
        
        # StoSP가 0에서 새로 시작된 경우 TPC 초기화
        new_active = cumulative_unit > 0
        if not prev_active and new_active:
            tpc_count = 0
        
        tpc_values.append(tpc_count)
        # TPCS는 StoSU 최종값(sticky rule 적용) 계산 후에 계산됨 (아래에서 처리)
        tpcs_values.append(0)  # 임시값, 나중에 재계산
    
    # TP 값 계산
    # - 기본: StoSP × (1 + 1%) = 1.01
    # - USDT: StoSP × (1 + 0.5%) = 1.005
    tp_percent = 0.005 if TICKER == "USDT" else 0.01
    tp_multiplier = 1 + tp_percent
    tp_values = []
    last_tp = np.nan
    for sp_value in stossp_values:
        if not pd.isna(sp_value) and isinstance(sp_value, (int, float)) and sp_value != 0:
            last_tp = float(sp_value) * tp_multiplier
        tp_values.append(last_tp)
    
    # StoSU 값 유지/갱신 (Sell 직후에는 직전 값 유지, 새로운 Buy 누적으로 갱신)
    stosu_final_values = []
    last_stosu = np.nan
    for su_value in stosu_values:
        if not pd.isna(su_value) and isinstance(su_value, (int, float)) and su_value != 0:
            last_stosu = float(su_value)
        if pd.isna(last_stosu):
            stosu_final_values.append(0.0)
        else:
            stosu_final_values.append(last_stosu)
    
    # TPCS 계산: =IF(TPC=1,StoSU*0.25,IF(TPC=2,StoSU*0.05,IF(TPC>=3,0,StoSU)))
    # StoSU 최종값(sticky rule 적용)을 사용하여 계산
    tpcs_final_values = []
    for idx in range(len(tpc_values)):
        tpc_val = tpc_values[idx]
        stosu_final = stosu_final_values[idx]
        
        if tpc_val == 1:
            tpcs_value = stosu_final * 0.25
        elif tpc_val == 2:
            tpcs_value = stosu_final * 0.05
        elif tpc_val >= 3:
            tpcs_value = 0.0
        else:
            # TPC == 0 또는 다른 값
            tpcs_value = stosu_final
        
        # 소수점 두 자리로 반올림
        tpcs_value = round(tpcs_value, 2)
        tpcs_final_values.append(tpcs_value)
    
    # StoSP, TP, StoSU, TPC, TPCS 열 추가
    df_15m_copy['StoSP'] = stossp_values
    df_15m_copy['TP'] = tp_values
    df_15m_copy['StoSU'] = stosu_final_values
    df_15m_copy['TPC'] = tpc_values
    df_15m_copy['TPCS'] = tpcs_final_values
    
    # ⚠️[수정됨] 최신→과거로 정렬 (엑셀 표시 순서와 동일)
    df_15m_copy = df_15m_copy.sort_values(sort_col, ascending=False).reset_index(drop=True)
    
    # ⚠️[수정됨] NBS 계산 로직
    # 정렬된 DataFrame의 StoSP 값을 가져와서 계산해야 순서가 맞습니다.
    # 기준: 현재 행(i)은 값이 있고, 바로 아래 행(i+1, 과거)은 값이 없으면 NBS=1 (새로운 스택 시작)
    
    sorted_stosp = df_15m_copy['StoSP'].values  # 정렬된 StoSP 값 추출
    nbs_values = []
    
    for idx in range(len(sorted_stosp)):
        # curr_val: 현재 행 (최신)
        curr_val = sorted_stosp[idx]
        curr_valid = not pd.isna(curr_val) and (isinstance(curr_val, (int, float)) and curr_val != 0)
        
        # prev_val: 바로 아래 행 (과거, idx+1)
        if idx + 1 < len(sorted_stosp):
            prev_val = sorted_stosp[idx + 1]
            prev_valid = not pd.isna(prev_val) and (isinstance(prev_val, (int, float)) and prev_val != 0)
        else:
            prev_valid = False  # 더 과거 데이터가 없으면 무효로 처리
        
        # NBS 조건: 현재(최신)는 유효하고, 과거(직전)는 무효할 때 1
        if curr_valid and not prev_valid:
            nbs_values.append(1)
        else:
            nbs_values.append(0)
    
    df_15m_copy['NBS'] = nbs_values
    
    return df_15m_copy

def check_today_snapshot_exists():
    """
    BINANCE_balance_history_detail.csv 파일을 확인하여 오늘 날짜(UTC 0시 기준)의 기록이 있는지 확인합니다.
    Returns: True(이미 있음), False(없음, 기록 필요)
    """
    try:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        log_dir = os.path.join(script_dir, LOG_DIR)
        csv_filename = os.path.join(log_dir, "BINANCE_balance_history_detail.csv")
        
        if not os.path.exists(csv_filename):
            return False  # 파일이 없으면 기록 필요
            
        # 오늘 날짜 (UTC 0시 기준, YYYY-MM-DD)
        today_str = dt.datetime.now(tz.UTC).strftime('%Y-%m-%d')
        
        # 파일 크기 확인 (빈 파일 체크)
        if os.path.getsize(csv_filename) == 0:
            return False
        
        with open(csv_filename, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            # 헤더 건너뛰기
            next(reader, None)
            
            # 모든 행을 리스트로 읽기 (역순 확인을 위해)
            rows = list(reader)
            
            if len(rows) == 0:
                return False
            
            # 마지막 행부터 역순으로 확인 (최신 기록이 마지막에 있으므로)
            for row in reversed(rows):
                if len(row) > 0:
                    # row[0]은 'YYYY-MM-DD HH:MM:SS' 형식임
                    timestamp_str = row[0].strip()
                    if timestamp_str.startswith(today_str):
                        return True  # 오늘 날짜 기록 발견!
                        
        return False  # 다 뒤져봤는데 오늘 날짜 없음
        
    except Exception as e:
        print(f"{get_timestamp()} ⚠️자산기록 확인 중 오류 (안전하게 False 반환): {e}")
        import traceback
        traceback.print_exc()
        return False  # 에러 나면 안전하게 기록 시도하도록 False 반환

def record_total_balance_snapshot(stage_prefix: str = "[자산기록]"):
    """
    [일일 자산 상세 기록 - Binance 버전]
    - 총자산, 총원금, 총손익, 총수익률, 현금(USDT)
    - 코인별: 매수금(Principal), 평가금(Val), 평가손익(PnL), 수익률(Rate) 모두 기록
    """
    try:
        # 파일 경로 설정 (logs 폴더에 저장)
        script_dir = os.path.dirname(os.path.abspath(__file__))
        log_dir = os.path.join(script_dir, LOG_DIR)
        if not os.path.exists(log_dir):
            os.makedirs(log_dir)
        csv_filename = os.path.join(log_dir, "BINANCE_balance_history_detail.csv")
        file_exists = os.path.isfile(csv_filename)
        
        # 1. 현금(USDT) 잔고 조회 (available + locked 모두 포함)
        # 자산기록 시에는 BNB 수수료 차감 안함 (subtract_bnb_fee=False)
        usdt_info = binance_get_account_balance("USDT", subtract_bnb_fee=False)
        total_usdt = float(usdt_info.get('total_precise', 0)) if usdt_info else 0.0  # free + locked
        
        # 2. 코인별 상세 데이터 수집
        total_coin_valuation = 0.0  # 총 코인 평가금액
        total_coin_principal = 0.0  # 총 코인 매수원금
        
        ticker_data = {}  # 코인별 데이터 저장소
        
        for ticker in ROTATION_TICKERS:
            # 잔고 조회 (자산기록 시에는 BNB 수수료 차감 안함)
            coin_info = binance_get_account_balance(ticker, subtract_bnb_fee=False)
            
            # balance = 보유량 (Binance는 total_precise 사용: free + locked 모두 포함)
            amount = float(coin_info.get('total_precise', 0)) if coin_info else 0.0
            
            # 직전행 TP에서 평균단가 역산
            avg_buy_price = 0.0
            prev_tp = None
            
            # 15분봉 파일에서 직전행 TP 읽기
            try:
                script_dir = os.path.dirname(os.path.abspath(__file__))
                base_dir = os.path.join(script_dir, "cryptodaily15min")
                ticker_folder_map = {
                    "BTC": "F BINANCE 1BTC",
                    "ETH": "F BINANCE 2ETH",
                    "XRP": "F BINANCE 3XRP",
                    "SOL": "F BINANCE 4SOL",
                    "BNB": "F BINANCE 5BNB"
                }
                ticker_folder = ticker_folder_map.get(ticker, f"F BINANCE {ticker}")
                ticker_dir = os.path.join(base_dir, ticker_folder)
                
                if os.path.exists(ticker_dir):
                    # 최신 after 파일 찾기
                    after_file_path = None
                    latest_mtime = 0
                    
                    for f in os.listdir(ticker_dir):
                        if f.startswith(f"after_F_{ticker}_BINANCE_") and f.endswith(".xlsx") and not f.startswith("~$"):
                            file_path = os.path.join(ticker_dir, f)
                            try:
                                mtime = os.path.getmtime(file_path)
                                if mtime > latest_mtime:
                                    latest_mtime = mtime
                                    after_file_path = file_path
                            except OSError:
                                continue
                    
                    # 직전행 TP 읽기
                    if after_file_path:
                        try:
                            df_15m = pd.read_excel(after_file_path, sheet_name=f"{ticker}USDT15M", nrows=2)
                            if len(df_15m) > 1 and 'TP' in df_15m.columns:
                                prev_tp_raw = df_15m.iloc[1].get('TP', np.nan)
                                if pd.notna(prev_tp_raw):
                                    try:
                                        prev_tp = float(prev_tp_raw)
                                        if prev_tp > 0:
                                            # 평균단가 역산: 직전행TP / (1 + tp_percent)
                                            tp_percent = 0.01  # 1%
                                            avg_buy_price = prev_tp / (1 + tp_percent)
                                    except (TypeError, ValueError):
                                        pass
                        except Exception as e:
                            print(f"{get_timestamp()} {stage_prefix} ⚠️{ticker} 15분봉 파일 읽기 실패: {e}")
            except Exception as e:
                print(f"{get_timestamp()} {stage_prefix} ⚠️{ticker} 직전행 TP 조회 실패: {e}")
            
            current_val = 0.0   # 평가금
            principal = 0.0     # 매수원금
            pnl = 0.0           # 평가손익
            profit_rate = 0.0   # 수익률
            
            if amount > 0:
                try:
                    # 현재가 조회 (Binance)
                    symbol = f"{ticker}USDT"
                    current_price = binance_ticker_price(symbol)
                    
                    # 1) 평가금액 (현재 가치)
                    current_val = amount * current_price
                    
                    # 2) 매수원금 (평단가 * 수량)
                    principal = amount * avg_buy_price
                    
                    # 3) 평가손익 (평가금 - 원금)
                    pnl = current_val - principal
                    
                    # 4) 수익률 계산 (직전행 TP로 역산한 평균단가 사용)
                    if avg_buy_price > 0:
                        profit_rate = ((current_price - avg_buy_price) / avg_buy_price) * 100
                    
                    # 총합 누적
                    total_coin_valuation += current_val
                    total_coin_principal += principal
                    
                except Exception as e:
                    print(f"{get_timestamp()} ⚠️{ticker} 가격 조회 실패: {e}")
            
            # 딕셔너리에 모든 정보 저장 (소수점 유지)
            ticker_data[ticker] = {
                'principal': round(principal, 2),  # 매수금 (소수점 2자리)
                'val': round(current_val, 2),      # 평가금 (소수점 2자리)
                'pnl': round(pnl, 2),              # 손익금 (소수점 2자리)
                'rate': round(profit_rate, 2)      # 수익률 (소수점 2자리)
            }
        
        # 3. 전체 자산 통계 계산
        total_equity = total_usdt + total_coin_valuation  # 총자산
        total_accumulated_pnl = total_coin_valuation - total_coin_principal  # 누적 손익 (Accumulated PnL)
        
        total_return_rate = 0.0
        if total_coin_principal > 0:
            total_return_rate = (total_accumulated_pnl / total_coin_principal) * 100
        
        # =================================================================
        # [Daily PnL 계산] 어제 자산(Equity)과 비교
        # =================================================================
        daily_pnl = 0.0
        
        if file_exists:
            try:
                with open(csv_filename, 'r', encoding='utf-8-sig') as f:
                    reader = csv.reader(f)
                    lines = list(reader)
                    
                    if len(lines) > 1:
                        # 헤더 제외 (첫 번째 행은 헤더)
                        data_lines = lines[1:]  # 헤더 제외
                        
                        if len(data_lines) > 0:
                            # 오늘 날짜 추출 (Time 컬럼의 날짜 부분)
                            now_utc = dt.datetime.now(tz.UTC)
                            today_date_str = now_utc.strftime('%Y-%m-%d')
                            
                            # 역순으로 읽으면서 전날(다른 날짜)의 마지막 행 찾기
                            last_equity = None
                            last_date_str = None
                            
                            for i in range(len(data_lines) - 1, -1, -1):  # 마지막 행부터 역순으로
                                row = data_lines[i]
                                if len(row) > 0:
                                    time_str = str(row[0]).strip()  # Time 컬럼
                                    
                                    # 날짜 부분 추출 (예: "2026-01-07 00:21" -> "2026-01-07")
                                    try:
                                        if ' ' in time_str:
                                            row_date_str = time_str.split(' ')[0].strip()
                                        elif len(time_str) >= 10:
                                            row_date_str = time_str[:10].strip()
                                        else:
                                            continue
                                        
                                        # 날짜 형식 검증 (YYYY-MM-DD)
                                        if len(row_date_str) == 10 and row_date_str.count('-') == 2:
                                            # 오늘 날짜가 아닌 첫 번째 행을 찾음 (전날의 마지막 기록)
                                            if row_date_str != today_date_str and len(row) > 1:
                                                try:
                                                    equity_value = row[1].strip()
                                                    if equity_value:
                                                        last_equity = float(equity_value)  # Total Equity (소수점 포함)
                                                        last_date_str = row_date_str
                                                        break
                                                except (ValueError, IndexError, TypeError):
                                                    continue
                                    except Exception as parse_error:
                                        continue
                            
                            # 전날 기록을 찾았으면 Daily PnL 계산
                            if last_equity is not None:
                                daily_pnl = total_equity - last_equity
                                print(f"{get_timestamp()} {stage_prefix} 📊 Daily PnL 계산: 오늘({today_date_str}) {total_equity:.2f} - 전날({last_date_str}) {last_equity:.2f} = {daily_pnl:.2f}")
                            else:
                                # 전날 기록이 없으면 0 (첫 기록이거나 같은 날짜만 있는 경우)
                                daily_pnl = 0.0
                                print(f"{get_timestamp()} {stage_prefix} ⚠️ Daily PnL 계산: 전날 기록을 찾지 못함 (첫 기록이거나 같은 날짜만 존재)")
            except Exception as e:
                print(f"{get_timestamp()} {stage_prefix} ⚠️ Daily PnL 계산 중 오류: {e}")
                import traceback
                print(f"{get_timestamp()} {stage_prefix} ⚠️ 오류 상세: {traceback.format_exc()}")
                daily_pnl = 0.0  # 읽기 실패 시 0 처리
        
        # 4. CSV 저장 (UTC 0시 기준)
        now_str = dt.datetime.now(tz.UTC).strftime('%Y-%m-%d %H:%M:%S')
        
        # [수정] 고정된 컬럼 순서와 개수 사용 (항상 동일한 컬럼 수 유지)
        # 무조건 이 순서와 개수대로 칸을 만듭니다 (고정석)
        fixed_columns = [
            'Time', 'Total Equity', 'Daily PnL', 'Cash', 'Total Coin Val', 
            'Total Principal', 'Total Acc PnL', 'Total Return(%)',
            'BTC Buy', 'BTC Val', 'BTC PnL', 'BTC %',
            'ETH Buy', 'ETH Val', 'ETH PnL', 'ETH %',
            'XRP Buy', 'XRP Val', 'XRP PnL', 'XRP %',
            'SOL Buy', 'SOL Val', 'SOL PnL', 'SOL %',
            'BNB Buy', 'BNB Val', 'BNB PnL', 'BNB %'
        ]
        
        # 데이터 딕셔너리 생성 (고정 컬럼 순서에 맞춰 값 채우기)
        row_dict = {
            'Time': now_str,
            'Total Equity': f"{total_equity:.2f}",
            'Daily PnL': f"{daily_pnl:.2f}",
            'Cash': f"{total_usdt:.2f}",
            'Total Coin Val': f"{total_coin_valuation:.2f}",
            'Total Principal': f"{total_coin_principal:.2f}",
            'Total Acc PnL': f"{total_accumulated_pnl:.2f}",
            'Total Return(%)': f"{total_return_rate:.2f}"
        }
        
        # 코인별 데이터 추가 (고정 순서: BTC, ETH, XRP, SOL, BNB)
        for ticker in ['BTC', 'ETH', 'XRP', 'SOL', 'BNB']:
            data = ticker_data.get(ticker, {'principal': 0, 'val': 0, 'pnl': 0, 'rate': 0})
            row_dict[f"{ticker} Buy"] = f"{data['principal']:.2f}"
            row_dict[f"{ticker} Val"] = f"{data['val']:.2f}"
            row_dict[f"{ticker} PnL"] = f"{data['pnl']:.2f}"
            row_dict[f"{ticker} %"] = f"{data['rate']:.2f}"
        
        # 고정 컬럼 순서에 맞춰 데이터 리스트 생성 (없는 값은 빈 문자열로 채움)
        row_data = [row_dict.get(col, '') for col in fixed_columns]
        
        # [수정] 고정 컬럼 구조 사용 - 헤더 불일치 또는 행 컬럼 수 불일치 시 재작성
        header_mismatch = False
        row_count_mismatch = False
        if file_exists:
            try:
                with open(csv_filename, 'r', encoding='utf-8-sig') as f:
                    reader = csv.reader(f)
                    existing_header = next(reader, None)
                    
                    # 헤더 불일치 확인
                    if existing_header != fixed_columns:
                        header_mismatch = True
                    
                    # 기존 데이터 모두 읽기
                    existing_rows = []
                    expected_col_count = len(fixed_columns)
                    for row in reader:
                        # 행의 컬럼 수가 고정 컬럼 수와 맞지 않으면 빈 값으로 채움
                        if len(row) != expected_col_count:
                            row_count_mismatch = True
                            # 부족한 컬럼은 빈 문자열로 채움, 초과하는 컬럼은 무시
                            if len(row) < expected_col_count:
                                row = row + [''] * (expected_col_count - len(row))
                            else:
                                row = row[:expected_col_count]
                        existing_rows.append(row)
                    
                    # 헤더 불일치 또는 행 컬럼 수 불일치 시 재작성
                    if header_mismatch or row_count_mismatch:
                        if header_mismatch:
                            print(f"{get_timestamp()} {stage_prefix} ⚠️ CSV 헤더 불일치 감지. 고정 컬럼 구조로 재작성합니다.")
                        if row_count_mismatch:
                            print(f"{get_timestamp()} {stage_prefix} ⚠️ CSV 행 컬럼 수 불일치 감지. 고정 컬럼 구조로 재작성합니다.")
                        
                        # 기존 데이터를 고정 컬럼 구조에 맞춰 변환
                        # 기존 헤더를 딕셔너리로 변환하여 매핑
                        existing_header_dict = {}
                        if existing_header:
                            existing_header_dict = {col: i for i, col in enumerate(existing_header) if col}
                        
                        # 새 헤더로 파일 재작성
                        with open(csv_filename, 'w', newline='', encoding='utf-8-sig') as f_write:
                            writer = csv.DictWriter(f_write, fieldnames=fixed_columns)
                            writer.writeheader()
                            
                            # 기존 데이터 행을 고정 컬럼 구조에 맞춰 변환
                            for row in existing_rows:
                                row_dict_converted = {}
                                for col in fixed_columns:
                                    if col in existing_header_dict:
                                        idx = existing_header_dict[col]
                                        if idx < len(row):
                                            row_dict_converted[col] = row[idx]
                                        else:
                                            row_dict_converted[col] = ''
                                    else:
                                        # 헤더에 없는 컬럼은 순서대로 매핑 시도
                                        if len(row) > len(existing_header_dict):
                                            # 행의 컬럼 수가 헤더보다 많으면 순서대로 매핑
                                            col_idx = fixed_columns.index(col) if col in fixed_columns else -1
                                            if 0 <= col_idx < len(row):
                                                row_dict_converted[col] = row[col_idx]
                                            else:
                                                row_dict_converted[col] = ''
                                        else:
                                            row_dict_converted[col] = ''
                                writer.writerow(row_dict_converted)
                            
                            # 새 데이터도 함께 작성
                            writer.writerow(row_dict)
                        
                        print(f"{get_timestamp()} {stage_prefix} ✅ CSV 고정 컬럼 구조로 재작성 완료. 기존 데이터 보존 및 새 데이터 추가됨.")
                        return  # 이미 데이터를 작성했으므로 함수 종료
            except Exception as e:
                print(f"{get_timestamp()} {stage_prefix} ⚠️ 기존 CSV 파일 확인 중 오류: {e}")
                import traceback
                traceback.print_exc()
                header_mismatch = False  # 오류 발생 시 새로 작성
        
        # 파일 쓰기 (고정 컬럼 구조 사용)
        with open(csv_filename, 'a', newline='', encoding='utf-8-sig') as f:
            writer = csv.DictWriter(f, fieldnames=fixed_columns)
            # 파일이 없으면 헤더 작성
            if not file_exists:
                writer.writeheader()
            # 데이터 작성 (고정 컬럼 순서에 맞춰, 없는 값은 빈 문자열로 채움)
            row_to_save = {col: row_dict.get(col, '') for col in fixed_columns}
            writer.writerow(row_to_save)
        
        # 상세 출력 메시지 (티커마다 한 줄씩)
        print(f"{get_timestamp()} {stage_prefix} 🏁🏁🏁🏁🏁📸자산 기록 저장 완료🏁🏁🏁🏁🏁")
        print(f"{get_timestamp()} {stage_prefix} 총자산: {total_equity:,.2f} USDT | 현금: {total_usdt:,.2f} USDT | 코인평가: {total_coin_valuation:,.2f} USDT")
        print(f"{get_timestamp()} {stage_prefix} 일일 PnL: {daily_pnl:+,.2f} USDT | 누적 PnL: {total_accumulated_pnl:+,.2f} USDT ({total_return_rate:+.2f}%)")
        
        # 코인별 정보 (티커마다 한 줄씩)
        for ticker in ROTATION_TICKERS:
            data = ticker_data.get(ticker, {'principal': 0, 'val': 0, 'pnl': 0, 'rate': 0})
            if data['val'] > 0:  # 보유 중인 코인만 표시
                print(f"{get_timestamp()} {stage_prefix} {ticker}: 평가금 {data['val']:,.2f} USDT (원금: {data['principal']:,.2f} USDT, 손익: {data['pnl']:+,.2f} USDT, {data['rate']:+.2f}%)")
        
        # 디스코드 메시지 전송 (콘솔 로그와 동일한 형식)
        try:
            # 콘솔 로그와 동일한 형식으로 디스코드 메시지 생성
            msg_lines = [
                f"{get_timestamp()} {stage_prefix} 🏁🏁🏁🏁🏁📸자산 기록 저장 완료🏁🏁🏁🏁🏁",
                f"{get_timestamp()} {stage_prefix} 총자산: {total_equity:,.2f} USDT | 현금: {total_usdt:,.2f} USDT | 코인평가: {total_coin_valuation:,.2f} USDT",
                f"{get_timestamp()} {stage_prefix} 일일 PnL: {daily_pnl:+,.2f} USDT | 누적 PnL: {total_accumulated_pnl:+,.2f} USDT ({total_return_rate:+.2f}%)"
            ]
            
            # 코인별 정보 추가
            for ticker in ROTATION_TICKERS:
                data = ticker_data.get(ticker, {'principal': 0, 'val': 0, 'pnl': 0, 'rate': 0})
                if data['val'] > 0:  # 보유 중인 코인만 표시
                    msg_lines.append(f"{get_timestamp()} {stage_prefix} {ticker}: 평가금 {data['val']:,.2f} USDT (원금: {data['principal']:,.2f} USDT, 손익: {data['pnl']:+,.2f} USDT, {data['rate']:+.2f}%)")
            
            msg = "\n".join(msg_lines)
            send_discord_message(msg)
        except Exception:
            pass  # 디스코드 전송 실패해도 자산기록은 성공으로 처리
        
    except Exception as e:
        print(f"{get_timestamp()} ⚠️자산기록 저장 실패: {e}")
        import traceback
        traceback.print_exc()

# ==========================================
# [체인 유지관리] 엑셀 파일의 최신 캔들 시간을 확인하여 체인 유효성 판단
# ==========================================
def check_recent_after_files_exist():
    """
    모든 로테이션 티커에 대해 최신 after_ 엑셀 파일을 확인하여
    현재 시점에서 '있어야 할' 최신 완성 캔들 데이터가 존재하는지 검증합니다.
    
    [로직]
    1. 현재 시간 기준으로 '예상되는 최신 완성 캔들 시간'을 계산합니다.
       예) 12:40 실행 -> 현재 구간(12:30~12:45) -> 12:30 캔들은 미완성 -> 12:15 캔들이 최신 완성본이어야 함.
    2. 엑셀 파일의 최신 행(Date)이 이 시간과 정확히 일치하는지 확인합니다.
    3. (성공 여부, 실패한 티커 목록) 튜플을 반환합니다.
    
    Returns:
        (bool, List[str]): (모든 티커 성공 여부, 실패한 티커 목록)
    """
    script_dir = os.path.dirname(os.path.abspath(__file__))
    base_dir = os.path.join(script_dir, "cryptodaily15min")
    
    # Binance용 티커 폴더 매핑 (공백 포함)
    ticker_folder_mapping = {
        "BTC": "F BINANCE 1BTC",
        "ETH": "F BINANCE 2ETH",
        "XRP": "F BINANCE 3XRP",
        "SOL": "F BINANCE 4SOL",
        "BNB": "F BINANCE 5BNB"
    }
    
    # 1. 현재 시간 기준 '예상 최신 캔들 시간(UTC)' 계산
    now_utc = dt.datetime.now(tz.UTC)
    
    # 15분 단위 내림 (현재 진행중인 캔들의 시작 시간)
    current_block_minute = (now_utc.minute // 15) * 15
    current_candle_start = now_utc.replace(minute=current_block_minute, second=0, microsecond=0)
    
    # 완성된 최신 캔들은 '현재 진행중인 캔들'의 바로 전 캔들 (15분 전)
    # 예: 12:40 실행 → 현재 진행중: 12:30~12:45 (12:30 시작) → 완성본: 12:15~12:30 (12:15 시작)
    expected_candle_time = current_candle_start - dt.timedelta(minutes=15)
    
    # 포맷팅 (비교용 문자열)
    expected_date_str = expected_candle_time.strftime("%y/%m/%d")
    expected_time_str = expected_candle_time.strftime("%H:%M")
    
    print(f"{get_timestamp()} [초기화] 🔍체인 검증 기준: {expected_date_str},{expected_time_str} 캔들이 있어야 합니다.")
    # 활성화된 티커만 검사 (ROTATION_TICKERS의 모든 티커 검사)
    active_tickers = ROTATION_TICKERS.copy()
    if not active_tickers:
        print(f"{get_timestamp()} [초기화] ⚠️활성화된 티커가 없습니다.")
        return (False, active_tickers)  # 모든 티커가 실패로 처리
    print(f"{get_timestamp()} [초기화] 📋체인 검증 대상 티커: {', '.join(active_tickers)} ({len(active_tickers)}개)")
    
    failed_tickers = []
    try:
        for ticker in active_tickers:
            print(f"{get_timestamp()} [초기화] 🔍{ticker} 체인 검증 중")
            folder_name = ticker_folder_mapping.get(ticker, f"F BINANCE_{ticker}")
            ticker_folder = os.path.join(base_dir, folder_name)
            
            if not os.path.exists(ticker_folder):
                print(f"{get_timestamp()} [초기화] ⚠️{ticker} 폴더가 없습니다.")
                failed_tickers.append(ticker)
                continue
                
            # 해당 폴더에서 가장 최신의 after_ 파일 찾기
            after_mtime = 0
            after_file_path = None
            
            try:
                for f in os.listdir(ticker_folder):
                    if f.startswith(f"after_F_{ticker}_BINANCE_") and f.endswith(".xlsx") and not f.startswith("~$"):
                        file_path = os.path.join(ticker_folder, f)
                        try:
                            mtime = os.path.getmtime(file_path)
                            if mtime > after_mtime:
                                after_mtime = mtime
                                after_file_path = file_path
                        except OSError:
                            continue
            except Exception:
                pass
            
            # 해당 폴더에서 가장 최신의 previous_ 파일 찾기
            previous_mtime = 0
            previous_file_path = None
            
            try:
                for f in os.listdir(ticker_folder):
                    if f.startswith(f"previous_F_{ticker}_BINANCE_") and f.endswith(".xlsx") and not f.startswith("~$"):
                        file_path = os.path.join(ticker_folder, f)
                        try:
                            mtime = os.path.getmtime(file_path)
                            if mtime > previous_mtime:
                                previous_mtime = mtime
                                previous_file_path = file_path
                        except OSError:
                            continue
            except Exception:
                pass
                
            if not after_file_path and not previous_file_path:
                print(f"{get_timestamp()} [초기화] ⚠️{ticker} After/Previous 파일이 없습니다.")
                failed_tickers.append(ticker)
                continue
            
            # 체인 검증: After 파일 먼저 확인, 일치하지 않으면 Previous 파일 확인
            found_match = False
            checked_files = []
            
            # After 파일 확인
            if after_file_path:
                try:
                    df = pd.read_excel(after_file_path, sheet_name=f"{ticker}USDT15M", nrows=2)
                    
                    if 'Date(UTC)' in df.columns and len(df) > 0:
                        latest_date_val = str(df.iloc[0]['Date(UTC)']).strip()
                        try:
                            clean_date_str = latest_date_val.replace(',', ' ')
                            file_dt = dt.datetime.strptime(clean_date_str, "%y/%m/%d %H:%M").replace(tzinfo=tz.UTC)
                            
                            if file_dt.replace(second=0, microsecond=0) == expected_candle_time.replace(second=0, microsecond=0):
                                found_match = True
                                print(f"{get_timestamp()} [초기화] ✅{ticker} 체인 검증 완료 (After: {file_dt.strftime('%H:%M')})")
                            else:
                                checked_files.append(f"After: {file_dt.strftime('%H:%M')}")
                        except ValueError:
                            checked_files.append("After: 파싱 실패")
                    else:
                        checked_files.append("After: 형식 오류")
                except Exception as e:
                    checked_files.append(f"After: 읽기 실패 ({str(e)[:30]})")
            
            # Previous 파일 확인 (After에서 일치하지 않은 경우에만)
            if not found_match and previous_file_path:
                try:
                    df_prev = pd.read_excel(previous_file_path, sheet_name=f"{ticker}USDT15M", nrows=2)
                    
                    if 'Date(UTC)' in df_prev.columns and len(df_prev) > 0:
                        prev_date_val = str(df_prev.iloc[0]['Date(UTC)']).strip()
                        try:
                            clean_date_str = prev_date_val.replace(',', ' ')
                            file_dt = dt.datetime.strptime(clean_date_str, "%y/%m/%d %H:%M").replace(tzinfo=tz.UTC)
                            
                            if file_dt.replace(second=0, microsecond=0) == expected_candle_time.replace(second=0, microsecond=0):
                                found_match = True
                                print(f"{get_timestamp()} [초기화] ✅{ticker} 체인 검증 완료 (Previous: {file_dt.strftime('%H:%M')})")
                            else:
                                checked_files.append(f"Previous: {file_dt.strftime('%H:%M')}")
                        except ValueError:
                            checked_files.append("Previous: 파싱 실패")
                    else:
                        checked_files.append("Previous: 형식 오류")
                except Exception as e:
                    checked_files.append(f"Previous: 읽기 실패 ({str(e)[:30]})")
            
            # 매칭되지 않은 경우
            if not found_match:
                files_info = ", ".join(checked_files) if checked_files else "파일 없음"
                print(f"{get_timestamp()} [초기화] ⚠️{ticker} 체인 불일치 ({files_info} != 예상: {expected_time_str})")
                failed_tickers.append(ticker)
    except KeyboardInterrupt:
        print(f"\n{get_timestamp()} [초기화] 🛑 체인 검증 중 사용자에 의해 중단되었습니다.")
        raise  # 상위로 전파하여 메인 예외 처리에서도 처리되도록
    
    # 모든 티커가 성공했는지 확인
    all_success = len(failed_tickers) == 0
    passed_tickers = [t for t in active_tickers if t not in failed_tickers]
    
    if passed_tickers:
        print(f"{get_timestamp()} [초기화] ✅체인 검증 통과 티커: {', '.join(passed_tickers)} ({len(passed_tickers)}개)")
    if not all_success:
        print(f"{get_timestamp()} [초기화] ⚠️체인 검증 실패 티커: {', '.join(failed_tickers)} ({len(failed_tickers)}개)")
    
    return (all_success, failed_tickers)

# --- 바이낸스 공용 GET: 대체 도메인 + 재시도/백오프 ---
BINANCE_BASES = [
    "https://api.binance.com",
    "https://api1.binance.com",
    "https://api2.binance.com",
    "https://api3.binance.com",
]

# 캔들 수집용: 바이낸스 USDT-M 선물 API 베이스
BINANCE_FUTURES_BASES = [
    "https://fapi.binance.com",
    "https://fapi1.binance.com",
    "https://fapi2.binance.com",
    "https://fapi3.binance.com",
]

def _binance_get(path, params, timeout=20, max_retries=5, pause=0.05):
    last_err = None
    backoff = pause
    for _ in range(max_retries):
        for base in BINANCE_BASES:
            url = f"{base}{path}"
            try:
                r = requests.get(
                    url, params=params, timeout=timeout,
                    headers={"User-Agent": "Mozilla/5.0"}
                )
                if r.status_code == 200:
                    return r
                if r.status_code in (418, 429) or 500 <= r.status_code < 600:
                    # 레이트리밋/서버 오류: 잠깐 대기 후 다음 베이스/재시도
                    time.sleep(backoff)
                    last_err = Exception(f"Binance HTTP {r.status_code}: {r.text[:200]}")
                    continue
                r.raise_for_status()
            except Exception as e:
                last_err = e
                time.sleep(backoff)
        backoff = min(backoff * 2, 1.5)  # 지수 백오프(최대 1.5초)
    raise last_err if last_err else RuntimeError("Binance request failed")

def _binance_futures_get(path, params, timeout=20, max_retries=5, pause=0.05):
    """캔들 수집 전용: 바이낸스 USDT-M 선물 API GET (대체 도메인 + 재시도)."""
    last_err = None
    backoff = pause
    for _ in range(max_retries):
        for base in BINANCE_FUTURES_BASES:
            url = f"{base}{path}"
            try:
                r = requests.get(
                    url, params=params, timeout=timeout,
                    headers={"User-Agent": "Mozilla/5.0"}
                )
                if r.status_code == 200:
                    return r
                if r.status_code in (418, 429) or 500 <= r.status_code < 600:
                    time.sleep(backoff)
                    last_err = Exception(f"Binance Futures HTTP {r.status_code}: {r.text[:200]}")
                    continue
                r.raise_for_status()
            except Exception as e:
                last_err = e
                time.sleep(backoff)
        backoff = min(backoff * 2, 1.5)
    raise last_err if last_err else RuntimeError("Binance Futures request failed")

# (업비트 API 제거)

# -------------------- 바이낸스 일봉 (선물 캔들) --------------------
BINANCE_KLINES = "https://fapi.binance.com/fapi/v1/klines"
BINANCE_LIMIT = 1000

def fetch_binance_daily(symbol: str, total_days: int, include_today: bool = False, fixed_end_time_ms: Optional[int] = None) -> pd.DataFrame:
    """
    바이낸스 1d klines, UTC 기준으로 통일하여 반환.
    페이징 방법 B (backward): endTime부터 과거로 진행
    
    바이낸스 API 문서:
    - 스팟: GET /api/v3/klines — limit 기본 500, 최대 1000
    - 페이징 방법 B: endTime부터 시작, 가장 오래된 캔들의 openTime - 1ms로 갱신하여 과거로 진행
    - 선물의 경우 startTime-endTime 간 최대 200일 제약 (스팟은 제약 없음)
    
    Args:
        symbol: 심볼 (예: "BTCUSDT")
        total_days: 수집할 일봉 개수
        include_today: True면 오늘 진행중 캔들 포함 (UTC 기준으로 판단)
        fixed_end_time_ms: 조회 기준 시간 (UTC milliseconds, None이면 현재 시간)
    
    Returns:
        DataFrame: Date(UTC), KST, 종, 시, 고, 저, Vol. 컬럼 포함
        - Date(UTC): UTC 기준 시간 (계산에 사용)
        - KST: 참고용 (계산 로직에는 사용하지 않음)
    """
    path = "/fapi/v1/klines"
    all_rows: List[list] = []
    seen = set()  # 이미 본 open_time 추적 (중복 방지)
    remaining = total_days
    current_end_time = fixed_end_time_ms  # 현재 페이징의 endTime (과거로 진행하면서 갱신됨)

    while remaining > 0:
        # 바이낸스 스팟 API: limit 최대 1000
        limit = min(1000, remaining)
        params = {"symbol": symbol, "interval": "1d", "limit": limit}
        
        # 페이징 방법 B: endTime 사용 (과거로 진행)
        if current_end_time is not None:
            params["endTime"] = current_end_time
        # endTime이 None이면 가장 최근 캔들 반환 (바이낸스 API 기본 동작)

        r = _binance_futures_get(path, params)
        batch = r.json()
        if not batch:
            break

        # 중복 방지: 이미 본 open_time은 제외
        new_rows_count = 0
        for row in batch:
            ot = row[0]  # open_time (ms)
            if ot in seen:
                continue
            seen.add(ot)
            all_rows.append(row)
            new_rows_count += 1
        
        # 페이징 방법 B: 가장 오래된 캔들의 openTime - 1ms로 endTime 갱신 (과거로 진행)
        # batch는 시간순 정렬되어 있으므로 [0]이 가장 오래된 캔들
        earliest_open = batch[0][0]  # 배치의 가장 오래된 봉 open_time
        current_end_time = earliest_open - 1  # 다음 요청: 더 과거로
        
        # remaining 감소: 실제로 추가된 새 행 개수만큼만 감소
        remaining -= new_rows_count
        
        # 더 이상 새 데이터가 없으면 종료
        if new_rows_count == 0:
            break
        
        time.sleep(0.05)  # API 호출 간 안전 여유 (최적화: 0.1 → 0.05)

    if not all_rows:
        return pd.DataFrame()

    cols = ["open_time","open","high","low","close","volume","close_time",
            "quote_asset_volume","trades","taker_buy_base","taker_buy_quote","ignore"]
    df = pd.DataFrame(all_rows, columns=cols)

    # 안정 정렬 + 중복제거(최종 안전벨트)
    df["open_time"] = df["open_time"].astype("int64", copy=False)
    df = df.sort_values("open_time", kind="mergesort", ignore_index=True)
    df = df.loc[~df["open_time"].duplicated(keep="last")].reset_index(drop=True)

    # UTC 기준 날짜 사용 - [수정] Timestamp 객체로 반환 (문자열 변환 제거)
    df["DateUTC"] = pd.to_datetime(df["open_time"], unit="ms", utc=True)
    # [수정] strftime 제거: Timestamp 객체 그대로 사용 (나중에 엑셀 저장 시에만 문자열로 변환)
    df["DateUTC_dt"] = df["DateUTC"].dt.tz_localize(None)  # timezone 제거하여 naive datetime으로 변환

    # 심볼에 따른 식별자 설정
    identifier = symbol.replace("USDT", "") + "USD"

    # KST 시간 계산 (UTC+9) - 참고용으로만 사용 (계산 로직에는 사용하지 않음)
    # 일봉은 UTC 00:00 = KST 09:00
    kst_dates = []
    for utc_dt in df["DateUTC"]:
        # KST로 변환 (일봉은 UTC 00:00 = KST 09:00)
        kst_dt = utc_dt.astimezone(KST)
        kst_dates.append(kst_dt.strftime("%y/%m/%d,09:00"))

    out = pd.DataFrame({
        "Date(UTC)": df["DateUTC_dt"],  # UTC 기준 시간 (계산에 사용, Timestamp 객체)
        "KST": kst_dates,  # KST 시간 (참고용, 계산 로직에는 사용하지 않음, YY/MM/DD,09:00 형식)
        "종": pd.to_numeric(df["close"], errors="coerce"),
        "시": pd.to_numeric(df["open"], errors="coerce"),
        "고": pd.to_numeric(df["high"], errors="coerce"),
        "저": pd.to_numeric(df["low"], errors="coerce"),
        "Vol.": pd.to_numeric(df["volume"], errors="coerce"),
    })

    if not include_today:
        # UTC 기준으로 오늘 날짜 판단 (미완성 캔들 제거)
        today_utc = dt.datetime.now(tz.UTC).date()
        # 문자열 날짜를 date 객체로 변환하여 비교 (UTC 기준, YY/MM/DD,00:00 형식)
        out["Date_only"] = out["Date(UTC)"].apply(lambda x: x.split(',')[0] if ',' in str(x) else str(x))
        out = out[pd.to_datetime(out["Date_only"], format="%y/%m/%d").dt.date < today_utc]
        out = out.drop("Date_only", axis=1)

    out = out.sort_values("Date(UTC)", ascending=False).reset_index(drop=True)
    return out

# (업비트 API 제거)



def fetch_binance_minutes1(symbol: str, total_count: int, include_today: bool = False, fixed_end_time_ms: Optional[int] = None, stage_prefix: str = "") -> pd.DataFrame:
    """
    바이낸스 1m klines, UTC 기준으로 통일하여 반환.
    페이징 방법 B (backward): endTime부터 과거로 진행
    
    바이낸스 API 문서:
    - 스팟: GET /api/v3/klines — limit 기본 500, 최대 1000
    - 페이징 방법 B: endTime부터 시작, 가장 오래된 캔들의 openTime - 1ms로 갱신하여 과거로 진행
    
    Args:
        symbol: 심볼 (예: "BTCUSDT")
        total_count: 수집할 1분봉 개수
        include_today: True면 오늘 진행중 캔들 포함 (UTC 기준으로 판단)
        fixed_end_time_ms: 조회 기준 시간 (UTC milliseconds, None이면 현재 시간)
    
    Returns:
        DataFrame: Date(UTC), KST, 종, 시, 고, 저, Vol. 컬럼 포함
        - Date(UTC): UTC 기준 시간 (계산에 사용)
        - KST: 참고용 (계산 로직에는 사용하지 않음)
    """
    path = "/fapi/v1/klines"
    all_rows: List[list] = []
    seen = set()  # 이미 본 open_time 추적 (중복 방지)
    remaining = total_count
    current_end_time = fixed_end_time_ms  # 현재 페이징의 endTime (과거로 진행하면서 갱신됨)

    while remaining > 0:
        # 바이낸스 스팟 API: limit 최대 1000
        limit = min(1000, remaining)
        params = {"symbol": symbol, "interval": "1m", "limit": limit}
        
        # 페이징 방법 B: endTime 사용 (과거로 진행)
        if current_end_time is not None:
            params["endTime"] = current_end_time
        # endTime이 None이면 가장 최근 캔들 반환 (바이낸스 API 기본 동작)

        r = _binance_futures_get(path, params)
        batch = r.json()
        if not batch:
            break

        # 중복 방지: 이미 본 open_time은 제외
        new_rows_count = 0
        for row in batch:
            ot = row[0]  # open_time (ms)
            if ot in seen:
                continue
            seen.add(ot)
            all_rows.append(row)
            new_rows_count += 1
        
        # 페이징 방법 B: 가장 오래된 캔들의 openTime - 1ms로 endTime 갱신 (과거로 진행)
        # batch는 시간순 정렬되어 있으므로 [0]이 가장 오래된 캔들
        earliest_open = batch[0][0]  # 배치의 가장 오래된 봉 open_time
        current_end_time = earliest_open - 1  # 다음 요청: 더 과거로
        
        # remaining 감소: 실제로 추가된 새 행 개수만큼만 감소
        remaining -= new_rows_count
        
        # 더 이상 새 데이터가 없으면 종료
        if new_rows_count == 0:
            break
        
        time.sleep(0.05)  # API 호출 간 안전 여유

    if not all_rows:
        return pd.DataFrame()

    cols = ["open_time","open","high","low","close","volume","close_time",
            "quote_asset_volume","trades","taker_buy_base","taker_buy_quote","ignore"]
    df = pd.DataFrame(all_rows, columns=cols)

    # 안정 정렬 + 중복제거(최종 안전벨트)
    df["open_time"] = df["open_time"].astype("int64", copy=False)
    df = df.sort_values("open_time", kind="mergesort", ignore_index=True)
    df = df.loc[~df["open_time"].duplicated(keep="last")].reset_index(drop=True)

    # UTC 기준 날짜 사용 - [수정] Timestamp 객체로 반환 (문자열 변환 제거)
    df["DateUTC"] = pd.to_datetime(df["open_time"], unit="ms", utc=True)
    # [수정] strftime 제거: Timestamp 객체 그대로 사용 (나중에 엑셀 저장 시에만 문자열로 변환)
    df["DateUTC_dt"] = df["DateUTC"].dt.tz_localize(None)  # timezone 제거하여 naive datetime으로 변환

    # KST 시간 계산 (UTC+9) - 참고용으로만 사용 (계산 로직에는 사용하지 않음)
    kst_dates = []
    for utc_dt in df["DateUTC"]:
        kst_dt = utc_dt.astimezone(KST)
        kst_dates.append(kst_dt.strftime("%y/%m/%d,%H:%M"))

    out = pd.DataFrame({
        "Date(UTC)": df["DateUTC_dt"],  # UTC 기준 시간 (계산에 사용, Timestamp 객체)
        "KST": kst_dates,  # KST 시간 (참고용, 계산 로직에는 사용하지 않음, YY/MM/DD,HH:MM 형식)
        "종": pd.to_numeric(df["close"], errors="coerce"),
        "시": pd.to_numeric(df["open"], errors="coerce"),
        "고": pd.to_numeric(df["high"], errors="coerce"),
        "저": pd.to_numeric(df["low"], errors="coerce"),
        "Vol.": pd.to_numeric(df["volume"], errors="coerce"),
    })

    if not include_today:
        # UTC 기준으로 오늘 날짜 판단 (미완성 캔들 제거)
        today_utc = dt.datetime.now(tz.UTC).date()
        # 문자열 날짜를 date 객체로 변환하여 비교 (UTC 기준, YY/MM/DD,HH:MM 형식)
        out["Date_only"] = out["Date(UTC)"].apply(lambda x: x.split(',')[0] if ',' in str(x) else str(x))
        out = out[pd.to_datetime(out["Date_only"], format="%y/%m/%d").dt.date < today_utc]
        out = out.drop("Date_only", axis=1)

    out = out.sort_values("Date(UTC)", ascending=False).reset_index(drop=True)
    
    # ⚠️중요: 1분봉 간격 검증 (페이징 누락 방지 확인)
    if len(out) > 1 and 'Date(UTC)' in out.columns:
        # Date(UTC)를 datetime으로 변환하여 간격 검증
        df_dt = pd.to_datetime(out['Date(UTC)'], format='%y/%m/%d,%H:%M', errors='coerce')
        # 내림차순 정렬 확인 (최신→과거)
        if df_dt.is_monotonic_decreasing:
            # 각 행과 다음 행의 시간 차이 계산 (분 단위)
            time_diffs = (df_dt.iloc[:-1].values - df_dt.iloc[1:].values) / np.timedelta64(1, 'm')
            # 1분이 아닌 간격이 있는지 확인
            non_1m_indices = np.where((time_diffs != 1) & (~np.isnan(time_diffs)))[0]
            if len(non_1m_indices) > 0:
                print(f"{get_timestamp()} [{stage_prefix}] ⚠️[1분봉 검증] {symbol}: 1분 간격이 아닌 구간 발견 ({len(non_1m_indices)}개)")
                for idx in non_1m_indices[:5]:  # 최대 5개만 출력
                    diff_minutes = time_diffs[idx]
                    print(f"{get_timestamp()} [{stage_prefix}]    인덱스 {idx}→{idx+1}: {out.iloc[idx]['Date(UTC)']} → {out.iloc[idx+1]['Date(UTC)']} ({diff_minutes:.1f}분 차이)")
                if len(non_1m_indices) > 5:
                    print(f"{get_timestamp()} [{stage_prefix}]    ... 외 {len(non_1m_indices)-5}개 구간")
            else:
                print(f"{get_timestamp()} [{stage_prefix}] ✅[1분봉 검증] {symbol}: 모든 간격이 1분입니다 ({len(out)}개 캔들)")
    
    return out

# -------------------- 바이낸스 15분봉 (선물) --------------------
BINANCE_KLINES_15M = "https://fapi.binance.com/fapi/v1/klines"
BINANCE_LIMIT = 1000

# -------------------- 바이낸스 주봉 (선물) --------------------
BINANCE_KLINES_WEEKLY = "https://fapi.binance.com/fapi/v1/klines"

def fetch_binance_minutes15(symbol: str, total_count: int, include_today: bool = False, fixed_end_time_ms: Optional[int] = None) -> pd.DataFrame:
    """
    바이낸스 15m klines, UTC 기준으로 통일하여 반환.
    페이징 방법 B (backward): endTime부터 과거로 진행
    
    바이낸스 API 문서:
    - 스팟: GET /api/v3/klines — limit 기본 500, 최대 1000
    - 페이징 방법 B: endTime부터 시작, 가장 오래된 캔들의 openTime - 1ms로 갱신하여 과거로 진행
    
    Args:
        symbol: 심볼 (예: "BTCUSDT")
        total_count: 수집할 15분봉 개수
        include_today: True면 오늘 진행중 캔들 포함 (UTC 기준으로 판단)
        fixed_end_time_ms: 조회 기준 시간 (UTC milliseconds, None이면 현재 시간)
    
    Returns:
        DataFrame: Date(UTC), KST, 종, 시, 고, 저, Vol. 컬럼 포함
        - Date(UTC): UTC 기준 시간 (계산에 사용)
        - KST: 참고용 (계산 로직에는 사용하지 않음)
    """
    path = "/fapi/v1/klines"
    all_rows: List[list] = []
    seen = set()  # 이미 본 open_time 추적 (중복 방지)
    remaining = total_count
    current_end_time = fixed_end_time_ms  # 현재 페이징의 endTime (과거로 진행하면서 갱신됨)

    while remaining > 0:
        # 바이낸스 스팟 API: limit 최대 1000
        limit = min(1000, remaining)
        params = {"symbol": symbol, "interval": "15m", "limit": limit}
        
        # 페이징 방법 B: endTime 사용 (과거로 진행)
        if current_end_time is not None:
            params["endTime"] = current_end_time
        # endTime이 None이면 가장 최근 캔들 반환 (바이낸스 API 기본 동작)

        r = _binance_futures_get(path, params)
        batch = r.json()
        if not batch:
            break

        # 중복 방지: 이미 본 open_time은 제외
        new_rows_count = 0
        for row in batch:
            ot = row[0]  # open_time (ms)
            if ot in seen:
                continue
            seen.add(ot)
            all_rows.append(row)
            new_rows_count += 1
        
        # 페이징 방법 B: 가장 오래된 캔들의 openTime - 1ms로 endTime 갱신 (과거로 진행)
        # batch는 시간순 정렬되어 있으므로 [0]이 가장 오래된 캔들
        earliest_open = batch[0][0]  # 배치의 가장 오래된 봉 open_time
        current_end_time = earliest_open - 1  # 다음 요청: 더 과거로
        
        # remaining 감소: 실제로 추가된 새 행 개수만큼만 감소
        remaining -= new_rows_count
        
        # 더 이상 새 데이터가 없으면 종료
        if new_rows_count == 0:
            break
        
        time.sleep(0.05)  # API 호출 간 안전 여유 (최적화: 0.1 → 0.05)

    if not all_rows:
        return pd.DataFrame()

    cols = ["open_time","open","high","low","close","volume","close_time",
            "quote_asset_volume","trades","taker_buy_base","taker_buy_quote","ignore"]
    df = pd.DataFrame(all_rows, columns=cols)

    # 시간순 정렬 및 중복 제거 (최신→과거 순서로 정렬)
    df['open_time'] = df['open_time'].astype('int64', copy=False)
    df = df.sort_values('open_time', kind='mergesort', ascending=False, ignore_index=True)
    df = df.loc[~df['open_time'].duplicated(keep='last')].reset_index(drop=True)

    # UTC 기준으로 통일 - [수정] Timestamp 객체로 반환 (문자열 변환 제거)
    df["DateUTC"] = pd.to_datetime(df["open_time"], unit="ms", utc=True)
    # [수정] strftime 제거: Timestamp 객체 그대로 사용 (나중에 엑셀 저장 시에만 문자열로 변환)
    df["DateUTC_dt"] = df["DateUTC"].dt.tz_localize(None)  # timezone 제거하여 naive datetime으로 변환

    # 심볼에 따른 식별자 설정
    identifier = symbol.replace("USDT", "") + "USD15"

    # KST 시간 계산 (UTC+9) - 참고용으로만 사용 (계산 로직에는 사용하지 않음)
    kst_dates = []
    for utc_dt in df["DateUTC"]:
        # KST로 변환 (참고용)
        kst_dt = utc_dt.astimezone(KST)
        kst_dates.append(kst_dt.strftime("%y/%m/%d,%H:%M"))

    out = pd.DataFrame({
        "Date(UTC)": df["DateUTC_dt"],  # UTC 기준 시간 (계산에 사용, Timestamp 객체)
        "KST": kst_dates,  # KST 시간 (참고용, 계산 로직에는 사용하지 않음)
        "종": pd.to_numeric(df["close"], errors="coerce"),
        "시": pd.to_numeric(df["open"], errors="coerce"),
        "고": pd.to_numeric(df["high"], errors="coerce"),
        "저": pd.to_numeric(df["low"], errors="coerce"),
        "Vol.": pd.to_numeric(df["volume"], errors="coerce"),
    })

    if not include_today:
        # UTC 기준으로 오늘 날짜 판단 (미완성 캔들 제거)
        today_utc = dt.datetime.now(tz.UTC).date()
        # Timestamp 객체에서 date만 추출하여 비교
        out = out[out["Date(UTC)"].dt.date < today_utc]

    out = out.sort_values("Date(UTC)", ascending=False).reset_index(drop=True)

    return out

def fetch_binance_minutes5(symbol: str, total_count: int, include_today: bool = False, fixed_end_time_ms: Optional[int] = None) -> pd.DataFrame:
    """
    바이낸스 5m klines, UTC 기준으로 통일하여 반환.
    페이징 방법 B (backward): endTime부터 과거로 진행
    
    바이낸스 API 문서:
    - 스팟: GET /api/v3/klines — limit 기본 500, 최대 1000
    - 페이징 방법 B: endTime부터 시작, 가장 오래된 캔들의 openTime - 1ms로 갱신하여 과거로 진행
    
    Args:
        symbol: 심볼 (예: "BTCUSDT")
        total_count: 수집할 5분봉 개수
        include_today: True면 오늘 진행중 캔들 포함 (UTC 기준으로 판단)
        fixed_end_time_ms: 조회 기준 시간 (UTC milliseconds, None이면 현재 시간)
    
    Returns:
        DataFrame: Date(UTC), KST, 종, 시, 고, 저, Vol. 컬럼 포함
        - Date(UTC): UTC 기준 시간 (계산에 사용)
        - KST: 참고용 (계산 로직에는 사용하지 않음)
    """
    path = "/fapi/v1/klines"
    all_rows: List[list] = []
    seen = set()  # 이미 본 open_time 추적 (중복 방지)
    remaining = total_count
    current_end_time = fixed_end_time_ms  # 현재 페이징의 endTime (과거로 진행하면서 갱신됨)

    while remaining > 0:
        # 바이낸스 스팟 API: limit 최대 1000
        limit = min(1000, remaining)
        params = {"symbol": symbol, "interval": "5m", "limit": limit}
        
        # 페이징 방법 B: endTime 사용 (과거로 진행)
        if current_end_time is not None:
            params["endTime"] = current_end_time
        # endTime이 None이면 가장 최근 캔들 반환 (바이낸스 API 기본 동작)

        r = _binance_futures_get(path, params)
        batch = r.json()
        if not batch:
            break

        # 중복 방지: 이미 본 open_time은 제외
        new_rows_count = 0
        for row in batch:
            ot = row[0]  # open_time (ms)
            if ot in seen:
                continue
            seen.add(ot)
            all_rows.append(row)
            new_rows_count += 1
        
        # 페이징 방법 B: 가장 오래된 캔들의 openTime - 1ms로 endTime 갱신 (과거로 진행)
        # batch는 시간순 정렬되어 있으므로 [0]이 가장 오래된 캔들
        earliest_open = batch[0][0]  # 배치의 가장 오래된 봉 open_time
        current_end_time = earliest_open - 1  # 다음 요청: 더 과거로
        
        # remaining 감소: 실제로 추가된 새 행 개수만큼만 감소
        remaining -= new_rows_count
        
        # 더 이상 새 데이터가 없으면 종료
        if new_rows_count == 0:
            break
        
        time.sleep(0.05)  # API 호출 간 안전 여유 (최적화: 0.1 → 0.05)

    if not all_rows:
        return pd.DataFrame()

    cols = ["open_time","open","high","low","close","volume","close_time",
            "quote_asset_volume","trades","taker_buy_base","taker_buy_quote","ignore"]
    df = pd.DataFrame(all_rows, columns=cols)

    df['open_time'] = df['open_time'].astype('int64', copy=False)
    df = df.sort_values('open_time', kind='mergesort', ascending=False, ignore_index=True)
    df = df.loc[~df['open_time'].duplicated(keep='last')].reset_index(drop=True)

    # [수정] Timestamp 객체로 반환 (문자열 변환 제거)
    df["DateUTC"] = pd.to_datetime(df["open_time"], unit="ms", utc=True)
    # [수정] strftime 제거: Timestamp 객체 그대로 사용 (나중에 엑셀 저장 시에만 문자열로 변환)
    df["DateUTC_dt"] = df["DateUTC"].dt.tz_localize(None)  # timezone 제거하여 naive datetime으로 변환

    # KST 시간 계산 (UTC+9) - 참고용으로만 사용 (계산 로직에는 사용하지 않음)
    kst_dates = []
    for utc_dt in df["DateUTC"]:
        kst_dt = utc_dt.astimezone(KST)  # 참고용
        kst_dates.append(kst_dt.strftime("%y/%m/%d,%H:%M"))

    out = pd.DataFrame({
        "Date(UTC)": df["DateUTC_dt"],  # UTC 기준 시간 (계산에 사용, Timestamp 객체)
        "KST": kst_dates,  # KST 시간 (참고용, 계산 로직에는 사용하지 않음)
        "종": pd.to_numeric(df["close"], errors="coerce"),
        "시": pd.to_numeric(df["open"], errors="coerce"),
        "고": pd.to_numeric(df["high"], errors="coerce"),
        "저": pd.to_numeric(df["low"], errors="coerce"),
        "Vol.": pd.to_numeric(df["volume"], errors="coerce"),
    })

    if not include_today:
        # UTC 기준으로 오늘 날짜 판단 (미완성 캔들 제거)
        today_utc = dt.datetime.now(tz.UTC).date()
        # Timestamp 객체에서 date만 추출하여 비교
        out = out[out["Date(UTC)"].dt.date < today_utc]

    out = out.sort_values("Date(UTC)", ascending=False).reset_index(drop=True)

    return out

def fetch_binance_hours1(symbol: str, total_count: int, include_today: bool = False, fixed_end_time_ms: Optional[int] = None) -> pd.DataFrame:
    """
    바이낸스 1h klines, UTC 기준으로 통일하여 반환.
    페이징 방법 B (backward): endTime부터 과거로 진행
    
    바이낸스 API 문서:
    - 스팟: GET /api/v3/klines — limit 기본 500, 최대 1000
    - 페이징 방법 B: endTime부터 시작, 가장 오래된 캔들의 openTime - 1ms로 갱신하여 과거로 진행
    
    Args:
        symbol: 심볼 (예: "BTCUSDT")
        total_count: 수집할 1시간봉 개수
        include_today: True면 오늘 진행중 캔들 포함 (UTC 기준으로 판단)
        fixed_end_time_ms: 조회 기준 시간 (UTC milliseconds, None이면 현재 시간)
    
    Returns:
        DataFrame: Date(UTC), KST, 종, 시, 고, 저, Vol. 컬럼 포함
        - Date(UTC): UTC 기준 시간 (계산에 사용, YY/MM/DD,00:00 형식)
        - KST: 참고용 (계산 로직에는 사용하지 않음, YY/MM/DD,09:00 형식)
    """
    path = "/fapi/v1/klines"
    all_rows: List[list] = []
    seen = set()  # 이미 본 open_time 추적 (중복 방지)
    remaining = total_count
    current_end_time = fixed_end_time_ms  # 현재 페이징의 endTime (과거로 진행하면서 갱신됨)

    while remaining > 0:
        # 바이낸스 스팟 API: limit 최대 1000
        limit = min(1000, remaining)
        params = {"symbol": symbol, "interval": "1h", "limit": limit}
        
        # 페이징 방법 B: endTime 사용 (과거로 진행)
        if current_end_time is not None:
            params["endTime"] = current_end_time
        # endTime이 None이면 가장 최근 캔들 반환 (바이낸스 API 기본 동작)

        r = _binance_futures_get(path, params)
        batch = r.json()
        if not batch:
            break

        # 중복 방지: 이미 본 open_time은 제외
        new_rows_count = 0
        for row in batch:
            ot = row[0]  # open_time (ms)
            if ot in seen:
                continue
            seen.add(ot)
            all_rows.append(row)
            new_rows_count += 1
        
        # 페이징 방법 B: 가장 오래된 캔들의 openTime - 1ms로 endTime 갱신 (과거로 진행)
        # batch는 시간순 정렬되어 있으므로 [0]이 가장 오래된 캔들
        earliest_open = batch[0][0]  # 배치의 가장 오래된 봉 open_time
        current_end_time = earliest_open - 1  # 다음 요청: 더 과거로
        
        # remaining 감소: 실제로 추가된 새 행 개수만큼만 감소
        remaining -= new_rows_count
        
        # 더 이상 새 데이터가 없으면 종료
        if new_rows_count == 0:
            break
        
        time.sleep(0.05)  # API 호출 간 안전 여유 (최적화: 0.1 → 0.05)

    if not all_rows:
        return pd.DataFrame()

    cols = ["open_time","open","high","low","close","volume","close_time",
            "quote_asset_volume","trades","taker_buy_base","taker_buy_quote","ignore"]
    df = pd.DataFrame(all_rows, columns=cols)

    df['open_time'] = df['open_time'].astype('int64', copy=False)
    df = df.sort_values('open_time', kind='mergesort', ascending=False, ignore_index=True)
    df = df.loc[~df['open_time'].duplicated(keep='last')].reset_index(drop=True)

    # [수정] Timestamp 객체로 반환 (문자열 변환 제거)
    df["DateUTC"] = pd.to_datetime(df["open_time"], unit="ms", utc=True)
    # 1시간봉은 각 캔들의 시작 시간 표시 (분, 초, 마이크로초 제거) - .dt.floor('h') 사용
    df["DateUTC_dt"] = df["DateUTC"].dt.floor('h').dt.tz_localize(None)  # timezone 제거하여 naive datetime으로 변환

    # KST 시간 계산 (UTC+9) - 참고용으로만 사용 (계산 로직에는 사용하지 않음)
    kst_dates = []
    for utc_dt in df["DateUTC"]:
        # 분, 초, 마이크로초 제거
        candle_start_time = utc_dt.replace(minute=0, second=0, microsecond=0)
        kst_dt = candle_start_time.astimezone(KST)  # 참고용
        kst_dates.append(kst_dt.strftime("%y/%m/%d,%H:00"))

    out = pd.DataFrame({
        "Date(UTC)": df["DateUTC_dt"],  # UTC 기준 시간 (계산에 사용, Timestamp 객체)
        "KST": kst_dates,  # KST 시간 (참고용, 계산 로직에는 사용하지 않음)
        "종": pd.to_numeric(df["close"], errors="coerce"),
        "시": pd.to_numeric(df["open"], errors="coerce"),
        "고": pd.to_numeric(df["high"], errors="coerce"),
        "저": pd.to_numeric(df["low"], errors="coerce"),
        "Vol.": pd.to_numeric(df["volume"], errors="coerce"),
    })

    if not include_today:
        # UTC 기준으로 오늘 날짜 판단 (미완성 캔들 제거)
        today_utc = dt.datetime.now(tz.UTC).date()
        # Timestamp 객체에서 date만 추출하여 비교
        out = out[out["Date(UTC)"].dt.date < today_utc]

    out = out.sort_values("Date(UTC)", ascending=False).reset_index(drop=True)

    return out

def fetch_binance_weekly(symbol: str, total_count: int, include_today: bool = False, fixed_end_time_ms: Optional[int] = None) -> pd.DataFrame:
    """
    바이낸스 주봉(1w) klines, UTC 기준으로 통일하여 반환.
    페이징 방법 B (backward): endTime부터 과거로 진행
    
    바이낸스 API 문서:
    - 스팟: GET /api/v3/klines — limit 기본 500, 최대 1000
    - 페이징 방법 B: endTime부터 시작, 가장 오래된 캔들의 openTime - 1ms로 갱신하여 과거로 진행
    
    Args:
        symbol: 심볼 (예: "BTCUSDT")
        total_count: 수집할 주봉 개수
        include_today: True면 오늘 진행중 캔들 포함 (UTC 기준으로 판단)
        fixed_end_time_ms: 조회 기준 시간 (UTC milliseconds, None이면 현재 시간)
    
    Returns:
        DataFrame: Date(UTC), KST, 종, 시, 고, 저, Vol. 컬럼 포함
        - Date(UTC): UTC 기준 시간 (계산에 사용)
        - KST: 참고용 (계산 로직에는 사용하지 않음)
    """
    path = "/fapi/v1/klines"
    all_rows: List[list] = []
    seen = set()  # 이미 본 open_time 추적 (중복 방지)
    remaining = total_count
    current_end_time = fixed_end_time_ms  # 현재 페이징의 endTime (과거로 진행하면서 갱신됨)

    while remaining > 0:
        # 바이낸스 스팟 API: limit 최대 1000
        limit = min(1000, remaining)
        params = {"symbol": symbol, "interval": "1w", "limit": limit}
        
        # 페이징 방법 B: endTime 사용 (과거로 진행)
        if current_end_time is not None:
            params["endTime"] = current_end_time
        # endTime이 None이면 가장 최근 캔들 반환 (바이낸스 API 기본 동작)

        r = _binance_futures_get(path, params)
        batch = r.json()
        if not batch:
            break

        # 중복 방지: 이미 본 open_time은 제외
        new_rows_count = 0
        for row in batch:
            ot = row[0]  # open_time (ms)
            if ot in seen:
                continue
            seen.add(ot)
            all_rows.append(row)
            new_rows_count += 1
        
        # 페이징 방법 B: 가장 오래된 캔들의 openTime - 1ms로 endTime 갱신 (과거로 진행)
        # batch는 시간순 정렬되어 있으므로 [0]이 가장 오래된 캔들
        earliest_open = batch[0][0]  # 배치의 가장 오래된 봉 open_time
        current_end_time = earliest_open - 1  # 다음 요청: 더 과거로
        
        # remaining 감소: 실제로 추가된 새 행 개수만큼만 감소
        remaining -= new_rows_count
        
        # 더 이상 새 데이터가 없으면 종료
        if new_rows_count == 0:
            break
        
        time.sleep(0.05)  # API 호출 간 안전 여유 (최적화: 0.1 → 0.05)

    if not all_rows:
        return pd.DataFrame()

    cols = ["open_time","open","high","low","close","volume","close_time",
            "quote_asset_volume","trades","taker_buy_base","taker_buy_quote","ignore"]
    df = pd.DataFrame(all_rows, columns=cols)

    # 안정 정렬 + 중복제거(최종 안전벨트)
    df["open_time"] = df["open_time"].astype("int64", copy=False)
    df = df.sort_values("open_time", kind="mergesort", ignore_index=True)
    df = df.loc[~df["open_time"].duplicated(keep="last")].reset_index(drop=True)

    # UTC 기준 날짜 사용 (업비트와 동일한 기준)
    df["DateUTC"] = pd.to_datetime(df["open_time"], unit="ms", utc=True)
    # 주봉은 주의 시작 시간 표시 (시간, 분, 초, 마이크로초 제거)
    date_list = []
    for _, row in df.iterrows():
        utc_time = pd.to_datetime(row["open_time"], unit="ms", utc=True)
        # 주의 시작 시간 (보통 월요일 00:00 UTC)
        candle_start_time = utc_time.replace(minute=0, second=0, microsecond=0)
        date_list.append(candle_start_time.strftime("%y/%m/%d,%H:00"))
    df["Date"] = date_list

    # 심볼에 따른 식별자 설정
    identifier = symbol.replace("USDT", "") + "USDW"

    # KST 시간 계산 (UTC+9) - 참고용으로만 사용 (계산 로직에는 사용하지 않음)
    kst_dates = []
    for utc_date_str in df["Date"]:
        # YY/MM/DD,HH:00 형식
        date_part, time_part = utc_date_str.split(',')
        utc_dt = dt.datetime.strptime(f"{date_part} {time_part}", "%y/%m/%d %H:%M")
        utc_dt = utc_dt.replace(tzinfo=tz.UTC)
        # KST로 변환 (참고용)
        kst_dt = utc_dt.astimezone(KST)
        kst_dates.append(kst_dt.strftime("%y/%m/%d,%H:00"))

    out = pd.DataFrame({
        "Date(UTC)": df["Date"],  # UTC 기준 시간 (계산에 사용)
        "KST": kst_dates,  # KST 시간 (참고용, 계산 로직에는 사용하지 않음)
        "종": pd.to_numeric(df["close"], errors="coerce"),
        "시": pd.to_numeric(df["open"], errors="coerce"),
        "고": pd.to_numeric(df["high"], errors="coerce"),
        "저": pd.to_numeric(df["low"], errors="coerce"),
        "Vol.": pd.to_numeric(df["volume"], errors="coerce"),
    })

    if not include_today:
        # UTC 기준으로 오늘 날짜 판단 (미완성 캔들 제거)
        today_utc = dt.datetime.now(tz.UTC).date()
        # 문자열 날짜를 date 객체로 변환하여 비교 (UTC 기준, YY/MM/DD,00:00 형식)
        out["Date_only"] = out["Date(UTC)"].apply(lambda x: x.split(',')[0] if ',' in str(x) else str(x))
        out = out[pd.to_datetime(out["Date_only"], format="%y/%m/%d").dt.date < today_utc]
        out = out.drop("Date_only", axis=1)

    out = out.sort_values("Date(UTC)", ascending=False).reset_index(drop=True)
    return out

# 일봉을 주봉으로 변환하는 함수 (API 호출 최적화)
def convert_daily_to_weekly(df_daily: pd.DataFrame) -> pd.DataFrame:
    """
    일봉 데이터를 주봉 데이터로 변환합니다.
    바이낸스 주봉은 월요일 00:00 UTC부터 시작합니다.
    
    Args:
        df_daily: 일봉 DataFrame (Date(UTC), 종, 시, 고, 저, Vol. 컬럼 포함)
    
    Returns:
        DataFrame: 주봉 DataFrame (Date(UTC), 종, 시, 고, 저, Vol. 컬럼 포함)
    """
    if df_daily.empty:
        return pd.DataFrame()
    
    df = df_daily.copy()
    
    # Date(UTC)를 datetime으로 변환
    df['Date(UTC)_dt'] = pd.to_datetime(df['Date(UTC)'], format='%y/%m/%d,%H:%M', errors='coerce')
    df = df[df['Date(UTC)_dt'].notna()].copy()
    
    if df.empty:
        return pd.DataFrame()
    
    # 주의 시작일 계산 (월요일 00:00 UTC)
    # 월요일 = 0, 일요일 = 6
    df['week_start'] = df['Date(UTC)_dt'] - pd.to_timedelta(df['Date(UTC)_dt'].dt.dayofweek, unit='D')
    df['week_start'] = df['week_start'].dt.normalize()  # 시간을 00:00:00으로 설정
    
    # 주 단위로 그룹화하여 주봉 생성
    weekly_data = []
    for week_start, group in df.groupby('week_start', sort=True):
        # 주의 첫 일봉 시가, 마지막 일봉 종가
        시가 = float(group.iloc[0]['시'])
        종가 = float(group.iloc[-1]['종'])
        고가 = float(group['고'].max())
        저가 = float(group['저'].min())
        거래량 = float(group['Vol.'].sum())
        
        # 주의 시작일을 Date(UTC) 형식으로 변환
        week_start_str = week_start.strftime('%y/%m/%d,00:00')
        
        # KST 계산 (UTC 00:00 = KST 09:00)
        kst_dt = week_start.astimezone(KST) if week_start.tz else week_start.replace(tzinfo=tz.UTC).astimezone(KST)
        kst_str = kst_dt.strftime('%y/%m/%d,09:00')
        
        weekly_data.append({
            'Date(UTC)': week_start_str,
            'KST': kst_str,
            '종': 종가,
            '시': 시가,
            '고': 고가,
            '저': 저가,
            'Vol.': 거래량
        })
    
    df_weekly = pd.DataFrame(weekly_data)
    
    # 최신 주봉이 위로 오도록 정렬 (최신 → 과거)
    df_weekly = df_weekly.sort_values('Date(UTC)', ascending=False).reset_index(drop=True)
    
    return df_weekly

# (업비트 주봉 생성 제거)
def calculate_buy(fore_or_one, sellside):
    """
    Buy 지표를 계산합니다.
    4or1과 sellside를 사용하여 계산합니다.
    """
    # NaN 체크
    if pd.isna(fore_or_one) or pd.isna(sellside):
        return ""
    
    # 조건: 4or1 < 4이고 sellside <= 0.05일 때 "buy" 반환
    if fore_or_one < 4 and sellside <= 0.05:
        return "buy"
    else:
        return ""

def calculate_sell(fore_or_one, buyside):
    """
    Sell 지표를 계산합니다. (일봉용)
    4or1과 buyside를 사용하여 계산합니다.
    """
    # NaN 체크
    if pd.isna(fore_or_one) or pd.isna(buyside):
        return ""
    
    # 조건: 4or1 >= 4이고 buyside <= 0.05일 때 "sell" 반환
    if fore_or_one >= 4 and buyside <= 0.05:
        return "sell"
    else:
        return ""

def calculate_sell_short(buyside):
    """
    Sell 지표를 계산합니다. (15분봉, 1시간봉, 4시간봉용)
    4or1 없이 buyside만 사용하여 계산합니다.
    """
    # NaN 체크
    if pd.isna(buyside):
        return ""
    
    # 조건: buyside <= 0.05일 때 "sell" 반환
    if buyside <= 0.05:
        return "sell"
    else:
        return ""



def calculate_sellside(sfast, fast, base):
    """
    sellside 지표를 계산합니다.
    SFast, Fast, Base를 사용하여 계산합니다.
    전처리(0.5 시프트)는 SFast에만 적용합니다: IF(sfast-0.5<1.5, sfast-0.5+6, sfast-0.5)
    """
    # NaN 체크
    if pd.isna(sfast) or pd.isna(fast) or pd.isna(base):
        return np.nan
    
    # sellside 전용 변환(SFast 한정): value에서 0.5를 빼고, 1.5 미만이면 6을 더함
    def transform_for_sellside(value):
        """sellside 계산용 값 변환"""
        adjusted = value - 0.5
        if adjusted < 1.5:
            return adjusted + 6
        else:
            return adjusted
    
    # SFast만 전처리, Fast/Base는 원값 사용
    sellside_sfast = transform_for_sellside(sfast)
    sellside_fast = fast
    sellside_base = base
    
    def calculate_component(value):
        """각 구성 요소의 계산 함수"""
        abs_value = abs(value)
        
        if 1.5 <= abs_value <= 7.5:
            result = (1/9) * (abs_value ** 2) - abs_value + (9/4)
            return result
        else:
            return 0
    
    # 변환된 값으로 각 구성 요소 계산
    sfast_component = calculate_component(sellside_sfast)
    fast_component = calculate_component(sellside_fast)
    base_component = calculate_component(sellside_base)
    
    # 최종 계산: (2/4.45 * sfast_component) + (1.15/4.45 * fast_component) + (1.3/4.45 * base_component)
    result = (2/4.45 * sfast_component) + (1.15/4.45 * fast_component) + (1.3/4.45 * base_component)
    
    return result

def calculate_buyside(sfast, fast, base):
    """
    buyside 지표를 계산합니다.
    SFast, Fast, Base를 사용하여 계산합니다.
    """
    # NaN 체크
    if pd.isna(sfast) or pd.isna(fast) or pd.isna(base):
        return np.nan
    
    def calculate_component(value):
        """각 구성 요소의 계산 함수"""
        abs_value = abs(value)
        
        if 1.5 <= abs_value <= 4.5:
            result = (-1/9) * (abs_value ** 2) + (1/3) * abs_value + (3/4)
            return result
        elif 4.5 <= abs_value <= 7.5:
            result = (-1/9) * (abs_value ** 2) + (5/3) * abs_value - (21/4)
            return result
        else:
            return 0
    
    # 각 구성 요소 계산
    sfast_component = calculate_component(sfast)
    fast_component = calculate_component(fast)
    base_component = calculate_component(base)
    
    # 최종 계산: 1 - (2/4.45 * sfast_component + 1.15/4.45 * fast_component + 1.3/4.45 * base_component)
    result = 1 - (2/4.45 * sfast_component + 1.15/4.45 * fast_component + 1.3/4.45 * base_component)
    
    return result

def calculate_base(sma5, sma10, sma20):
    """
    Base 지표를 계산합니다.
    Excel 수식을 파이썬으로 이식.
    """
    # NaN 체크
    if pd.isna(sma5) or pd.isna(sma10) or pd.isna(sma20):
        return np.nan
    
    sma5_val, sma10_val, sma20_val = float(sma5), float(sma10), float(sma20)
    eps = 0.0  # 정확한 비교를 위한 epsilon 값

    def gt(a, b):
        return a > b + eps

    def eq(a, b):
        return abs(a - b) <= eps

    # 1) Strict phase
    phase_strict = 0
    if gt(sma5_val, sma10_val) and gt(sma10_val, sma20_val):
        phase_strict = 1
    elif gt(sma10_val, sma5_val) and gt(sma5_val, sma20_val):
        phase_strict = 2
    elif gt(sma10_val, sma20_val) and gt(sma20_val, sma5_val):
        phase_strict = 3
    elif gt(sma20_val, sma10_val) and gt(sma10_val, sma5_val):
        phase_strict = 4
    elif gt(sma20_val, sma5_val) and gt(sma5_val, sma10_val):
        phase_strict = 5
    elif gt(sma5_val, sma20_val) and gt(sma20_val, sma10_val):
        phase_strict = 6

    # 2) Beta
    beta = 0.0
    if gt(sma5_val, sma10_val) and gt(sma10_val, sma20_val):
        beta = (sma10_val - sma20_val) / (sma5_val - sma20_val) if not eq(sma5_val, sma20_val) else 0.0
    elif gt(sma10_val, sma5_val) and gt(sma5_val, sma20_val):
        beta = 1.0 - (sma5_val - sma20_val) / (sma10_val - sma20_val) if not eq(sma10_val, sma20_val) else 0.0
    elif gt(sma10_val, sma20_val) and gt(sma20_val, sma5_val):
        beta = (sma20_val - sma5_val) / (sma10_val - sma5_val) if not eq(sma10_val, sma5_val) else 0.0
    elif gt(sma20_val, sma10_val) and gt(sma10_val, sma5_val):
        beta = 1.0 - (sma10_val - sma5_val) / (sma20_val - sma5_val) if not eq(sma20_val, sma5_val) else 0.0
    elif gt(sma20_val, sma5_val) and gt(sma5_val, sma10_val):
        beta = (sma5_val - sma10_val) / (sma20_val - sma10_val) if not eq(sma20_val, sma10_val) else 0.0
    elif gt(sma5_val, sma20_val) and gt(sma20_val, sma10_val):
        beta = 1.0 - (sma20_val - sma10_val) / (sma5_val - sma10_val) if not eq(sma5_val, sma10_val) else 0.0

    # 3) add6 조건
    phase_plus_beta = phase_strict + beta
    add6 = 6 if (phase_plus_beta > 0 and phase_plus_beta < 1.5) else 0

    # 4) Equal phase
    equal_phase = 0
    if eq(sma5_val, sma10_val) and gt(sma5_val, sma20_val):
        equal_phase = 2
    elif eq(sma5_val, sma20_val) and gt(sma10_val, sma5_val):
        equal_phase = 3
    elif eq(sma10_val, sma20_val) and gt(sma10_val, sma5_val):
        equal_phase = 4
    elif eq(sma10_val, sma5_val) and gt(sma20_val, sma10_val):
        equal_phase = 5
    elif eq(sma20_val, sma5_val) and gt(sma20_val, sma10_val):
        equal_phase = 6
    elif eq(sma20_val, sma10_val) and gt(sma5_val, sma20_val):
        equal_phase = 7

    # 5) 최종값
    final_value = phase_plus_beta + add6 + equal_phase

    # 반올림 처리 제거 - 원본 값 그대로 반환
    return final_value

def calculate_fast(sma5, sma7, sma10):
    """
    Fast 지표를 계산합니다.
    Excel 수식을 파이썬으로 이식.
    """
    # NaN 체크
    if pd.isna(sma5) or pd.isna(sma7) or pd.isna(sma10):
        return np.nan
    
    sma5_val, sma7_val, sma10_val = float(sma5), float(sma7), float(sma10)
    eps = 0.0  # 정확한 비교를 위한 epsilon 값

    def gt(a, b):
        return a > b + eps

    def eq(a, b):
        return abs(a - b) <= eps

    # 1) Strict phase
    phase_strict = 0
    if gt(sma5_val, sma7_val) and gt(sma7_val, sma10_val):
        phase_strict = 1
    elif gt(sma7_val, sma5_val) and gt(sma5_val, sma10_val):
        phase_strict = 2
    elif gt(sma7_val, sma10_val) and gt(sma10_val, sma5_val):
        phase_strict = 3
    elif gt(sma10_val, sma7_val) and gt(sma7_val, sma5_val):
        phase_strict = 4
    elif gt(sma10_val, sma5_val) and gt(sma5_val, sma7_val):
        phase_strict = 5
    elif gt(sma5_val, sma10_val) and gt(sma10_val, sma7_val):
        phase_strict = 6

    # 2) Beta
    beta = 0.0
    if gt(sma5_val, sma7_val) and gt(sma7_val, sma10_val):
        beta = (sma7_val - sma10_val) / (sma5_val - sma10_val) if not eq(sma5_val, sma10_val) else 0.0
    elif gt(sma7_val, sma5_val) and gt(sma5_val, sma10_val):
        beta = 1.0 - (sma5_val - sma10_val) / (sma7_val - sma10_val) if not eq(sma7_val, sma10_val) else 0.0
    elif gt(sma7_val, sma10_val) and gt(sma10_val, sma5_val):
        beta = (sma10_val - sma5_val) / (sma7_val - sma5_val) if not eq(sma7_val, sma5_val) else 0.0
    elif gt(sma10_val, sma7_val) and gt(sma7_val, sma5_val):
        beta = 1.0 - (sma7_val - sma5_val) / (sma10_val - sma5_val) if not eq(sma10_val, sma5_val) else 0.0
    elif gt(sma10_val, sma5_val) and gt(sma5_val, sma7_val):
        beta = (sma5_val - sma7_val) / (sma10_val - sma7_val) if not eq(sma10_val, sma7_val) else 0.0
    elif gt(sma5_val, sma10_val) and gt(sma10_val, sma7_val):
        beta = 1.0 - (sma10_val - sma7_val) / (sma5_val - sma7_val) if not eq(sma5_val, sma7_val) else 0.0

    # 3) add6 조건
    phase_plus_beta = phase_strict + beta
    add6 = 6 if (phase_plus_beta > 0 and phase_plus_beta < 1.5) else 0

    # 4) Equal phase
    equal_phase = 0
    if eq(sma5_val, sma7_val) and gt(sma5_val, sma10_val):
        equal_phase = 2
    elif eq(sma5_val, sma10_val) and gt(sma7_val, sma5_val):
        equal_phase = 3
    elif eq(sma7_val, sma10_val) and gt(sma7_val, sma5_val):
        equal_phase = 4
    elif eq(sma7_val, sma5_val) and gt(sma10_val, sma7_val):
        equal_phase = 5
    elif eq(sma10_val, sma5_val) and gt(sma10_val, sma7_val):
        equal_phase = 6
    elif eq(sma10_val, sma7_val) and gt(sma5_val, sma10_val):
        equal_phase = 7

    # 5) 최종값
    final_value = phase_plus_beta + add6 + equal_phase

    # 반올림 처리 제거 - 원본 값 그대로 반환
    return final_value

def calculate_superfast(sma3, sma5, sma7):
    """
    SuperFast 지표를 계산합니다.
    Excel 수식을 파이썬으로 이식.
    """
    # NaN 체크
    if pd.isna(sma3) or pd.isna(sma5) or pd.isna(sma7):
        return np.nan
    
    sma3_val, sma5_val, sma7_val = float(sma3), float(sma5), float(sma7)
    eps = 0.0  # 정확한 비교를 위한 epsilon 값

    def gt(a, b):
        return a > b + eps

    def eq(a, b):
        return abs(a - b) <= eps

    # 1) Strict phase
    phase_strict = 0
    if gt(sma3_val, sma5_val) and gt(sma5_val, sma7_val):
        phase_strict = 1
    elif gt(sma5_val, sma3_val) and gt(sma3_val, sma7_val):
        phase_strict = 2
    elif gt(sma5_val, sma7_val) and gt(sma7_val, sma3_val):
        phase_strict = 3
    elif gt(sma7_val, sma5_val) and gt(sma5_val, sma3_val):
        phase_strict = 4
    elif gt(sma7_val, sma3_val) and gt(sma3_val, sma5_val):
        phase_strict = 5
    elif gt(sma3_val, sma7_val) and gt(sma7_val, sma5_val):
        phase_strict = 6

    # 2) Beta
    beta = 0.0
    if gt(sma3_val, sma5_val) and gt(sma5_val, sma7_val):
        beta = (sma5_val - sma7_val) / (sma3_val - sma7_val) if not eq(sma3_val, sma7_val) else 0.0
    elif gt(sma5_val, sma3_val) and gt(sma3_val, sma7_val):
        beta = 1.0 - (sma3_val - sma7_val) / (sma5_val - sma7_val) if not eq(sma5_val, sma7_val) else 0.0
    elif gt(sma5_val, sma7_val) and gt(sma7_val, sma3_val):
        beta = (sma7_val - sma3_val) / (sma5_val - sma3_val) if not eq(sma5_val, sma3_val) else 0.0
    elif gt(sma7_val, sma5_val) and gt(sma5_val, sma3_val):
        beta = 1.0 - (sma5_val - sma3_val) / (sma7_val - sma3_val) if not eq(sma7_val, sma3_val) else 0.0
    elif gt(sma7_val, sma3_val) and gt(sma3_val, sma5_val):
        beta = (sma3_val - sma5_val) / (sma7_val - sma5_val) if not eq(sma7_val, sma5_val) else 0.0
    elif gt(sma3_val, sma7_val) and gt(sma7_val, sma5_val):
        beta = 1.0 - (sma7_val - sma5_val) / (sma3_val - sma5_val) if not eq(sma3_val, sma5_val) else 0.0

    # 3) add6 조건
    phase_plus_beta = phase_strict + beta
    add6 = 6 if (phase_plus_beta > 0 and phase_plus_beta < 1.5) else 0

    # 4) Equal phase
    equal_phase = 0
    if eq(sma3_val, sma5_val) and gt(sma3_val, sma7_val):
        equal_phase = 2
    elif eq(sma3_val, sma7_val) and gt(sma5_val, sma3_val):
        equal_phase = 3
    elif eq(sma5_val, sma7_val) and gt(sma5_val, sma3_val):
        equal_phase = 4
    elif eq(sma5_val, sma3_val) and gt(sma7_val, sma5_val):
        equal_phase = 5
    elif eq(sma7_val, sma3_val) and gt(sma7_val, sma5_val):
        equal_phase = 6
    elif eq(sma7_val, sma5_val) and gt(sma3_val, sma7_val):
        equal_phase = 7

    # 5) 최종값
    final_value = phase_plus_beta + add6 + equal_phase

    # 반올림 처리 제거 - 원본 값 그대로 반환
    return final_value

def calculate_4or1(하단, 상단):
    """
    4or1 지표를 계산합니다.
    하단/(하단+상단) >= 0.666인 경우: 4 + 비율
    하단/(하단+상단) < 0.666인 경우: 1 + 비율
    """
    # NaN 체크
    if pd.isna(하단) or pd.isna(상단):
        return np.nan
    
    # 상단과 하단이 모두 0인 경우 (현재가가 Max20과 Min20 사이에 정확히 중간에 있는 경우)
    if 하단 == 0 and 상단 == 0:
        return 1.0  # 기본값 1.0 반환
    
    # 분모가 0인 경우 방지
    denominator = 상단 + 하단
    if denominator == 0:
        return 1.0  # 기본값 1.0 반환
    
    # 정상 계산
    ratio = 하단 / denominator
    if ratio >= 0.666:
        return 4.0 + ratio  # 4.666 ~ 4.999
    else:
        return 1.0 + ratio  # 1.000 ~ 1.665

def calculate_all_indicators(df, market_type):
    """
    모든 지표를 한 번에 계산합니다.
    """
    if df.empty:
        return df
    
    # 과거→현재 순서로 정렬 (계산을 위해)
    df = df.sort_values("Date(UTC)").reset_index(drop=True)
    
    # SMA 계산 (3, 5, 7, 10, 20일)
    df["SMA3"] = df["종"].rolling(window=3, min_periods=3).mean()
    df["SMA5"] = df["종"].rolling(window=5, min_periods=5).mean()
    df["SMA7"] = df["종"].rolling(window=7, min_periods=7).mean()
    df["SMA10"] = df["종"].rolling(window=10, min_periods=10).mean()
    df["SMA20"] = df["종"].rolling(window=20, min_periods=20).mean()
    
    # Max15, Min15 계산 (15일 동안의 시고저종에서 최고가와 최저가)
    df["Max15"] = df[["시", "고", "저", "종"]].rolling(window=15, min_periods=15).max().max(axis=1)
    df["Min15"] = df[["시", "고", "저", "종"]].rolling(window=15, min_periods=15).min().min(axis=1)
    
    # 하단, 상단 계산 (Max15/Min15이 NaN이면 NaN)
    df["하단"] = df.apply(lambda row: abs((row["종"] - row["Min15"]) / row["Min15"]) if not pd.isna(row["Min15"]) else np.nan, axis=1)
    df["상단"] = df.apply(lambda row: abs((row["종"] - row["Max15"]) / row["Max15"]) if not pd.isna(row["Max15"]) else np.nan, axis=1)
    
    # SFast 계산
    df["SFast"] = df.apply(lambda row: calculate_superfast(row["SMA3"], row["SMA5"], row["SMA7"]), axis=1)
    
    # Fast 계산
    df["Fast"] = df.apply(lambda row: calculate_fast(row["SMA5"], row["SMA7"], row["SMA10"]), axis=1)
    
    # Base 계산
    df["Base"] = df.apply(lambda row: calculate_base(row["SMA5"], row["SMA10"], row["SMA20"]), axis=1)
    
    # 4or1 계산
    df["4or1"] = df.apply(lambda row: calculate_4or1(row["하단"], row["상단"]), axis=1)
    
    # buyside 계산
    df["buyside"] = df.apply(lambda row: calculate_buyside(row["SFast"], row["Fast"], row["Base"]), axis=1)
    
    # sellside 계산
    df["sellside"] = df.apply(lambda row: calculate_sellside(row["SFast"], row["Fast"], row["Base"]), axis=1)
    
    # Buy 계산
    df["Buy"] = df.apply(lambda row: calculate_buy(row["4or1"], row["sellside"]), axis=1)
    
    # Sell 계산
    df["Sell"] = df.apply(lambda row: calculate_sell(row["4or1"], row["buyside"]), axis=1)
    
    # Samount1D 계산: (1-buyside) * 1unit (티커별 USDT 정밀도 적용)
    symbol = f"{TICKER}USDT"
    usdt_precision = SYMBOL_USDT_PRECISION.get(symbol, 5)
    df["Samount1D"] = df.apply(lambda row: round((1 - row["buyside"]) * TRADING_UNIT, usdt_precision) if not pd.isna(row["buyside"]) else np.nan, axis=1)
    
    # Bamount1D 계산: (1-sellside) * 1unit (티커별 USDT 정밀도 적용)
    df["Bamount1D"] = df.apply(lambda row: round((1 - row["sellside"]) * TRADING_UNIT, usdt_precision) if not pd.isna(row["sellside"]) else np.nan, axis=1)
    
    # 최신→과거로 재정렬
    df = df.sort_values("Date(UTC)", ascending=False).reset_index(drop=True)
    
    # 숫자 컬럼 정리
    num_cols = ["종", "시", "고", "저", "Vol.", "SMA3", "SMA5", "SMA7", "SMA10", "SMA20", "Max15", "Min15", "하단", "상단", "SFast", "Fast", "Base", "4or1", "buyside", "sellside", "Samount1D", "Bamount1D"]
    df[num_cols] = df[num_cols].apply(pd.to_numeric, errors="coerce")
    
    return df

def calculate_all_indicators_5m(df, market_type):
    """
    5분봉용 모든 지표를 계산합니다. (Max200/Min200 사용)
    """
    if df.empty:
        return df
    
    # 과거→현재 순서로 정렬 (계산을 위해)
    df = df.sort_values("Date(UTC)").reset_index(drop=True)
    
    # SMA 계산 (3, 5, 7, 10, 20일)
    df["SMA3"] = df["종"].rolling(window=3, min_periods=3).mean()
    df["SMA5"] = df["종"].rolling(window=5, min_periods=5).mean()
    df["SMA7"] = df["종"].rolling(window=7, min_periods=7).mean()
    df["SMA10"] = df["종"].rolling(window=10, min_periods=10).mean()
    df["SMA20"] = df["종"].rolling(window=20, min_periods=20).mean()
    
    # Max200, Min200 계산 (200일 동안의 시고저종에서 최고가와 최저가)
    df["Max200"] = df[["시", "고", "저", "종"]].rolling(window=200, min_periods=200).max().max(axis=1)
    df["Min200"] = df[["시", "고", "저", "종"]].rolling(window=200, min_periods=200).min().min(axis=1)
    
    # 하단, 상단 계산 (Max200/Min200이 NaN이면 NaN)
    df["하단"] = df.apply(lambda row: abs((row["종"] - row["Min200"]) / row["Min200"]) if not pd.isna(row["Min200"]) else np.nan, axis=1)
    df["상단"] = df.apply(lambda row: abs((row["종"] - row["Max200"]) / row["Max200"]) if not pd.isna(row["Max200"]) else np.nan, axis=1)
    
    # SFast 계산
    df["SFast"] = df.apply(lambda row: calculate_superfast(row["SMA3"], row["SMA5"], row["SMA7"]), axis=1)
    
    # Fast 계산
    df["Fast"] = df.apply(lambda row: calculate_fast(row["SMA5"], row["SMA7"], row["SMA10"]), axis=1)
    
    # Base 계산
    df["Base"] = df.apply(lambda row: calculate_base(row["SMA5"], row["SMA10"], row["SMA20"]), axis=1)
    
    # 4or1 계산
    df["4or1"] = df.apply(lambda row: calculate_4or1(row["하단"], row["상단"]), axis=1)
    
    # buyside 계산
    df["buyside"] = df.apply(lambda row: calculate_buyside(row["SFast"], row["Fast"], row["Base"]), axis=1)
    
    # sellside 계산
    df["sellside"] = df.apply(lambda row: calculate_sellside(row["SFast"], row["Fast"], row["Base"]), axis=1)
    
    # Buy 계산
    df["Buy"] = df.apply(lambda row: calculate_buy(row["4or1"], row["sellside"]), axis=1)
    
    # Sell 계산 (15분봉용 - 4or1 없이 buyside만 사용)
    df["Sell"] = df.apply(lambda row: calculate_sell_short(row["buyside"]), axis=1)
    
    # 최신→과거 순서로 다시 정렬
    df = df.sort_values("Date(UTC)", ascending=False).reset_index(drop=True)
    
    return df


def calculate_all_indicators_15m(df, market_type):
    """
    15분봉용 모든 지표를 한 번에 계산합니다.
    Source 기준: Max70, Min70을 사용합니다.
    
    Args:
        df: 15분봉 DataFrame (Date(UTC) 컬럼 포함, UTC 기준)
        market_type: 시장 타입 (사용하지 않음, 호환성 유지)
    
    Returns:
        DataFrame: 모든 지표가 계산된 15분봉 DataFrame
    
    Note:
        - Date(UTC) 컬럼을 기준으로 정렬 및 계산 (UTC 기준)
    """
    if df.empty:
        return df
    
    # 현재→과거 순서로 정렬 (계산을 위해, AFTER 단계와 동일한 순서)
    df = df.sort_values("Date(UTC)", ascending=False).reset_index(drop=True)
    
    # SMA 계산: 각 행(idx)에서 그 행부터 앞으로(과거로) window개까지의 평균
    # idx=0 (최신)에서 idx=0~window-1까지의 평균
    # PREVIOUS와 AFTER 단계 모두 동일한 계산 방식 사용
    # 
    # SMA200 계산식 (PREVIOUS 단계):
    #   각 행의 SMA200 = 해당 행(idx)부터 idx+200까지의 종가 평균
    #   df.loc[idx, "SMA200"] = df.iloc[idx:idx+200]["종"].mean()
    # 
    # 예시 (현재→과거 순서):
    #   - idx=0 (최신): SMA200 = (종가[0] + 종가[1] + ... + 종가[199]) / 200
    #   - idx=1: SMA200 = (종가[1] + 종가[2] + ... + 종가[200]) / 200
    #   - idx=2: SMA200 = (종가[2] + 종가[3] + ... + 종가[201]) / 200
    # 
    # PREVIOUS와 AFTER 단계의 일관성:
    #   PREVIOUS: for idx in range(len(df)): df.loc[idx, "SMA200"] = df.iloc[idx:idx+200]["종"].mean()
    #   AFTER:    df.loc[0, "SMA200"] = df.iloc[0:200]["종"].mean()
    #   → 두 방식 모두 동일한 결과: 각 행(idx)을 포함한 최근 200개 캔들의 종가 평균
    for idx in range(len(df)):
        df.loc[idx, "SMA3"] = df.iloc[idx:idx+3]["종"].mean() if idx + 3 <= len(df) else np.nan
        df.loc[idx, "SMA5"] = df.iloc[idx:idx+5]["종"].mean() if idx + 5 <= len(df) else np.nan
        df.loc[idx, "SMA7"] = df.iloc[idx:idx+7]["종"].mean() if idx + 7 <= len(df) else np.nan
        df.loc[idx, "SMA10"] = df.iloc[idx:idx+10]["종"].mean() if idx + 10 <= len(df) else np.nan
        df.loc[idx, "SMA12"] = df.iloc[idx:idx+12]["종"].mean() if idx + 12 <= len(df) else np.nan
        df.loc[idx, "SMA20"] = df.iloc[idx:idx+20]["종"].mean() if idx + 20 <= len(df) else np.nan
        df.loc[idx, "SMA25"] = df.iloc[idx:idx+25]["종"].mean() if idx + 25 <= len(df) else np.nan
        df.loc[idx, "SMA50"] = df.iloc[idx:idx+50]["종"].mean() if idx + 50 <= len(df) else np.nan
        df.loc[idx, "SMA100"] = df.iloc[idx:idx+100]["종"].mean() if idx + 100 <= len(df) else np.nan
        df.loc[idx, "SMA200"] = df.iloc[idx:idx+200]["종"].mean() if idx + 200 <= len(df) else np.nan
        df.loc[idx, "SMA400"] = df.iloc[idx:idx+400]["종"].mean() if idx + 400 <= len(df) else np.nan
        df.loc[idx, "SMA800"] = df.iloc[idx:idx+800]["종"].mean() if idx + 800 <= len(df) else np.nan
    
    # SMAF: SMA3·SMA12 6:4 가중평균
    df["SMAF"] = df["SMA3"] * 0.6 + df["SMA12"] * 0.4
    
    # Max70, Min70 계산: 각 행(idx)에서 그 행부터 앞으로(과거로) 70개까지의 최고가/최저가
    for idx in range(len(df)):
        if idx + 70 <= len(df):
            window_data = df.iloc[idx:idx+70][["시", "고", "저", "종"]]
            df.loc[idx, "Max70"] = window_data.values.max()
            df.loc[idx, "Min70"] = window_data.values.min()
        else:
            df.loc[idx, "Max70"] = np.nan
            df.loc[idx, "Min70"] = np.nan
    
    # 하단, 상단 계산 (Max70/Min70이 NaN이면 NaN)
    df["하단"] = df.apply(lambda row: abs((row["종"] - row["Min70"]) / row["Min70"]) if not pd.isna(row["Min70"]) else np.nan, axis=1)
    df["상단"] = df.apply(lambda row: abs((row["종"] - row["Max70"]) / row["Max70"]) if not pd.isna(row["Max70"]) else np.nan, axis=1)
    
    # SFast 계산
    df["SFast"] = df.apply(lambda row: calculate_superfast(row["SMA3"], row["SMA5"], row["SMA7"]), axis=1)
    
    # Fast 계산
    df["Fast"] = df.apply(lambda row: calculate_fast(row["SMA5"], row["SMA7"], row["SMA10"]), axis=1)
    
    # Base 계산
    df["Base"] = df.apply(lambda row: calculate_base(row["SMA5"], row["SMA10"], row["SMA20"]), axis=1)
    
    # 4or1 계산
    df["4or1"] = df.apply(lambda row: calculate_4or1(row["하단"], row["상단"]), axis=1)
    
    # buyside 계산
    df["buyside"] = df.apply(lambda row: calculate_buyside(row["SFast"], row["Fast"], row["Base"]), axis=1)
    
    # sellside 계산
    df["sellside"] = df.apply(lambda row: calculate_sellside(row["SFast"], row["Fast"], row["Base"]), axis=1)
    
    # Buy 계산
    df["Buy"] = df.apply(lambda row: calculate_buy(row["4or1"], row["sellside"]), axis=1)
    
    # Sell 계산 (15분봉용: buyside만 사용)
    df["Sell"] = df.apply(lambda row: calculate_sell_short(row["buyside"]), axis=1)
    
    # SamountW, BamountW 열 초기화 (주봉에서 복사될 예정)
    df["SamountW"] = np.nan
    df["BamountW"] = np.nan
    
    # Samount1D, Bamount1D 열 초기화 (일봉에서 복사될 예정)
    df["Samount1D"] = np.nan
    df["Bamount1D"] = np.nan
    
    # SPRD 계산: (max(sma25,sma100,sma200)-min(sma25,sma100,sma200))/min(sma25,sma100,sma200)
    df["SPRD"] = df.apply(lambda row: (
        (max(row["SMA25"], row["SMA100"], row["SMA200"]) - min(row["SMA25"], row["SMA100"], row["SMA200"])) / min(row["SMA25"], row["SMA100"], row["SMA200"])
        if not pd.isna(row["SMA25"]) and not pd.isna(row["SMA100"]) and not pd.isna(row["SMA200"]) and min(row["SMA25"], row["SMA100"], row["SMA200"]) > 0
        else np.nan
    ), axis=1)
    
    # SPRD2 계산: (max(저가,sma100,sma200)-min(저가,sma100,sma200))/min(저가,sma100,sma200)
    df["SPRD2"] = df.apply(lambda row: (
        (max(row["저"], row["SMA100"], row["SMA200"]) - min(row["저"], row["SMA100"], row["SMA200"])) / min(row["저"], row["SMA100"], row["SMA200"])
        if not pd.isna(row["저"]) and not pd.isna(row["SMA100"]) and not pd.isna(row["SMA200"]) and min(row["저"], row["SMA100"], row["SMA200"]) > 0
        else np.nan
    ), axis=1)
    
    # KSC 열 초기화 (숫자만 저장)
    df["KSC"] = 0
    # Bomb 열 초기화 (Bomb 발생 시 "Bomb" 저장) - object 타입으로 명시적 설정
    df["Bomb"] = ""
    df["Bomb"] = df["Bomb"].astype('object')
    # BombCount 열 초기화 (내부 계산용, Bomb이 되는 시점의 카운트 값)
    df["BombCount"] = 0
    # PRFT 열 초기화 (내용은 나중에 채움)
    df["PRFT"] = np.nan
    # KSC stack 열 초기화 (내부 계산용, Bomb이 되는 시점의 스택 카운트 값)
    df["KSC stack"] = 0
    
    # 1HMSFast 계산 (15M 시트: SMAF, SMA100, SMA200 사용, 종가 없음)
    # 각 행은 자신의 SMAF, SMA100, SMA200으로 계산 (shift 없음)
    df["1HMSFast"] = df.apply(lambda row: calculate_1hmsfast_15m(row["SMAF"], row["SMA100"], row["SMA200"]), axis=1)
    
    # LS 열: -1 = (현재 2<1HMSF<3 AND 직전 1.5<1HMSF<=2) OR (현재 4<1HMSF<=5 AND 직전 5<1HMSF<6) / 1 = (5<현재<6 AND 직전 4~5) OR (1<현재<=2 AND 직전 2~3)
    # 각 행(idx) = 2행, 다음 행(idx+1) = 3행. 최신→과거 순서이므로 idx+1이 과거(Excel 아래행).
    df["LS"] = np.nan
    for i in range(len(df)):
        if i + 1 >= len(df):
            df.at[i, "LS"] = np.nan
        else:
            h2 = df.iloc[i]["1HMSFast"]
            h3 = df.iloc[i + 1]["1HMSFast"]
            if pd.isna(h2) or pd.isna(h3):
                df.at[i, "LS"] = np.nan
            else:
                h2_f, h3_f = float(h2), float(h3)
                if (2 < h2_f < 3 and 1.5 < h3_f <= 2) or (4 < h2_f <= 5 and 5 < h3_f < 6):
                    df.at[i, "LS"] = -1
                elif (5 < h2_f < 6 and 4 < h3_f <= 5) or (1 < h2_f <= 2 and 2 < h3_f < 3):
                    df.at[i, "LS"] = 1
                else:
                    df.at[i, "LS"] = np.nan
    
    # 숫자 컬럼 정리 (Source 기준: Max70, Min70 사용, SMA400, SMA800 추가)
    num_cols = ["종", "시", "고", "저", "Vol.", "SMA3", "SMA5", "SMA7", "SMA10", "SMA12", "SMAF", "SMA20", "SMA25", "SMA50", "SMA100", "SMA200", "SMA400", "SMA800", "Max70", "Min70", "하단", "상단", "SFast", "Fast", "Base", "4or1", "buyside", "sellside", "1HMSFast", "SPRD", "SPRD2", "SamountW", "BamountW", "Samount1D", "Bamount1D"]
    df[num_cols] = df[num_cols].apply(pd.to_numeric, errors="coerce")
    
    # 이미 최신→과거 순서이므로 재정렬 불필요 (정렬은 위에서 이미 완료)
    
    return df

def calculate_all_indicators_1h(df, market_type):
    """
    1시간봉용 모든 지표를 한 번에 계산합니다.
    Source 기준: SMA25, SMA100, SMA200, SMA400, SMA800, Max200, Min200, 1HCLASS, -1HCLASS
    """
    if df.empty:
        return df
    
    # Date(UTC) 컬럼 정규화 (Timestamp와 문자열 혼합 방지)
    if 'Date(UTC)' in df.columns:
        if df['Date(UTC)'].dtype == 'object':
            # format 명시하여 파싱 시도 (연도/일 혼동 방지)
            try:
                df['Date(UTC)'] = pd.to_datetime(df['Date(UTC)'], format='%y/%m/%d,%H:%M', errors='coerce')
            except:
                # 쉼표 제거 후 형식 시도
                try:
                    df['Date(UTC)'] = df['Date(UTC)'].astype(str).str.replace(',', ' ', regex=False).str.strip()
                    df['Date(UTC)'] = pd.to_datetime(df['Date(UTC)'], format='%y/%m/%d %H:%M', errors='coerce')
                except:
                    # fallback: format 없이 파싱
                    df['Date(UTC)'] = pd.to_datetime(df['Date(UTC)'], errors='coerce')
        else:
            # 이미 datetime 타입이면 그대로 사용
            pass
    
    # ⚠️중요: 입력 데이터를 확실하게 현재→과거 순서로 정렬 (SMA 계산 일관성 보장)
    # 데이터 수집 과정에서 정렬이 여러 번 섞일 수 있으므로, 계산 직전에 확실하게 정렬
    df = df.sort_values("Date(UTC)", ascending=False).reset_index(drop=True)
    
    # Buy/Sell 컬럼이 없으면 생성
    if "Buy" not in df.columns:
        df["Buy"] = ""
    if "Sell" not in df.columns:
        df["Sell"] = ""
    
    # Previous 데이터의 Buy/Sell이 NaN이면 빈 문자열로 변환
    df["Buy"] = df["Buy"].fillna("")
    df["Sell"] = df["Sell"].fillna("")
    
    # 모든 지표 컬럼 초기화 (NaN으로 시작)
    indicator_cols = ["SMA25", "SMA100", "SMA200", "SMA400", "SMA800", "Max200", "Min200", "하단", "상단", "SFast", "Fast", "Base", "1HMSFast", "4or1", "buyside", "sellside", "1HCLASS", "-1HCLASS", "p1H"]
    for col in indicator_cols:
        if col not in df.columns:
            df[col] = np.nan
    
    # SMA 계산: 각 행(idx)에서 그 행부터 앞으로(과거로) window개까지의 평균
    for idx in range(len(df)):
        df.loc[idx, "SMA25"] = df.iloc[idx:idx+25]["종"].mean() if idx + 25 <= len(df) else np.nan
        df.loc[idx, "SMA100"] = df.iloc[idx:idx+100]["종"].mean() if idx + 100 <= len(df) else np.nan
        df.loc[idx, "SMA200"] = df.iloc[idx:idx+200]["종"].mean() if idx + 200 <= len(df) else np.nan
        df.loc[idx, "SMA400"] = df.iloc[idx:idx+400]["종"].mean() if idx + 400 <= len(df) else np.nan
        df.loc[idx, "SMA800"] = df.iloc[idx:idx+800]["종"].mean() if idx + 800 <= len(df) else np.nan
    
    # Max200, Min200 계산: 각 행(idx)에서 그 행부터 앞으로(과거로) 200개까지의 최고가/최저가
    for idx in range(len(df)):
        if idx + 200 <= len(df):
            window_data = df.iloc[idx:idx+200][["시", "고", "저", "종"]]
            df.loc[idx, "Max200"] = window_data.values.max()
            df.loc[idx, "Min200"] = window_data.values.min()
        else:
            df.loc[idx, "Max200"] = np.nan
            df.loc[idx, "Min200"] = np.nan
    
    # 하단, 상단 계산 (Max200/Min200이 NaN이면 NaN)
    df["하단"] = df.apply(lambda row: abs((row["종"] - row["Min200"]) / row["Min200"]) if not pd.isna(row["Min200"]) else np.nan, axis=1)
    df["상단"] = df.apply(lambda row: abs((row["종"] - row["Max200"]) / row["Max200"]) if not pd.isna(row["Max200"]) else np.nan, axis=1)
    
    # SFast 계산 (SMA25, SMA100, SMA200 사용)
    df["SFast"] = df.apply(lambda row: calculate_superfast(row["SMA25"], row["SMA100"], row["SMA200"]), axis=1)
    
    # Fast 계산 (SMA25, SMA200, SMA400 사용)
    df["Fast"] = df.apply(lambda row: calculate_fast(row["SMA25"], row["SMA200"], row["SMA400"]), axis=1)
    
    # Base 계산 (SMA25, SMA400, SMA800 사용)
    df["Base"] = df.apply(lambda row: calculate_base(row["SMA25"], row["SMA400"], row["SMA800"]), axis=1)
    
    # 1HMSFast 계산 (종가, SMA25, SMA100 사용)
    # 각 행은 자신의 종가, SMA25, SMA100으로 계산 (shift 없음)
    df["1HMSFast"] = df.apply(lambda row: calculate_1hmsfast(row["종"], row["SMA25"], row["SMA100"]), axis=1)
    
    # 4or1 계산
    df["4or1"] = df.apply(lambda row: calculate_4or1(row["하단"], row["상단"]), axis=1)
    
    # buyside 계산
    df["buyside"] = df.apply(lambda row: calculate_buyside(row["SFast"], row["Fast"], row["Base"]), axis=1)
    
    # sellside 계산
    df["sellside"] = df.apply(lambda row: calculate_sellside(row["SFast"], row["Fast"], row["Base"]), axis=1)
    
    # Buy 계산
    df["Buy"] = df.apply(lambda row: calculate_buy(row["4or1"], row["sellside"]), axis=1)
    
    # Sell 계산 (1시간봉용: buyside만 사용)
    df["Sell"] = df.apply(lambda row: calculate_sell_short(row["buyside"]), axis=1)
    
    # 1HCLASS 계산: 1H 캔들에서 SFast/Fast/Base가 and(2 <= 값 < 5)이면 각각 +1, 총합(0~3)
    df["1HCLASS"] = df.apply(
        lambda row:
        (1 if pd.notna(row["SFast"]) and 2 <= row["SFast"] < 5 else 0) +
        (1 if pd.notna(row["Fast"]) and 2 <= row["Fast"] < 5 else 0) +
        (1 if pd.notna(row["Base"]) and 2 <= row["Base"] < 5 else 0),
        axis=1
    )
    
    # -1HCLASS 계산: 1H 캔들에서 SFast/Fast/Base가 or(값 < 2, 값 >= 7)이면 각각 -1, 총합(0~-3)
    df["-1HCLASS"] = df.apply(
        lambda row:
        (-1 if pd.notna(row["SFast"]) and (row["SFast"] < 2 or row["SFast"] >= 7) else 0) +
        (-1 if pd.notna(row["Fast"]) and (row["Fast"] < 2 or row["Fast"] >= 7) else 0) +
        (-1 if pd.notna(row["Base"]) and (row["Base"] < 2 or row["Base"] >= 7) else 0),
        axis=1
    )
    
    # p1H 계산: 1H 시트 SFast/Fast/Base 각각 4 <= 값 < 5 인 경우 1로 카운트, 총합(0,1,2,3)
    df["p1H"] = df.apply(
        lambda row:
        (1 if pd.notna(row["SFast"]) and 4 <= row["SFast"] < 5 else 0) +
        (1 if pd.notna(row["Fast"]) and 4 <= row["Fast"] < 5 else 0) +
        (1 if pd.notna(row["Base"]) and 4 <= row["Base"] < 5 else 0),
        axis=1
    )
    
    # 이미 최신→과거 순서이므로 재정렬 불필요 (정렬은 위에서 이미 완료)
    
    # 숫자 컬럼 정리 (존재하는 컬럼만 변환)
    num_cols = ["종", "시", "고", "저", "Vol.", "SMA25", "SMA100", "SMA200", "SMA400", "SMA800", "Max200", "Min200", "하단", "상단", "SFast", "Fast", "Base", "1HMSFast", "4or1", "buyside", "sellside", "1HCLASS", "-1HCLASS", "p1H"]
    for col in num_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    
    return df
def calculate_all_indicators_weekly(df, market_type):
    """
    주봉용 모든 지표를 한 번에 계산합니다.
    Max25, Min25를 사용합니다 (25개 기준).
    """
    if df.empty:
        return df
    
    # 과거→현재 순서로 정렬 (계산을 위해)
    df = df.sort_values("Date(UTC)").reset_index(drop=True)
    
    # SMA 계산 (3, 5, 7, 10, 20일)
    df["SMA3"] = df["종"].rolling(window=3, min_periods=3).mean()
    df["SMA5"] = df["종"].rolling(window=5, min_periods=5).mean()
    df["SMA7"] = df["종"].rolling(window=7, min_periods=7).mean()
    df["SMA10"] = df["종"].rolling(window=10, min_periods=10).mean()
    df["SMA20"] = df["종"].rolling(window=20, min_periods=20).mean()
    
    # Max25, Min25 계산 (25개 캔들 동안의 시고저종에서 최고가와 최저가)
    df["Max25"] = df[["시", "고", "저", "종"]].rolling(window=25, min_periods=25).max().max(axis=1)
    df["Min25"] = df[["시", "고", "저", "종"]].rolling(window=25, min_periods=25).min().min(axis=1)
    
    # 하단, 상단 계산 (Max25/Min25이 NaN이면 NaN)
    df["하단"] = df.apply(lambda row: abs((row["종"] - row["Min25"]) / row["Min25"]) if not pd.isna(row["Min25"]) else np.nan, axis=1)
    df["상단"] = df.apply(lambda row: abs((row["종"] - row["Max25"]) / row["Max25"]) if not pd.isna(row["Max25"]) else np.nan, axis=1)
    
    # SFast 계산
    df["SFast"] = df.apply(lambda row: calculate_superfast(row["SMA3"], row["SMA5"], row["SMA7"]), axis=1)
    
    # Fast 계산
    df["Fast"] = df.apply(lambda row: calculate_fast(row["SMA5"], row["SMA7"], row["SMA10"]), axis=1)
    
    # Base 계산
    df["Base"] = df.apply(lambda row: calculate_base(row["SMA5"], row["SMA10"], row["SMA20"]), axis=1)
    
    # 4or1 계산
    df["4or1"] = df.apply(lambda row: calculate_4or1(row["하단"], row["상단"]), axis=1)
    
    # buyside 계산
    df["buyside"] = df.apply(lambda row: calculate_buyside(row["SFast"], row["Fast"], row["Base"]), axis=1)
    
    # sellside 계산
    df["sellside"] = df.apply(lambda row: calculate_sellside(row["SFast"], row["Fast"], row["Base"]), axis=1)
    
    # Buy 계산
    df["Buy"] = df.apply(lambda row: calculate_buy(row["4or1"], row["sellside"]), axis=1)
    
    # Sell 계산 (주봉용: buyside만 사용)
    df["Sell"] = df.apply(lambda row: calculate_sell_short(row["buyside"]), axis=1)
    
    # SamountW 계산: (1-buyside) * 1unit (티커별 USDT 정밀도 적용)
    symbol = f"{TICKER}USDT"
    usdt_precision = SYMBOL_USDT_PRECISION.get(symbol, 5)
    df["SamountW"] = df.apply(lambda row: round((1 - row["buyside"]) * TRADING_UNIT, usdt_precision) if not pd.isna(row["buyside"]) else np.nan, axis=1)
    
    # BamountW 계산: (1-sellside) * 1unit (티커별 USDT 정밀도 적용)
    df["BamountW"] = df.apply(lambda row: round((1 - row["sellside"]) * TRADING_UNIT, usdt_precision) if not pd.isna(row["sellside"]) else np.nan, axis=1)
    
    # 최신→과거로 재정렬
    df = df.sort_values("Date(UTC)", ascending=False).reset_index(drop=True)
    
    # 숫자 컬럼 정리
    num_cols = ["종", "시", "고", "저", "Vol.", "SMA3", "SMA5", "SMA7", "SMA10", "SMA20", "Max25", "Min25", "하단", "상단", "SFast", "Fast", "Base", "4or1", "buyside", "sellside", "SamountW", "BamountW"]
    df[num_cols] = df[num_cols].apply(pd.to_numeric, errors="coerce")
    
    return df

def copy_weekly_amounts_to_15m(df_15m: pd.DataFrame, df_weekly: pd.DataFrame) -> pd.DataFrame:
    """
    15분봉 시트에 주봉 시트의 SamountW, BamountW 값을 복사합니다.
    주봉의 날짜를 기준으로 해당 주에 속하는 15분봉에 값을 매칭합니다.
    
    Args:
        df_15m: 15분봉 DataFrame (Date(UTC) 컬럼 포함, UTC 기준)
        df_weekly: 주봉 DataFrame (Date(UTC) 컬럼 포함, UTC 기준)
    
    Returns:
        DataFrame: SamountW, BamountW 컬럼이 추가된 15분봉 DataFrame
    
    Note:
        - 모든 시간 매칭은 UTC 기준으로 수행 (KST 변환 없음, VLOOKUP 방식)
        - 주봉 시작일부터 7일간의 15분봉에 UTC 시간 기준으로 매칭
        - 인덱스나 순서가 아닌 UTC 시간 자체로 매칭
    """
    df_15m = df_15m.copy()
    df_weekly = df_weekly.copy()
    
    # SamountW, BamountW 열 초기화
    df_15m["SamountW"] = np.nan
    df_15m["BamountW"] = np.nan
    
    # 15분봉 날짜를 datetime으로 변환 (UTC 기준)
    df_15m["Date(UTC)_dt"] = pd.to_datetime(df_15m["Date(UTC)"], format="%y/%m/%d,%H:%M", errors='coerce')
    
    # 주봉 날짜를 datetime으로 변환 (UTC 기준, YY/MM/DD,HH:00 형식)
    df_weekly["Date(UTC)_dt"] = pd.to_datetime(df_weekly["Date(UTC)"], format="%y/%m/%d,%H:%M", errors='coerce')
    
    # 15분봉 날짜만 추출 (시간 제거, UTC 기준 날짜)
    df_15m["Date_only"] = df_15m["Date(UTC)_dt"].dt.date
    
    # 주봉 데이터를 날짜순으로 정렬 (과거→최신, UTC 기준)
    df_weekly = df_weekly.sort_values("Date(UTC)_dt", ascending=True).reset_index(drop=True)
    
    for _, weekly_row in df_weekly.iterrows():
        if pd.isna(weekly_row["Date(UTC)_dt"]):
            continue
        weekly_start_date = weekly_row["Date(UTC)_dt"].date()  # UTC 기준 날짜
        weekly_end_date = (weekly_row["Date(UTC)_dt"] + pd.Timedelta(days=7)).date()  # UTC 기준 날짜
        
        # 해당 주에 속하는 15분봉 찾기 (UTC 기준 날짜만 비교)
        mask = (df_15m["Date_only"] >= weekly_start_date) & (df_15m["Date_only"] < weekly_end_date)
        
        if mask.any():
            df_15m.loc[mask, "SamountW"] = weekly_row["SamountW"]
            df_15m.loc[mask, "BamountW"] = weekly_row["BamountW"]
    
    # 임시 컬럼 제거
    df_15m = df_15m.drop(["Date_only", "Date(UTC)_dt"], axis=1)
    
    return df_15m

def copy_daily_amounts_to_15m(df_15m: pd.DataFrame, df_daily: pd.DataFrame) -> pd.DataFrame:
    """
    15분봉 시트에 일봉 시트의 Samount1D, Bamount1D 값을 복사합니다.
    일봉의 날짜를 기준으로 해당 일에 속하는 15분봉에 값을 매칭합니다.
    
    Args:
        df_15m: 15분봉 DataFrame (Date(UTC) 컬럼 포함, UTC 기준)
        df_daily: 일봉 DataFrame (Date(UTC) 컬럼 포함, UTC 기준)
    
    Returns:
        DataFrame: Samount1D, Bamount1D 컬럼이 추가된 15분봉 DataFrame
    
    Note:
        - 모든 시간 매칭은 UTC 기준으로 수행 (KST 변환 없음, VLOOKUP 방식)
        - 일봉 날짜와 15분봉 날짜를 UTC 기준으로 비교하여 매칭
        - 인덱스나 순서가 아닌 UTC 시간 자체로 매칭
    """
    df_15m = df_15m.copy()
    df_daily = df_daily.copy()
    
    # Samount1D, Bamount1D 열 초기화
    df_15m["Samount1D"] = np.nan
    df_15m["Bamount1D"] = np.nan
    
    # 15분봉 날짜를 datetime으로 변환 (UTC 기준) - format 지정 안 함 (자동 인식)
    df_15m["Date(UTC)_dt"] = pd.to_datetime(df_15m["Date(UTC)"], errors='coerce')
    
    # 일봉 날짜를 datetime으로 변환 (UTC 기준) - format 지정 안 함 (자동 인식)
    df_daily["Date(UTC)_dt"] = pd.to_datetime(df_daily["Date(UTC)"], errors='coerce')
    
    # 15분봉 날짜만 추출 (시간 제거, UTC 기준 날짜)
    df_15m["Date_only"] = df_15m["Date(UTC)_dt"].dt.date
    
    # 일봉 데이터를 날짜순으로 정렬 (과거→최신, UTC 기준)
    df_daily = df_daily.sort_values("Date(UTC)_dt", ascending=True).reset_index(drop=True)
    
    for _, daily_row in df_daily.iterrows():
        if pd.isna(daily_row["Date(UTC)_dt"]):
            continue
        daily_date = daily_row["Date(UTC)_dt"].date()  # UTC 기준 날짜
        
        # 해당 일에 속하는 15분봉 찾기 (UTC 기준 날짜만 비교)
        mask = (df_15m["Date_only"] == daily_date)
        
        if mask.any():
            df_15m.loc[mask, "Samount1D"] = daily_row["Samount1D"]
            df_15m.loc[mask, "Bamount1D"] = daily_row["Bamount1D"]
    
    # 임시 컬럼 제거
    df_15m = df_15m.drop(["Date_only", "Date(UTC)_dt"], axis=1)
    
    return df_15m

def calculate_final_amounts(df_15m: pd.DataFrame) -> pd.DataFrame:
    """
    15분봉에 최종 Samount, Bamount를 계산합니다.
    Samount = 0.7 * SamountW + 0.3 * Samount1D
    Bamount = 0.7 * BamountW + 0.3 * Bamount1D
    티커별 USDT 정밀도 적용
    """
    df_15m = df_15m.copy()
    
    # 현재 티커의 USDT 정밀도 가져오기
    symbol = f"{TICKER}USDT"
    usdt_precision = SYMBOL_USDT_PRECISION.get(symbol, 5)  # 기본값 5자리
    
    # 최종 Samount 계산: 0.7 * SamountW + 0.3 * Samount1D (티커별 정밀도)
    df_15m["Samount"] = df_15m.apply(
        lambda row: round(0.7 * row["SamountW"] + 0.3 * row["Samount1D"], usdt_precision) 
        if not pd.isna(row["SamountW"]) and not pd.isna(row["Samount1D"]) 
        else np.nan, axis=1
    )
    
    # 최종 Bamount 계산: 0.7 * BamountW + 0.3 * Bamount1D (티커별 정밀도)
    df_15m["Bamount"] = df_15m.apply(
        lambda row: round(0.7 * row["BamountW"] + 0.3 * row["Bamount1D"], usdt_precision) 
        if not pd.isna(row["BamountW"]) and not pd.isna(row["Bamount1D"]) 
        else np.nan, axis=1
    )
    
    return df_15m

def _extract_latest_close_from_15m(df_15m: pd.DataFrame) -> Optional[float]:
    """
    15분봉 최신 캔들의 종가를 추출하여 반환합니다.
    엑셀 시트 기준 C2 셀(헤더 제외 첫 행)에 해당하는 값입니다.
    """
    if df_15m.empty or '종' not in df_15m.columns:
        return None
    
    latest_close = df_15m.iloc[0]['종']
    if pd.isna(latest_close):
        return None
    
    if isinstance(latest_close, str):
        latest_close = latest_close.replace(',', '').strip()
    
    try:
        latest_close = float(latest_close)
    except (ValueError, TypeError):
        latest_close = pd.to_numeric(pd.Series([latest_close]), errors='coerce').iloc[0]
    
    if pd.isna(latest_close):
        return None
    return float(latest_close)

def _override_latest_close(df_target: pd.DataFrame, latest_close: float, sheet_label: str, stage_prefix: str = "") -> pd.DataFrame:
    """
    대상 시트 DataFrame의 최신 행(엑셀 2행)의 종가를 15분봉 최신 종가로 대체합니다.
    """
    if df_target.empty or '종' not in df_target.columns or latest_close is None:
        return df_target
    
    first_index = df_target.index[0]
    current_value = df_target.at[first_index, '종']
    
    # 값이 이미 동일하면 로그만 출력하지 않고 반환
    is_same = False
    try:
        if not pd.isna(current_value):
            cmp_value = float(str(current_value).replace(',', '').strip())
            is_same = math.isclose(cmp_value, latest_close, rel_tol=1e-9, abs_tol=1e-6)
    except (ValueError, TypeError):
        pass
    
    df_target.at[first_index, '종'] = latest_close
    
    if not is_same:
        # XRP는 소수점 넷째자리, 그 외는 둘째자리로 표시
        ticker_from_label = sheet_label.replace("USDT1H", "").replace("USDT1D", "").replace("USDTW", "").replace("USDT5M", "").replace("USDT15M", "")
        if ticker_from_label == "XRP":
            formatted_close = f"{latest_close:,.4f}" if not math.isnan(latest_close) else str(latest_close)
        else:
            formatted_close = f"{latest_close:,.2f}" if not math.isnan(latest_close) else str(latest_close)
        print(f"{get_timestamp()} [{stage_prefix}] 🔁 {sheet_label} 최신 종가를 15분봉 값 {formatted_close}으로 대체")
    
    return df_target

def calculate_sb1h_for_15m(df_15m: pd.DataFrame, df_1h4x: pd.DataFrame) -> pd.DataFrame:
    """
    15분봉 시트에 SB1H 열을 계산합니다.
    1H4x 시트의 시간대가 15분 간격으로 되어 있어 15분봉 시트와 행마다 정확히 매칭합니다.
    ⚠️중요: Date(UTC) 기준으로만 수행 (KST는 기록용일 뿐, 계산에는 사용하지 않음)
    """
    if df_15m.empty or df_1h4x.empty:
        return df_15m
    
    df_15m_copy = df_15m.copy()
    
    # ⚠️중요: SB1H 계산은 Date(UTC) 기준으로만 수행 (KST는 기록용일 뿐, 계산에는 사용하지 않음)
    # Date(UTC) 컬럼이 있으면 우선 사용, 없으면 KST 사용 (하위 호환성)
    time_col_15m = 'Date(UTC)' if 'Date(UTC)' in df_15m_copy.columns else ('KST' if 'KST' in df_15m_copy.columns else None)
    time_col_1h4x = 'Date(UTC)' if 'Date(UTC)' in df_1h4x.columns else ('KST' if 'KST' in df_1h4x.columns else None)
    
    if time_col_15m is None or time_col_1h4x is None:
        return df_15m_copy
    
    # 1H4x 시트 데이터를 딕셔너리로 변환 (정확한 시간으로 매칭)
    hourly_dict = {}
    
    for _, row in df_1h4x.iterrows():
        time_val = row.get(time_col_1h4x)
        # 문자열 여부와 상관없이 pd.to_datetime으로 안전하게 변환
        hour_dt = pd.to_datetime(time_val, errors='coerce')
        if pd.notna(hour_dt):
            # 정확한 시간을 키로 사용
            time_key = hour_dt.strftime("%y/%m/%d,%H:%M")
            hourly_dict[time_key] = {
                'Sell': str(row.get('Sell', '')).strip().lower(),
                'Buy': str(row.get('Buy', '')).strip().lower()
            }
    
    def get_sb1h_signal(time_val):
        # 문자열 여부와 상관없이 pd.to_datetime으로 안전하게 변환
        dt_obj = pd.to_datetime(time_val, errors='coerce')
        if pd.notna(dt_obj):
            time_key = dt_obj.strftime("%y/%m/%d,%H:%M")
            if time_key in hourly_dict:
                data = hourly_dict[time_key]
                # 우선순위: Sell > Buy > 빈값
                if data['Sell'] == 'sell':
                    return 'sell'
                elif data['Buy'] == 'buy':
                    return 'buy'
        return np.nan  # 빈 문자열 대신 np.nan 사용
    
    # SB1H 열 추가
    df_15m_copy['SB1H'] = df_15m_copy[time_col_15m].apply(get_sb1h_signal)
    
    return df_15m_copy

def calculate_daysb_15m(df_15m, df_daily, market_type):
    """
    15분봉 시트에 SB1D 열을 계산합니다.
    성능 최적화된 버전:
    - 15분봉 날짜에서 시분 제거하여 날짜만 추출
    - 일봉 시트에서 같은 날짜의 Sell/Buy 신호 검색
    - 우선순위: Sell > Buy > 빈값
    - 벡터화 연산으로 성능 대폭 향상
    - ⚠️중요: Date(UTC) 기준으로만 수행 (KST는 기록용일 뿐, 계산에는 사용하지 않음)
    """
    if df_15m.empty or df_daily.empty:
        return df_15m
    
    # Sell/Buy 컬럼이 없으면 SB1D 계산 불가
    if 'Sell' not in df_daily.columns or 'Buy' not in df_daily.columns:
        print(f"{get_timestamp()} [SB1D] ⚠️ 일봉에 Sell/Buy 컬럼 없음, SB1D 계산 건너뜀")
        return df_15m
    
    df_15m_copy = df_15m.copy()
    df_daily_copy = df_daily.copy()
    
    # ⚠️중요: SB1D 계산은 Date(UTC) 기준으로만 수행 (KST는 기록용일 뿐, 계산에는 사용하지 않음)
    # Date(UTC) 컬럼이 있으면 우선 사용, 없으면 KST를 UTC로 변환
    time_col_15m = 'Date(UTC)' if 'Date(UTC)' in df_15m_copy.columns else ('KST' if 'KST' in df_15m_copy.columns else None)
    time_col_daily = 'Date(UTC)'  # 일봉은 항상 UTC 기준
    
    if time_col_15m is None:
        print(f"{get_timestamp()} [SB1D] ⚠️ 15분봉에 Date(UTC)/KST 컬럼 없음, SB1D 계산 건너뜀")
        return df_15m_copy
    
    try:
        # 1. 15분봉 날짜를 안전하게 date 객체로 변환 (UTC 기준)
        # format을 지정하지 않아야 이미 Timestamp 객체인 경우나 다른 날짜 형식에도 대응 가능합니다.
        df_15m_copy['Date_only'] = pd.to_datetime(df_15m_copy[time_col_15m], errors='coerce').dt.date
        
        # 2. 일봉 날짜를 안전하게 date 객체로 변환
        # split(',') 같은 문자열 처리를 제거하고 바로 datetime으로 변환합니다.
        df_daily_copy['Date_only'] = pd.to_datetime(df_daily_copy[time_col_daily], errors='coerce').dt.date
    except Exception as e:
        print(f"{get_timestamp()} [SB1D] ⚠️ 날짜 변환 중 예외 발생: {e}")
        return df_15m_copy
    
    # 3. 일봉 데이터를 딕셔너리로 변환 (빠른 조회를 위해)
    daily_dict = {}
    for _, row in df_daily_copy.iterrows():
        date_key = row['Date_only']
        if pd.notna(date_key):
            daily_dict[date_key] = {
                'Sell': str(row.get('Sell', '')).strip().lower(),
                'Buy': str(row.get('Buy', '')).strip().lower()
            }
    
    # 4. 매칭 로직
    def get_sb1d_signal(date_only):
        if pd.notna(date_only) and date_only in daily_dict:
            daily_data = daily_dict[date_only]
            # 우선순위: Sell > Buy
            if daily_data['Sell'] == 'sell':
                return 'sell'
            elif daily_data['Buy'] == 'buy':
                return 'buy'
        return np.nan  # 빈 문자열 대신 np.nan 사용
    
    # 5. SB1D 열 업데이트
    df_15m_copy['SB1D'] = df_15m_copy['Date_only'].apply(get_sb1d_signal)
    
    # 임시 컬럼 제거
    df_15m_copy = df_15m_copy.drop('Date_only', axis=1)
    
    return df_15m_copy

def calculate_sb5m_for_15m(df_15m, df_5m):
    """
    5분봉 데이터를 기반으로 15분봉에 SB5M 신호를 추가합니다.
    5분봉 3개씩 그룹화해서 Buy/Sell 개수를 세어 판정합니다.
    날짜 기준은 Date(UTC)로 처리합니다.
    """
    if df_5m.empty or df_15m.empty:
        return df_15m
    
    # 원본 보존
    df_5m = df_5m.copy()
    df_15m = df_15m.copy()
    
    # ⚠️중요: 모든 그룹화와 정렬은 Date(UTC) 기준으로만 수행 (KST는 기록용일 뿐, 정렬/그룹화에는 사용하지 않음)
    # Date(UTC) 기준 그룹화
    if 'Date(UTC)' not in df_5m.columns or 'Date(UTC)' not in df_15m.columns:
        if 'SB5M' not in df_15m.columns:
            df_15m['SB5M'] = ''
        return df_15m
    
    df_5m['Date(UTC)_dt'] = pd.to_datetime(df_5m['Date(UTC)'], format='%y/%m/%d,%H:%M', errors='coerce')
    df_5m = df_5m[df_5m['Date(UTC)_dt'].notna()].copy()
    
    if df_5m.empty:
        if 'SB5M' not in df_15m.columns:
            df_15m['SB5M'] = ''
        if 'Date(UTC)_dt' in df_5m.columns:
            df_5m = df_5m.drop('Date(UTC)_dt', axis=1)
        return df_15m
    
    # 시간 정렬 (과거 → 현재)
    df_5m = df_5m.sort_values('Date(UTC)_dt', ascending=True).reset_index(drop=True)
    
    # 5분봉: 15분 그룹 생성
    df_5m['15min_group'] = df_5m['Date(UTC)_dt'].dt.floor('15min')
    
    # 15분봉: 15min_group 생성
    if df_15m['Date(UTC)'].dtype == 'object':
        df_15m['Date(UTC)_dt'] = pd.to_datetime(df_15m['Date(UTC)'], format='%y/%m/%d,%H:%M', errors='coerce')
    else:
        df_15m['Date(UTC)_dt'] = pd.to_datetime(df_15m['Date(UTC)'], errors='coerce')
    
    if '15min_group' not in df_15m.columns:
        df_15m['15min_group'] = df_15m['Date(UTC)_dt'].dt.floor('15min')
    else:
        mask_na = df_15m['15min_group'].isna()
        if mask_na.any():
            df_15m.loc[mask_na, '15min_group'] = df_15m.loc[mask_na, 'Date(UTC)_dt'].dt.floor('15min')
    
    # --- 15분 그룹별 SB5M 신호 계산 ---
    sb5m_data = []
    for group_time, group_df in df_5m.groupby('15min_group', sort=True):
        buy_count = 0
        sell_count = 0
        
        for _, row in group_df.iterrows():
            buy_val = row.get('Buy', '')
            sell_val = row.get('Sell', '')
            
            if pd.notna(buy_val) and isinstance(buy_val, str) and buy_val.strip().lower() == 'buy':
                buy_count += 1
            elif pd.notna(sell_val) and isinstance(sell_val, str) and sell_val.strip().lower() == 'sell':
                sell_count += 1
        
        if buy_count == 3:
            sb5m_signal = 'buy3'
        elif buy_count == 2:
            sb5m_signal = 'buy2'
        elif buy_count == 1:
            sb5m_signal = 'buy1'
        elif sell_count == 3:
            sb5m_signal = 'sell3'
        elif sell_count == 2:
            sb5m_signal = 'sell2'
        elif sell_count == 1:
            sb5m_signal = 'sell1'
        else:
            sb5m_signal = ''
        
        sb5m_data.append({
            '15min_group': group_time,
            'SB5M': sb5m_signal
        })
    
    # --- 15분봉에 SB5M 매칭 ---
    if sb5m_data:
        df_sb5m = pd.DataFrame(sb5m_data)
        df_15m = df_15m.merge(df_sb5m, on='15min_group', how='left')
        df_15m['SB5M'] = df_15m['SB5M'].fillna('')
    else:
        if 'SB5M' not in df_15m.columns:
            df_15m['SB5M'] = ''
    
    # 임시 컬럼 제거
    if 'Date(UTC)_dt' in df_15m.columns:
        df_15m = df_15m.drop('Date(UTC)_dt', axis=1)
    if '15min_group' in df_15m.columns:
        df_15m = df_15m.drop('15min_group', axis=1)
    
    return df_15m

def copy_1hmsfast_to_5m(df_5m: pd.DataFrame, df_15m: pd.DataFrame) -> pd.DataFrame:
    """
    5분봉 시트에 15분봉 시트의 1HMSFast 값을 시간 매칭하여 복사합니다 (1HMSF로 저장).
    15분 구간을 기준으로 15분봉의 1HMSFast 값을 매칭합니다.
    
    Args:
        df_5m: 5분봉 DataFrame (Date(UTC) 컬럼 포함, UTC 기준)
        df_15m: 15분봉 DataFrame (Date(UTC) 컬럼 포함, UTC 기준)
    
    Returns:
        DataFrame: 1HMSF 컬럼이 추가된 5분봉 DataFrame
    
    Note:
        - 모든 시간 매칭은 UTC 기준으로 수행 (KST 변환 없음, VLOOKUP 방식)
        - 15분 구간의 시작 시간(0분, 15분, 30분, 45분)을 UTC 기준으로 계산하여 매칭
        - 인덱스나 순서가 아닌 UTC 시간 자체로 매칭
    """
    if df_5m.empty or df_15m.empty:
        if '1HMSF' not in df_5m.columns:
            df_5m['1HMSF'] = np.nan
        return df_5m
    
    df_5m_copy = df_5m.copy()
    
    # 15분봉 데이터를 딕셔너리로 변환 (UTC 기준)
    fifteen_min_dict = {}
    for _, row in df_15m.iterrows():
        date_val = row.get('Date(UTC)', '')
        # Timestamp 객체인 경우 직접 처리
        if isinstance(date_val, pd.Timestamp) or hasattr(date_val, 'strftime'):
            dt_val = pd.to_datetime(date_val, errors='coerce')
            if pd.notna(dt_val):
                if isinstance(dt_val, pd.Timestamp):
                    dt_val = dt_val.to_pydatetime()
                if dt_val.tzinfo is None:
                    dt_val = dt_val.replace(tzinfo=tz.UTC)
                # 15분 구간의 시작 시간 계산 (0분, 15분, 30분, 45분으로 정렬)
                minute = dt_val.minute
                minute_aligned = (minute // 15) * 15
                fifteen_min_start = dt_val.replace(minute=minute_aligned, second=0, microsecond=0)
                key = fifteen_min_start.strftime("%y/%m/%d,%H:%M")
                fifteen_min_dict[key] = row.get('1HMSFast', np.nan)
        else:
            date_str = str(date_val)
            if ',' in date_str:
                date_part, time_part = date_str.split(',')
                try:
                    dt_val = dt.datetime.strptime(f"{date_part} {time_part}", "%y/%m/%d %H:%M")
                    dt_val = dt_val.replace(tzinfo=tz.UTC)  # UTC 기준
                    # 15분 구간의 시작 시간 계산 (0분, 15분, 30분, 45분으로 정렬)
                    minute = dt_val.minute
                    minute_aligned = (minute // 15) * 15
                    fifteen_min_start = dt_val.replace(minute=minute_aligned, second=0, microsecond=0)
                    key = fifteen_min_start.strftime("%y/%m/%d,%H:%M")
                    fifteen_min_dict[key] = row.get('1HMSFast', np.nan)
                except:
                    pass
    
    def get_1hmsf_value(utc_time_str):
        # Timestamp 객체인 경우 문자열로 변환
        if hasattr(utc_time_str, 'strftime'):
            utc_time_str = utc_time_str.strftime("%y/%m/%d,%H:%M")
        
        if ',' in str(utc_time_str):
            date_part, time_part = str(utc_time_str).split(',')
            utc_dt = dt.datetime.strptime(f"{date_part} {time_part}", "%y/%m/%d %H:%M")
            utc_dt = utc_dt.replace(tzinfo=tz.UTC)  # UTC 기준
            
            # 해당 15분 구간의 시작 시간 계산 (UTC 기준, 0분, 15분, 30분, 45분으로 정렬)
            minute = utc_dt.minute
            minute_aligned = (minute // 15) * 15
            fifteen_min_start = utc_dt.replace(minute=minute_aligned, second=0, microsecond=0)
            key = fifteen_min_start.strftime("%y/%m/%d,%H:%M")
            
            if key in fifteen_min_dict:
                return fifteen_min_dict[key]
        return np.nan
    
    # 1HMSF 열 추가/업데이트 (숫자로 유지, 엑셀 저장 시 포맷팅 적용)
    df_5m_copy['1HMSF'] = df_5m_copy['Date(UTC)'].apply(get_1hmsf_value)
    # 숫자 타입으로 유지 (문자열 변환 제거 - 엑셀에서 숫자로 인식되도록)
    
    return df_5m_copy

def recalculate_buy_for_5m(df_5m: pd.DataFrame) -> pd.DataFrame:
    """
    5분봉 시트의 Buy 열을 재계산합니다.
    1HMSF와 SB1M 열을 고려하여 gear1/gear2 조건에 따라 계산합니다.
    
    gear1 (2 <= 1HMSF < 7): 기존 조건
    - 4or1 < 4 AND sellside <= 0.05
    
    gear2 (1HMSF >= 7 또는 1HMSF < 2): 변경 조건
    - 4or1 < 4.9 AND sellside <= 0.05 AND SB1M이 sell1~sell5가 아님
    """
    if df_5m.empty:
        return df_5m
    
    df_5m_copy = df_5m.copy()
    
    def calculate_buy_with_gear(row):
        """gear 조건에 따라 Buy 신호 계산"""
        four_or_one = row.get('4or1', np.nan)
        sellside_val = row.get('sellside', np.nan)
        hmsf_val = row.get('1HMSF', np.nan)
        sb1m_val = row.get('SB1M', '')
        
        # NaN 체크
        if pd.isna(four_or_one) or pd.isna(sellside_val):
            return ""
        
        # sellside 조건 체크 (공통)
        if sellside_val > 0.05:
            return ""
        
        # 1HMSF 값 확인 (gear 분리용) - 문자열인 경우 숫자로 변환
        hmsf_float = None
        if pd.notna(hmsf_val):
            try:
                # 문자열인 경우 숫자로 변환
                if isinstance(hmsf_val, str):
                    hmsf_float = float(hmsf_val)
                else:
                    hmsf_float = float(hmsf_val)
            except (TypeError, ValueError):
                hmsf_float = None
        
        # gear 분리: gear1 (2 <= 1HMSF < 7), gear2 (1HMSF >= 7 또는 1HMSF < 2)
        is_gear1 = (hmsf_float is not None and 2.0 <= hmsf_float < 7.0)
        is_gear2 = (hmsf_float is not None and (hmsf_float >= 7.0 or hmsf_float < 2.0))
        
        # gear1: 기존 조건 (4or1 < 4 AND sellside <= 0.05)
        if is_gear1:
            if four_or_one < 4:
                return "buy"
            else:
                return ""
        
        # gear2: 변경 조건 (4or1 < 4.9 AND sellside <= 0.05 AND SB1M이 sell1~sell5가 아님)
        elif is_gear2:
            # 4or1 조건 체크
            if four_or_one >= 4.9:
                return ""
            
            # SB1M 조건 체크 (sell1, sell2, sell3, sell4, sell5가 아니어야 함)
            if isinstance(sb1m_val, str):
                sb1m_lower = sb1m_val.strip().lower()
                if sb1m_lower in ['sell1', 'sell2', 'sell3', 'sell4', 'sell5']:
                    return ""  # SB1M이 sell1~sell5이면 Buy 신호 없음
            
            # 모든 조건 만족
            return "buy"
        
        # 1HMSF가 NaN이거나 gear 범위 밖인 경우: 기존 조건만 체크 (gear1과 동일)
        else:
            if four_or_one < 4:
                return "buy"
            else:
                return ""
    
    # Buy 열 재계산
    df_5m_copy['Buy'] = df_5m_copy.apply(calculate_buy_with_gear, axis=1)
    
    return df_5m_copy

def copy_1hmsfast_from_1h_to_15m(df_15m: pd.DataFrame, df_1h: pd.DataFrame) -> pd.DataFrame:
    """
    15분봉 시트에 1H 시트의 1HMSFast 값을 시간 매칭하여 복사합니다.
    1시간 구간을 기준으로 1시간봉의 1HMSFast 값을 매칭합니다.
    
    Args:
        df_15m: 15분봉 DataFrame (Date(UTC) 컬럼 포함, UTC 기준)
        df_1h: 1시간봉 DataFrame (Date(UTC) 컬럼 포함, UTC 기준)
    
    Returns:
        DataFrame: 1HMSFast 컬럼이 업데이트된 15분봉 DataFrame
    
    Note:
        - 모든 시간 매칭은 UTC 기준으로 수행 (KST 변환 없음, VLOOKUP 방식)
        - 1시간 구간의 시작 시간(0분)을 UTC 기준으로 계산하여 매칭
        - 인덱스나 순서가 아닌 UTC 시간 자체로 매칭
    """
    if df_15m.empty or df_1h.empty:
        return df_15m
    
    df_15m_copy = df_15m.copy()
    
    # 1시간봉 데이터를 딕셔너리로 변환 (UTC 기준)
    hourly_dict = {}
    for _, row in df_1h.iterrows():
        date_val = row.get('Date(UTC)', '')
        # Timestamp 객체인 경우 직접 처리
        if isinstance(date_val, pd.Timestamp) or hasattr(date_val, 'strftime'):
            hour_dt = pd.to_datetime(date_val, errors='coerce')
            if pd.notna(hour_dt):
                if isinstance(hour_dt, pd.Timestamp):
                    hour_dt = hour_dt.to_pydatetime()
                if hour_dt.tzinfo is None:
                    hour_dt = hour_dt.replace(tzinfo=tz.UTC)
                hour_start = hour_dt.replace(minute=0, second=0, microsecond=0)
                hour_key = hour_start.strftime("%y/%m/%d,%H:%M")
                hourly_dict[hour_key] = row.get('1HMSFast', np.nan)
        else:
            date_str = str(date_val)
            if ',' in date_str:
                date_part, time_part = date_str.split(',')
                try:
                    hour_dt = dt.datetime.strptime(f"{date_part} {time_part}", "%y/%m/%d %H:%M")
                    hour_dt = hour_dt.replace(tzinfo=tz.UTC)  # UTC 기준
                    # 1시간 구간의 시작 시간을 키로 사용 (UTC 기준)
                    hour_key = hour_dt.strftime("%y/%m/%d,%H:%M")
                    hourly_dict[hour_key] = row.get('1HMSFast', np.nan)
                except:
                    pass
    
    def get_1hmsfast_value(utc_time_str):
        # Timestamp 객체인 경우 문자열로 변환
        if hasattr(utc_time_str, 'strftime'):
            utc_time_str = utc_time_str.strftime("%y/%m/%d,%H:%M")
        
        if ',' in str(utc_time_str):
            date_part, time_part = str(utc_time_str).split(',')
            utc_dt = dt.datetime.strptime(f"{date_part} {time_part}", "%y/%m/%d %H:%M")
            utc_dt = utc_dt.replace(tzinfo=tz.UTC)  # UTC 기준
            
            # 해당 1시간 구간의 시작 시간 계산 (UTC 기준, 0분으로 정렬)
            hour_start = utc_dt.replace(minute=0, second=0, microsecond=0)
            hour_key = hour_start.strftime("%y/%m/%d,%H:%M")
            
            if hour_key in hourly_dict:
                return hourly_dict[hour_key]
        return np.nan
    
    # 1HMSFast 열 업데이트
    df_15m_copy['1HMSFast'] = df_15m_copy['Date(UTC)'].apply(get_1hmsfast_value)
    
    return df_15m_copy

def calculate_ksc_for_15m(df_15m: pd.DataFrame) -> pd.DataFrame:
    """
    15분봉 시트에 KillStackCount (KSC) 열을 계산합니다.
    
    계산 로직:
    - 15M 캔들을 맨 아래(과거)부터 순회 (UTC 기준)
    - ORDER가 Buy5 또는 Buy10이고, 1HMSFast가 2 <= 1HMSFast < 7이고, spread >= 티커별 SPRD2 400개 평균의 30%일 때만 KSC 카운트 스택 증가
    - spread = (max(sma25, sma100, sma200) - min(sma25, sma100, sma200)) / min(sma25, sma100, sma200)
    - 스택 쌓임: ORDER가 Buy5/Buy10이고, 2 <= 1HMSFast <= 4.4 또는 4.6 <= 1HMSFast < 7 (spread >= 티커별 SPRD2 400개 평균의 30% 조건 만족 시)
    - Bomb 처리: ORDER가 Buy5/Buy10이고, 4.4 < 1HMSFast < 4.6 (스택 유지, spread >= 티커별 SPRD2 400개 평균의 30% 조건 만족 시)
    - 이전 값 유지: ORDER가 Buy5/Buy10이 아니지만, 2 <= 1HMSFast < 7이고 spread >= 티커별 SPRD2 400개 평균의 30%이면 이전 값 유지
    - spread < 티커별 SPRD2 400개 평균의 30% 또는 1HMSFast < 2 또는 1HMSFast >= 7이면 0으로 스택 초기화
    
    Args:
        df_15m: 15분봉 DataFrame (Date(UTC) 컬럼 포함, UTC 기준)
    
    Returns:
        DataFrame: KSC 컬럼이 추가된 15분봉 DataFrame
    
    Note:
        - Date(UTC) 컬럼 기준으로 정렬 및 계산 (UTC 기준)
    """
    if df_15m.empty:
        return df_15m
    
    df_15m_copy = df_15m.copy()
    
    # Date(UTC) 기준으로 정렬
    sort_col = 'Date(UTC)'
    
    # 과거→현재 순서로 정렬 (맨 아래부터 계산)
    df_15m_copy = df_15m_copy.sort_values(sort_col, ascending=True).reset_index(drop=True)
    
    # 티커별 SPRD2 열의 최근 400개 평균의 30%를 threshold로 계산
    sprd2_threshold = None  # SPRD2가 없거나 데이터가 부족한 경우 None
    if 'SPRD2' in df_15m_copy.columns:
        # 최신→과거 순서로 정렬된 상태에서 최근 400개 추출
        df_sorted_desc = df_15m_copy.sort_values(sort_col, ascending=False).reset_index(drop=True)
        sprd2_recent = df_sorted_desc['SPRD2'].head(400)
        sprd2_valid = sprd2_recent[pd.notna(sprd2_recent)]
        if len(sprd2_valid) > 0:
            sprd2_avg = sprd2_valid.mean()
            sprd2_threshold = sprd2_avg * 0.3  # 평균의 30%
    
    # KSC 열 초기화 (숫자만 저장)
    ksc_values = []
    # Bomb 열 초기화 (Bomb 발생 시 "Bomb" 저장)
    bomb_values = []
    bomb_count_values = []  # Bomb이 되는 시점의 카운트 값 저장 (BombCount 열용)
    ksc_stack_values = []  # Bomb이 되는 시점의 스택 카운트 값 저장
    prev_kill_count = 0
    
    for idx, row in df_15m_copy.iterrows():
        order = str(row.get('ORDER', '')).strip()
        hmsfast = row.get('1HMSFast', np.nan)
        sma25 = row.get('SMA25', np.nan)
        sma100 = row.get('SMA100', np.nan)
        sma200 = row.get('SMA200', np.nan)
        ksc_value = prev_kill_count  # 기본값은 이전 카운트 (숫자만)
        bomb_value = ""  # Bomb이 아닌 경우 빈 문자열
        bomb_count = 0  # Bomb이 아닌 경우 0 (BombCount 열용)
        ksc_stack = 0  # Bomb이 아닌 경우 0
        
        # spread 계산: (max(sma25, sma100, sma200) - min(sma25, sma100, sma200)) / min(sma25, sma100, sma200)
        spread = np.nan
        if not pd.isna(sma25) and not pd.isna(sma100) and not pd.isna(sma200):
            sma_values = [float(sma25), float(sma100), float(sma200)]
            sma_max = max(sma_values)
            sma_min = min(sma_values)
            if sma_min > 0:
                spread = (sma_max - sma_min) / sma_min
        
        # ORDER가 Sell5 또는 Sell10이면 초기화
        if order in ['Sell5', 'Sell10']:
            if prev_kill_count > 0:
                prev_kill_count = 0
            ksc_value = prev_kill_count
            bomb_value = ""
            bomb_count = 0
            ksc_values.append(ksc_value)
            bomb_values.append(bomb_value)
            bomb_count_values.append(bomb_count)
            ksc_stack_values.append(ksc_stack)
            continue
        
        # ORDER가 Buy5 또는 Buy10인 경우
        if order in ['Buy5', 'Buy10']:
            # 1HMSFast 값이 유효한 경우
            if not pd.isna(hmsfast):
                hmsfast_val = float(hmsfast)
                
                # 기본 조건: 2 <= 1HMSFast < 7이고 spread >= 티커별 SPRD2 400개 평균의 30%일 때만 KSC 카운트 스택 관련 로직 적용
                if 2.0 <= hmsfast_val < 7.0 and not pd.isna(spread) and sprd2_threshold is not None and spread >= sprd2_threshold:
                    # Bomb 처리: 4.4 < 1HMSFast < 4.6 (새로운 수열 규칙: Bomb 발생해도 카운트 +1, 스택 유지)
                    if 4.4 < hmsfast_val < 4.6:
                        # Bomb 표시, 카운트 증가 (스택 유지)
                        # 새로운 수열 규칙: Bomb 발생해도 카운트 +1, 스택 쌓는게 유지됨
                        prev_kill_count = prev_kill_count + 1
                        # KSC는 숫자만 저장, Bomb 열에 "Bomb" 저장
                        ksc_value = prev_kill_count  # 증가된 값을 KSC에 저장
                        bomb_value = "Bomb"  # Bomb 열에 "Bomb" 저장
                        bomb_count = prev_kill_count  # Bomb이 되는 시점의 카운트 값 저장 (BombCount 열용)
                        ksc_stack = prev_kill_count  # Bomb이 되는 시점의 스택 카운트 값 저장 (증가된 값)
                    # 스택 쌓임: 2 <= 1HMSFast <= 4.4 또는 4.6 <= 1HMSFast < 7
                    elif (2.0 <= hmsfast_val <= 4.4) or (4.6 <= hmsfast_val < 7.0):
                        # kill 카운트 증가 (제한 없음)
                        prev_kill_count = prev_kill_count + 1
                        ksc_value = prev_kill_count
                    else:
                        # 이론적으로 도달하지 않지만 안전을 위해
                        ksc_value = prev_kill_count
                # spread < 티커별 SPRD2 400개 평균의 30% 또는 1HMSFast < 2 또는 1HMSFast >= 7이면 카운트 스택 로직 적용 안 함 → 0으로 초기화
                elif hmsfast_val >= 7.0:
                    # 스택 쌓이다가 bomb 신호 없이 7 초과하면 0으로 스택 초기화
                    if prev_kill_count > 0:
                        prev_kill_count = 0
                    ksc_value = prev_kill_count
                else:
                    # 1HMSFast < 2 또는 spread < 티커별 SPRD2 400개 평균의 30%인 경우 0으로 초기화
                    if prev_kill_count > 0:
                        prev_kill_count = 0
                    ksc_value = prev_kill_count
            else:
                # 1HMSFast가 NaN이면 0으로 초기화
                if prev_kill_count > 0:
                    prev_kill_count = 0
                ksc_value = prev_kill_count
        else:
            # ORDER가 Buy5/Buy10이 아니면
            if not pd.isna(hmsfast):
                hmsfast_val = float(hmsfast)
                # 2 <= 1HMSFast < 7이고 spread >= 티커별 SPRD2 400개 평균의 30%이면 이전 값 유지
                if 2.0 <= hmsfast_val < 7.0 and not pd.isna(spread) and sprd2_threshold is not None and spread >= sprd2_threshold:
                    # 이전 값 유지 (스택 쌓지 않음)
                    ksc_value = prev_kill_count
                # spread < 티커별 SPRD2 400개 평균의 30% 또는 1HMSFast < 2 또는 1HMSFast >= 7이면 0으로 초기화
                elif hmsfast_val >= 7.0 or hmsfast_val < 2.0:
                    # 7 초과 또는 2 미만이면 스택 초기화
                    if prev_kill_count > 0:
                        prev_kill_count = 0
                    ksc_value = prev_kill_count
                else:
                    # spread < 티커별 SPRD2 400개 평균의 30%인 경우 0으로 초기화
                    if prev_kill_count > 0:
                        prev_kill_count = 0
                    ksc_value = prev_kill_count
            else:
                # 1HMSFast가 NaN이면 0으로 초기화
                if prev_kill_count > 0:
                    prev_kill_count = 0
                ksc_value = prev_kill_count
        
        ksc_values.append(ksc_value)
        bomb_values.append(bomb_value)
        bomb_count_values.append(bomb_count)
        ksc_stack_values.append(ksc_stack)
    
    # KSC 열 추가 (숫자만)
    df_15m_copy['KSC'] = ksc_values
    # Bomb 열 추가 (Bomb 발생 시 "Bomb" 저장) - object 타입으로 명시적 설정
    df_15m_copy['Bomb'] = bomb_values
    df_15m_copy['Bomb'] = df_15m_copy['Bomb'].astype('object')
    # BombCount 열 추가 (Bomb이 되는 시점의 카운트 값, Bomb이 아닌 경우 0)
    df_15m_copy['BombCount'] = bomb_count_values
    # KSC stack 열 추가 (Bomb이 되는 시점의 스택 카운트 값, Bomb이 아닌 경우 0)
    df_15m_copy['KSC stack'] = ksc_stack_values
    
    # 최신→과거로 다시 정렬
    df_15m_copy = df_15m_copy.sort_values(sort_col, ascending=False).reset_index(drop=True)
    
    return df_15m_copy

def calculate_prft_for_15m(df_15m: pd.DataFrame) -> pd.DataFrame:
    """
    15분봉 시트에 PRFT (Profit) 열을 계산합니다.
    
    계산 로직:
    - 15M 캔들을 맨 아래(과거)부터 순회
    - PRFT = 스택값 (숫자) - 항상 표시 (조건 없음)
    
    TPOVER (Target Price Over) 로직:
    - Sell5 또는 Sell10이면서 종가가 TP 열 값 이상인 경우
    - 종가 >= TP 인 경우 PRFT = 'TPOVER' (스택 초기화)
    
    PRFT 스택 로직:
    - StoSP, StoSU가 누적될 때 (이전 행보다 증가) 스택 +1
    - StoSP, StoSU가 유지될 때 (이전 행과 동일) 스택 유지
    - PRFT에 TPOVER 발생 시 스택 초기화
    
    Args:
        df_15m: 15분봉 DataFrame (Date(UTC) 컬럼 포함, UTC 기준)
    
    Returns:
        DataFrame: PRFT 컬럼이 추가된 15분봉 DataFrame
    
    Note:
        - Date(UTC) 컬럼 기준으로 정렬 및 계산 (UTC 기준)
    """
    if df_15m.empty:
        return df_15m
    
    df_15m_copy = df_15m.copy()
    
    # ⚠️중요: 정렬은 Date(UTC) 기준으로만 수행 (KST는 기록용 컬럼일 뿐, 정렬에는 사용하지 않음)
    sort_col = 'Date(UTC)'
    
    # 과거→현재 순서로 정렬 (맨 아래부터 계산)
    df_15m_copy = df_15m_copy.sort_values(sort_col, ascending=True).reset_index(drop=True)
    
    # PRFT 열 초기화
    prft_values = []
    prft_stack = 0  # PRFT 스택 값 (0부터 시작)
    prev_stosp = np.nan  # 이전 행의 StoSP 값
    prev_stosu = np.nan  # 이전 행의 StoSU 값
    
    for idx, row in df_15m_copy.iterrows():
        order = str(row.get('ORDER', '')).strip()
        order_lower = order.lower()  # 대소문자 무시 비교
        close = row.get('종', np.nan)
        stossp = row.get('StoSP', np.nan)
        stosu = row.get('StoSU', np.nan)
        tp_value = row.get('TP', np.nan)
        prft_value = 0  # 기본값은 0
        
        # TPOVER 조건 확인 (최우선, 스택 초기화) - 대소문자 무시
        if order_lower in ['sell5', 'sell10']:
            if not pd.isna(close) and not pd.isna(tp_value) and tp_value > 0:
                # 종가 >= TP 인 경우 TPOVER
                if float(close) >= float(tp_value):
                    prft_value = 'TPOVER'
                    prft_stack = 0  # TPOVER 발생 시 스택 초기화
                    prev_stosp = np.nan  # 이전 값 초기화
                    prev_stosu = np.nan
                    prft_values.append(prft_value)
                    continue
        
        # StoSP/StoSU 변화 확인 (스택 관리)
        # StoSP/StoSU가 유효한 값인지 확인 (NaN이 아니고 0이 아니면 유효)
        stossp_valid = not pd.isna(stossp) and (isinstance(stossp, (int, float)) and stossp != 0)
        stosu_valid = not pd.isna(stosu) and (isinstance(stosu, (int, float)) and stosu != 0)
        
        if stossp_valid and stosu_valid:
            # StoSP/StoSU가 초기화되지 않은 경우 (유효한 값)
            if not pd.isna(prev_stosp) and not pd.isna(prev_stosu):
                # 이전 값이 있고, StoSP 또는 StoSU가 증가했으면 스택 +1
                if stossp > prev_stosp or stosu > prev_stosu:
                    prft_stack += 1
                # StoSP와 StoSU가 모두 동일하거나 감소한 경우는 스택 유지 (변경 없음)
            
            # 현재 값을 이전 값으로 저장
            prev_stosp = stossp
            prev_stosu = stosu
        # StoSP/StoSU가 초기화된 경우 (NaN이거나 0): 스택 유지 (초기화하지 않음)
        
        # PRFT = 스택값 (항상 표시, 조건 없음)
        prft_value = prft_stack
        
        prft_values.append(prft_value)
    
    # PRFT 열 추가
    df_15m_copy['PRFT'] = prft_values
    
    # 최신→과거로 다시 정렬
    df_15m_copy = df_15m_copy.sort_values(sort_col, ascending=False).reset_index(drop=True)
    
    return df_15m_copy

def get_prft_count_from_prft(df_15m: pd.DataFrame, target_idx: int) -> int:
    """
    PRFT가 "PRFT"인 행에서 이전 행들을 역추적하여 prft 카운트를 계산합니다.
    
    Args:
        df_15m: 15분봉 DataFrame (최신→과거 순서, 즉 iloc[0]이 최신)
        target_idx: "PRFT"인 행의 인덱스 (iloc 기준)
    
    Returns:
        prft 카운트 (이전까지 증가했던 prft 카운트)
    """
    if target_idx >= len(df_15m):
        return 0
    
    # "PRFT"인 행의 PRFT 확인
    if str(df_15m.iloc[target_idx].get('PRFT', '')).strip() != 'PRFT':
        return 0
    
    # 이전 행들을 역추적하면서 카운트 계산
    # 최신→과거 순서이므로, target_idx 이후(과거 방향)로 확인
    max_count = 0
    for i in range(target_idx + 1, len(df_15m)):
        prft_value = df_15m.iloc[i].get('PRFT', 0)
        
        # PRFT가 숫자인 경우
        if isinstance(prft_value, (int, float)):
            max_count = max(max_count, int(prft_value))
        # PRFT가 "PRFT"이거나 0이면 이전 카운트 체인이 끝남
        elif str(prft_value).strip() == 'PRFT' or prft_value == 0:
            break
    
    return max_count

def get_ksc_stack_from_ksc(df_15m: pd.DataFrame, target_idx: int) -> int:
    """
    KSC가 "Bomb"인 행에서 이전 행들을 역추적하여 KSC stack을 계산합니다.
    
    Args:
        df_15m: 15분봉 DataFrame (최신→과거 순서, 즉 iloc[0]이 최신)
        target_idx: "Bomb"인 행의 인덱스 (iloc 기준)
    
    Returns:
        KSC stack (이전까지 증가했던 kill 카운트)
    """
    if target_idx >= len(df_15m):
        return 0
    
    # "Bomb"인 행의 KSC 확인
    if str(df_15m.iloc[target_idx].get('KSC', '')).strip() != 'Bomb':
        return 0
    
    # 이전 행들을 역추적하면서 카운트 계산
    # 최신→과거 순서이므로, target_idx 이후(과거 방향)로 확인
    max_count = 0
    for i in range(target_idx + 1, len(df_15m)):
        ksc_value = df_15m.iloc[i].get('KSC', 0)
        
        # KSC가 숫자인 경우
        if isinstance(ksc_value, (int, float)):
            max_count = max(max_count, int(ksc_value))
        # KSC가 "Bomb"이거나 0이면 이전 카운트 체인이 끝남
        elif str(ksc_value).strip() == 'Bomb' or ksc_value == 0:
            break
    
    return max_count

def calculate_latest_row_only_15m(df, market_type):
    """
    15분봉 After 단계 최적화: 최신 1개 행만 지표 계산 (previous 지표 유지)
    
    입력: [새 데이터(idx=0), Previous(idx=1~)] (최신→과거 순서, UTC 기준)
    출력: [새 데이터(지표 계산됨), Previous(그대로)] (최신→과거 순서 유지)
    
    계산 방식:
    - SMA10: idx 0~9 (2행 + 3~11행) 총 10개를 더해서 10으로 나눔
    - 모든 지표는 2행(idx=0) + previous 데이터를 사용
    
    Args:
        df: 15분봉 DataFrame (Date(UTC) 컬럼 포함, UTC 기준)
        market_type: 시장 타입 (사용하지 않음, 호환성 유지)
    
    Returns:
        DataFrame: 최신 행만 지표가 계산된 15분봉 DataFrame
    
    Note:
        - Date(UTC) 컬럼 기준으로 정렬 및 계산 (UTC 기준)
    """
    if df.empty:
        return df
    
    # [Cursor 패치] 엑셀에서 읽은 데이터의 쉼표 제거 및 숫자/날짜 강제 변환
    # 이 과정이 없으면 데이터가 제대로 읽히지 않아 len(df)가 줄어들고 fallback으로 튕김
    df = clean_df_display_format(df)
    
    # 데이터가 조금이라도 있으면 최대한 latest_row_only 모드 유지
    # 200개 미만이어도 새 데이터만 계산 시도 (previous 데이터가 있으면 활용)
    if len(df) < 25:
        # 최소한 SMA25 계산을 위해 25개는 필요
        return calculate_all_indicators_15m(df, market_type)
    
    # ⚠️중요: 엑셀에서 읽은 데이터의 숫자 컬럼을 강제로 숫자로 변환 (쉼표 제거 후 변환)
    # 엑셀에서 불러온 숫자가 문자열("88,123.45")로 저장되어 있을 수 있음
    # 이 과정이 없으면 SMA 계산 시 문자열과 숫자가 섞여서 NaN이 발생합니다.
    numeric_cols = ['종', '시', '고', '저', 'Vol.']
    for col in numeric_cols:
        if col in df.columns:
            # 문자열인 경우에만 쉼표 제거 시도
            if df[col].dtype == 'object':
                df[col] = df[col].astype(str).str.replace(',', '')
            # 강제 숫자 변환
            df[col] = pd.to_numeric(df[col], errors='coerce')
    
    # Buy/Sell 컬럼이 없으면 생성
    if "Buy" not in df.columns:
        df["Buy"] = ""
    if "Sell" not in df.columns:
        df["Sell"] = ""
    
    # Previous 데이터의 Buy/Sell이 NaN이면 빈 문자열로 변환
    df["Buy"] = df["Buy"].fillna("")
    df["Sell"] = df["Sell"].fillna("")
    
    # idx=0만 계산
    idx = 0
    
    # SMA 계산: 2행(idx=0) 포함하여 계산
    # SMA3: idx 0,1,2 (2행+3행+4행) = 3개
    # SMA10: idx 0~9 (2행~11행) = 10개
    # SMA200: idx 0~199 (2행~201행) = 200개
    # 
    # SMA200 계산식 (AFTER 단계):
    #   SMA200[idx=0] = (종가[idx=0] + 종가[idx=1] + ... + 종가[idx=199]) / 200
    #   = df.iloc[0:200]["종"].mean()
    # 
    # PREVIOUS 단계와의 일관성:
    #   PREVIOUS: df["SMA200"] = df["종"].rolling(window=200, min_periods=200).mean()
    #   AFTER:    df.loc[0, "SMA200"] = df.iloc[0:200]["종"].mean()
    #   → 두 방식 모두 동일한 결과: 최신 행(idx=0)을 포함한 최근 200개 캔들의 종가 평균
    df.loc[idx, "SMA3"] = df.iloc[idx:idx+3]["종"].mean()
    df.loc[idx, "SMA5"] = df.iloc[idx:idx+5]["종"].mean()
    df.loc[idx, "SMA7"] = df.iloc[idx:idx+7]["종"].mean()
    df.loc[idx, "SMA10"] = df.iloc[idx:idx+10]["종"].mean()
    df.loc[idx, "SMA12"] = df.iloc[idx:idx+12]["종"].mean()
    sma3_v = df.loc[idx, "SMA3"]
    sma12_v = df.loc[idx, "SMA12"]
    # SMAF: SMA3·SMA12 6:4 가중평균
    df.loc[idx, "SMAF"] = (float(sma3_v) * 0.6 + float(sma12_v) * 0.4) if pd.notna(sma3_v) and pd.notna(sma12_v) else np.nan
    df.loc[idx, "SMA20"] = df.iloc[idx:idx+20]["종"].mean()
    df.loc[idx, "SMA25"] = df.iloc[idx:idx+25]["종"].mean()
    df.loc[idx, "SMA50"] = df.iloc[idx:idx+50]["종"].mean()
    df.loc[idx, "SMA100"] = df.iloc[idx:idx+100]["종"].mean()
    df.loc[idx, "SMA200"] = df.iloc[idx:idx+200]["종"].mean()
    df.loc[idx, "SMA400"] = df.iloc[idx:idx+400]["종"].mean() if idx + 400 <= len(df) else np.nan
    df.loc[idx, "SMA800"] = df.iloc[idx:idx+800]["종"].mean() if idx + 800 <= len(df) else np.nan
    
    # Max70, Min70 계산: 2행 포함 70개 캔들 (idx 0~69) - Source 기준
    if idx + 70 <= len(df):
        window_data = df.iloc[idx:idx+70][["시", "고", "저", "종"]]
        df.loc[idx, "Max70"] = window_data.values.max()
        df.loc[idx, "Min70"] = window_data.values.min()
    else:
        df.loc[idx, "Max70"] = np.nan
        df.loc[idx, "Min70"] = np.nan
    
    # 하단, 상단 계산 (Max70, Min70 사용)
    current_price = df.loc[idx, "종"]
    min70 = df.loc[idx, "Min70"]
    max70 = df.loc[idx, "Max70"]
    
    if pd.notna(min70) and min70 != 0:
        df.loc[idx, "하단"] = abs((current_price - min70) / min70)
    else:
        df.loc[idx, "하단"] = np.nan
        
    if pd.notna(max70) and max70 != 0:
        df.loc[idx, "상단"] = abs((current_price - max70) / max70)
    else:
        df.loc[idx, "상단"] = np.nan
    
    # SFast, Fast, Base 계산
    df.loc[idx, "SFast"] = calculate_superfast(df.loc[idx, "SMA3"], df.loc[idx, "SMA5"], df.loc[idx, "SMA7"])
    df.loc[idx, "Fast"] = calculate_fast(df.loc[idx, "SMA5"], df.loc[idx, "SMA7"], df.loc[idx, "SMA10"])
    df.loc[idx, "Base"] = calculate_base(df.loc[idx, "SMA5"], df.loc[idx, "SMA10"], df.loc[idx, "SMA20"])
    
    # 4or1 계산
    df.loc[idx, "4or1"] = calculate_4or1(df.loc[idx, "하단"], df.loc[idx, "상단"])
    
    # buyside, sellside 계산
    df.loc[idx, "buyside"] = calculate_buyside(df.loc[idx, "SFast"], df.loc[idx, "Fast"], df.loc[idx, "Base"])
    df.loc[idx, "sellside"] = calculate_sellside(df.loc[idx, "SFast"], df.loc[idx, "Fast"], df.loc[idx, "Base"])
    
    # Buy, Sell 계산
    df.loc[idx, "Buy"] = calculate_buy(df.loc[idx, "4or1"], df.loc[idx, "sellside"])
    df.loc[idx, "Sell"] = calculate_sell_short(df.loc[idx, "buyside"])
    
    # 1HMSFast 계산 (15M: idx 행 자신의 SMAF, SMA100, SMA200 사용, shift 없음)
    df.loc[idx, "1HMSFast"] = calculate_1hmsfast_15m(
        df.loc[idx, "SMAF"], df.loc[idx, "SMA100"], df.loc[idx, "SMA200"]
    )
    
    # LS 열: -1 = (현재 2<1HMSF<3 AND 직전 1.5<1HMSF<=2) OR (현재 4<1HMSF<=5 AND 직전 5<1HMSF<6) / 1 = (5<현재<6 AND 직전 4~5) OR (1<현재<=2 AND 직전 2~3)
    if idx + 1 < len(df):
        h2 = df.loc[idx, "1HMSFast"]
        h3 = df.loc[idx + 1, "1HMSFast"]
        if pd.notna(h2) and pd.notna(h3):
            h2_f, h3_f = float(h2), float(h3)
            if (2 < h2_f < 3 and 1.5 < h3_f <= 2) or (4 < h2_f <= 5 and 5 < h3_f < 6):
                df.loc[idx, "LS"] = -1
            elif (5 < h2_f < 6 and 4 < h3_f <= 5) or (1 < h2_f <= 2 and 2 < h3_f < 3):
                df.loc[idx, "LS"] = 1
            else:
                df.loc[idx, "LS"] = np.nan
        else:
            df.loc[idx, "LS"] = np.nan
    else:
        df.loc[idx, "LS"] = np.nan
    
    # SPRD 계산: (max(sma25,sma100,sma200)-min(sma25,sma100,sma200))/min(sma25,sma100,sma200)
    sma25_val = df.loc[idx, "SMA25"]
    sma100_val = df.loc[idx, "SMA100"]
    sma200_val = df.loc[idx, "SMA200"]
    if not pd.isna(sma25_val) and not pd.isna(sma100_val) and not pd.isna(sma200_val):
        spread_min = min(sma25_val, sma100_val, sma200_val)
        if spread_min > 0:
            df.loc[idx, "SPRD"] = (max(sma25_val, sma100_val, sma200_val) - spread_min) / spread_min
        else:
            df.loc[idx, "SPRD"] = np.nan
    else:
        df.loc[idx, "SPRD"] = np.nan
    
    # SPRD2 계산: (max(저가,sma100,sma200)-min(저가,sma100,sma200))/min(저가,sma100,sma200)
    low_val = df.loc[idx, "저"]
    if not pd.isna(low_val) and not pd.isna(sma100_val) and not pd.isna(sma200_val):
        spread_min = min(low_val, sma100_val, sma200_val)
        if spread_min > 0:
            df.loc[idx, "SPRD2"] = (max(low_val, sma100_val, sma200_val) - spread_min) / spread_min
        else:
            df.loc[idx, "SPRD2"] = np.nan
    else:
        df.loc[idx, "SPRD2"] = np.nan
    
    return df

def calculate_latest_row_only_5m(df, market_type):
    """
    5분봉 After 단계 최적화: 최신 1개 행만 지표 계산 (previous 지표 유지)
    5분봉은 Max200/Min200 사용
    
    입력: [새 데이터(idx=0), Previous 1200개(idx=1~1200)] (최신→과거 순서)
    출력: [새 데이터(지표 계산됨), Previous 1200개(그대로)] (최신→과거 순서 유지)
    """
    if df.empty:
        return df
    
    # [Cursor 패치] 엑셀에서 읽은 데이터의 쉼표 제거 및 숫자/날짜 강제 변환
    # 이 과정이 없으면 데이터가 제대로 읽히지 않아 len(df)가 줄어들고 fallback으로 튕김
    df = clean_df_display_format(df)
    
    # 데이터가 조금이라도 있으면 최대한 latest_row_only 모드 유지
    # 200개 미만이어도 새 데이터만 계산 시도 (previous 데이터가 있으면 활용)
    if len(df) < 25:
        # 최소한 SMA25 계산을 위해 25개는 필요
        return calculate_all_indicators_5m(df, market_type)
    
    # ⚠️중요: 엑셀에서 읽은 데이터의 숫자 컬럼을 강제로 숫자로 변환 (쉼표 제거 후 변환)
    # 엑셀에서 불러온 숫자가 문자열("88,123.45")로 저장되어 있을 수 있음
    # 이 과정이 없으면 SMA 계산 시 문자열과 숫자가 섞여서 NaN이 발생합니다.
    numeric_cols = ['종', '시', '고', '저', 'Vol.']
    for col in numeric_cols:
        if col in df.columns:
            # 문자열인 경우에만 쉼표 제거 시도
            if df[col].dtype == 'object':
                df[col] = df[col].astype(str).str.replace(',', '')
            # 강제 숫자 변환
            df[col] = pd.to_numeric(df[col], errors='coerce')
    
    # Buy/Sell 컬럼이 없으면 생성
    if "Buy" not in df.columns:
        df["Buy"] = ""
    if "Sell" not in df.columns:
        df["Sell"] = ""
    
    # Previous 데이터의 Buy/Sell이 NaN이면 빈 문자열로 변환
    df["Buy"] = df["Buy"].fillna("")
    df["Sell"] = df["Sell"].fillna("")
    
    # 최신→과거 순서 유지, idx=0만 계산 (과거 데이터는 idx=1~199 사용)
    latest_idx = 0
    
    # idx=0 계산 가능 여부 확인 (Max200을 위해 최소 200개 필요)
    if len(df) >= 200:
        # SMA 계산 (idx 1~20 사용하여 idx 0 계산)
        df.loc[latest_idx, "SMA3"] = df.iloc[latest_idx+1:latest_idx+4]["종"].mean()
        df.loc[latest_idx, "SMA5"] = df.iloc[latest_idx+1:latest_idx+6]["종"].mean()
        df.loc[latest_idx, "SMA7"] = df.iloc[latest_idx+1:latest_idx+8]["종"].mean()
        df.loc[latest_idx, "SMA10"] = df.iloc[latest_idx+1:latest_idx+11]["종"].mean()
        df.loc[latest_idx, "SMA20"] = df.iloc[latest_idx+1:latest_idx+21]["종"].mean()
        
        # Max200, Min200 계산 (200개 캔들 동안의 시고저종에서 최고가와 최저가)
        window_data = df.iloc[latest_idx+1:latest_idx+201][["시", "고", "저", "종"]]
        df.loc[latest_idx, "Max200"] = window_data.values.max()
        df.loc[latest_idx, "Min200"] = window_data.values.min()
        
        # 하단, 상단 계산 (5분봉은 Max200/Min200)
        current_price = df.loc[latest_idx, "종"]
        min200 = df.loc[latest_idx, "Min200"]
        max200 = df.loc[latest_idx, "Max200"]
        
        if pd.notna(min200) and min200 != 0:
            df.loc[latest_idx, "하단"] = abs((current_price - min200) / min200)
        else:
            df.loc[latest_idx, "하단"] = np.nan
            
        if pd.notna(max200) and max200 != 0:
            df.loc[latest_idx, "상단"] = abs((current_price - max200) / max200)
        else:
            df.loc[latest_idx, "상단"] = np.nan
        
        # SFast, Fast, Base 계산
        sma3 = df.loc[latest_idx, "SMA3"]
        sma5 = df.loc[latest_idx, "SMA5"]
        sma7 = df.loc[latest_idx, "SMA7"]
        sma10 = df.loc[latest_idx, "SMA10"]
        sma20 = df.loc[latest_idx, "SMA20"]
        
        df.loc[latest_idx, "SFast"] = calculate_superfast(sma3, sma5, sma7)
        df.loc[latest_idx, "Fast"] = calculate_fast(sma5, sma7, sma10)
        df.loc[latest_idx, "Base"] = calculate_base(sma5, sma10, sma20)
        
        # 4or1 계산
        하단 = df.loc[latest_idx, "하단"]
        상단 = df.loc[latest_idx, "상단"]
        df.loc[latest_idx, "4or1"] = calculate_4or1(하단, 상단)
        
        # buyside, sellside 계산
        sfast = df.loc[latest_idx, "SFast"]
        fast = df.loc[latest_idx, "Fast"]
        base = df.loc[latest_idx, "Base"]
        
        df.loc[latest_idx, "buyside"] = calculate_buyside(sfast, fast, base)
        df.loc[latest_idx, "sellside"] = calculate_sellside(sfast, fast, base)
        
        # Buy, Sell 계산
        four_or_one = df.loc[latest_idx, "4or1"]
        sellside_val = df.loc[latest_idx, "sellside"]
        buyside_val = df.loc[latest_idx, "buyside"]
        
        df.loc[latest_idx, "Buy"] = calculate_buy(four_or_one, sellside_val)
        df.loc[latest_idx, "Sell"] = calculate_sell_short(buyside_val)
    
    # 순서 유지 (이미 최신→과거 순서)
    return df

def calculate_latest_3rows_only_5m(df, market_type):
    """
    5분봉 After 단계 최적화: 최신 3개 행(2-4행)만 지표 계산 (previous 지표 유지)
    5분봉은 Max200/Min200 사용
    
    입력: [새 데이터 3개(idx=0-2), Previous 1200개(idx=3~1202)] (최신→과거 순서)
    출력: [새 데이터 3개(지표 계산됨), Previous 1200개(그대로)] (최신→과거 순서 유지)
    """
    if df.empty or len(df) < 200:
        # 데이터가 부족하면 전체 계산
        return calculate_all_indicators_5m(df, market_type)
    
    # ⚠️중요: 엑셀에서 읽은 데이터의 숫자 컬럼을 강제로 숫자로 변환 (쉼표 제거 후 변환)
    # 엑셀에서 불러온 숫자가 문자열("88,123.45")로 저장되어 있을 수 있음
    # 이 과정이 없으면 SMA 계산 시 문자열과 숫자가 섞여서 NaN이 발생합니다.
    numeric_cols = ['종', '시', '고', '저', 'Vol.']
    for col in numeric_cols:
        if col in df.columns:
            # 문자열인 경우에만 쉼표 제거 시도
            if df[col].dtype == 'object':
                df[col] = df[col].astype(str).str.replace(',', '')
            # 강제 숫자 변환
            df[col] = pd.to_numeric(df[col], errors='coerce')
    
    # Buy/Sell 컬럼이 없으면 생성
    if "Buy" not in df.columns:
        df["Buy"] = ""
    if "Sell" not in df.columns:
        df["Sell"] = ""
    
    # Previous 데이터의 Buy/Sell이 NaN이면 빈 문자열로 변환
    df["Buy"] = df["Buy"].fillna("")
    df["Sell"] = df["Sell"].fillna("")
    
    # 최신→과거 순서 유지, idx=0, 1, 2만 계산
    # previous 데이터는 idx 3부터 시작
    # idx=0, 1, 2 각각 계산 (각 인덱스는 그 아래(과거)의 데이터를 사용)
    for latest_idx in [0, 1, 2]:
        if latest_idx >= len(df):
            break
        
        # idx 계산 가능 여부 확인 (Max200을 위해 최소 200개 필요)
        # 각 인덱스는 자신 포함하여 계산
        # idx=0: idx 0, 1, 2... 사용 (2행 + 3~201행)
        # idx=1: idx 1, 2, 3... 사용 (3행 + 4~202행)
        # idx=2: idx 2, 3, 4... 사용 (4행 + 5~203행)
        if len(df) >= latest_idx + 200:  # 해당 인덱스 포함하여 최소 200개 있어야 함
            # SMA 계산: 각 인덱스 포함하여 계산 (해당 행 + previous)
            df.loc[latest_idx, "SMA3"] = df.iloc[latest_idx:latest_idx+3]["종"].mean() if len(df) >= latest_idx+3 else np.nan
            df.loc[latest_idx, "SMA5"] = df.iloc[latest_idx:latest_idx+5]["종"].mean() if len(df) >= latest_idx+5 else np.nan
            df.loc[latest_idx, "SMA7"] = df.iloc[latest_idx:latest_idx+7]["종"].mean() if len(df) >= latest_idx+7 else np.nan
            df.loc[latest_idx, "SMA10"] = df.iloc[latest_idx:latest_idx+10]["종"].mean() if len(df) >= latest_idx+10 else np.nan
            df.loc[latest_idx, "SMA20"] = df.iloc[latest_idx:latest_idx+20]["종"].mean() if len(df) >= latest_idx+20 else np.nan
            
            # Max200, Min200 계산 (200개 캔들: 해당 행 포함 + previous)
            window_end = min(latest_idx + 200, len(df))
            if window_end > latest_idx:
                window_data = df.iloc[latest_idx:window_end][["시", "고", "저", "종"]]
                df.loc[latest_idx, "Max200"] = window_data.values.max()
                df.loc[latest_idx, "Min200"] = window_data.values.min()
            else:
                df.loc[latest_idx, "Max200"] = np.nan
                df.loc[latest_idx, "Min200"] = np.nan
            
            # 하단, 상단 계산 (5분봉은 Max200/Min200)
            current_price = df.loc[latest_idx, "종"]
            min200 = df.loc[latest_idx, "Min200"]
            max200 = df.loc[latest_idx, "Max200"]
            
            if pd.notna(min200) and min200 != 0:
                df.loc[latest_idx, "하단"] = abs((current_price - min200) / min200)
            else:
                df.loc[latest_idx, "하단"] = np.nan
                
            if pd.notna(max200) and max200 != 0:
                df.loc[latest_idx, "상단"] = abs((current_price - max200) / max200)
            else:
                df.loc[latest_idx, "상단"] = np.nan
            
            # SFast, Fast, Base 계산
            sma3 = df.loc[latest_idx, "SMA3"]
            sma5 = df.loc[latest_idx, "SMA5"]
            sma7 = df.loc[latest_idx, "SMA7"]
            sma10 = df.loc[latest_idx, "SMA10"]
            sma20 = df.loc[latest_idx, "SMA20"]
            
            df.loc[latest_idx, "SFast"] = calculate_superfast(sma3, sma5, sma7)
            df.loc[latest_idx, "Fast"] = calculate_fast(sma5, sma7, sma10)
            df.loc[latest_idx, "Base"] = calculate_base(sma5, sma10, sma20)
            
            # 4or1 계산
            하단 = df.loc[latest_idx, "하단"]
            상단 = df.loc[latest_idx, "상단"]
            df.loc[latest_idx, "4or1"] = calculate_4or1(하단, 상단)
            
            # buyside, sellside 계산
            sfast = df.loc[latest_idx, "SFast"]
            fast = df.loc[latest_idx, "Fast"]
            base = df.loc[latest_idx, "Base"]
            
            df.loc[latest_idx, "buyside"] = calculate_buyside(sfast, fast, base)
            df.loc[latest_idx, "sellside"] = calculate_sellside(sfast, fast, base)
            
            # Buy, Sell 계산
            four_or_one = df.loc[latest_idx, "4or1"]
            sellside_val = df.loc[latest_idx, "sellside"]
            buyside_val = df.loc[latest_idx, "buyside"]
            
            df.loc[latest_idx, "Buy"] = calculate_buy(four_or_one, sellside_val)
            df.loc[latest_idx, "Sell"] = calculate_sell_short(buyside_val)
    
    # 순서 유지 (이미 최신→과거 순서)
    return df

def calculate_latest_row_only_1h(df, market_type):
    """
    1시간봉 After 단계 최적화: 최신 1개 행만 지표 계산 (previous 지표 유지)
    
    입력: [새 데이터(idx=0), Previous(idx=1~)] (최신→과거 순서)
    출력: [새 데이터(지표 계산됨), Previous(그대로)] (최신→과거 순서 유지)
    
    계산 방식:
    - SMA: 25, 100, 200, 400, 800
    - SFast: SMA25, SMA100, SMA200
    - Fast: SMA25, SMA200, SMA400
    - Base: SMA25, SMA400, SMA800
    - Max200, Min200
    - 1HCLASS, -1HCLASS
    - 모든 지표는 2행(idx=0) + previous 데이터를 사용
    """
    if df.empty:
        return df
    
    # [Cursor 패치] 엑셀에서 읽은 데이터의 쉼표 제거 및 숫자/날짜 강제 변환
    # 이 과정이 없으면 데이터가 제대로 읽히지 않아 len(df)가 줄어들고 fallback으로 튕김
    df = clean_df_display_format(df)
    
    # Date(UTC) 컬럼 정규화 (Timestamp와 문자열 혼합 방지)
    if 'Date(UTC)' in df.columns:
        if df['Date(UTC)'].dtype == 'object':
            # format 명시하여 파싱 시도 (연도/일 혼동 방지)
            try:
                df['Date(UTC)'] = pd.to_datetime(df['Date(UTC)'], format='%y/%m/%d,%H:%M', errors='coerce')
            except:
                # 쉼표 제거 후 형식 시도
                try:
                    df['Date(UTC)'] = df['Date(UTC)'].astype(str).str.replace(',', ' ', regex=False).str.strip()
                    df['Date(UTC)'] = pd.to_datetime(df['Date(UTC)'], format='%y/%m/%d %H:%M', errors='coerce')
                except:
                    # fallback: format 없이 파싱
                    df['Date(UTC)'] = pd.to_datetime(df['Date(UTC)'], errors='coerce')
        else:
            # 이미 datetime 타입이면 그대로 사용
            pass
    
    # ⚠️중요: 입력 데이터를 확실하게 현재→과거 순서로 정렬 (SMA 계산 일관성 보장)
    # 데이터 수집 과정에서 정렬이 여러 번 섞일 수 있으므로, 계산 직전에 확실하게 정렬
    df = df.sort_values("Date(UTC)", ascending=False).reset_index(drop=True)
    
    # 데이터가 조금이라도 있으면 최대한 latest_row_only 모드 유지
    # 800개 미만이어도 새 데이터만 계산 시도 (previous 데이터가 있으면 활용)
    if len(df) < 25:
        # 최소한 SMA25 계산을 위해 25개는 필요
        return calculate_all_indicators_1h(df, market_type)
    
    # ⚠️중요: 엑셀에서 읽은 데이터의 숫자 컬럼을 강제로 숫자로 변환 (쉼표 제거 후 변환)
    # 엑셀에서 불러온 숫자가 문자열("88,123.45")로 저장되어 있을 수 있음
    # 이 과정이 없으면 SMA 계산 시 문자열과 숫자가 섞여서 NaN이 발생합니다.
    numeric_cols = ['종', '시', '고', '저', 'Vol.']
    for col in numeric_cols:
        if col in df.columns:
            # 문자열인 경우에만 쉼표 제거 시도
            if df[col].dtype == 'object':
                df[col] = df[col].astype(str).str.replace(',', '')
            # 강제 숫자 변환
            df[col] = pd.to_numeric(df[col], errors='coerce')
    
    # ⚠️중요: 숫자 지표 컬럼을 명시적으로 숫자 타입으로 변환 (previous 데이터의 지표가 문자열로 읽혔을 수 있음)
    num_cols = ["SMA25", "SMA100", "SMA200", "SMA400", "SMA800", "Max200", "Min200", "하단", "상단", "SFast", "Fast", "Base", "1HMSFast", "4or1", "buyside", "sellside", "1HCLASS", "-1HCLASS", "p1H"]
    for col in num_cols:
        if col in df.columns:
            # 문자열인 경우에만 쉼표 제거 시도
            if df[col].dtype == 'object':
                df[col] = df[col].astype(str).str.replace(',', '')
            df[col] = pd.to_numeric(df[col], errors='coerce')
    if 'p1H' not in df.columns:
        df['p1H'] = np.nan
    
    # Buy/Sell 컬럼이 없으면 생성
    if "Buy" not in df.columns:
        df["Buy"] = ""
    if "Sell" not in df.columns:
        df["Sell"] = ""
    
    # Previous 데이터의 Buy/Sell이 NaN이면 빈 문자열로 변환
    df["Buy"] = df["Buy"].fillna("")
    df["Sell"] = df["Sell"].fillna("")
    
    # ⚠️중요: 새 데이터(지표가 NaN인 행)를 모두 계산
    # previous 데이터는 이미 지표가 계산되어 있으므로 건드리지 않음
    # 새 데이터는 SMA800이 NaN이므로, NaN인 행들을 찾아서 계산
    new_data_indices = []
    if "SMA800" in df.columns:
        new_data_indices = df[df["SMA800"].isna()].index.tolist()
    elif "SMA25" in df.columns:
        # SMA800이 없으면 SMA25로 판단
        new_data_indices = df[df["SMA25"].isna()].index.tolist()
    else:
        # 지표 컬럼이 없으면 최신 1개만 계산 (하위 호환성)
        new_data_indices = [0]
    
    # 새 데이터가 없으면 그대로 반환
    if not new_data_indices:
        return df
    
    # 새 데이터의 각 행에 대해 지표 계산
    for idx in new_data_indices:
        # SMA 계산: idx 포함하여 계산
        df.loc[idx, "SMA25"] = df.iloc[idx:idx+25]["종"].mean() if idx + 25 <= len(df) else np.nan
        df.loc[idx, "SMA100"] = df.iloc[idx:idx+100]["종"].mean() if idx + 100 <= len(df) else np.nan
        df.loc[idx, "SMA200"] = df.iloc[idx:idx+200]["종"].mean() if idx + 200 <= len(df) else np.nan
        df.loc[idx, "SMA400"] = df.iloc[idx:idx+400]["종"].mean() if idx + 400 <= len(df) else np.nan
        df.loc[idx, "SMA800"] = df.iloc[idx:idx+800]["종"].mean() if idx + 800 <= len(df) else np.nan
    
        # Max200, Min200 계산: idx 포함 200개 캔들
        if idx + 200 <= len(df):
            window_data = df.iloc[idx:idx+200][["시", "고", "저", "종"]]
            df.loc[idx, "Max200"] = window_data.values.max()
            df.loc[idx, "Min200"] = window_data.values.min()
        else:
            df.loc[idx, "Max200"] = np.nan
            df.loc[idx, "Min200"] = np.nan
    
    # 하단, 상단 계산
    current_price = df.loc[idx, "종"]
    min200 = df.loc[idx, "Min200"]
    max200 = df.loc[idx, "Max200"]
    
    if pd.notna(min200) and min200 != 0:
        df.loc[idx, "하단"] = abs((current_price - min200) / min200)
    else:
        df.loc[idx, "하단"] = np.nan
        
    if pd.notna(max200) and max200 != 0:
        df.loc[idx, "상단"] = abs((current_price - max200) / max200)
    else:
        df.loc[idx, "상단"] = np.nan
    
    # SFast, Fast, Base 계산
    df.loc[idx, "SFast"] = calculate_superfast(df.loc[idx, "SMA25"], df.loc[idx, "SMA100"], df.loc[idx, "SMA200"])
    df.loc[idx, "Fast"] = calculate_fast(df.loc[idx, "SMA25"], df.loc[idx, "SMA200"], df.loc[idx, "SMA400"])
    df.loc[idx, "Base"] = calculate_base(df.loc[idx, "SMA25"], df.loc[idx, "SMA400"], df.loc[idx, "SMA800"])
    
    # 4or1 계산
    df.loc[idx, "4or1"] = calculate_4or1(df.loc[idx, "하단"], df.loc[idx, "상단"])
    
    # buyside, sellside 계산
    df.loc[idx, "buyside"] = calculate_buyside(df.loc[idx, "SFast"], df.loc[idx, "Fast"], df.loc[idx, "Base"])
    df.loc[idx, "sellside"] = calculate_sellside(df.loc[idx, "SFast"], df.loc[idx, "Fast"], df.loc[idx, "Base"])
    
    # Buy, Sell 계산
    df.loc[idx, "Buy"] = calculate_buy(df.loc[idx, "4or1"], df.loc[idx, "sellside"])
    df.loc[idx, "Sell"] = calculate_sell_short(df.loc[idx, "buyside"])
    
    # 1HCLASS 계산: 1H 캔들에서 SFast/Fast/Base가 and(2 <= 값 < 5)이면 각각 +1, 총합(0~3)
    sfast_val = df.loc[idx, "SFast"]
    fast_val = df.loc[idx, "Fast"]
    base_val = df.loc[idx, "Base"]
    df.loc[idx, "1HCLASS"] = (
        (1 if pd.notna(sfast_val) and 2 <= sfast_val < 5 else 0) +
        (1 if pd.notna(fast_val) and 2 <= fast_val < 5 else 0) +
        (1 if pd.notna(base_val) and 2 <= base_val < 5 else 0)
    )
    
    # -1HCLASS 계산: 1H 캔들에서 SFast/Fast/Base가 or(값 < 2, 값 >= 7)이면 각각 -1, 총합(0~-3)
    df.loc[idx, "-1HCLASS"] = (
            (-1 if pd.notna(sfast_val) and (sfast_val < 2 or sfast_val >= 7) else 0) +
            (-1 if pd.notna(fast_val) and (fast_val < 2 or fast_val >= 7) else 0) +
            (-1 if pd.notna(base_val) and (base_val < 2 or base_val >= 7) else 0)
        )
    
    # p1H 계산: SFast/Fast/Base 각각 4 <= 값 < 5 인 경우 1로 카운트, 총합(0~3)
    df.loc[idx, "p1H"] = (
        (1 if pd.notna(sfast_val) and 4 <= sfast_val < 5 else 0) +
        (1 if pd.notna(fast_val) and 4 <= fast_val < 5 else 0) +
        (1 if pd.notna(base_val) and 4 <= base_val < 5 else 0)
    )
    
    # 1HMSFast 계산: idx의 종가, SMA25, SMA100 사용 (shift 없음)
    df.loc[idx, "1HMSFast"] = calculate_1hmsfast(
        df.loc[idx, "종"],
        df.loc[idx, "SMA25"],
        df.loc[idx, "SMA100"]
    )
    
    return df

def calculate_latest_row_only_1d(df, market_type):
    """
    일봉 After 단계 최적화: 최신 1개 행만 지표 계산 (previous 지표 유지)
    일봉은 Max15/Min15 사용
    
    입력: [새 데이터(idx=0), Previous(idx=1~)] (최신→과거 순서)
    출력: [새 데이터(지표 계산됨), Previous(그대로)] (최신→과거 순서 유지)
    
    계산 방식:
    - SMA10: idx 0~9 (2행 + 3~11행) 총 10개를 더해서 10으로 나눔
    - 모든 지표는 2행(idx=0) + previous 데이터를 사용
    """
    if df.empty:
        return df
    
    # [Cursor 패치] 엑셀에서 읽은 데이터의 쉼표 제거 및 숫자/날짜 강제 변환
    # 이 과정이 없으면 데이터가 제대로 읽히지 않아 len(df)가 줄어들고 fallback으로 튕김
    df = clean_df_display_format(df)
    
    # 데이터가 조금이라도 있으면 최대한 latest_row_only 모드 유지
    # 20개 미만이어도 새 데이터만 계산 시도 (previous 데이터가 있으면 활용)
    if len(df) < 3:
        # 최소한 SMA3 계산을 위해 3개는 필요
        return calculate_all_indicators(df, market_type)
    
    # Buy/Sell 컬럼이 없으면 생성
    if "Buy" not in df.columns:
        df["Buy"] = ""
    if "Sell" not in df.columns:
        df["Sell"] = ""
    
    # Previous 데이터의 Buy/Sell이 NaN이면 빈 문자열로 변환
    df["Buy"] = df["Buy"].fillna("")
    df["Sell"] = df["Sell"].fillna("")
    
    # idx=0만 계산
    idx = 0
    
    # SMA 계산: 2행(idx=0) 포함하여 계산
    df.loc[idx, "SMA3"] = df.iloc[idx:idx+3]["종"].mean()
    df.loc[idx, "SMA5"] = df.iloc[idx:idx+5]["종"].mean()
    df.loc[idx, "SMA7"] = df.iloc[idx:idx+7]["종"].mean()
    df.loc[idx, "SMA10"] = df.iloc[idx:idx+10]["종"].mean()
    df.loc[idx, "SMA20"] = df.iloc[idx:idx+20]["종"].mean()
    
    # Max15, Min15 계산: 2행 포함 15개 캔들 (idx 0~14)
    window_data = df.iloc[idx:idx+15][["시", "고", "저", "종"]]
    df.loc[idx, "Max15"] = window_data.values.max()
    df.loc[idx, "Min15"] = window_data.values.min()
    
    # 하단, 상단 계산
    current_price = df.loc[idx, "종"]
    min15 = df.loc[idx, "Min15"]
    max15 = df.loc[idx, "Max15"]
    
    if pd.notna(min15) and min15 != 0:
        df.loc[idx, "하단"] = abs((current_price - min15) / min15)
    else:
        df.loc[idx, "하단"] = np.nan
        
    if pd.notna(max15) and max15 != 0:
        df.loc[idx, "상단"] = abs((current_price - max15) / max15)
    else:
        df.loc[idx, "상단"] = np.nan
    
    # SFast, Fast, Base 계산
    df.loc[idx, "SFast"] = calculate_superfast(df.loc[idx, "SMA3"], df.loc[idx, "SMA5"], df.loc[idx, "SMA7"])
    df.loc[idx, "Fast"] = calculate_fast(df.loc[idx, "SMA5"], df.loc[idx, "SMA7"], df.loc[idx, "SMA10"])
    df.loc[idx, "Base"] = calculate_base(df.loc[idx, "SMA5"], df.loc[idx, "SMA10"], df.loc[idx, "SMA20"])
    
    # 4or1 계산
    df.loc[idx, "4or1"] = calculate_4or1(df.loc[idx, "하단"], df.loc[idx, "상단"])
    
    # buyside, sellside 계산
    df.loc[idx, "buyside"] = calculate_buyside(df.loc[idx, "SFast"], df.loc[idx, "Fast"], df.loc[idx, "Base"])
    df.loc[idx, "sellside"] = calculate_sellside(df.loc[idx, "SFast"], df.loc[idx, "Fast"], df.loc[idx, "Base"])
    
    # Buy, Sell 계산
    df.loc[idx, "Buy"] = calculate_buy(df.loc[idx, "4or1"], df.loc[idx, "sellside"])
    df.loc[idx, "Sell"] = calculate_sell(df.loc[idx, "4or1"], df.loc[idx, "buyside"])  # 일봉은 calculate_sell 사용
    
    # Samount1D, Bamount1D 계산
    if pd.notna(df.loc[idx, "buyside"]):
        df.loc[idx, "Samount1D"] = (1 - df.loc[idx, "buyside"]) * TRADING_UNIT
    if pd.notna(df.loc[idx, "sellside"]):
        df.loc[idx, "Bamount1D"] = (1 - df.loc[idx, "sellside"]) * TRADING_UNIT
    
    return df

def calculate_latest_row_only_dateM(df_15m: pd.DataFrame) -> pd.DataFrame:
    """
    15분봉 데이터의 2행(idx=0)만 dateM을 계산합니다.
    3행 이후(idx=1~)는 previous 데이터 유지.
    
    Args:
        df_15m: 15분봉 DataFrame (Date(UTC) 컬럼 포함, UTC 기준)
    
    Returns:
        DataFrame: dateM 컬럼이 계산된 15분봉 DataFrame
    
    Note:
        - Date(UTC) 컬럼 기준으로 계산 (UTC 기준)
    """
    df = df_15m.copy()
    
    # dateM 열이 없으면 초기화
    if 'dateM' not in df.columns:
        df['dateM'] = 0
    
    # 2행(idx=0)만 계산
    i = 0
    if len(df) > 0:
        # 현재 캔들부터 200개 캔들 범위에서 Max 값 찾기
        end_idx = min(i + 200, len(df))
        max_value = df.iloc[i:end_idx]['고'].max()
        
        # 현재 시점부터 미래로 순회하면서 Max 값과 같은 고가를 가진 캔들을 찾기
        max_date = None
        for j in range(i, end_idx):  # 200개 캔들 범위 내에서만 검색
            if df.iloc[j]['고'] == max_value:
                max_date = df.iloc[j]['Date(UTC)']
                break
        
        if max_date is not None:
            # 캔들 개수로 계산 (j - i + 1) - 현재부터 Max 날짜까지의 총 캔들 개수
            candle_count = j - i + 1
            
            # 음수가 되지 않도록 처리
            df.iloc[i, df.columns.get_loc('dateM')] = max(0, candle_count)
        else:
            # Max 값과 같은 가격을 가진 캔들이 없는 경우 0
            df.iloc[i, df.columns.get_loc('dateM')] = 0
    
    return df

def calculate_latest_row_only_LD(df_15m: pd.DataFrame) -> pd.DataFrame:
    """
    15분봉 데이터의 2행(idx=0)만 LD를 계산합니다.
    3행 이후(idx=1~)는 previous 데이터 유지.
    
    Args:
        df_15m: 15분봉 DataFrame (Date(UTC) 컬럼 포함, UTC 기준)
    
    Returns:
        DataFrame: LD 컬럼이 계산된 15분봉 DataFrame
    
    Note:
        - Date(UTC) 컬럼 기준으로 계산 (UTC 기준)
    """
    import math
    
    df = df_15m.copy()
    
    # LD 열이 없으면 초기화
    if 'LD' not in df.columns:
        df['LD'] = 0.0
    
    # 2행(idx=0)만 계산
    i = 0
    if len(df) > 0:
        # 종, 시, 고, 저, sma3, sma5, sma7, sma10, sma20 값들
        values = [
            df.iloc[i]['종'],
            df.iloc[i]['시'], 
            df.iloc[i]['고'],
            df.iloc[i]['저'],
            df.iloc[i]['SMA3'],
            df.iloc[i]['SMA5'],
            df.iloc[i]['SMA7'],
            df.iloc[i]['SMA10'],
            df.iloc[i]['SMA20']
        ]
        
        # Spread2 계산
        max_val = max(values)
        min_val = min(values)
        spread2 = (max_val - min_val) / min_val * 1000
        
        # dateM 계산 (200개 캔들 기준)
        end_idx = min(i + 200, len(df))
        max_value = df.iloc[i:end_idx]['고'].max()
        
        # 현재 시점부터 미래로 순회하면서 Max 값과 같은 고가를 가진 캔들을 찾기
        dateM = 0
        for j in range(i, end_idx):  # 200개 캔들 범위 내에서만 검색
            if df.iloc[i:end_idx].iloc[j-i]['고'] == max_value:
                dateM = j - i + 1  # 캔들 개수 계산
                break
        
        # Spread1 계산 (Python에서 200개 캔들로 상단+하단 계산)
        # 200개 캔들 범위에서 Max200, Min200 계산
        max200 = df.iloc[i:end_idx][["시", "고", "저", "종"]].max().max()
        min200 = df.iloc[i:end_idx][["시", "고", "저", "종"]].min().min()
        
        # 현재 가격
        current_price = df.iloc[i]['종']
        
        # 상단, 하단 계산 (200개 캔들 기준)
        if min200 != 0:
            하단 = abs((current_price - min200) / min200)
        else:
            하단 = 0
            
        if max200 != 0:
            상단 = abs((current_price - max200) / max200)
        else:
            상단 = 0
        
        spread1 = (상단 + 하단) * 1000
        
        # 이차방정식 t^2 + (dateM)t - Spread2*Spread1 = 0 의 계수
        a = 1
        b = dateM
        c = -spread2 * spread1
        
        # 판별식 계산 (항상 양수)
        discriminant = b * b - 4 * a * c
        
        # 양의 근 계산 (t1이 항상 양수)
        t1 = (-b + math.sqrt(discriminant)) / (2 * a)
        
        # 양의 근을 원래 값으로 저장 (셀 서식에서 소수점 3자리 표시)
        df.iloc[i, df.columns.get_loc('LD')] = t1
    
    return df

def calculate_latest_row_only_ksc(df_15m: pd.DataFrame) -> pd.DataFrame:
    """
    15분봉 After 단계 최적화: 최신 1개 행(idx=0)만 KSC 계산 (previous 지표 유지)
    
    입력: [새 데이터(idx=0), Previous(idx=1~)] (최신→과거 순서)
    출력: [새 데이터(KSC 계산됨), Previous(그대로)] (최신→과거 순서 유지)
    """
    if df_15m.empty or len(df_15m) < 1:
        return df_15m
    
    # 티커별 SPRD2 열의 최근 400개 평균의 30%를 threshold로 계산
    sprd2_threshold = None  # SPRD2가 없거나 데이터가 부족한 경우 None
    if 'SPRD2' in df_15m.columns:
        # Date(UTC) 기준으로 정렬
        sort_col = 'Date(UTC)'
        # 최신→과거 순서로 정렬된 상태에서 최근 400개 추출
        df_sorted_desc = df_15m.sort_values(sort_col, ascending=False).reset_index(drop=True)
        sprd2_recent = df_sorted_desc['SPRD2'].head(400)
        sprd2_valid = sprd2_recent[pd.notna(sprd2_recent)]
        if len(sprd2_valid) > 0:
            sprd2_avg = sprd2_valid.mean()
            sprd2_threshold = sprd2_avg * 0.3  # 평균의 30%
    
    # 2행(인덱스 0)의 데이터
    row = df_15m.iloc[0]
    order = str(row.get('ORDER', '')).strip()
    hmsfast = row.get('1HMSFast', np.nan)
    sma25 = row.get('SMA25', np.nan)
    sma100 = row.get('SMA100', np.nan)
    sma200 = row.get('SMA200', np.nan)
    
    # 3행(인덱스 1)의 KSC 값을 이전 값으로 사용 (KSC는 숫자만 저장)
    prev_kill_count = 0
    if len(df_15m) > 1:
        prev_ksc = df_15m.iloc[1].get('KSC', 0)
        # KSC는 숫자만 저장하므로 숫자로 읽기
        if isinstance(prev_ksc, (int, float)):
            prev_kill_count = int(prev_ksc)
        else:
            try:
                prev_kill_count = int(float(prev_ksc))
            except:
                prev_kill_count = 0
    
    # spread 계산
    spread = np.nan
    if not pd.isna(sma25) and not pd.isna(sma100) and not pd.isna(sma200):
        sma_values = [float(sma25), float(sma100), float(sma200)]
        sma_max = max(sma_values)
        sma_min = min(sma_values)
        if sma_min > 0:
            spread = (sma_max - sma_min) / sma_min
    
    ksc_value = prev_kill_count  # 기본값은 이전 카운트 (숫자만)
    bomb_value = ""  # Bomb이 아닌 경우 빈 문자열
    bomb_count = 0  # Bomb이 아닌 경우 0 (BombCount 열용)
    ksc_stack = 0
    
    # ORDER가 Sell5 또는 Sell10이면 초기화
    if order in ['Sell5', 'Sell10']:
        ksc_value = 0
        bomb_value = ""
        bomb_count = 0
        df_15m = df_15m.copy()
        df_15m.loc[0, 'KSC'] = ksc_value
        if 'Bomb' not in df_15m.columns:
            df_15m['Bomb'] = ""
            df_15m['Bomb'] = df_15m['Bomb'].astype('object')
        df_15m.loc[0, 'Bomb'] = bomb_value
        if 'BombCount' not in df_15m.columns:
            df_15m['BombCount'] = 0
        df_15m.loc[0, 'BombCount'] = bomb_count
        if 'KSC stack' not in df_15m.columns:
            df_15m['KSC stack'] = 0
        df_15m.loc[0, 'KSC stack'] = ksc_stack
        return df_15m
    
    # ORDER가 Buy5 또는 Buy10인 경우
    if order in ['Buy5', 'Buy10']:
        if not pd.isna(hmsfast):
            hmsfast_val = float(hmsfast)
            # 기본 조건: 2 <= 1HMSFast < 7이고 spread >= 티커별 SPRD2 400개 평균의 30%일 때만 KSC 카운트 스택 관련 로직 적용
            if 2.0 <= hmsfast_val < 7.0 and not pd.isna(spread) and sprd2_threshold is not None and spread >= sprd2_threshold:
                # Bomb 처리: 4.4 < 1HMSFast < 4.6 (새로운 수열 규칙: Bomb 발생해도 카운트 +1, 스택 유지)
                if 4.4 < hmsfast_val < 4.6:
                    # 새로운 수열 규칙: Bomb 발생해도 카운트 +1, 스택 쌓는게 유지됨
                    # 단일 행 계산 함수이므로, Bomb 발생 시에도 카운트를 증가시켜 표시
                    # (실제 스택은 전체 계산 함수에서만 유지됨)
                    prev_kill_count = prev_kill_count + 1
                    # KSC는 숫자만 저장, Bomb 열에 "Bomb" 저장
                    ksc_value = prev_kill_count  # 증가된 값을 KSC에 저장
                    bomb_value = "Bomb"  # Bomb 열에 "Bomb" 저장
                    bomb_count = prev_kill_count  # Bomb이 되는 시점의 카운트 값 저장 (BombCount 열용)
                    ksc_stack = prev_kill_count  # Bomb이 되는 시점의 스택 카운트 값 저장 (증가된 값)
                # 스택 쌓임: 2 <= 1HMSFast <= 4.4 또는 4.6 <= 1HMSFast < 7
                elif (2.0 <= hmsfast_val <= 4.4) or (4.6 <= hmsfast_val < 7.0):
                    # kill 카운트 증가 (제한 없음)
                    prev_kill_count = prev_kill_count + 1
                    ksc_value = prev_kill_count
                else:
                    ksc_value = prev_kill_count
            # spread < 티커별 SPRD2 400개 평균의 30% 또는 1HMSFast < 2 또는 1HMSFast >= 7이면 카운트 스택 로직 적용 안 함 → 0으로 초기화
            elif hmsfast_val >= 7.0:
                # 스택 쌓이다가 bomb 신호 없이 7 초과하면 0으로 스택 초기화
                ksc_value = 0
            else:
                # 1HMSFast < 2 또는 spread < 티커별 SPRD2 400개 평균의 30%인 경우 0으로 초기화
                ksc_value = 0
        else:
            # 1HMSFast가 NaN이면 0으로 초기화
            ksc_value = 0
    else:
        # ORDER가 Buy5/Buy10이 아니면
        if not pd.isna(hmsfast):
            hmsfast_val = float(hmsfast)
            # 2 <= 1HMSFast < 7이고 spread >= 티커별 SPRD2 400개 평균의 30%이면 이전 값 유지
            if 2.0 <= hmsfast_val < 7.0 and not pd.isna(spread) and sprd2_threshold is not None and spread >= sprd2_threshold:
                # 이전 값 유지 (스택 쌓지 않음)
                ksc_value = prev_kill_count
            # spread < 티커별 SPRD2 400개 평균의 30% 또는 1HMSFast < 2 또는 1HMSFast >= 7이면 0으로 초기화
            elif hmsfast_val >= 7.0 or hmsfast_val < 2.0:
                # 7 초과 또는 2 미만이면 스택 초기화
                ksc_value = 0
            else:
                # spread < 티커별 SPRD2 400개 평균의 30%인 경우 0으로 초기화
                ksc_value = 0
        else:
            # 1HMSFast가 NaN이면 0으로 초기화
            ksc_value = 0
    
    # 2행(인덱스 0)에만 KSC, Bomb, BombCount, KSC stack 값 설정
    df_15m = df_15m.copy()
    df_15m.loc[0, 'KSC'] = ksc_value
    if 'Bomb' not in df_15m.columns:
        df_15m['Bomb'] = ""
        df_15m['Bomb'] = df_15m['Bomb'].astype('object')
    else:
        df_15m['Bomb'] = df_15m['Bomb'].astype('object')
    df_15m.loc[0, 'Bomb'] = bomb_value
    if 'BombCount' not in df_15m.columns:
        df_15m['BombCount'] = 0
    df_15m.loc[0, 'BombCount'] = bomb_count
    if 'KSC stack' not in df_15m.columns:
        df_15m['KSC stack'] = 0
    df_15m.loc[0, 'KSC stack'] = ksc_stack
    
    return df_15m

def calculate_latest_row_only_prft(df_15m: pd.DataFrame) -> pd.DataFrame:
    """
    15분봉 After 단계 최적화: 최신 1개 행(idx=0)만 PRFT 계산 (previous 지표 유지)
    
    입력: [새 데이터(idx=0), Previous(idx=1~)] (최신→과거 순서)
    출력: [새 데이터(PRFT 계산됨), Previous(그대로)] (최신→과거 순서 유지)
    
    계산 로직:
    - PRFT = 스택값 (숫자) - 항상 표시 (조건 없음)
    
    TPOVER (Target Price Over) 로직:
    - Sell5 또는 Sell10이면서 종가가 TP 열 값 이상인 경우
    - 종가 >= TP 인 경우 PRFT = 'TPOVER' (스택 초기화)
    """
    if df_15m.empty or len(df_15m) < 1:
        return df_15m
    
    df_15m_copy = df_15m.copy()
    
    # 현재 행(idx=0)의 데이터
    row = df_15m_copy.iloc[0]
    order = str(row.get('ORDER', '')).strip()
    order_lower = order.lower()  # 대소문자 무시 비교
    close = row.get('종', np.nan)
    stossp = row.get('StoSP', np.nan)
    stosu = row.get('StoSU', np.nan)
    tp_value = row.get('TP', np.nan)
    
    # 이전 행(idx=1)의 데이터 (previous에서 가져옴)
    prev_prft_stack = 0
    prev_stosp = np.nan
    prev_stosu = np.nan
    
    if len(df_15m_copy) > 1:
        prev_row = df_15m_copy.iloc[1]
        prev_prft = prev_row.get('PRFT', 0)
    
        # 이전 행의 PRFT 값이 숫자면 스택으로 사용
        if isinstance(prev_prft, (int, float)) and pd.notna(prev_prft):
            prev_prft_stack = int(prev_prft)
        elif isinstance(prev_prft, str) and prev_prft.strip() == 'TPOVER':
            # TPOVER면 스택은 0
            prev_prft_stack = 0
        else:
            prev_prft_stack = 0
        
        # 이전 행의 StoSP, StoSU 값
        prev_stosp = prev_row.get('StoSP', np.nan)
        prev_stosu = prev_row.get('StoSU', np.nan)
    
    # PRFT 스택 초기값: 이전 행의 스택 값
    prft_stack = prev_prft_stack
    
    # TPOVER 조건 확인 (최우선, 스택 초기화) - 대소문자 무시
    if order_lower in ['sell5', 'sell10']:
        if not pd.isna(close) and not pd.isna(tp_value) and tp_value > 0:
            # 종가 >= TP 인 경우 TPOVER
            if float(close) >= float(tp_value):
                prft_value = 'TPOVER'
                df_15m_copy.loc[0, 'PRFT'] = prft_value
                return df_15m_copy
    
    # StoSP/StoSU 변화 확인 (스택 관리)
    # StoSP/StoSU가 유효한 값인지 확인 (NaN이 아니고 0이 아니면 유효)
    stossp_valid = not pd.isna(stossp) and (isinstance(stossp, (int, float)) and stossp != 0)
    stosu_valid = not pd.isna(stosu) and (isinstance(stosu, (int, float)) and stosu != 0)
    
    if stossp_valid and stosu_valid:
        # StoSP/StoSU가 초기화되지 않은 경우 (유효한 값)
        if not pd.isna(prev_stosp) and not pd.isna(prev_stosu):
            # 이전 값이 있고, StoSP 또는 StoSU가 증가했으면 스택 +1
            if stossp > prev_stosp or stosu > prev_stosu:
                prft_stack += 1
            # StoSP와 StoSU가 모두 동일하거나 감소한 경우는 스택 유지 (변경 없음)
    
    # PRFT = 스택값 (항상 표시, 조건 없음)
    prft_value = prft_stack
    
    # 2행(인덱스 0)에만 PRFT 값 설정
    df_15m_copy.loc[0, 'PRFT'] = prft_value
    
    return df_15m_copy

def calculate_sb5m_for_15m(df_15m, df_5m):
    """
    5분봉 데이터를 기반으로 15분봉에 SB5M 신호를 추가합니다.
    5분봉 3개씩 그룹화해서 Buy/Sell 개수를 세어 판정합니다.
    날짜 기준은 UTC로 처리합니다 (바이낸스 기준).
    """
    if df_5m.empty or df_15m.empty:
        return df_15m
    
    # 원본 보존
    df_5m = df_5m.copy()
    df_15m = df_15m.copy()
    
    # --- Date(UTC) 기준 그룹화 (바이낸스는 UTC 기준) ---
    # 5분봉: 이미 datetime64면 그대로 사용, 문자열만 파싱
    if pd.api.types.is_datetime64_any_dtype(df_5m['Date(UTC)']):
        df_5m['Date(UTC)_dt'] = df_5m['Date(UTC)']
    else:
        # 문자열인 경우에만 파싱 (UserWarning 억제)
        s = df_5m['Date(UTC)'].astype(str).str.strip().str.replace(',', ' ', regex=False)
        import warnings
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", UserWarning)
            df_5m['Date(UTC)_dt'] = pd.to_datetime(s, errors='coerce')
    df_5m = df_5m[df_5m['Date(UTC)_dt'].notna()].copy()
    
    if df_5m.empty:
        if 'SB5M' not in df_15m.columns:
            df_15m['SB5M'] = ''
        return df_15m
    
    # 시간 정렬 (과거 → 현재)
    df_5m = df_5m.sort_values('Date(UTC)_dt', ascending=True).reset_index(drop=True)
    
    # 5분봉: 15분 그룹 생성
    df_5m['15min_group'] = df_5m['Date(UTC)_dt'].dt.floor('15min')
    
    # 15분봉: Date(UTC) 파싱 (이미 datetime64면 그대로 사용, 문자열만 파싱)
    if pd.api.types.is_datetime64_any_dtype(df_15m['Date(UTC)']):
        df_15m['Date(UTC)_dt'] = df_15m['Date(UTC)']
    else:
        # 문자열인 경우에만 파싱 (UserWarning 억제)
        s = df_15m['Date(UTC)'].astype(str).str.strip().str.replace(',', ' ', regex=False)
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", UserWarning)
            df_15m['Date(UTC)_dt'] = pd.to_datetime(s, errors='coerce')
    
    # 15분봉: 15min_group 없으면 생성, 있으면 NaN만 채움
    if '15min_group' not in df_15m.columns:
        df_15m['15min_group'] = df_15m['Date(UTC)_dt'].dt.floor('15min')
    else:
        mask_na = df_15m['15min_group'].isna()
        if mask_na.any():
            df_15m.loc[mask_na, '15min_group'] = df_15m.loc[mask_na, 'Date(UTC)_dt'].dt.floor('15min')
    
    # --- 15분 그룹별 SB5M 신호 계산 ---
    sb5m_data = []
    for group_time, group_df in df_5m.groupby('15min_group', sort=True):
        buy_count = 0
        sell_count = 0
        
        for _, row in group_df.iterrows():
            buy_val = row.get('Buy', '')
            sell_val = row.get('Sell', '')
            
            if pd.notna(buy_val) and isinstance(buy_val, str) and buy_val.strip().lower() == 'buy':
                buy_count += 1
            elif pd.notna(sell_val) and isinstance(sell_val, str) and sell_val.strip().lower() == 'sell':
                sell_count += 1
        
        if buy_count == 3:
            sb5m_signal = 'buy3'
        elif buy_count == 2:
            sb5m_signal = 'buy2'
        elif buy_count == 1:
            sb5m_signal = 'buy1'
        elif sell_count == 3:
            sb5m_signal = 'sell3'
        elif sell_count == 2:
            sb5m_signal = 'sell2'
        elif sell_count == 1:
            sb5m_signal = 'sell1'
        else:
            sb5m_signal = ''
        
        sb5m_data.append({
            '15min_group': group_time,
            'SB5M': sb5m_signal
        })
    
    # --- 여기부터가 핵심 변경: merge 제거, map 사용 ---
    if sb5m_data:
        sb5m_df = pd.DataFrame(sb5m_data)
        
        # 타입 통일 (datetime64[ns])
        try:
            if sb5m_df['15min_group'].dtype != 'datetime64[ns]':
                sb5m_df['15min_group'] = pd.to_datetime(sb5m_df['15min_group'], errors='coerce')
            if df_15m['15min_group'].dtype != 'datetime64[ns]':
                df_15m['15min_group'] = pd.to_datetime(df_15m['15min_group'], errors='coerce')
        except Exception:
            pass
        
        # 15min_group → SB5M 매핑 딕셔너리
        sb5m_map = dict(zip(sb5m_df['15min_group'], sb5m_df['SB5M']))
        
        # 기존 SB5M은 무시하고 새로 덮어씀
        df_15m['SB5M'] = df_15m['15min_group'].map(sb5m_map)
    else:
        # SB5M 데이터가 하나도 없을 때
        if 'SB5M' not in df_15m.columns:
            df_15m['SB5M'] = ''
    
    # 최종: 빈 값은 ''로 통일
    df_15m['SB5M'] = df_15m['SB5M'].fillna('')
    
    # 임시 컬럼 제거
    if '15min_group' in df_15m.columns:
        df_15m = df_15m.drop('15min_group', axis=1)
    
    return df_15m

# (김프 계산 제거)

# (15분봉 김프 계산 제거)

def calculate_order_column(df_15m: pd.DataFrame, sheet_name: str = "") -> pd.DataFrame:
    """
    15분봉 시트에 ORDER 열을 계산합니다.
    새로운 매매 신호 로직:
    
    매도 신호 (Sell) - Gear 분리 적용:
    
    판정 순서:
    1. 공통 규칙 (gear1/gear2 공통) - 최우선
    2. gear1 기존 로직
    3. gear3 새로운 규칙 (OR 조건)
    4. gear2 새로운 규칙
    
    Gear 분리 기준:
    - Gear1: 2 <= 1HMSFast < 7
    - Gear2: 1HMSFast >= 7 또는 1HMSFast < 2
    - Gear3: 2 <= 1HMSFast < 7
    
    1. 공통 규칙 (gear1/gear2 공통):
    - 15분 Sell + SB1H Sell → Sell5/Sell10
      * SB1D = "sell" → Sell10 (10000원 매도)
      * SB1D ≠ "sell" → Sell5 (5000원 매도)
    
    2. gear1 기존 로직 (2 <= 1HMSFast < 7):
    - 15분 Sell + SB5M sell3: 1 unit 매도 (Sell5) - SB1H/SB1D 무관
    
    3. gear3 새로운 규칙 (OR 조건, 2 <= 1HMSFast < 7):
    - SB5M sell3 + SB1M sell15: 1 unit 매도 (Sell5) - 15분 Sell 무관
    
    4. gear2 새로운 규칙 (1HMSFast >= 7 또는 < 2):
    - SB5M sell3: 1 unit 매도 (Sell5) - 15분 Sell, buyside, TP, SB1H, SB1D 무관
    
    매수 신호 (Buy) - Gear 분리 적용:
    
    Gear 분리 기준:
    - Gear1: 2 <= 1HMSFast < 7
    - Gear2: 1HMSFast >= 7 또는 1HMSFast < 2
    - Gear3: 2 <= 1HMSFast < 2.2
    
    공통 필수 조건:
    - 15분 Buy = "buy" (4or1 < 4 AND sellside <= 0.05)
    - SB5M = "buy1", "buy2", "buy3" 중 하나
    
    Gear3 로직 (OR 조건):
    - SB5M buy3 + 2 <= 1HMSFast < 2.2 → Buy5 (15분 Buy 조건 불필요)
    
    Gear1 로직:
    - SB5M buy2/buy3: LD 게이트 무관
      * SB1H Buy + SB1D Buy → Buy10
      * 그 외 → Buy5
    - SB5M buy1: LD 게이트 적용
      * SB1H Buy + SB1D Buy → Buy10 (LD 무관)
      * SB1H Buy + SB1D Sell + LD ≤ 0.333 → Buy5
      * SB1H Buy + SB1D 없음 + LD ≤ 0.333 → Buy5
    
    Gear2 로직 (OR 조건):
    - 조건 1: 기존 조건 (Gear1과 동일) - 우선 적용, Buy10 가능
      * 필수: 15분 Buy = "buy" AND SB5M in ["buy1", "buy2", "buy3"]
      * 결과: SB1H, SB1D, LD 값에 따라 Buy5/Buy10 결정
    - 조건 2: SB5M buy1/buy2/buy3 + SB1M이 sell이 아닐 것 - 조건 1 불만족 시 적용, 항상 Buy5만 생성
      * 필수: SB5M in ["buy1", "buy2", "buy3"] AND SB1M not in ["sell01", "sell02", ..., "sell15"]
      * 결과: 항상 Buy5 생성 (15분 Buy 조건 불필요)
    
    LD 게이트 제외 조건:
    - SB5M buy2/buy3: LD 값과 무관
    - Gear2에서 조건 2로 생성된 Buy5: LD 값과 무관
    
    Args:
        df_15m: 15분봉 DataFrame (Date(UTC) 컬럼 포함, UTC 기준)
        sheet_name: 시트 이름 (사용하지 않음, 호환성 유지)
    
    Returns:
        DataFrame: ORDER 컬럼이 추가된 15분봉 DataFrame
    
    Note:
        - 시간 직접 사용 없음, 신호 기반 계산만 수행
    """
    if df_15m.empty:
        return df_15m
    
    df_15m_copy = df_15m.copy()
    
    # [수정] TP 참조를 위해 직전 행의 TP를 현재 행으로 당겨옴 (최신순 정렬 기준)
    # 현재 행(0)에서 이전 행(1)의 TP를 확인하기 위함
    if 'TP' in df_15m_copy.columns:
        df_15m_copy['prev_TP_val'] = df_15m_copy['TP'].shift(-1)
    else:
        df_15m_copy['prev_TP_val'] = np.nan
    
    # LD 값을 숫자로 표준화 (문자열, %, 콤마 제거)
    if 'LD' in df_15m_copy.columns:
        s = df_15m_copy['LD'].astype(str).str.replace('%','', regex=False).str.replace(',','', regex=False)
        df_15m_copy['LD_num'] = pd.to_numeric(s, errors='coerce')
    else:
        df_15m_copy['LD_num'] = 0.0
    
    # ORDER 열 계산: 새로운 매매 신호 로직 (nan 방지 처리 포함)
    def get_order_signal(row):
        try:
            # 안전한 값 추출 (nan, None, 빈 문자열 처리)
            sell_signal = str(row.get('Sell', '')).strip().lower()
            buy_signal = str(row.get('Buy', '')).strip().lower()
            sb1h_signal = str(row.get('SB1H', '')).strip().lower()
            sb1d_signal = str(row.get('SB1D', '')).strip().lower()
            sb5m_signal = str(row.get('SB5M', '')).strip().lower()
            sb1m_signal = str(row.get('SB1M', '')).strip().lower()
            
            # nan, None, 빈 값 처리
            if sell_signal in ['nan', 'none', '']:
                sell_signal = ''
            if buy_signal in ['nan', 'none', '']:
                buy_signal = ''
            if sb1h_signal in ['nan', 'none', '']:
                sb1h_signal = ''
            if sb1d_signal in ['nan', 'none', '']:
                sb1d_signal = ''
            if sb5m_signal in ['nan', 'none', '']:
                sb5m_signal = ''
            if sb1m_signal in ['nan', 'none', '']:
                sb1m_signal = ''
        
            # 매도 신호 (Sell) - Gear 분리 적용
            # ⚠️ 중요: 매도 신호를 매수 신호보다 우선 처리 (충돌 시 매도 우선)
            # 판정 순서: 공통 규칙 → gear1 → gear2
            
            # 1HMSFast 값 추출 (Gear 판정용)
            try:
                hmsfast_val = float(row.get('1HMSFast', np.nan))
            except (TypeError, ValueError):
                hmsfast_val = np.nan
            
            # Gear 분리
            is_gear1 = (not pd.isna(hmsfast_val) and hmsfast_val >= 2.0 and hmsfast_val < 7.0)
            is_gear2 = (not pd.isna(hmsfast_val) and (hmsfast_val >= 7.0 or hmsfast_val < 2.0))
            
            # ========== 1. 공통 규칙 (gear1/gear2 공통) - 최우선 ==========
            # 15분 Sell + SB1H Sell → Sell5/Sell10
            if sell_signal == 'sell' and sb1h_signal == 'sell':
                if sb1d_signal == 'sell':
                    return 'Sell10'  # 공통 규칙: 15분 Sell + SB1H Sell + SB1D Sell → Sell10 (10000원 매도)
                else:
                    return 'Sell5'   # 공통 규칙: 15분 Sell + SB1H Sell → Sell5 (5000원 매도)
            
            # ========== 2. gear1 기존 로직 (2 <= 1HMSFast < 7) ==========
            if is_gear1 and sell_signal == 'sell' and sb5m_signal == 'sell3':
                return 'Sell5'   # gear1: 15분 Sell + SB5M sell3 → Sell5 (SB1H/SB1D 무관)
            
            # ========== 3. gear3 새로운 규칙 (OR 조건, 2 <= 1HMSFast < 7) ==========
            # SB5M sell3 + SB1M sell15 → Sell5 (15분 Sell 조건 없이도 발생)
            if is_gear1 and sb5m_signal == 'sell3' and sb1m_signal == 'sell15':
                return 'Sell5'   # gear3: SB5M sell3 + SB1M sell15 → Sell5 (15분 Sell 무관)
            
            # ========== 4. gear2 새로운 규칙 (1HMSFast >= 7 또는 < 2) ==========
            # SB5M sell3 → Sell5 (15분 Sell 조건 없이도 발생)
            if is_gear2 and sb5m_signal == 'sell3':
                return 'Sell5'   # gear2: SB5M sell3 → Sell5 (15분 Sell, buyside, TP, SB1H, SB1D 무관)
            
            # [매도 우선 원칙] 매도 신호(Sell = "sell")가 있으면 매수 로직 실행 안함
            if sell_signal == 'sell':
                return ''  # 매도 조건은 만족하지 않았지만 매도 신호가 있으므로 매수하지 않음
            
            # 매수 신호 (Buy) - Gear 분리 적용
            # [수정] SB5M 신호가 있으면 먼저 진입 (15분 Buy 조건과 분리)
            if sb5m_signal in ['buy1', 'buy2', 'buy3']:
                # 1HMSFast, sellside 값 추출
                try:
                    hmsfast_val = float(row.get('1HMSFast', np.nan))
                    sellside_val = float(row.get('sellside', np.nan))
                except (TypeError, ValueError):
                    hmsfast_val = np.nan
                    sellside_val = np.nan
                
                # Gear 분리
                is_gear1 = (not pd.isna(hmsfast_val) and hmsfast_val >= 2.0 and hmsfast_val < 7.0)
                is_gear2 = (not pd.isna(hmsfast_val) and (hmsfast_val >= 7.0 or hmsfast_val < 2.0))
                is_gear3 = (not pd.isna(hmsfast_val) and hmsfast_val >= 2.0 and hmsfast_val < 2.2)
                
                # Gear3 로직 (2 <= 1HMSFast < 2.2) - OR 조건, 15분 Buy 불필요
                if is_gear3 and sb5m_signal == 'buy3':
                    return 'Buy5'   # gear3: SB5M buy3 + 2 <= 1HMSFast < 2.2 → Buy5 (15분 Buy 조건 불필요)
                
                # Gear1 로직 (2 <= 1HMSFast < 7) - 15분 Buy 필요
                if is_gear1:
                    # Gear1은 15분 Buy 신호가 있어야 함
                    if buy_signal != 'buy':
                        return ''
                    
                    # 추가 조건: 15분 Buy + SB5M buy1/buy2/buy3 + SB1M buy10 이상 → Buy5
                    if sb5m_signal in ['buy1', 'buy2', 'buy3']:
                        sb1m_buy10_list = ['buy10', 'buy11', 'buy12', 'buy13', 'buy14', 'buy15']
                        if sb1m_signal in sb1m_buy10_list:
                            return 'Buy5'    # 15분 Buy + SB5M buy1/buy2/buy3 + SB1M buy10 이상 → Buy5
                    
                    # SB5M buy2/buy3인 경우 - LD 게이트 무관
                    if sb5m_signal in ['buy2', 'buy3']:
                        if sb1h_signal == 'buy' and sb1d_signal == 'buy':
                            return 'Buy10'   # 15분 Buy + SB1H Buy + SB1D Buy + SB5M buy2/buy3: 1 unit 매수
                        else:
                            return 'Buy5'    # 15분 Buy + SB5M buy2/buy3: 1 unit 매수 (LD 무관)
                    else:
                        # SB5M buy1인 경우 - SB1H Buy 필요, LD 게이트 적용
                        if sb1h_signal == 'buy':
                            if sb1d_signal == 'sell':
                                # SB1D Sell이 있으면 LD 값에 따라 판단
                                ld_value = row.get('LD_num', 0)
                                if pd.isna(ld_value):
                                    ld_value = 0
                                if ld_value <= 0.333:
                                    return 'Buy5'    # LD <= 0.333이면 1 unit 매수
                                else:
                                    return ''        # LD > 0.333이면 매수하지 않음
                            elif sb1d_signal == 'buy':
                                return 'Buy10'   # 15분 Buy + SB1H Buy + SB1D Buy + SB5M buy1: 1 unit 매수
                            else:
                                # SB1D 없음 - LD 게이트 적용
                                ld_value = row.get('LD_num', 0)
                                if pd.isna(ld_value):
                                    ld_value = 0
                                if ld_value <= 0.333:
                                    return 'Buy5'    # LD <= 0.333이면 1 unit 매수
                                else:
                                    return ''        # LD > 0.333이면 매수하지 않음
                        else:
                            return ''        # SB5M buy1이면 SB1H Buy 필요
                
                # Gear2 로직 (1HMSFast >= 7 또는 < 2) - OR 조건
                # ⚠️ 참고: 매도 신호가 우선 처리되므로, 매도 신호가 없을 때만 이 조건들이 체크됨
                elif is_gear2:
                    # 조건 1: 기존 조건 (Gear1과 동일) - 우선 적용, 15분 Buy 필요
                    # 필수: 15분 Buy = "buy" AND SB5M in ["buy1", "buy2", "buy3"]
                    if buy_signal == 'buy':
                        gear1_result = None
                        if sb5m_signal in ['buy2', 'buy3']:
                            if sb1h_signal == 'buy' and sb1d_signal == 'buy':
                                gear1_result = 'Buy10'
                            else:
                                gear1_result = 'Buy5'
                        else:
                            # SB5M buy1인 경우
                            if sb1h_signal == 'buy':
                                if sb1d_signal == 'sell':
                                    ld_value = row.get('LD_num', 0)
                                    if pd.isna(ld_value):
                                        ld_value = 0
                                    if ld_value <= 0.333:
                                        gear1_result = 'Buy5'
                                elif sb1d_signal == 'buy':
                                    gear1_result = 'Buy10'
                                else:
                                    # SB1D 없음
                                    ld_value = row.get('LD_num', 0)
                                    if pd.isna(ld_value):
                                        ld_value = 0
                                    if ld_value <= 0.333:
                                        gear1_result = 'Buy5'
                        
                        # 조건 1이 만족되면 반환
                        if gear1_result:
                            return gear1_result
                        # 조건 1이 만족되지 않으면 조건 2로 넘어감 (15분 Buy가 있어도 조건 1 미충족 시 조건 2 적용)
                    
                    # 조건 2: SB5M buy1/buy2/buy3 + SB1M이 sell이 아닐 것 - 15분 Buy 조건 불필요, 항상 Buy5 생성
                    # ⚠️ 주의: 매도 신호(Sell10/Sell5)가 우선 처리되므로, 매도 신호가 없을 때만 실행됨
                    # 조건 1이 만족되지 않았을 때 실행됨 (15분 Buy가 있든 없든 상관없이 SB5M + SB1M 조건으로 Buy5 생성)
                    if sb5m_signal in ['buy1', 'buy2', 'buy3']:
                        # SB1M이 sell01~sell15가 아닐 때만 Buy5 생성
                        sb1m_sell_list = ['sell01', 'sell02', 'sell03', 'sell04', 'sell05', 
                                         'sell06', 'sell07', 'sell08', 'sell09', 'sell10',
                                         'sell11', 'sell12', 'sell13', 'sell14', 'sell15']
                        if sb1m_signal not in sb1m_sell_list:
                            return 'Buy5'
                    
                    return ''
                
                # 기타 (1HMSFast NaN 또는 범위 밖) - Gear1과 동일한 로직 적용 (15분 Buy 필요)
                else:
                    # 기타 경우도 15분 Buy 신호가 있어야 함
                    if buy_signal != 'buy':
                        return ''
                    
                    # SB5M buy2/buy3인 경우 - LD 게이트 무관
                    if sb5m_signal in ['buy2', 'buy3']:
                        if sb1h_signal == 'buy' and sb1d_signal == 'buy':
                            return 'Buy10'   # 15분 Buy + SB1H Buy + SB1D Buy + SB5M buy2/buy3: 1 unit 매수
                        else:
                            return 'Buy5'    # 15분 Buy + SB5M buy2/buy3: 1 unit 매수 (LD 무관)
                    else:
                        # SB5M buy1인 경우 - SB1H Buy 필요, LD 게이트 적용
                        if sb1h_signal == 'buy':
                            if sb1d_signal == 'sell':
                                # SB1D Sell이 있으면 LD 값에 따라 판단
                                ld_value = row.get('LD_num', 0)
                                if pd.isna(ld_value):
                                    ld_value = 0
                                if ld_value <= 0.333:
                                    return 'Buy5'    # LD <= 0.333이면 1 unit 매수
                                else:
                                    return ''        # LD > 0.333이면 매수하지 않음
                            elif sb1d_signal == 'buy':
                                return 'Buy10'   # 15분 Buy + SB1H Buy + SB1D Buy + SB5M buy1: 1 unit 매수
                            else:
                                # SB1D 없음 - LD 게이트 적용
                                ld_value = row.get('LD_num', 0)
                                if pd.isna(ld_value):
                                    ld_value = 0
                                if ld_value <= 0.333:
                                    return 'Buy5'    # LD <= 0.333이면 1 unit 매수
                                else:
                                    return ''        # LD > 0.333이면 매수하지 않음
                        else:
                            return ''        # SB5M buy1이면 SB1H Buy 필요
            else:
                return ''  # 신호 없음
        except Exception as e:
            # 예외 발생 시 빈 문자열 반환 (nan 방지)
            return ''
    
    # ORDER 열 추가 (nan 방지: 빈 문자열로 초기화)
    df_15m_copy['ORDER'] = df_15m_copy.apply(get_order_signal, axis=1)
    # nan 값 처리: nan이면 빈 문자열로 변환
    df_15m_copy['ORDER'] = df_15m_copy['ORDER'].fillna('')
    df_15m_copy['ORDER'] = df_15m_copy['ORDER'].astype(str).replace('nan', '').replace('None', '')
    
    # 사후 검증: Buy5 LD 게이트 위반 체크
    # LD 게이트 제외 조건:
    # 1. SB5M buy2/buy3: LD 값과 무관
    # 2. Gear2에서 조건 2로 생성된 Buy5: LD 값과 무관 (SB5M buy1/buy2/buy3 모두 포함)
    # (Gear2 판정은 1HMSFast >= 7 또는 < 2)
    try:
        # Gear2 판정을 위한 1HMSFast 값 확인
        df_15m_copy['_is_gear2'] = False
        if '1HMSFast' in df_15m_copy.columns:
            hmsfast_vals = pd.to_numeric(df_15m_copy['1HMSFast'], errors='coerce')
            df_15m_copy['_is_gear2'] = (
                (hmsfast_vals >= 7.0) | (hmsfast_vals < 2.0)
            ) & (~pd.isna(hmsfast_vals))
        
        # LD 게이트 제외 조건: SB5M buy2/buy3 및 Gear2에서 조건 2로 생성된 Buy5
        # SB5M buy2/buy3는 항상 LD 게이트 제외
        # Gear2에서 조건 2로 생성된 Buy5도 LD 게이트 제외 (SB5M buy1/buy2/buy3 모두 포함)
        sb5m_buy2_buy3 = df_15m_copy['SB5M'].isin(['buy2', 'buy3'])
        # Gear2 조건 2: 15분 Buy가 없거나 조건 1이 만족되지 않았을 때 SB5M만으로 Buy5 생성
        # 이 경우는 ORDER가 Buy5이고 Gear2이고 SB5M이 buy1/buy2/buy3인 경우
        gear2_condition2 = (
            (df_15m_copy['ORDER'] == 'Buy5') & 
            df_15m_copy['_is_gear2'] & 
            df_15m_copy['SB5M'].isin(['buy1', 'buy2', 'buy3'])
        )
        ld_gate_excluded = sb5m_buy2_buy3 | gear2_condition2
        
        # LD 게이트 위반 체크 (제외 조건 제외)
        bad = df_15m_copy[
            (df_15m_copy['ORDER'] == 'Buy5') & 
            (df_15m_copy['LD_num'] > 0.333) &
            (~ld_gate_excluded)  # LD 게이트 제외 조건 제외
        ]
        
        if not bad.empty:
            print('❌ Buy5 LD gate violated rows (SB5M buy2/buy3 제외):')
            print(bad[['Buy','SB1H','SB1D','SB5M','LD','LD_num','ORDER', '1HMSFast']].head(10))
        
        # 임시 컬럼 제거
        df_15m_copy = df_15m_copy.drop(['_is_gear2'], axis=1, errors='ignore')
    except Exception as e:
        # 검증 실패 시 경고만 출력하고 계속 진행
        print(f'⚠️ Buy5 LD gate validation error: {e}')
    
    # 임시 컬럼 제거 (prev_TP_val)
    if 'prev_TP_val' in df_15m_copy.columns:
        df_15m_copy = df_15m_copy.drop('prev_TP_val', axis=1)
    
    return df_15m_copy

# (김프 복사 제거)

def _force_date_text(df):
    """엑셀 쓰기 직전에 Date(UTC), KST 컬럼을 강제로 문자열로 변환
    
    Args:
        df: DataFrame (Date(UTC), KST 컬럼 포함)
    
    Returns:
        DataFrame: Date(UTC), KST 컬럼이 문자열로 변환된 DataFrame
    
    Note:
        - Date(UTC) 컬럼은 UTC 기준 시간을 문자열로 변환 (계산에 사용)
        - KST 컬럼은 참고용으로만 표시 (계산 로직에는 사용하지 않음)
        - 엑셀에서 날짜형으로 자동 변환되지 않도록 문자열로 변환
    """
    import numpy as np
    import re
    
    FMT_OUT = "%y/%m/%d,%H:%M"  # UTC 기준 시간 포맷
    
    def normalize_col(s: pd.Series) -> pd.Series:
        # 1) 이미 datetime이면 바로 포맷 (UTC 기준)
        if np.issubdtype(s.dtype, np.datetime64):
            return pd.to_datetime(s, utc=True).dt.strftime(FMT_OUT)  # UTC 기준으로 명시

        # 2) 문자열이면 케이스별 명시 포맷 적용
        v = s.astype(str)

        # 결과 버퍼(초기값: 원문 유지)
        out = v.copy()

        # a) YYYY-MM-DD HH:MM:SS (UTC 기준으로 파싱)
        m1 = v.str.match(r"^\d{4}-\d{2}-\d{2}[ T]\d{2}:\d{2}:\d{2}$")
        if m1.any():
            out.loc[m1] = pd.to_datetime(
                v.loc[m1].str.replace("T", " ", regex=False),
                format="%Y-%m-%d %H:%M:%S",
                utc=True  # UTC 기준으로 명시
            ).dt.strftime(FMT_OUT)

        # b) ISO8601(마이크로초/타임존 꼬리 포함) → 앞 19자리만 사용 (UTC 기준으로 파싱)
        m2 = v.str.match(r"^\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}.*$")
        if m2.any():
            out.loc[m2] = pd.to_datetime(
                v.loc[m2].str.slice(0, 19),
                format="%Y-%m-%dT%H:%M:%S",
                utc=True  # UTC 기준으로 명시
            ).dt.strftime(FMT_OUT)

        # c) 이미 목표 포맷이면 그대로 둠 (YY/MM/DD,HH:MM)
        #    패턴: 25/09/12,13:45
        # (아무 처리 안 함)

        return out

    for col in ("Date(UTC)", "KST"):
        if col in df.columns:
            df[col] = normalize_col(df[col])
    return df
# (김프 복사 제거)


# -------------------- 메인 --------------------
def main(polling_start_time=None, skip_first_row=False, pre_fetched_data=None):
    """
    메인 함수: 바이낸스 데이터 수집, 지표 계산, Excel 저장을 수행합니다.
    
    Args:
        polling_start_time: 실행 시작 시간 (UTC 기준, 스케줄러에서 전달)
        skip_first_row: True면 1단계(previous 파일 생성), False면 2단계(after 파일 생성 및 주문 전송)
        pre_fetched_data: 2단계 실행 시 미리 수집한 데이터 (선택적, 성능 최적화용)
    
    Returns:
        DataFrame: 15분봉 DataFrame (2단계에서 previous 파일이 없으면 None 반환)
    
    Note:
        - 모든 시간 처리는 UTC 기준으로 수행
        - 1단계: 전체 데이터 수집 및 지표 계산, previous 파일 생성
        - 2단계: 최신 데이터만 수집, previous 파일과 병합, 지표 계산, 주문 실행, after 파일 생성
    """
    # 시작 시간 기록
    start_time = time.time()
    # PREVIOUS/AFTER 표시용 prefix
    stage_prefix = "PREVIOUS" if skip_first_row else "AFTER"
    print(f"{get_timestamp()} [{stage_prefix}] 🚀 스크립트 시작" + (" (1단계: previous 파일 생성)" if skip_first_row else " (2단계: after 파일 생성)"))
    
    # 메모리 누수 방지를 위한 가비지 컬렉션
    gc.collect()
    
    # 설정
    binance_symbol_ticker = f"{TICKER}USDT"
    
    # 캔들 개수 계산 (Y = 400 기준)
    CANDLE_COUNT = calculate_candle_count(Y)
    COLLECTION_COUNT = calculate_collection_count(Y)
    
    # 현재 시간 (UTC 기준)
    current_time_utc = dt.datetime.now(tz.UTC)
    current_minute = current_time_utc.minute
    current_hour = current_time_utc.hour
    current_date = current_time_utc.date()
    
    # ⚠️중요: 모든 API 호출이 동일한 시점을 바라보도록 기준 시간 고정
    # 초와 마이크로초를 0으로 만들어 "정각" 기준으로 설정 (예: 05:00:05 -> 05:00:00)
    # "직전 완성된 봉"까지만 가져오기 위해 1밀리초를 뺌 (04:59:59.999)
    sync_time = current_time_utc.replace(second=0, microsecond=0)
    fixed_fixed_end_time_ms = int(sync_time.timestamp() * 1000) - 1
    
    print(f"{get_timestamp()} [{stage_prefix}] 🕐 API 기준 시간 고정: {sync_time.strftime('%Y-%m-%d %H:%M:%S')} UTC")
    
    # 1단계(previous)와 2단계(after)에 따라 다른 개수 사용
    if skip_first_row:
        # 1단계(previous 파일): 수집 개수 사용 (미완성 1개 제거 후 최종 개수로 제한)
        # Previous 파일 생성 시 처리 과정:
        # - 15분봉: 수집 1601개(801+800) → 미완성 1개 제거 → 800개 제거 → 800개 남김
        # - 5분봉: 수집 2601개(2401+200) → 미완성 1개 제거 → 200개 제거 → 2400개 남김
        minute5_count = COLLECTION_COUNT['5m']   # 5분봉: 2601개
        minute15_count = COLLECTION_COUNT['15m']  # 15분봉: 1601개
    else:
        # 2단계(after 파일): 최신 수집 → 미완성 1개 제거
        # 5분봉: 4개 수집 → 1개 제거 → 3개 사용 (15분 구간 채우기: 0~5분, 5~10분, 10~15분)
        # 15분봉: 2개 수집 → 1개 제거 → 1개 사용
        minute5_count = 4   # 5분봉 최신 4개 (미완성 1개 제거 후 3개)
        minute15_count = 2  # 15분봉 최신 2개 (미완성 1개 제거 후 1개)
    
    # 1시간봉, 일봉, 주봉은 수집 개수 사용
    hour1_count = COLLECTION_COUNT['1h']  # 1시간봉: roundup((Y+200)/4/200) × 200 = 200개
    daily_count = COLLECTION_COUNT['1d']  # 일봉: roundup((Y+200)/4/24/200) × 200 = 200개
    weekly_count = COLLECTION_COUNT['1w']  # 주봉: roundup((Y+200)/4/24/7/200) × 200 = 200개
    
    # 1분봉 수집 개수 (병렬 수집에서 사용)
    minute1_count = COLLECTION_COUNT['1m'] if skip_first_row else 16  # 1단계: 12400개, 2단계: 16개
    
    include_today = True  # 오늘 진행중 캔들 포함 (현재 시점 시고저종 필요)

    # 저장 경로: 스크립트 폴더의 cryptodaily15min 하위폴더
    script_dir = os.path.dirname(os.path.abspath(__file__))
    base_save_dir = os.path.join(script_dir, "cryptodaily15min")
    
    # 티커별 폴더 매핑 (바이낸스)
    ticker_folder_mapping = {
        "BTC": "F BINANCE 1BTC",
        "ETH": "F BINANCE 2ETH", 
        "XRP": "F BINANCE 3XRP",
        "SOL": "F BINANCE 4SOL",
        "BNB": "F BINANCE 5BNB"
    }
    
    # 티커별 폴더명 가져오기
    ticker_folder = ticker_folder_mapping.get(TICKER, f"F BINANCE {TICKER}")
    save_dir = os.path.join(base_save_dir, ticker_folder)
    
    # cryptodaily15min 폴더가 없으면 생성
    if not os.path.exists(base_save_dir):
        os.makedirs(base_save_dir)
        print(f"{get_timestamp()} 📁 cryptodaily15min 폴더 생성됨: {base_save_dir}")
    
    # 티커별 폴더가 없으면 생성
    if not os.path.exists(save_dir):
        os.makedirs(save_dir)
        print(f"{get_timestamp()} 📁 {ticker_folder} 폴더 생성됨: {save_dir}")
    
    # 현재 시간으로 타임스탬프 생성 (UTC 기준, 연월일시분초)
    timestamp = dt.datetime.now(tz.UTC).strftime("%Y%m%d_%H%M%S")
    
    # 데이터 개수를 동적으로 표시 (10K, 1M 단위)
    def format_count(count):
        if count >= 1000000:
            return f"{count//1000000}M"
        elif count >= 10000:
            k_count = count // 1000
            return f"{k_count}K"
        else:
            return str(count)
    
    daily_count_formatted = format_count(daily_count)
    minute5_count_formatted = format_count(minute5_count)
    minute15_count_formatted = format_count(minute15_count)
    hour1_count_formatted = format_count(hour1_count)
    weekly_count_formatted = format_count(weekly_count)
    
    # 1단계 실행 시 파일명에 previous 추가, 2단계 실행 시 after 추가
    if skip_first_row:
        prefix = "previous_"
    else:
        prefix = "after_"
    filename = f"{prefix}F_{TICKER}_BINANCE_DAILY{daily_count_formatted}_5MIN{minute5_count_formatted}_15MIN{minute15_count_formatted}_1H{hour1_count_formatted}_WEEKLY{weekly_count_formatted}_{timestamp}.xlsx"
    save_path = os.path.join(save_dir, filename)

    # 2단계 실행 시 previous 파일과 합치기
    df_prev_1m = pd.DataFrame()
    df_prev_5m = pd.DataFrame()
    df_prev_15m = pd.DataFrame()
    df_prev_1h = pd.DataFrame()
    df_prev_1h4x = pd.DataFrame()
    df_prev_1d = pd.DataFrame()
    df_prev_weekly = pd.DataFrame()  # ⚠️중요: 초기화 필수 (2단계에서 주봉 시트가 없을 수 있음)
    if not skip_first_row:
        # 2단계(After 생성) 실행 시 이전 파일 읽기
        # 실행 흐름:
        # 1. 최초: 1단계(previous 생성) → previous 파일만 존재
        # 2. 그 다음: 2단계(previous 읽어서 after 생성) → previous가 최신이므로 previous 읽음
        # 3. 그 다음부터: 2단계(after 읽어서 after 생성) → after가 최신이므로 after 읽음
        # 따라서 previous와 after 모두 검색하여 최신 파일을 선택
        candidates = []
        if os.path.exists(save_dir):
            for f in os.listdir(save_dir):
                # 파일명 필터링: previous_... 또는 after_... 로 시작하는 엑셀 파일
                if (f.startswith(f"previous_F_{TICKER}_BINANCE_") or f.startswith(f"after_F_{TICKER}_BINANCE_")) \
                   and f.endswith('.xlsx') and not f.startswith('~$'):
                    file_path = os.path.join(save_dir, f)
                    try:
                        mtime = os.path.getmtime(file_path)
                        candidates.append((mtime, file_path))
                    except OSError:
                        continue
        
        if candidates:
            # 수정 시간 역순 정렬 (가장 최근 파일이 0번)
            candidates.sort(reverse=True)
            latest_previous_file = candidates[0][1]  # 변수명은 previous지만 실제로는 after일 수도 있음
            print(f"{get_timestamp()} [{stage_prefix}] 📂 이어쓰기 대상 파일 로드: {os.path.basename(latest_previous_file)}")
            
            # 파일이 완전히 저장되고 읽을 수 있을 때까지 대기
            if not wait_for_file_ready(latest_previous_file, max_wait_seconds=5):
                print(f"{get_timestamp()} [{stage_prefix}] ❌파일 준비 대기 시간 초과: {os.path.basename(latest_previous_file)}")
                return None  # 파일 준비 실패 시 중단 (데이터 끊김 방지)
            
            # 이전 파일(previous 또는 after)에서 5분봉, 15분봉, 1시간봉, 일봉 데이터 읽기
            read_success = False
            for read_attempt in range(5):  # 5회 재시도
                try:
                    if read_attempt > 0:
                        time.sleep(0.5)
                    import openpyxl
                    wb_prev = openpyxl.load_workbook(latest_previous_file, data_only=True)
                    read_success = True
                    break
                except Exception as e:
                    if read_attempt < 4:
                        continue
                    print(f"{get_timestamp()} [{stage_prefix}] ❌파일 읽기 실패: {e}")
                    return None  # 읽기 실패 시 중단 (데이터 끊김 방지)
            
            if read_success:
                try:
                    # 5분봉 시트 읽기
                    if f"{TICKER}USDT5M" in wb_prev.sheetnames:
                        df_prev_5m = pd.read_excel(latest_previous_file, sheet_name=f"{TICKER}USDT5M", header=0)
                        # [Source 방식] 엑셀 읽은 직후 데이터 세척 (쉼표 제거, 숫자/날짜 강제 변환)
                        df_prev_5m = clean_df_display_format(df_prev_5m, sheet_type='5m')
                        # KST는 엑셀 저장 직전에만 문자열로 변환 (지금은 유지)
                        if 'KST' in df_prev_5m.columns:
                            df_prev_5m['KST'] = df_prev_5m['KST'].astype(str)
                            df_prev_5m['KST'] = df_prev_5m['KST'].str.replace('nan', '').str.replace('NaT', '')
                        # Date(UTC)는 Timestamp로 유지하여 정렬 (데이터 파괴 방지)
                        if 'Date(UTC)' in df_prev_5m.columns:
                            df_prev_5m = df_prev_5m.sort_values('Date(UTC)', ascending=False, na_position='last').reset_index(drop=True)
                        print(f"{get_timestamp()} [{stage_prefix}] 📖 5분봉 {len(df_prev_5m)}개 읽기 완료")
                    
                    # 15분봉 시트 읽기
                    if f"{TICKER}USDT15M" in wb_prev.sheetnames:
                        df_prev_15m = pd.read_excel(latest_previous_file, sheet_name=f"{TICKER}USDT15M", header=0)
                        # [Source 방식] 엑셀 읽은 직후 데이터 세척 (쉼표 제거, 숫자/날짜 강제 변환)
                        df_prev_15m = clean_df_display_format(df_prev_15m, sheet_type='15m')
                        # 중요 컬럼 타입 변환
                        if 'Bomb' in df_prev_15m.columns:
                            df_prev_15m['Bomb'] = df_prev_15m['Bomb'].astype('object')
                        if '1HCLASS' in df_prev_15m.columns:  # 구버전 호환
                            df_prev_15m = df_prev_15m.rename(columns={'1HCLASS': '1HCL'})
                        if '-1HCL' not in df_prev_15m.columns:
                            df_prev_15m['-1HCL'] = np.nan
                        # KST는 엑셀 저장 직전에만 문자열로 변환 (지금은 유지)
                        if 'KST' in df_prev_15m.columns:
                            df_prev_15m['KST'] = df_prev_15m['KST'].astype(str)
                            df_prev_15m['KST'] = df_prev_15m['KST'].str.replace('nan', '').str.replace('NaT', '')
                        # Date(UTC)는 Timestamp로 유지하여 정렬 (데이터 파괴 방지)
                        if 'Date(UTC)' in df_prev_15m.columns:
                            df_prev_15m = df_prev_15m.sort_values('Date(UTC)', ascending=False, na_position='last').reset_index(drop=True)
                        print(f"{get_timestamp()} [{stage_prefix}] 📖 15분봉 {len(df_prev_15m)}개 읽기 완료")
                    
                    # 1시간봉 시트 읽기
                    if f"{TICKER}USDT1H" in wb_prev.sheetnames:
                        df_prev_1h = pd.read_excel(latest_previous_file, sheet_name=f"{TICKER}USDT1H", header=0)
                        # [Source 방식] 엑셀 읽은 직후 데이터 세척 (쉼표 제거, 숫자/날짜 강제 변환)
                        df_prev_1h = clean_df_display_format(df_prev_1h, sheet_type='1h')
                        # KST는 엑셀 저장 직전에만 문자열로 변환 (지금은 유지)
                        if 'KST' in df_prev_1h.columns:
                            df_prev_1h['KST'] = df_prev_1h['KST'].astype(str)
                            df_prev_1h['KST'] = df_prev_1h['KST'].str.replace('nan', '').str.replace('NaT', '')
                        # 숫자형 변환 (안전장치)
                        num_cols_1h = ["종", "시", "고", "저", "Vol.", "SMA25", "SMA100", "SMA200", "SMA400", "SMA800", "Max200", "Min200", "하단", "상단", "SFast", "Fast", "Base", "1HMSFast", "4or1", "buyside", "sellside", "1HCLASS", "-1HCLASS"]
                        for col in num_cols_1h:
                            if col in df_prev_1h.columns:
                                df_prev_1h[col] = pd.to_numeric(df_prev_1h[col], errors='coerce')
                        # Date(UTC)는 Timestamp로 유지하여 정렬 (데이터 파괴 방지)
                        if 'Date(UTC)' in df_prev_1h.columns:
                            df_prev_1h = df_prev_1h.sort_values('Date(UTC)', ascending=False, na_position='last').reset_index(drop=True)
                        print(f"{get_timestamp()} [{stage_prefix}] 📖 1시간봉 {len(df_prev_1h)}개 읽기 완료")
                    
                    # 1H4x 시트 읽기
                    df_prev_1h4x = pd.DataFrame()
                    if f"{TICKER}USDT1H4x" in wb_prev.sheetnames:
                        df_prev_1h4x = pd.read_excel(latest_previous_file, sheet_name=f"{TICKER}USDT1H4x", header=0)
                        # [Source 방식] 엑셀 읽은 직후 데이터 세척 (쉼표 제거, 숫자/날짜 강제 변환)
                        df_prev_1h4x = clean_df_display_format(df_prev_1h4x, sheet_type='1h4x')
                        # KST는 엑셀 저장 직전에만 문자열로 변환 (지금은 유지)
                        if 'KST' in df_prev_1h4x.columns:
                            df_prev_1h4x['KST'] = df_prev_1h4x['KST'].astype(str)
                            df_prev_1h4x['KST'] = df_prev_1h4x['KST'].str.replace('nan', '').str.replace('NaT', '')
                        print(f"{get_timestamp()} [{stage_prefix}] 📖 1H4x {len(df_prev_1h4x)}개 읽기 완료")
                    
                    # 일봉 시트 읽기
                    if f"{TICKER}USDT1D" in wb_prev.sheetnames:
                        df_prev_1d = pd.read_excel(latest_previous_file, sheet_name=f"{TICKER}USDT1D", header=0)
                        # [Source 방식] 엑셀 읽은 직후 데이터 세척 (쉼표 제거, 숫자/날짜 강제 변환)
                        df_prev_1d = clean_df_display_format(df_prev_1d, sheet_type='1d')
                        # KST는 엑셀 저장 직전에만 문자열로 변환 (지금은 유지)
                        if 'KST' in df_prev_1d.columns:
                            df_prev_1d['KST'] = df_prev_1d['KST'].astype(str)
                            df_prev_1d['KST'] = df_prev_1d['KST'].str.replace('nan', '').str.replace('NaT', '')
                        # Date(UTC)는 Timestamp로 유지하여 정렬 (데이터 파괴 방지)
                        if 'Date(UTC)' in df_prev_1d.columns:
                            df_prev_1d = df_prev_1d.sort_values('Date(UTC)', ascending=False, na_position='last').reset_index(drop=True)
                        print(f"{get_timestamp()} [{stage_prefix}] 📖 일봉 {len(df_prev_1d)}개 읽기 완료")
                    
                    # 주봉 시트 읽기 (PREVIOUS 단계에서만 읽음, AFTER 단계에서는 일봉으로부터 새로 생성)
                    if skip_first_row and f"{TICKER}USDTW" in wb_prev.sheetnames:
                        df_prev_weekly = pd.read_excel(latest_previous_file, sheet_name=f"{TICKER}USDTW", header=0)
                        # [Source 방식] 엑셀 읽은 직후 데이터 세척 (쉼표 제거, 숫자/날짜 강제 변환)
                        df_prev_weekly = clean_df_display_format(df_prev_weekly)
                        # KST는 엑셀 저장 직전에만 문자열로 변환 (지금은 유지)
                        if 'KST' in df_prev_weekly.columns:
                            df_prev_weekly['KST'] = df_prev_weekly['KST'].astype(str)
                            df_prev_weekly['KST'] = df_prev_weekly['KST'].str.replace('nan', '').str.replace('NaT', '')
                        # Date(UTC)는 Timestamp로 유지하여 정렬 (데이터 파괴 방지)
                        if 'Date(UTC)' in df_prev_weekly.columns:
                            df_prev_weekly = df_prev_weekly.sort_values('Date(UTC)', ascending=False, na_position='last').reset_index(drop=True)
                        print(f"{get_timestamp()} [{stage_prefix}] 📖 주봉 {len(df_prev_weekly)}개 읽기 완료")
                        
                        # ⚠️중요: previous 데이터의 숫자 컬럼을 강제로 숫자로 변환 (콤마 제거 후 변환)
                        # 엑셀에서 불러온 숫자가 문자열("89,000.00")로 저장되어 있을 수 있음
                        numeric_cols = ['종', '시', '고', '저', 'Vol.']
                        # df_prev_weekly는 2단계(skip_first_row=False)에서 읽지 않으므로 조건부로 추가
                        df_prev_list = [df_prev_5m, df_prev_15m, df_prev_1h, df_prev_1h4x, df_prev_1d]
                        if skip_first_row and 'df_prev_weekly' in locals() and not df_prev_weekly.empty:
                            df_prev_list.append(df_prev_weekly)
                        for df_prev in df_prev_list:
                            if not df_prev.empty:
                                for col in numeric_cols:
                                    if col in df_prev.columns:
                                        # 문자열로 변환 후 콤마 제거, 그 다음 숫자로 변환
                                        df_prev[col] = pd.to_numeric(df_prev[col].astype(str).str.replace(',', ''), errors='coerce')
                    
                    # previous 파일의 컬럼 순서를 표준화 (나중에 엑셀 저장 시와 동일한 순서로)
                    # 5분봉 컬럼 순서 맞춤 (SB1M, 1HMSF 추가)
                    binance_cols_5m_standard = ["Date(UTC)", "KST", "종", "시", "고", "저", "Vol.", "SMA3", "SMA5", "SMA7", "SMA10", "SMA20", "Max200", "Min200", "하단", "상단", "SFast", "Fast", "Base", "4or1", "buyside", "sellside", "Sell", "Buy", "SB1M", "1HMSF"]
                    if not df_prev_5m.empty:
                        binance_cols_5m_standard = [col for col in binance_cols_5m_standard if col in df_prev_5m.columns]
                        df_prev_5m = df_prev_5m[binance_cols_5m_standard]
                    
                    # 15분봉 컬럼 순서 맞춤
                    # 계산 컬럼(SB5M, SB1M, SB1H, SB1D, ORDER, KSC, PRFT, dateM, LD)은 없어도 유지 (나중에 계산됨)
                    binance_cols_15m_standard = ["Date(UTC)", "KST", "종", "시", "고", "저", "Vol.", "SMA3", "SMA5", "SMA7", "SMA10", "SMA12", "SMAF", "SMA20", "SMA25", "SMA100", "SMA200", "SMA400", "SMA800", "Max70", "Min70", "하단", "상단", "SFast", "Fast", "Base", "4or1", "buyside", "sellside", "Sell", "Buy", "SB1M", "SB5M", "SB1H", "SB1D", "ORDER", "1HMSFast", "1HCL", "-1HCL", "p", "KSC", "Bomb", "PRFT", "StoSP", "TP", "StoSU", "TPC", "TPCS", "NBS", "LS", "SamountW", "BamountW", "Samount1D", "Bamount1D", "Samount", "Bamount", "dateM", "LD", "SPRD", "SPRD2"]
                    if not df_prev_15m.empty:
                        # 표준 컬럼 중 존재하는 것만 선택
                        existing_cols = [col for col in binance_cols_15m_standard if col in df_prev_15m.columns]
                        # 계산 컬럼은 없어도 빈 값으로 추가
                        calc_cols = ["SB5M", "SB1M", "SB1H", "SB1D", "ORDER", "1HCL", "-1HCL", "p", "KSC", "Bomb", "PRFT", "StoSP", "TP", "StoSU", "TPC", "TPCS", "NBS", "LS", "SPRD", "dateM", "LD"]
                        for calc_col in calc_cols:
                            if calc_col not in df_prev_15m.columns and calc_col in binance_cols_15m_standard:
                                if calc_col in ["SB5M", "SB1M", "SB1H", "SB1D", "ORDER", "Bomb"]:
                                    df_prev_15m[calc_col] = ''
                                elif calc_col == "LS":
                                    df_prev_15m[calc_col] = ''  # 헤더만 추가, 내용 채우지 않음
                                elif calc_col in ["TPC", "TPCS", "NBS"]:
                                    df_prev_15m[calc_col] = 0
                                else:
                                    df_prev_15m[calc_col] = np.nan
                        # 2단계 이어붙이기: 구 previous(또는 after)에 SMA400/SMA800 없으면 NaN으로 추가 (컬럼 순서·구조 통일)
                        for col in ['SMA400', 'SMA800']:
                            if col not in df_prev_15m.columns and col in binance_cols_15m_standard:
                                df_prev_15m[col] = np.nan
                        # bomb → Bomb로 변경 (대소문자 통일)
                        if 'bomb' in df_prev_15m.columns and 'Bomb' not in df_prev_15m.columns:
                            df_prev_15m['Bomb'] = df_prev_15m['bomb']
                            df_prev_15m = df_prev_15m.drop(columns=['bomb'], errors='ignore')
                        # 최종 컬럼 순서 맞추기
                        final_cols = [col for col in binance_cols_15m_standard if col in df_prev_15m.columns]
                        df_prev_15m = df_prev_15m[final_cols]
                    
                    # 1분봉 시트 읽기 (2단계에서만 필요)
                    if not skip_first_row and f"{TICKER}USDT1M" in wb_prev.sheetnames:
                        df_prev_1m = pd.read_excel(latest_previous_file, sheet_name=f"{TICKER}USDT1M", header=0)
                        # [Source 방식] 엑셀 읽은 직후 데이터 세척 (쉼표 제거, 숫자/날짜 강제 변환)
                        df_prev_1m = clean_df_display_format(df_prev_1m, sheet_type='1m')
                        # KST는 엑셀 저장 직전에만 문자열로 변환 (지금은 유지)
                        if 'KST' in df_prev_1m.columns:
                            df_prev_1m['KST'] = df_prev_1m['KST'].astype(str)
                            df_prev_1m['KST'] = df_prev_1m['KST'].str.replace('nan', '').str.replace('NaT', '')
                        # Date(UTC)는 Timestamp로 유지하여 정렬 (데이터 파괴 방지)
                        if 'Date(UTC)' in df_prev_1m.columns:
                            df_prev_1m = df_prev_1m.sort_values('Date(UTC)', ascending=False, na_position='last').reset_index(drop=True)
                        print(f"{get_timestamp()} [{stage_prefix}] 📖 1분봉 {len(df_prev_1m)}개 읽기 완료")
                    
                    # 1시간봉 컬럼 순서 맞춤 (Source 기준: SMA25, SMA100, SMA200, SMA400, SMA800, Max200, Min200, 1HCLASS, -1HCLASS)
                    binance_cols_1h_standard = ["Date(UTC)", "KST", "종", "시", "고", "저", "Vol.", "SMA25", "SMA100", "SMA200", "SMA400", "SMA800", "Max200", "Min200", "하단", "상단", "SFast", "Fast", "Base", "1HMSFast", "4or1", "buyside", "sellside", "Sell", "Buy", "1HCLASS", "-1HCLASS", "p1H"]
                    if not df_prev_1h.empty:
                        # p1H 열 없으면 추가 (내용 채우지 않음)
                        if 'p1H' not in df_prev_1h.columns:
                            df_prev_1h['p1H'] = np.nan
                        # 표준 컬럼 중 존재하는 것만 선택
                        existing_cols = [col for col in binance_cols_1h_standard if col in df_prev_1h.columns]
                        df_prev_1h = df_prev_1h[existing_cols]
                    
                    # 일봉 컬럼 순서 맞춤
                    binance_cols_1d_standard = ["Date(UTC)", "KST", "종", "시", "고", "저", "Vol.", "SMA3", "SMA5", "SMA7", "SMA10", "SMA20", "Max15", "Min15", "하단", "상단", "SFast", "Fast", "Base", "4or1", "buyside", "sellside", "Sell", "Buy", "Samount1D", "Bamount1D"]
                    if not df_prev_1d.empty:
                        binance_cols_1d_standard = [col for col in binance_cols_1d_standard if col in df_prev_1d.columns]
                        df_prev_1d = df_prev_1d[binance_cols_1d_standard]
                    
                    # 주봉 컬럼 순서 맞춤 (PREVIOUS 단계에서만 실행)
                    if skip_first_row:
                        binance_cols_weekly_standard = ["Date(UTC)", "KST", "종", "시", "고", "저", "Vol.", "SMA3", "SMA5", "SMA7", "SMA10", "SMA20", "Max25", "Min25", "하단", "상단", "SFast", "Fast", "Base", "4or1", "buyside", "sellside", "Sell", "Buy", "SamountW", "BamountW"]
                        if 'df_prev_weekly' in locals() and not df_prev_weekly.empty:
                            binance_cols_weekly_standard = [col for col in binance_cols_weekly_standard if col in df_prev_weekly.columns]
                            df_prev_weekly = df_prev_weekly[binance_cols_weekly_standard]
                except Exception as e:
                    print(f"{get_timestamp()} [{stage_prefix}] ❌데이터 파싱 실패: {e}")
                    return None
        else:
            print(f"{get_timestamp()} [{stage_prefix}] ❌이어쓸 이전 파일(Previous/After)을 찾을 수 없습니다.")
            print(f"{get_timestamp()} [{stage_prefix}] 💡 팁: 최초 1회는 반드시 1단계(Previous 생성)가 실행되어야 합니다.")
            return None

    # endTime 계산: 실행 시점에 맞는 endTime 설정
    fixed_end_time_ms = None
    fixed_end_time_ms_1h = None  # 1시간봉 전용 endTime (15분 단위로 내림)
    if polling_start_time is not None:
        # 실행 시점의 분과 초 추출
        exec_minute = polling_start_time.minute
        exec_second = polling_start_time.second
        
        # 1단계: 7분, 22분, 37분, 52분에 실행 → 해당 15분 구간의 시작 시간 - 5초
        # 2단계: 15분1초, 30분1초, 45분1초, 0분1초에 실행 → 해당 분의 1초 시점
        if skip_first_row:
            # 1단계: 7분, 22분, 37분, 52분 → 해당 15분 구간의 시작 시간 - 5초
            exec_utc = polling_start_time.astimezone(tz.UTC) if polling_start_time.tzinfo else polling_start_time.replace(tzinfo=tz.UTC)
            if exec_minute == 7:
                # 7분: XX시 0분 05초
                base_minute = 0
                base_second = 5
            elif exec_minute == 22:
                # 22분: XX시 15분 05초
                base_minute = 15
                base_second = 5
            elif exec_minute == 37:
                # 37분: XX시 30분 05초
                base_minute = 30
                base_second = 5
            elif exec_minute == 52:
                # 52분: XX시 45분 05초
                base_minute = 45
                base_second = 5
            else:
                # 초회실행: 현재 시간에 따라 결정
                if 0 <= exec_minute < 15:
                    # 0-15분: XX시 00분 05초
                    base_minute = 0
                    base_second = 5
                elif 15 <= exec_minute < 30:
                    # 15-30분: XX시 16분 00초 (사용자 요구사항)
                    base_minute = 16
                    base_second = 0
                elif 30 <= exec_minute < 45:
                    # 30-45분: XX시 31분 00초 (사용자 요구사항)
                    base_minute = 31
                    base_second = 0
                else:  # 45-60분
                    # 45-60분: XX시 45분 05초
                    base_minute = 45
                    base_second = 5
            
            end_time_dt = exec_utc.replace(minute=base_minute, second=base_second, microsecond=0)
            fixed_end_time_ms = int(end_time_dt.timestamp() * 1000)
            # 1시간봉: 15분 단위로 내림 (00분, 15분, 30분, 45분)
            end_time_dt_1h = exec_utc.replace(minute=base_minute, second=0, microsecond=0)
            fixed_end_time_ms_1h = int(end_time_dt_1h.timestamp() * 1000)
        else:
            # 2단계: 15분1초, 30분1초, 45분1초, 0분1초
            if exec_minute in [15, 30, 45, 0] and exec_second == 1:
                exec_utc = polling_start_time.astimezone(tz.UTC) if polling_start_time.tzinfo else polling_start_time.replace(tzinfo=tz.UTC)
                # ⚠️중요: 실행 시간 그대로 사용 (예: 30분 1초 실행 → fixed_end_time_ms = 03:30:01)
                # 바이낸스 API는 endTime **이하**의 캔들을 반환하므로, 03:30:01을 전달하면:
                #   - 03:30 캔들 (03:30~03:45 구간, 1초짜리 미완성)
                #   - 03:15 캔들 (03:15~03:30 구간, 15분짜리 완성)
                # → 미완성 제거 로직에서 03:30을 제거 → 03:15만 남음
                end_time_dt = exec_utc.replace(second=1, microsecond=0)
                fixed_end_time_ms = int(end_time_dt.timestamp() * 1000)
                # 1시간봉: 15분 단위로 내림 (00분, 15분, 30분, 45분)
                end_time_dt_1h = exec_utc.replace(second=0, microsecond=0)
                fixed_end_time_ms_1h = int(end_time_dt_1h.timestamp() * 1000)

    # 데이터 수집
    # 1분봉 데이터 수집 (minute1_count는 이미 위에서 정의됨)
    df_binance_ticker_1m = pd.DataFrame()
    
    if not skip_first_row and not df_prev_1m.empty:
        # 2단계: 최신 캔들 조회 (5분봉 3개 = 최대 15개 필요 + 여유분)
        # ⚠️중요: 5분봉 최신 시간까지 커버하려면 충분한 1분봉 데이터 필요
        # minute1_count 사용 (기본 16개, 경계선 데이터 누락 방지)
        if pre_fetched_data and '1m' in pre_fetched_data and not pre_fetched_data['1m'].empty:
            df_binance_ticker_1m_new = pre_fetched_data['1m'].copy()
        else:
            df_binance_ticker_1m_new = fetch_binance_minutes1(binance_symbol_ticker, minute1_count, include_today=include_today, fixed_end_time_ms=fixed_end_time_ms, stage_prefix=stage_prefix)
        
        # 2단계: 16개 수집 → 미완성 1개 제거 → 15개 사용
        if not skip_first_row and len(df_binance_ticker_1m_new) > 0:
            try:
                # ⚠️중요: 최신 캔들(첫 번째 행)의 '분'이 현재 실행 시간의 '분'과 같으면 미완성 캔들
                # 예: 15분에 실행하면 현재 시간=00:15, 최신 캔들=00:15 (미완성) → 제거
                #     30분에 실행하면 현재 시간=00:30, 최신 캔들=00:29 (완성) → 유지
                
                if len(df_binance_ticker_1m_new) > 0:
                    # 최신 캔들 시간 확인 (이미 Timestamp 객체일 수 있음)
                    latest_date_val = df_binance_ticker_1m_new.iloc[0]['Date(UTC)']
                    
                    # Timestamp 객체인 경우 그대로 사용, 문자열인 경우 파싱
                    if isinstance(latest_date_val, pd.Timestamp):
                        latest_date = latest_date_val
                    elif pd.api.types.is_datetime64_any_dtype(pd.Series([latest_date_val])):
                        latest_date = pd.to_datetime(latest_date_val)
                    else:
                        # 문자열인 경우 format 없이 자동 인식
                        latest_date = pd.to_datetime(latest_date_val, errors='coerce')
                
                if pd.notna(latest_date):
                        # 현재 실행 시간의 '분' 확인 (Source 방식: 함수 내부에서 다시 계산)
                        current_time_utc = dt.datetime.now(tz.UTC)
                        current_minute = current_time_utc.minute
                        current_hour = current_time_utc.hour
                        current_date = current_time_utc.date()
                        
                    # 최신 캔들의 '분'이 현재 시간의 '분'과 같고, 날짜/시간도 같으면 미완성 -> 제거
                        # NaT 체크: latest_date가 NaT가 아닌 경우에만 .date() 호출
                        if pd.notna(latest_date) and (latest_date.minute == current_minute and 
                        latest_date.hour == current_hour and 
                        latest_date.date() == current_date):
                            df_binance_ticker_1m_new = df_binance_ticker_1m_new.iloc[1:].reset_index(drop=True)
                            latest_date_str = str(latest_date_val) if not isinstance(latest_date_val, pd.Timestamp) else latest_date_val.strftime('%y/%m/%d,%H:%M')
                            print(f"{get_timestamp()} [{stage_prefix}] ✅ 1분봉 미완성 캔들 제거: {latest_date_str} (현재 시간: {current_time_utc.strftime('%y/%m/%d,%H:%M')})")
            except Exception as e:
                print(f"{get_timestamp()} [{stage_prefix}] ⚠️{ticker} 1분봉 미완성 캔들 제거 실패: {e}")
                import traceback
                traceback.print_exc()
        
        # after 엑셀 구조: 1행 헤더, 2행 새 데이터(계산 필요), 3행부터 previous 데이터(변경 없음)
        base_cols_1m = ['Date(UTC)', 'KST', '종', '시', '고', '저', 'Vol.']
        cols_1m_new = [col for col in base_cols_1m if col in df_binance_ticker_1m_new.columns]
        
        if cols_1m_new and 'Date(UTC)' in cols_1m_new and len(df_binance_ticker_1m_new) >= 15:
            # previous 데이터의 모든 컬럼 유지 (지표 포함) - previous 엑셀의 2행부터
            df_prev_1m_all_cols = df_prev_1m.copy()
            
            # 새 데이터는 최신 15개 사용 (2행, 3행, ..., 16행에 배치할 데이터)
            # ⚠️중요: 2단계는 15개만 사용하고 나머지는 previous에서 가져옴
            # 예: 30분 실행 시 previous는 xx:14까지, 새 데이터는 xx:15~xx:29 (15개) → 정확히 이어짐
            # 5분봉 3개 그룹 보장을 위해 최소 15개 필요 (xx:15, xx:20, xx:25 그룹)
            use_count = min(15, len(df_binance_ticker_1m_new))  # 15개 사용 (2단계 규칙)
            df_new_1m_basic = df_binance_ticker_1m_new.iloc[0:use_count][cols_1m_new].copy()
            
            # 새 데이터에 previous와 동일한 컬럼 구조 만들기 (지표는 NaN으로)
            for col in df_prev_1m_all_cols.columns:
                if col not in df_new_1m_basic.columns:
                    df_new_1m_basic[col] = np.nan
            
            # 컬럼 순서 맞추기
            df_new_1m_basic = df_new_1m_basic[df_prev_1m_all_cols.columns]
            
            # [중요] 병합 직전 타입 강제 통일 (Timestamp와 str 혼합 방지)
            if 'Date(UTC)' in df_new_1m_basic.columns:
                df_new_1m_basic['Date(UTC)'] = pd.to_datetime(df_new_1m_basic['Date(UTC)'], errors='coerce')
            if 'Date(UTC)' in df_prev_1m_all_cols.columns:
                df_prev_1m_all_cols['Date(UTC)'] = pd.to_datetime(df_prev_1m_all_cols['Date(UTC)'], errors='coerce')
            
            # ⚠️중요: 중복 제거는 Date(UTC) 기준으로만 수행 (KST는 기록용일 뿐, 중복 제거에는 사용하지 않음)
            # 타입 통일 후 Timestamp 객체로 직접 비교
            if 'Date(UTC)' in df_new_1m_basic.columns and 'Date(UTC)' in df_prev_1m_all_cols.columns:
                # 새 데이터의 Date(UTC) 값들을 Timestamp 객체로 변환하여 set 생성
                new_date_utc_set = set()
                for idx in range(len(df_new_1m_basic)):
                    new_date_utc = df_new_1m_basic.iloc[idx]['Date(UTC)']
                    if pd.notna(new_date_utc):
                        # Timestamp 객체를 직접 사용 (floor로 정규화하여 비교)
                        if isinstance(new_date_utc, pd.Timestamp):
                            new_date_utc_set.add(new_date_utc.floor('1min'))
                        else:
                            dt_obj = pd.to_datetime(new_date_utc, errors='coerce')
                            if pd.notna(dt_obj):
                                new_date_utc_set.add(dt_obj.floor('1min'))
                
                if new_date_utc_set:
                    # previous 데이터에서 새 데이터의 Date(UTC) 시간과 중복되는 행 제거 (Timestamp 객체로 직접 비교)
                    df_prev_1m_all_cols['Date(UTC)_floor'] = pd.to_datetime(df_prev_1m_all_cols['Date(UTC)'], errors='coerce').dt.floor('1min')
                    df_prev_1m_all_cols = df_prev_1m_all_cols[
                        ~df_prev_1m_all_cols['Date(UTC)_floor'].isin(new_date_utc_set)
                    ].copy()
                    df_prev_1m_all_cols = df_prev_1m_all_cols.drop('Date(UTC)_floor', axis=1)
            
            # 합치기: 새 데이터 15개(2-16행) + previous 데이터(17행부터)
            # 1행 헤더는 나중에 엑셀 저장 시 자동 생성됨
            df_binance_ticker_1m = pd.concat([
                df_new_1m_basic,  # 2-16행: 새 데이터 15개 (계산 필요)
                df_prev_1m_all_cols  # 17행부터: previous 데이터 (변경 없음)
            ], ignore_index=True)
            
            # [중요] 병합 직후 타입 통일 (Timestamp와 str 혼합 방지)
            df_binance_ticker_1m = clean_df_display_format(df_binance_ticker_1m)
            
            # 메모리 정리: 원본 DataFrame 삭제
            del df_new_1m_basic, df_prev_1m_all_cols
            
            # ⚠️중요: 시간 기준 정렬은 Date(UTC) 기준으로만 수행 (KST는 기록용일 뿐, 정렬에는 사용하지 않음)
            if 'Date(UTC)' in df_binance_ticker_1m.columns:
                # 정렬 직전 타입 일치 여부 최종 확인
                df_binance_ticker_1m = clean_df_display_format(df_binance_ticker_1m)
                # 정렬 실행
                df_binance_ticker_1m = df_binance_ticker_1m.sort_values('Date(UTC)', ascending=False, na_position='last').reset_index(drop=True)
        else:
            df_binance_ticker_1m = df_binance_ticker_1m_new
    else:
        # 1단계 또는 previous 파일이 없는 경우: 전체 조회
        if skip_first_row:
            # [최적화] 병렬 수집 로직 (1분봉/5분봉/15분봉/1시간봉/일봉 동시 수집)
            print(f"{get_timestamp()} [{stage_prefix}] 🚀 캔들 병렬 수집 시작")
            
            # ThreadPoolExecutor를 사용해 5개의 API 호출을 동시에 실행
            with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
                # 동시에 작업 요청: fetch 함수들과 필요한 매개변수를 executor에 제출
                future_1m = executor.submit(fetch_binance_minutes1, binance_symbol_ticker, minute1_count, include_today, fixed_fixed_end_time_ms, stage_prefix)
                future_5m = executor.submit(fetch_binance_minutes5, binance_symbol_ticker, minute5_count, include_today, fixed_fixed_end_time_ms)
                future_15m = executor.submit(fetch_binance_minutes15, binance_symbol_ticker, minute15_count, include_today, fixed_fixed_end_time_ms)
                future_1h = executor.submit(fetch_binance_hours1, binance_symbol_ticker, hour1_count, include_today, fixed_fixed_end_time_ms)
                # 일봉은 미완성 캔들 제거 로직이 없으므로 include_today=True로 설정하여 200개 수집 보장
                future_1d = executor.submit(fetch_binance_daily, binance_symbol_ticker, daily_count, include_today, fixed_fixed_end_time_ms)
                
                # 결과 수집 (모든 future가 완료될 때까지 대기)
                df_binance_ticker_1m = future_1m.result()
                df_binance_ticker_5m = future_5m.result()
                df_binance_ticker_15m = future_15m.result()
                df_binance_ticker_1h = future_1h.result()
                df_binance_ticker_1d = future_1d.result()
            
            print(f"{get_timestamp()} [{stage_prefix}] ✅모든 캔들 수집 완료 (병렬 수집)")
        
        # 1단계 실행 시: 수집 단계에서 미완성 캔들 제거
        if len(df_binance_ticker_1m) > 0:
            # 최신 1개 제거 (미완성 캔들)
            df_binance_ticker_1m = df_binance_ticker_1m.iloc[1:].reset_index(drop=True)
        if len(df_binance_ticker_5m) > 0:
                # 최신 1개 제거 (미완성 캔들)
            df_binance_ticker_5m = df_binance_ticker_5m.iloc[1:].reset_index(drop=True)
        if len(df_binance_ticker_15m) > 0:
                # 최신 1개 제거 (미완성 캔들)
                df_binance_ticker_15m = df_binance_ticker_15m.iloc[1:].reset_index(drop=True)
        else:
            # 2단계: 순차 수집 (병렬 수집 사용 안 함)
            print(f"{get_timestamp()} [{stage_prefix}] 📥 1분봉 캔들 수집 중...")
            df_binance_ticker_1m = fetch_binance_minutes1(binance_symbol_ticker, minute1_count, include_today=include_today, fixed_end_time_ms=fixed_end_time_ms, stage_prefix=stage_prefix)
    
    # 5분봉 데이터 수집
    if not skip_first_row and not df_prev_5m.empty:
        # 2단계: 최신 4개 캔들만 조회 (미완성 1개 제거 후 3개)
        if pre_fetched_data and '5m' in pre_fetched_data and not pre_fetched_data['5m'].empty:
            df_binance_ticker_5m_new = pre_fetched_data['5m'].copy()
        else:
            df_binance_ticker_5m_new = fetch_binance_minutes5(binance_symbol_ticker, 4, include_today=include_today, fixed_end_time_ms=fixed_end_time_ms)
        
        # 캐싱 시 이미 미완성 캔들 제거됨 - 추가 처리 불필요
        
        # after 엑셀 구조: 1행 헤더, 2-4행 새 데이터 3개(계산 필요), 5행부터 previous 데이터(변경 없음)
        base_cols_5m = ['Date(UTC)', 'KST', '종', '시', '고', '저', 'Vol.']
        cols_5m_new = [col for col in base_cols_5m if col in df_binance_ticker_5m_new.columns]
        
        if cols_5m_new and 'Date(UTC)' in cols_5m_new and len(df_binance_ticker_5m_new) >= 3:
            # previous 데이터의 모든 컬럼 유지 (지표 포함) - previous 엑셀의 2행부터
            df_prev_5m_all_cols = df_prev_5m.copy()
            
            # 새 데이터는 최신 3개 사용 (2-4행에 배치할 데이터)
            df_new_5m_basic = df_binance_ticker_5m_new.iloc[0:3][cols_5m_new].copy()
            
            # 새 데이터에 previous와 동일한 컬럼 구조 만들기 (지표는 NaN으로)
            for col in df_prev_5m_all_cols.columns:
                if col not in df_new_5m_basic.columns:
                    df_new_5m_basic[col] = np.nan
            
            # 컬럼 순서 맞추기
            df_new_5m_basic = df_new_5m_basic[df_prev_5m_all_cols.columns]
            
            # [중요] 병합 직전 타입 강제 통일 (Timestamp와 str 혼합 방지)
            if 'Date(UTC)' in df_new_5m_basic.columns:
                df_new_5m_basic['Date(UTC)'] = pd.to_datetime(df_new_5m_basic['Date(UTC)'], errors='coerce')
            if 'Date(UTC)' in df_prev_5m_all_cols.columns:
                df_prev_5m_all_cols['Date(UTC)'] = pd.to_datetime(df_prev_5m_all_cols['Date(UTC)'], errors='coerce')
            
            # ⚠️중요: 중복 제거는 Date(UTC) 기준으로만 수행 (KST는 기록용일 뿐, 중복 제거에는 사용하지 않음)
            # 타입 통일 후 Timestamp 객체로 직접 비교
            if 'Date(UTC)' in df_new_5m_basic.columns and 'Date(UTC)' in df_prev_5m_all_cols.columns:
                # 새 데이터의 Date(UTC) 값들을 Timestamp 객체로 변환하여 set 생성
                new_date_utc_set = set()
                for idx in range(len(df_new_5m_basic)):
                    new_date_utc = df_new_5m_basic.iloc[idx]['Date(UTC)']
                    if pd.notna(new_date_utc):
                        # Timestamp 객체를 직접 사용 (floor로 정규화하여 비교)
                        if isinstance(new_date_utc, pd.Timestamp):
                            new_date_utc_set.add(new_date_utc.floor('5min'))
                        else:
                            dt_obj = pd.to_datetime(new_date_utc, errors='coerce')
                            if pd.notna(dt_obj):
                                new_date_utc_set.add(dt_obj.floor('5min'))
                
                if new_date_utc_set:
                    # previous 데이터에서 새 데이터의 Date(UTC) 시간과 중복되는 행 제거 (Timestamp 객체로 직접 비교)
                    df_prev_5m_all_cols['Date(UTC)_floor'] = pd.to_datetime(df_prev_5m_all_cols['Date(UTC)'], errors='coerce').dt.floor('5min')
                    df_prev_5m_all_cols = df_prev_5m_all_cols[
                        ~df_prev_5m_all_cols['Date(UTC)_floor'].isin(new_date_utc_set)
                    ].copy()
                    df_prev_5m_all_cols = df_prev_5m_all_cols.drop('Date(UTC)_floor', axis=1)
            
            # 합치기: 새 데이터 3개(2-4행) + previous 데이터(5행부터)
            # 1행 헤더는 나중에 엑셀 저장 시 자동 생성됨
            df_binance_ticker_5m = pd.concat([
                df_new_5m_basic,  # 2-4행: 새 데이터 3개 (계산 필요)
                df_prev_5m_all_cols  # 5행부터: previous 데이터 (변경 없음)
            ], ignore_index=True)
            
            # [중요] 병합 직후 타입 통일 (Timestamp와 str 혼합 방지)
            df_binance_ticker_5m = clean_df_display_format(df_binance_ticker_5m)
            
            # 메모리 정리: 원본 DataFrame 삭제
            del df_new_5m_basic, df_prev_5m_all_cols
            
            # ⚠️중요: 시간 기준 정렬은 Date(UTC) 기준으로만 수행 (KST는 기록용일 뿐, 정렬에는 사용하지 않음)
            if 'Date(UTC)' in df_binance_ticker_5m.columns:
                # 정렬 직전 타입 일치 여부 최종 확인
                df_binance_ticker_5m = clean_df_display_format(df_binance_ticker_5m)
                # 정렬 실행
                df_binance_ticker_5m = df_binance_ticker_5m.sort_values('Date(UTC)', ascending=False, na_position='last').reset_index(drop=True)
        else:
            df_binance_ticker_5m = df_binance_ticker_5m_new
    else:
        # 1단계는 이미 병렬 수집에서 처리됨
        if not skip_first_row:
            # 2단계: 순차 수집 (병렬 수집 사용 안 함)
            print(f"{get_timestamp()} [{stage_prefix}] 📥 5분봉 캔들 수집 중...")
            df_binance_ticker_5m = fetch_binance_minutes5(binance_symbol_ticker, minute5_count, include_today=include_today, fixed_end_time_ms=fixed_end_time_ms)

    # 15분봉 데이터 수집
    if not skip_first_row and not df_prev_15m.empty:
        # 2단계: 최신 2개 캔들만 조회 (미완성 1개 제거 후 1개)
        if pre_fetched_data and '15m' in pre_fetched_data and not pre_fetched_data['15m'].empty:
            df_binance_ticker_15m_new = pre_fetched_data['15m'].copy()
        else:
            df_binance_ticker_15m_new = fetch_binance_minutes15(binance_symbol_ticker, 2, include_today=include_today, fixed_end_time_ms=fixed_end_time_ms)
        
        # 캐싱 시 이미 미완성 캔들 제거됨 - 추가 처리 불필요
        
        # after 엑셀 구조: 1행 헤더, 2행 새 데이터(계산 필요), 3행부터 previous 데이터(변경 없음)
        base_cols_15m = ['Date(UTC)', 'KST', '종', '시', '고', '저', 'Vol.']
        cols_15m_new = [col for col in base_cols_15m if col in df_binance_ticker_15m_new.columns]
        
        if cols_15m_new and 'Date(UTC)' in cols_15m_new and len(df_binance_ticker_15m_new) > 0:
            # previous 데이터의 모든 컬럼 유지 (지표 포함) - previous 엑셀의 2행부터
            df_prev_15m_all_cols = df_prev_15m.copy()
            
            # 새 데이터는 최신 1개만 사용 (2행에 배치할 데이터)
            df_new_15m_basic = df_binance_ticker_15m_new.iloc[0:1][cols_15m_new].copy()
            
            # 새 데이터에 previous와 동일한 컬럼 구조 만들기 (지표는 NaN으로)
            for col in df_prev_15m_all_cols.columns:
                if col not in df_new_15m_basic.columns:
                    df_new_15m_basic[col] = np.nan
            
            # 컬럼 순서 맞추기
            df_new_15m_basic = df_new_15m_basic[df_prev_15m_all_cols.columns]
            
            # [중요] 병합 직전 타입 강제 통일 (Timestamp와 str 혼합 방지)
            if 'Date(UTC)' in df_new_15m_basic.columns:
                df_new_15m_basic['Date(UTC)'] = pd.to_datetime(df_new_15m_basic['Date(UTC)'], errors='coerce')
            if 'Date(UTC)' in df_prev_15m_all_cols.columns:
                df_prev_15m_all_cols['Date(UTC)'] = pd.to_datetime(df_prev_15m_all_cols['Date(UTC)'], errors='coerce')
            
            # ⚠️중요: 중복 제거는 Date(UTC) 기준으로만 수행 (KST는 기록용일 뿐, 중복 제거에는 사용하지 않음)
            # 타입 통일 후 Timestamp 객체로 직접 비교
            if 'Date(UTC)' in df_new_15m_basic.columns and 'Date(UTC)' in df_prev_15m_all_cols.columns:
                new_date_utc = df_new_15m_basic.iloc[0]['Date(UTC)'] if len(df_new_15m_basic) > 0 and pd.notna(df_new_15m_basic.iloc[0]['Date(UTC)']) else None
                if pd.notna(new_date_utc):
                    # Timestamp 객체로 직접 비교 (floor로 정규화)
                    if isinstance(new_date_utc, pd.Timestamp):
                        new_date_utc_floor = new_date_utc.floor('15min')
                    else:
                        dt_obj = pd.to_datetime(new_date_utc, errors='coerce')
                        new_date_utc_floor = dt_obj.floor('15min') if pd.notna(dt_obj) else None
                    
                    if pd.notna(new_date_utc_floor):
                        df_prev_15m_all_cols['Date(UTC)_floor'] = pd.to_datetime(df_prev_15m_all_cols['Date(UTC)'], errors='coerce').dt.floor('15min')
                        df_prev_15m_all_cols = df_prev_15m_all_cols[
                            df_prev_15m_all_cols['Date(UTC)_floor'] != new_date_utc_floor
                        ].copy()
                        df_prev_15m_all_cols = df_prev_15m_all_cols.drop('Date(UTC)_floor', axis=1)
            
            # 합치기: 새 데이터(2행) + previous 데이터(3행부터)
            # 1행 헤더는 나중에 엑셀 저장 시 자동 생성됨
            df_binance_ticker_15m = pd.concat([
                df_new_15m_basic,  # 2행: 새 데이터 (계산 필요)
                df_prev_15m_all_cols  # 3행부터: previous 데이터 (변경 없음)
            ], ignore_index=True)
            
            # [중요] 병합 직후 타입 통일 (Timestamp와 str 혼합 방지)
            df_binance_ticker_15m = clean_df_display_format(df_binance_ticker_15m)
            
            # 메모리 정리: 원본 DataFrame 삭제
            del df_new_15m_basic, df_prev_15m_all_cols
            
            # ⚠️중요: 시간 기준 정렬은 Date(UTC) 기준으로만 수행 (KST는 기록용일 뿐, 정렬에는 사용하지 않음)
            if 'Date(UTC)' in df_binance_ticker_15m.columns:
                # 정렬 직전 타입 일치 여부 최종 확인
                df_binance_ticker_15m = clean_df_display_format(df_binance_ticker_15m)
                # 정렬 실행
                df_binance_ticker_15m = df_binance_ticker_15m.sort_values('Date(UTC)', ascending=False, na_position='last').reset_index(drop=True)
        else:
            # 새 데이터가 없거나 조건이 맞지 않으면 previous 데이터만 사용
            if not df_prev_15m.empty:
                df_binance_ticker_15m = df_prev_15m.copy()
                print(f"{get_timestamp()} [{stage_prefix}] ⚠️ 새 15분봉 데이터가 없어 previous 데이터만 사용합니다.")
            else:
                df_binance_ticker_15m = df_binance_ticker_15m_new
    else:
        # 1단계는 이미 병렬 수집에서 처리됨
        if not skip_first_row:
            # 2단계: 순차 수집 (병렬 수집 사용 안 함)
            print(f"{get_timestamp()} [{stage_prefix}] 📥 15분봉 캔들 수집 중...")
            df_binance_ticker_15m = fetch_binance_minutes15(binance_symbol_ticker, minute15_count, include_today=include_today, fixed_end_time_ms=fixed_end_time_ms)

    # 일봉 데이터 수집
    if not skip_first_row and not df_prev_1d.empty:
        # 2단계: UTC 기준 조건부 조회
        # - xx일 00시 00분 1초 실행 시: 2개 수집 (완성 캔들 xx-1일 + 미완 캔들 xx일, 미완 1개 날림)
        # - 그 외 시간 실행 시: 1개 수집 (미완 캔들 xx-1일, 미완 유지)
        # 겹침 발생 시 2단계 캔들로 최신화 (drop_duplicates keep='first'로 처리)
        is_00hour_00min_01sec_1d = (current_hour == 0 and current_minute == 0 and current_time_utc.second == 1)
        daily_count_main = 2 if is_00hour_00min_01sec_1d else 1
        
        if pre_fetched_data and '1d' in pre_fetched_data and not pre_fetched_data['1d'].empty:
            df_binance_ticker_1d_new = pre_fetched_data['1d'].copy()
        else:
            df_binance_ticker_1d_new = fetch_binance_daily(binance_symbol_ticker, daily_count_main, include_today=include_today, fixed_end_time_ms=fixed_end_time_ms)
            # 미완성 캔들 삭제 안 함 (2단계에서는 수집 후 처리)
        
        # after 엑셀 구조: 1행 헤더, 2행 새 데이터(계산 필요), 3행부터 previous 데이터(변경 없음)
        base_cols_1d = ['Date(UTC)', 'KST', '종', '시', '고', '저', 'Vol.']
        cols_1d_new = [col for col in base_cols_1d if col in df_binance_ticker_1d_new.columns]
        
        if cols_1d_new and 'Date(UTC)' in cols_1d_new and len(df_binance_ticker_1d_new) > 0:
            df_prev_1d_all_cols = df_prev_1d.copy()
            
            # 새 데이터 처리
            # 중복 제거: Date(UTC) 컬럼 기준으로 중복 제거 (동일 시간 캔들 제거)
            if 'Date(UTC)' in df_binance_ticker_1d_new.columns:
                df_binance_ticker_1d_new = df_binance_ticker_1d_new.drop_duplicates(subset=['Date(UTC)'], keep='first')
            
            # UTC 00:00:01 실행 시: 2개 수집 → 미완성 1개 제거 → 완성된 캔들 1개만 사용
            if is_00hour_00min_01sec_1d and len(df_binance_ticker_1d_new) >= 2:
                # 최신 1개는 미완성 (UTC xx일 00:00~xx일 00:00:01, 시간표시 xx일), 2번째는 완성 (UTC xx-1일 00:00~xx일 00:00, 시간표시 xx-1일)
                # 완성된 캔들(2번째)만 사용
                df_new_1d_basic = df_binance_ticker_1d_new.iloc[1:2][cols_1d_new].copy()
            else:
                # 그 외 시간 실행 시: 최신 1개만 사용 (미완성 캔들, 시간표시 xx-1일)
                # previous와 겹침 발생 시 2단계 캔들로 최신화됨 (drop_duplicates keep='first')
                df_new_1d_basic = df_binance_ticker_1d_new.iloc[0:1][cols_1d_new].copy()
            
            # 새 데이터에 previous와 동일한 컬럼 구조 만들기 (지표는 NaN으로)
            for col in df_prev_1d_all_cols.columns:
                if col not in df_new_1d_basic.columns:
                    df_new_1d_basic[col] = np.nan
            
            # 컬럼 순서 맞추기
            df_new_1d_basic = df_new_1d_basic[df_prev_1d_all_cols.columns]
            
            # [중요] 병합 직전 타입 강제 통일 (Timestamp와 str 혼합 방지)
            if 'Date(UTC)' in df_new_1d_basic.columns:
                df_new_1d_basic['Date(UTC)'] = pd.to_datetime(df_new_1d_basic['Date(UTC)'], errors='coerce')
            if 'Date(UTC)' in df_prev_1d_all_cols.columns:
                df_prev_1d_all_cols['Date(UTC)'] = pd.to_datetime(df_prev_1d_all_cols['Date(UTC)'], errors='coerce')
            
            # ⚠️중요: 중복 제거는 Date(UTC) 기준으로만 수행 (KST는 기록용일 뿐, 중복 제거에는 사용하지 않음)
            # 타입 통일 후 Timestamp 객체로 직접 비교
            if 'Date(UTC)' in df_new_1d_basic.columns and 'Date(UTC)' in df_prev_1d_all_cols.columns:
                new_date_utc = df_new_1d_basic.iloc[0]['Date(UTC)'] if len(df_new_1d_basic) > 0 and pd.notna(df_new_1d_basic.iloc[0]['Date(UTC)']) else None
                if pd.notna(new_date_utc):
                    # Timestamp 객체로 직접 비교 (floor로 정규화)
                    if isinstance(new_date_utc, pd.Timestamp):
                        new_date_utc_floor = new_date_utc.floor('1D')
                    else:
                        dt_obj = pd.to_datetime(new_date_utc, errors='coerce')
                        new_date_utc_floor = dt_obj.floor('1D') if pd.notna(dt_obj) else None
                    
                    if pd.notna(new_date_utc_floor):
                        df_prev_1d_all_cols['Date(UTC)_floor'] = pd.to_datetime(df_prev_1d_all_cols['Date(UTC)'], errors='coerce').dt.floor('1D')
                        df_prev_1d_all_cols = df_prev_1d_all_cols[
                            df_prev_1d_all_cols['Date(UTC)_floor'] != new_date_utc_floor
                        ].copy()
                        df_prev_1d_all_cols = df_prev_1d_all_cols.drop('Date(UTC)_floor', axis=1)
            
            # 합치기: 새 데이터(2행) + previous 데이터(3행부터)
            # 1행 헤더는 나중에 엑셀 저장 시 자동 생성됨
            df_binance_ticker_1d = pd.concat([
                df_new_1d_basic,  # 2행: 새 데이터 (계산 필요)
                df_prev_1d_all_cols  # 3행부터: previous 데이터 (변경 없음)
            ], ignore_index=True)
            
            # [중요] 병합 직후 타입 통일 (Timestamp와 str 혼합 방지)
            df_binance_ticker_1d = clean_df_display_format(df_binance_ticker_1d)
            
            # 메모리 정리: 원본 DataFrame 삭제
            del df_new_1d_basic, df_prev_1d_all_cols
            
            # ⚠️중요: 시간 기준 정렬은 Date(UTC) 기준으로만 수행 (KST는 기록용일 뿐, 정렬에는 사용하지 않음)
            if 'Date(UTC)' in df_binance_ticker_1d.columns:
                # 정렬 직전 타입 일치 여부 최종 확인
                df_binance_ticker_1d = clean_df_display_format(df_binance_ticker_1d)
                # 정렬 실행
                df_binance_ticker_1d = df_binance_ticker_1d.sort_values('Date(UTC)', ascending=False, na_position='last').reset_index(drop=True)
            
            # 최대 개수 제한 (previous 데이터 포함)
            df_binance_ticker_1d = df_binance_ticker_1d.iloc[:daily_count + 1].reset_index(drop=True)  # +1은 새 데이터 포함
        else:
            df_binance_ticker_1d = df_binance_ticker_1d_new
    else:
        # 1단계는 이미 병렬 수집에서 처리됨
        if not skip_first_row:
            # 2단계: 순차 수집 (병렬 수집 사용 안 함)
            print(f"{get_timestamp()} [{stage_prefix}] 📥 일봉 캔들 수집 중...")
            df_binance_ticker_1d = fetch_binance_daily(binance_symbol_ticker, daily_count, include_today=include_today, fixed_end_time_ms=fixed_end_time_ms)

    # 1시간봉 데이터 수집
    if not skip_first_row and not df_prev_1h.empty:
        # 2단계: 최신 1개만 수집 (미완성 캔들 유지)
        # - 모든 시간 실행 시: 1개 수집 (미완성 캔들, 시간표시 xx시)
        # - 처리: 그대로 사용 (미완성 캔들 제거 안 함)
        # - previous와 합치기: 새 데이터(2행) + previous(3행부터)
        # 겹침 발생 시 2단계 캔들로 최신화 (drop_duplicates keep='first'로 처리)
        hour1_count_main = 1
        
        if pre_fetched_data and '1h' in pre_fetched_data and not pre_fetched_data['1h'].empty:
            df_binance_ticker_1h_new = pre_fetched_data['1h'].copy()
        else:
            # 1시간봉: fixed_end_time_ms_1h 사용 (15분 단위로 내림 처리)
            df_binance_ticker_1h_new = fetch_binance_hours1(binance_symbol_ticker, hour1_count_main, include_today=include_today, fixed_end_time_ms=fixed_end_time_ms_1h if fixed_end_time_ms_1h is not None else fixed_end_time_ms)
        
        # after 엑셀 구조: 1행 헤더, 2행 새 데이터(계산 필요), 3행부터 previous 데이터(변경 없음)
        base_cols_1h = ['Date(UTC)', 'KST', '종', '시', '고', '저', 'Vol.']
        cols_1h_new = [col for col in base_cols_1h if col in df_binance_ticker_1h_new.columns]
        
        if cols_1h_new and 'Date(UTC)' in cols_1h_new and len(df_binance_ticker_1h_new) > 0:
            df_prev_1h_all_cols = df_prev_1h.copy()
            
            # 새 데이터 처리
            # 중복 제거: Date(UTC) 컬럼 기준으로 중복 제거 (동일 시간 캔들 제거)
            if 'Date(UTC)' in df_binance_ticker_1h_new.columns:
                df_binance_ticker_1h_new = df_binance_ticker_1h_new.drop_duplicates(subset=['Date(UTC)'], keep='first')
            
            # 모든 시간 실행 시: 최신 1개만 사용 (미완성 캔들 유지)
            # - 수집: 1개
            # - 최신 1개: 미완성 캔들 (xx시 00분~xx시 15분/30분/45분/00분 1초, 시간표시 xx시)
            # - 처리: 그대로 사용 (미완성 캔들 제거 안 함)
            # - previous와 합치기: 새 데이터(2행) + previous(3행부터)
            # previous와 겹침 발생 시 2단계 캔들로 최신화됨 (drop_duplicates keep='first')
            df_new_1h_basic = df_binance_ticker_1h_new.iloc[0:1][cols_1h_new].copy()
            
            # 새 데이터에 previous와 동일한 컬럼 구조 만들기 (지표는 NaN으로)
            for col in df_prev_1h_all_cols.columns:
                if col not in df_new_1h_basic.columns:
                    df_new_1h_basic[col] = np.nan
            
            # 컬럼 순서 맞추기
            df_new_1h_basic = df_new_1h_basic[df_prev_1h_all_cols.columns]
            
            # [중요] 병합 직전 타입 강제 통일 (Timestamp와 str 혼합 방지)
            if 'Date(UTC)' in df_new_1h_basic.columns:
                df_new_1h_basic['Date(UTC)'] = pd.to_datetime(df_new_1h_basic['Date(UTC)'], errors='coerce')
            if 'Date(UTC)' in df_prev_1h_all_cols.columns:
                df_prev_1h_all_cols['Date(UTC)'] = pd.to_datetime(df_prev_1h_all_cols['Date(UTC)'], errors='coerce')
            
            # 합치기: 새 데이터(2행) + previous 데이터(3행부터) (UTC 기준으로 매칭)
            df_binance_ticker_1h = pd.concat([
                df_new_1h_basic,  # 2행: 새 데이터 (계산 필요)
                df_prev_1h_all_cols  # 3행부터: previous 데이터 (변경 없음)
            ], ignore_index=True)
            
            # [중요] 병합 직후 타입 통일 (Timestamp와 str 혼합 방지)
            df_binance_ticker_1h = clean_df_display_format(df_binance_ticker_1h)
            
            # UTC 시간 기준으로 정렬 (VLOOKUP처럼 UTC 시간으로 매칭)
            if 'Date(UTC)' in df_binance_ticker_1h.columns:
                # 정렬 직전 타입 일치 여부 최종 확인
                df_binance_ticker_1h = clean_df_display_format(df_binance_ticker_1h)
                # Date(UTC)를 datetime으로 변환하여 정렬 (UTC 기준)
                df_binance_ticker_1h['Date(UTC)_dt'] = pd.to_datetime(df_binance_ticker_1h['Date(UTC)'], format='%y/%m/%d,%H:%M', errors='coerce')
                # UTC 기준으로 정렬 (최신→과거)
                df_binance_ticker_1h = df_binance_ticker_1h.sort_values('Date(UTC)_dt', ascending=False, na_position='last').reset_index(drop=True)
                # Date(UTC) 기준 중복 제거 (keep='first' - 새 데이터 우선, UTC 시간으로 매칭)
                # 겹침 발생 시 2단계 캔들로 최신화 (새 데이터가 먼저 concat되므로 keep='first'로 새 데이터 유지)
                df_binance_ticker_1h = df_binance_ticker_1h.drop_duplicates(subset=['Date(UTC)'], keep='first').reset_index(drop=True)
                # 임시 컬럼 제거
                df_binance_ticker_1h = df_binance_ticker_1h.drop('Date(UTC)_dt', axis=1)
                # 최종 정렬 (UTC 기준, 최신→과거)
                df_binance_ticker_1h['Date(UTC)_dt'] = pd.to_datetime(df_binance_ticker_1h['Date(UTC)'], format='%y/%m/%d,%H:%M', errors='coerce')
                df_binance_ticker_1h = df_binance_ticker_1h.sort_values('Date(UTC)_dt', ascending=False, na_position='last').reset_index(drop=True)
                df_binance_ticker_1h = df_binance_ticker_1h.drop('Date(UTC)_dt', axis=1)
        else:
            df_binance_ticker_1h = df_binance_ticker_1h_new
    else:
        # 1단계는 이미 병렬 수집에서 처리됨
        if not skip_first_row:
            # 2단계: 순차 수집 (병렬 수집 사용 안 함)
            print(f"{get_timestamp()} [{stage_prefix}] 📥 1시간봉 캔들 수집 중...")
            # 1시간봉: fixed_end_time_ms_1h 사용 (15분 단위로 내림 처리)
            df_binance_ticker_1h = fetch_binance_hours1(binance_symbol_ticker, hour1_count, include_today=include_today, fixed_end_time_ms=fixed_end_time_ms_1h if fixed_end_time_ms_1h is not None else fixed_end_time_ms)

    # 주봉 데이터 생성 (일봉에서 변환 - API 호출 최적화)
    # 일봉 200개면 주봉 약 28개 생성 가능 (200일 ÷ 7일 ≈ 28주)
    df_binance_ticker_weekly = convert_daily_to_weekly(df_binance_ticker_1d)
    
    # 주봉은 일봉에서 변환하므로 별도 API 호출 불필요
    # 일봉 200개면 충분히 주봉 28개를 만들 수 있음
    
    # 15분봉 최신 종가를 1시간봉/일봉/주봉 최신 행에 주입 (지표 계산 일관성 확보)
    latest_close_15m = _extract_latest_close_from_15m(df_binance_ticker_15m)
    if latest_close_15m is not None:
        df_binance_ticker_1h = _override_latest_close(df_binance_ticker_1h, latest_close_15m, f"{TICKER}USDT1H", stage_prefix)
        df_binance_ticker_1d = _override_latest_close(df_binance_ticker_1d, latest_close_15m, f"{TICKER}USDT1D", stage_prefix)
        df_binance_ticker_weekly = _override_latest_close(df_binance_ticker_weekly, latest_close_15m, f"{TICKER}USDTW", stage_prefix)
    
    # 업비트 API/김프 사용 제거됨: 바이낸스 데이터만 사용

    # API 수집 단계에서 이미 미완성 캔들을 제거했으므로 추가 제거 불필요
    # 1단계(previous): API 수집 시 미완성 캔들 1개 제거됨
    # 2단계(after): API 수집 시 미완성 캔들 1개 제거됨

    # 1분봉 지표 계산
    if not df_binance_ticker_1m.empty:
        if skip_first_row:
            # 1단계: 전체 계산
            print(f"{get_timestamp()} [{stage_prefix}] 📊지표 계산 중")
            df_binance_ticker_1m = calculate_all_indicators_1m(df_binance_ticker_1m, "USD")
        else:
            # 2단계: 최신 행만 계산 (previous 지표 유지) - calculate_latest_row_only_1m 사용
            df_binance_ticker_1m = calculate_latest_row_only_1m(df_binance_ticker_1m, "USD")

    # 일봉 지표 계산
    if skip_first_row:
        # 1단계: 전체 계산
        df_binance_ticker_1d = calculate_all_indicators(df_binance_ticker_1d, "USD")
    else:
        # 2단계: 2행(인덱스 0)만 계산
        df_binance_ticker_1d = calculate_latest_row_only_1d(df_binance_ticker_1d, "USD")
    
    # 5분봉 지표 계산
    if skip_first_row:
        # 1단계: 전체 계산
        df_binance_ticker_5m = calculate_all_indicators_5m(df_binance_ticker_5m, "USD")
    else:
        # 2단계: 2-4행(인덱스 0-2)만 계산
        df_binance_ticker_5m = calculate_latest_3rows_only_5m(df_binance_ticker_5m, "USD")
    
    # SB1M 열 계산 (5분봉 시트에만) - 1분봉 데이터 기반
    if not df_binance_ticker_1m.empty:
        if skip_first_row:
            # 1단계: 전체 계산
            df_binance_ticker_5m = calculate_sb1m_for_5m(df_binance_ticker_5m, df_binance_ticker_1m)
        else:
            # 2단계: 2-4행(인덱스 0-2) 3개 계산
            if len(df_binance_ticker_5m) > 0:
                # 최신 3개 행 사용
                df_5m_temp = df_binance_ticker_5m.iloc[0:3].copy()
                
                try:
                    if 'Date(UTC)' in df_5m_temp.columns and 'Date(UTC)' in df_binance_ticker_1m.columns:
                        # 5분봉 날짜 파싱 (이미 datetime64면 그대로 사용, 문자열만 파싱)
                        if pd.api.types.is_datetime64_any_dtype(df_5m_temp['Date(UTC)']):
                            df_5m_temp['Date(UTC)_dt'] = df_5m_temp['Date(UTC)']
                        else:
                            # 문자열인 경우에만 파싱 (UserWarning 억제)
                            s = df_5m_temp['Date(UTC)'].astype(str).str.strip().str.replace(',', ' ', regex=False)
                            import warnings
                            with warnings.catch_warnings():
                                warnings.simplefilter("ignore", UserWarning)
                                df_5m_temp['Date(UTC)_dt'] = pd.to_datetime(s, errors='coerce')
                        
                        df_1m_copy = df_binance_ticker_1m.copy()
                        # 1분봉 날짜 파싱 (이미 datetime64면 그대로 사용, 문자열만 파싱)
                        if pd.api.types.is_datetime64_any_dtype(df_1m_copy['Date(UTC)']):
                            df_1m_copy['Date(UTC)_dt'] = df_1m_copy['Date(UTC)']
                        else:
                            # 문자열인 경우에만 파싱 (UserWarning 억제)
                            s = df_1m_copy['Date(UTC)'].astype(str).str.strip().str.replace(',', ' ', regex=False)
                            with warnings.catch_warnings():
                                warnings.simplefilter("ignore", UserWarning)
                                df_1m_copy['Date(UTC)_dt'] = pd.to_datetime(s, errors='coerce')
                        
                        # 5분 그룹핑
                        df_1m_copy['5min_group'] = df_1m_copy['Date(UTC)_dt'].dt.floor('5min')
                        
                        # 각 5분봉 행에 대해 계산
                        for idx in range(len(df_5m_temp)):
                            row = df_5m_temp.iloc[idx]
                            if pd.notna(row['Date(UTC)_dt']):
                                target_5min_group = row['Date(UTC)_dt'].floor('5min')
                                
                                # 해당 그룹의 1분봉 필터링
                                df_1m_group = df_1m_copy[df_1m_copy['5min_group'] == target_5min_group].copy()
                                
                                # 개수 검증 및 경고
                                if len(df_1m_group) < 5:
                                    print(f"{get_timestamp()} ⚠️SB1M 시간 매칭 경고: 5분봉 {target_5min_group}에 해당하는 1분봉이 {len(df_1m_group)}개 (예상: 5개)")
                                
                                # 1개라도 있으면 계산 시도
                                if not df_1m_group.empty:
                                    # 단일 행 계산을 위해 함수 호출
                                    res_df = calculate_sb1m_for_5m(df_5m_temp.iloc[[idx]], df_1m_group)
                                    sb1m_val = res_df.iloc[0].get('SB1M', '')
                                    
                                    # 원본에 반영
                                    original_idx = df_5m_temp.index[idx]
                                    if 'SB1M' not in df_binance_ticker_5m.columns:
                                        df_binance_ticker_5m['SB1M'] = ''
                                    df_binance_ticker_5m['SB1M'] = df_binance_ticker_5m['SB1M'].astype('object')
                                    df_binance_ticker_5m.loc[original_idx, 'SB1M'] = sb1m_val

                except Exception as e:
                    print(f"{get_timestamp()} ⚠️SB1M 2단계 계산 중 오류: {e}")
                    import traceback
                    traceback.print_exc()
    
    # 15분봉 지표 계산
    if skip_first_row:
        # 1단계: 전체 계산
        df_binance_ticker_15m = calculate_all_indicators_15m(df_binance_ticker_15m, "USD")
    else:
        # 2단계: 최신 1개만 계산 (previous 지표 유지)
        print(f"{get_timestamp()} [{stage_prefix}]    → 2행(최신)만 계산, 3행 이후는 previous 유지")
        # 2단계일 때 previous 데이터 백업 (3행 이후 복원용)
        if not df_prev_15m.empty and len(df_binance_ticker_15m) > 1:
            df_prev_15m_backup = df_prev_15m.copy()  # previous 데이터 백업 (3행~)
        df_binance_ticker_15m = calculate_latest_row_only_15m(df_binance_ticker_15m, "USD")
    
    # SB1M 열 계산 (15분봉 시트에만) - 1분봉 데이터 기반
    if not df_binance_ticker_1m.empty:
        if skip_first_row:
            # 1단계: 전체 계산
            df_binance_ticker_15m = calculate_sb1m_for_15m(df_binance_ticker_15m, df_binance_ticker_1m)
        else:
            # 2단계: 최신 1개 행만 계산
            if len(df_binance_ticker_15m) > 0 and len(df_binance_ticker_1m) > 0:
                try:
                    if 'Date(UTC)' in df_binance_ticker_15m.columns and 'Date(UTC)' in df_binance_ticker_1m.columns:
                        # 최신 1개 행 사용
                        df_15m_temp = df_binance_ticker_15m.iloc[0:1].copy()
                        
                        # 날짜 파싱 (이미 datetime64면 그대로 사용, 문자열만 파싱)
                        if pd.api.types.is_datetime64_any_dtype(df_15m_temp['Date(UTC)']):
                            df_15m_temp['Date(UTC)_dt'] = df_15m_temp['Date(UTC)']
                        else:
                            # 문자열인 경우에만 파싱 (UserWarning 억제)
                            s = df_15m_temp['Date(UTC)'].astype(str).str.strip().str.replace(',', ' ', regex=False)
                            import warnings
                            with warnings.catch_warnings():
                                warnings.simplefilter("ignore", UserWarning)
                                df_15m_temp['Date(UTC)_dt'] = pd.to_datetime(s, errors='coerce')
                        
                        df_1m_copy = df_binance_ticker_1m.copy()
                        # 1분봉 날짜 파싱 (이미 datetime64면 그대로 사용, 문자열만 파싱)
                        if pd.api.types.is_datetime64_any_dtype(df_1m_copy['Date(UTC)']):
                            df_1m_copy['Date(UTC)_dt'] = df_1m_copy['Date(UTC)']
                        else:
                            # 문자열인 경우에만 파싱 (UserWarning 억제)
                            s = df_1m_copy['Date(UTC)'].astype(str).str.strip().str.replace(',', ' ', regex=False)
                            with warnings.catch_warnings():
                                warnings.simplefilter("ignore", UserWarning)
                                df_1m_copy['Date(UTC)_dt'] = pd.to_datetime(s, errors='coerce')
                        
                        # 15분 그룹핑
                        df_1m_copy['15min_group'] = df_1m_copy['Date(UTC)_dt'].dt.floor('15min')
                        
                        # 각 15분봉 행에 대해 계산
                        for idx in range(len(df_15m_temp)):
                            row = df_15m_temp.iloc[idx]
                            if pd.notna(row['Date(UTC)_dt']):
                                target_15min_group = row['Date(UTC)_dt'].floor('15min')
                                
                                # 해당 그룹의 1분봉 필터링
                                df_1m_group = df_1m_copy[df_1m_copy['15min_group'] == target_15min_group].copy()
                                
                                # 1개라도 있으면 계산 시도
                                if not df_1m_group.empty:
                                    # 단일 행 계산을 위해 함수 호출
                                    res_df = calculate_sb1m_for_15m(df_15m_temp.iloc[[idx]], df_1m_group)
                                    sb1m_val = res_df.iloc[0].get('SB1M', '')
                                    
                                    # 원본에 반영
                                    original_idx = df_15m_temp.index[idx]
                                    if 'SB1M' not in df_binance_ticker_15m.columns:
                                        df_binance_ticker_15m['SB1M'] = ''
                                    df_binance_ticker_15m['SB1M'] = df_binance_ticker_15m['SB1M'].astype('object')
                                    df_binance_ticker_15m.loc[original_idx, 'SB1M'] = sb1m_val
                except Exception as e:
                    print(f"{get_timestamp()} ⚠️SB1M 15분봉 2단계 계산 중 오류: {e}")
                    import traceback
                    traceback.print_exc()
    
    # 5분봉에 15분봉의 1HMSFast 값을 시간 매칭하여 복사 (1HMSF 열 추가)
    if not df_binance_ticker_5m.empty and not df_binance_ticker_15m.empty:
        df_binance_ticker_5m = copy_1hmsfast_to_5m(df_binance_ticker_5m, df_binance_ticker_15m)
    
    # 5분봉 Buy 재계산 (1HMSF 복사 후, gear1/gear2 조건 적용)
    if not df_binance_ticker_5m.empty:
        df_binance_ticker_5m = recalculate_buy_for_5m(df_binance_ticker_5m)
    
    # 1시간봉 지표 계산
    if skip_first_row:
        # 1단계: 전체 계산
        df_binance_ticker_1h = calculate_all_indicators_1h(df_binance_ticker_1h, "USD")
    else:
        # 2단계: 2행(인덱스 0)만 계산
        df_binance_ticker_1h = calculate_latest_row_only_1h(df_binance_ticker_1h, "USD")
    
    # 주봉 지표 계산 (주봉은 28개로 고정이므로 항상 전체 계산)
    df_binance_ticker_weekly = calculate_all_indicators_weekly(df_binance_ticker_weekly, "USD")
    
    # 1H4x 시트 생성 (15분봉 데이터에서 기본 컬럼만 추출)
    base_cols_1h4x = ['Date(UTC)', 'KST', '종', '시', '고', '저', 'Vol.']
    if not skip_first_row and not df_prev_15m.empty:
        # 2단계: previous 파일에서 1H4x 시트 읽기 (이미 위에서 읽었으므로 df_prev_1h4x 사용)
        # df_prev_1h4x는 이미 위에서 읽었음 (7627-7637번 줄)
        
        # 2단계: 15분봉 데이터에서 기본 컬럼만 가져와서 previous와 병합
        cols_1h4x_new = [col for col in base_cols_1h4x if col in df_binance_ticker_15m.columns]
        
        if cols_1h4x_new and 'Date(UTC)' in cols_1h4x_new and len(df_binance_ticker_15m) > 0:
            # previous 데이터의 모든 컬럼 유지 (지표 포함)
            df_prev_1h4x_all_cols = df_prev_1h4x.copy() if not df_prev_1h4x.empty else pd.DataFrame()
            
            # 새 데이터는 최신 1개만 사용 (2행에 배치할 데이터)
            df_new_1h4x_basic = df_binance_ticker_15m.iloc[0:1][cols_1h4x_new].copy()
            
            # 새 데이터에 previous와 동일한 컬럼 구조 만들기 (지표는 NaN으로)
            if not df_prev_1h4x_all_cols.empty:
                for col in df_prev_1h4x_all_cols.columns:
                    if col not in df_new_1h4x_basic.columns:
                        df_new_1h4x_basic[col] = np.nan
                
                # 컬럼 순서 맞추기
                df_new_1h4x_basic = df_new_1h4x_basic[df_prev_1h4x_all_cols.columns]
            
            # ⚠️중요: 중복 제거는 Date(UTC) 기준으로만 수행
            if 'Date(UTC)' in df_new_1h4x_basic.columns and 'Date(UTC)' in df_prev_1h4x_all_cols.columns:
                new_date_utc = df_new_1h4x_basic.iloc[0]['Date(UTC)'] if len(df_new_1h4x_basic) > 0 and pd.notna(df_new_1h4x_basic.iloc[0]['Date(UTC)']) else None
                if new_date_utc and str(new_date_utc).strip() != '':
                    df_prev_1h4x_all_cols = df_prev_1h4x_all_cols[
                        df_prev_1h4x_all_cols['Date(UTC)'].astype(str).str.strip() != str(new_date_utc).strip()
                    ].copy()
            
            # [중요] 병합 직전 타입 강제 통일 (Timestamp와 str 혼합 방지)
            if 'Date(UTC)' in df_new_1h4x_basic.columns:
                df_new_1h4x_basic['Date(UTC)'] = pd.to_datetime(df_new_1h4x_basic['Date(UTC)'], errors='coerce')
            if 'Date(UTC)' in df_prev_1h4x_all_cols.columns:
                df_prev_1h4x_all_cols['Date(UTC)'] = pd.to_datetime(df_prev_1h4x_all_cols['Date(UTC)'], errors='coerce')
            
            # 합치기: 새 데이터(2행) + previous 데이터(3행부터)
            df_binance_ticker_1h4x = pd.concat([
                df_new_1h4x_basic,      # 새 데이터
                df_prev_1h4x_all_cols   # previous 데이터
            ], ignore_index=True)
            
            # [중요] 병합 직후 타입 통일 (Timestamp와 str 혼합 방지)
            df_binance_ticker_1h4x = clean_df_display_format(df_binance_ticker_1h4x)
            
            # 메모리 정리
            del df_new_1h4x_basic, df_prev_1h4x_all_cols
            
            # ⚠️중요: 시간 기준 정렬은 Date(UTC) 기준으로만 수행 (KST는 기록용일 뿐, 정렬에는 사용하지 않음)
            if not df_binance_ticker_1h4x.empty:
                # 정렬 전 타입 일치 여부 최종 확인
                df_binance_ticker_1h4x = clean_df_display_format(df_binance_ticker_1h4x)
                # 정렬 실행
                df_binance_ticker_1h4x = df_binance_ticker_1h4x.sort_values('Date(UTC)', ascending=False, na_position='last').reset_index(drop=True)
            
            # 최대 개수 제한 (previous 데이터 포함)
            df_binance_ticker_1h4x = df_binance_ticker_1h4x.iloc[:400].reset_index(drop=True)  # 최종 400개
        else:
            # previous 파일이 없는 경우: 15분봉 데이터만 사용
            df_binance_ticker_1h4x = df_new_1h4x_basic if 'df_new_1h4x_basic' in locals() else pd.DataFrame()
    else:
        # 1단계 또는 previous 파일이 없는 경우: 15분봉 데이터에서 기본 컬럼만 복사
        if skip_first_row:
            # 1단계: 15분봉 데이터에서 1601개 가져오기 (COLLECTION_COUNT['15m'])
            if len(df_binance_ticker_15m) >= COLLECTION_COUNT['15m']:
                df_binance_ticker_1h4x = df_binance_ticker_15m[base_cols_1h4x].iloc[:COLLECTION_COUNT['15m']].copy()
            else:
                df_binance_ticker_1h4x = df_binance_ticker_15m[base_cols_1h4x].copy()
        else:
            # previous 파일이 없는 경우: 15분봉 데이터에서 최대 400개
            if len(df_binance_ticker_15m) > 0:
                df_binance_ticker_1h4x = df_binance_ticker_15m[base_cols_1h4x].iloc[:400].copy()
            else:
                df_binance_ticker_1h4x = pd.DataFrame()
    
    # 1H4x 시트용 지표 계산
    if not df_binance_ticker_1h4x.empty:
        if skip_first_row:
            # 1단계: 전체 계산 후 최종 400개로 제한
            df_binance_ticker_1h4x = calculate_all_indicators_1h4x(df_binance_ticker_1h4x, "USD")
            # 지표 계산 후 최종 400개로 제한
            df_binance_ticker_1h4x = df_binance_ticker_1h4x.iloc[:400].reset_index(drop=True)
        else:
            # 2단계: 최신 1개만 계산 (previous 지표 유지)
            if len(df_binance_ticker_1h4x) > 0:
                df_binance_ticker_1h4x = calculate_latest_row_only_1h4x(df_binance_ticker_1h4x, "USD")
    
    # 15분봉에 새로운 열들 추가 (1H4x 시트에서 SB1H 값 매칭)
    if skip_first_row:
        # 1단계: 전체 계산
        df_binance_ticker_15m = calculate_sb1h_for_15m(df_binance_ticker_15m, df_binance_ticker_1h4x)
        
        # SB1D 계산 (일봉에서 15분봉으로)
        if not df_binance_ticker_1d.empty:
            df_binance_ticker_15m = calculate_daysb_15m(df_binance_ticker_15m, df_binance_ticker_1d, "USD")
        
        # SB5M 계산 (5분봉에서 15분봉으로, 개수 카운팅)
        if not df_binance_ticker_5m.empty:
            df_binance_ticker_15m = calculate_sb5m_for_15m(df_binance_ticker_15m, df_binance_ticker_5m)
        
        # 1HCL, -1HCL 복사 (1시간봉에서 15분봉으로)
        if not df_binance_ticker_1h.empty:
            df_binance_ticker_15m = copy_1hclass_to_15m(df_binance_ticker_15m, df_binance_ticker_1h)
            df_binance_ticker_15m = copy_minus_1hclass_to_15m(df_binance_ticker_15m, df_binance_ticker_1h)
            df_binance_ticker_15m = copy_p1h_to_15m_and_set_p(df_binance_ticker_15m, df_binance_ticker_1h)
    else:
        # ⚠️중요: 2단계 계산은 Date(UTC) 기준으로만 수행 (KST는 기록용일 뿐, 필터링/매칭에는 사용하지 않음)
        if len(df_binance_ticker_15m) > 0:
            # Date(UTC) 컬럼 정규화 (Timestamp와 문자열 혼합 방지)
            if 'Date(UTC)' in df_binance_ticker_15m.columns:
                # 이미 datetime 타입이면 그대로 유지, 문자열인 경우만 변환
                if not pd.api.types.is_datetime64_any_dtype(df_binance_ticker_15m['Date(UTC)']):
                    # format 명시하지 않고 자동 인식 (4자리 연도도 파싱 가능)
                    df_binance_ticker_15m['Date(UTC)'] = pd.to_datetime(df_binance_ticker_15m['Date(UTC)'], errors='coerce')
            
            # SB1H 컬럼이 없으면 생성 (object 타입으로 명시)
            if 'SB1H' not in df_binance_ticker_15m.columns:
                df_binance_ticker_15m['SB1H'] = ''
                df_binance_ticker_15m['SB1H'] = df_binance_ticker_15m['SB1H'].astype('object')
            
            # 최신 Date(UTC) 시간으로 필터링 (VLOOKUP 방식)
            if 'Date(UTC)' in df_binance_ticker_15m.columns:
                # 최신 Date(UTC) 시간 사용 (이미 정렬되어 있으므로 iloc[0] 사용)
                df_15m_temp = df_binance_ticker_15m.iloc[0:1].copy()
            else:
                df_15m_temp = df_binance_ticker_15m.iloc[0:1].copy()
            
            df_15m_temp = calculate_sb1h_for_15m(df_15m_temp, df_binance_ticker_1h4x)
            
            # SB1D 계산
            if not df_binance_ticker_1d.empty:
                df_15m_temp = calculate_daysb_15m(df_15m_temp, df_binance_ticker_1d, "USD")
            
            # SB5M 계산 (5분봉 전체 사용)
            if not df_binance_ticker_5m.empty:
                df_15m_temp = calculate_sb5m_for_15m(df_15m_temp, df_binance_ticker_5m)
            
            # Date(UTC) 시간으로 매칭하여 업데이트
            if len(df_15m_temp) > 0 and 'Date(UTC)' in df_15m_temp.columns:
                target_date_utc = df_15m_temp.iloc[0]['Date(UTC)']
                mask = df_binance_ticker_15m['Date(UTC)'] == target_date_utc
                if mask.any():
                    # SB1H 업데이트
                    val = df_15m_temp.iloc[0].get('SB1H', np.nan)
                    df_binance_ticker_15m['SB1H'] = df_binance_ticker_15m['SB1H'].astype('object')
                    df_binance_ticker_15m.loc[mask, 'SB1H'] = np.nan if (pd.isna(val) or val == '') else val
                    
                    # SB1D 업데이트
                    if 'SB1D' in df_15m_temp.columns:
                        val_sb1d = df_15m_temp.iloc[0].get('SB1D', np.nan)
                        if 'SB1D' not in df_binance_ticker_15m.columns:
                            df_binance_ticker_15m['SB1D'] = ''
                        df_binance_ticker_15m['SB1D'] = df_binance_ticker_15m['SB1D'].astype('object')
                        df_binance_ticker_15m.loc[mask, 'SB1D'] = np.nan if (pd.isna(val_sb1d) or val_sb1d == '') else val_sb1d
                    
                    # SB5M 업데이트
                    if 'SB5M' in df_15m_temp.columns:
                        val_sb5m = df_15m_temp.iloc[0].get('SB5M', '')
                        if 'SB5M' not in df_binance_ticker_15m.columns:
                            df_binance_ticker_15m['SB5M'] = ''
                        df_binance_ticker_15m['SB5M'] = df_binance_ticker_15m['SB5M'].astype('object')
                        df_binance_ticker_15m.loc[mask, 'SB5M'] = '' if (pd.isna(val_sb5m) or val_sb5m == '') else val_sb5m
            else:
                # SB1H 업데이트
                val = df_15m_temp.iloc[0].get('SB1H', np.nan) if len(df_15m_temp) > 0 else np.nan
                df_binance_ticker_15m['SB1H'] = df_binance_ticker_15m['SB1H'].astype('object')
                df_binance_ticker_15m.loc[0, 'SB1H'] = np.nan if (pd.isna(val) or val == '') else val
    
                # SB1D 업데이트
                if 'SB1D' in df_15m_temp.columns:
                    val_sb1d = df_15m_temp.iloc[0].get('SB1D', np.nan) if len(df_15m_temp) > 0 else np.nan
            if 'SB1D' not in df_binance_ticker_15m.columns:
                df_binance_ticker_15m['SB1D'] = ''
                df_binance_ticker_15m['SB1D'] = df_binance_ticker_15m['SB1D'].astype('object')
                df_binance_ticker_15m.loc[0, 'SB1D'] = np.nan if (pd.isna(val_sb1d) or val_sb1d == '') else val_sb1d
                
                # SB5M 업데이트
                if 'SB5M' in df_15m_temp.columns:
                    val_sb5m = df_15m_temp.iloc[0].get('SB5M', '') if len(df_15m_temp) > 0 else ''
                    if 'SB5M' not in df_binance_ticker_15m.columns:
                        df_binance_ticker_15m['SB5M'] = ''
                    df_binance_ticker_15m['SB5M'] = df_binance_ticker_15m['SB5M'].astype('object')
                    df_binance_ticker_15m.loc[0, 'SB5M'] = '' if (pd.isna(val_sb5m) or val_sb5m == '') else val_sb5m
    
    # 15분봉의 1HMSFast는 calculate_all_indicators_15m 또는 calculate_latest_row_only_15m에서 직접 계산됨
    
    # Samount, Bamount 계산 (1단계 vs 2단계 분기, UTC 시간 기준으로 매칭, VLOOKUP 방식)
    if skip_first_row:
        # 1단계: 전체 계산 (UTC 시간 기준으로 매칭)
    # 15분봉에 주봉 SamountW, BamountW 추가 (바이낸스 주봉 기준, UTC 시간 기준으로 매칭)
        df_binance_ticker_15m = copy_weekly_amounts_to_15m(df_binance_ticker_15m, df_binance_ticker_weekly)
    
    # 15분봉에 일봉 Samount1D, Bamount1D 추가 (바이낸스 일봉 기준, UTC 시간 기준으로 매칭)
        df_binance_ticker_15m = copy_daily_amounts_to_15m(df_binance_ticker_15m, df_binance_ticker_1d)
    
    # 15분봉에 최종 Samount, Bamount 계산
        df_binance_ticker_15m = calculate_final_amounts(df_binance_ticker_15m)
    else:
        # 2단계: 최신 1개만 계산 (previous에서 SamountW/BamountW/Samount1D/Bamount1D 유지)
        # Previous에서 가져온 SamountW/BamountW/Samount1D/Bamount1D가 이미 있음
        # 최신 1개(인덱스 0)만 계산
        if len(df_binance_ticker_15m) > 0:
            row_0 = df_binance_ticker_15m.iloc[0]
            samountW = row_0.get("SamountW", np.nan)
            bamountW = row_0.get("BamountW", np.nan)
            samount1D = row_0.get("Samount1D", np.nan)
            bamount1D = row_0.get("Bamount1D", np.nan)
            
            # SamountW/BamountW가 없으면 일봉/주봉에서 가져오기
            # ⚠️중요: 주봉의 SamountW, BamountW는 buyside/sellside가 실시간으로 변하므로
            # 매번 최신 값을 가져와야 함 (pd.isna 조건 제거)
            if len(df_binance_ticker_weekly) > 0:
                samountW = df_binance_ticker_weekly.iloc[0].get("SamountW", np.nan)
                bamountW = df_binance_ticker_weekly.iloc[0].get("BamountW", np.nan)
                df_binance_ticker_15m.loc[0, "SamountW"] = samountW
                df_binance_ticker_15m.loc[0, "BamountW"] = bamountW
            
            # ⚠️중요: 일봉의 Samount1D, Bamount1D는 buyside/sellside가 실시간으로 변하므로
            # 매번 최신 값을 가져와야 함 (pd.isna 조건 제거)
            if len(df_binance_ticker_1d) > 0:
                samount1D = df_binance_ticker_1d.iloc[0].get("Samount1D", np.nan)
                bamount1D = df_binance_ticker_1d.iloc[0].get("Bamount1D", np.nan)
                df_binance_ticker_15m.loc[0, "Samount1D"] = samount1D
                df_binance_ticker_15m.loc[0, "Bamount1D"] = bamount1D
            
            # 최종 Samount, Bamount 계산 (2행만)
            if not pd.isna(samountW) and not pd.isna(samount1D):
                df_binance_ticker_15m.loc[0, "Samount"] = 0.7 * samountW + 0.3 * samount1D
            else:
                df_binance_ticker_15m.loc[0, "Samount"] = np.nan
            
            if not pd.isna(bamountW) and not pd.isna(bamount1D):
                df_binance_ticker_15m.loc[0, "Bamount"] = 0.7 * bamountW + 0.3 * bamount1D
            else:
                df_binance_ticker_15m.loc[0, "Bamount"] = np.nan
    
    # dateM, LD 계산 (1단계 vs 2단계 분기)
    # 2단계일 때 previous 데이터 백업 (3행 이후 복원용)
    if not skip_first_row and not df_prev_15m.empty:
        df_prev_15m_backup = df_prev_15m.copy()  # previous 데이터 백업 (3행~)
    
    # dateM 열 계산 (15분봉 시트에만) - ORDER 계산 전에 먼저 실행
    if skip_first_row:
        # 1단계: 전체 계산
        df_binance_ticker_15m = calculate_dateM(df_binance_ticker_15m)
    else:
        # 2단계: 2행(인덱스 0)만 계산
        df_binance_ticker_15m = calculate_latest_row_only_dateM(df_binance_ticker_15m)
    
    # LD 열 계산 (15분봉 시트에만) - ORDER 계산 전에 먼저 실행
    if skip_first_row:
        # 1단계: 전체 계산
        df_binance_ticker_15m = calculate_LD(df_binance_ticker_15m)
    else:
        # 2단계: 2행(인덱스 0)만 계산
        df_binance_ticker_15m = calculate_latest_row_only_LD(df_binance_ticker_15m)
    
    # SB5M 열 계산 (15분봉 시트에만) - 5분봉 데이터 기반 (UTC 시간 기준으로 매칭, VLOOKUP 방식)
    if skip_first_row:
        # 1단계: 전체 계산 (UTC 시간 기준으로 매칭)
        df_binance_ticker_15m = calculate_sb5m_for_15m(df_binance_ticker_15m, df_binance_ticker_5m)
    else:
        # 2단계: 최신 행(인덱스 0, UTC 기준으로 정렬된 상태에서 최신)만 계산
        if len(df_binance_ticker_15m) > 0:
            # SB5M 컬럼이 없으면 생성 (object 타입으로 명시)
            if 'SB5M' not in df_binance_ticker_15m.columns:
                df_binance_ticker_15m['SB5M'] = ''
                df_binance_ticker_15m['SB5M'] = df_binance_ticker_15m['SB5M'].astype('object')
            
            # UTC 기준으로 정렬된 상태에서 최신 데이터 가져오기 (인덱스 0 = UTC 기준 최신)
            df_15m_temp = df_binance_ticker_15m.iloc[0:1].copy()
            
            try:
                # Date(UTC) 기준으로 그룹화 (바이낸스는 UTC 기준)
                if 'Date(UTC)' in df_15m_temp.columns:
                    if df_15m_temp['Date(UTC)'].dtype == 'object':
                        df_15m_temp['Date(UTC)'] = pd.to_datetime(df_15m_temp['Date(UTC)'], format='%y/%m/%d,%H:%M', errors='coerce')
                    else:
                        # 이미 datetime 타입이면 format 명시하여 재변환 (일관성 보장)
                        try:
                            df_15m_temp['Date(UTC)'] = pd.to_datetime(df_15m_temp['Date(UTC)'], format='%y/%m/%d,%H:%M', errors='coerce')
                        except:
                            import warnings
                            with warnings.catch_warnings():
                                warnings.simplefilter("ignore", UserWarning)
                        df_15m_temp['Date(UTC)'] = pd.to_datetime(df_15m_temp['Date(UTC)'], errors='coerce')
                    
                    df_15m_temp['15min_group'] = df_15m_temp['Date(UTC)'].dt.floor('15min')
                    target_15min_group = df_15m_temp.iloc[0]['15min_group']
                    
                    # 전체 5분봉 데이터에서 해당 15분 그룹에 속하는 5분봉만 필터링
                    df_5m_copy = df_binance_ticker_5m.copy()
                    if 'Date(UTC)' in df_5m_copy.columns:
                        # 이미 datetime64면 그대로 사용, 문자열만 파싱
                        if pd.api.types.is_datetime64_any_dtype(df_5m_copy['Date(UTC)']):
                            pass  # 이미 Timestamp 객체이면 그대로 사용
                        else:
                            # 문자열인 경우에만 파싱 (UserWarning 억제)
                            s = df_5m_copy['Date(UTC)'].astype(str).str.strip().str.replace(',', ' ', regex=False)
                            import warnings
                            with warnings.catch_warnings():
                                warnings.simplefilter("ignore", UserWarning)
                                df_5m_copy['Date(UTC)'] = pd.to_datetime(s, errors='coerce')
                        
                        df_5m_copy['15min_group'] = df_5m_copy['Date(UTC)'].dt.floor('15min')
                        
                        # 해당 15분 그룹에 속하는 5분봉만 필터링
                        df_5m_temp = df_5m_copy[df_5m_copy['15min_group'] == target_15min_group].copy()
                        
                        # 시간 순서대로 정렬 (과거→현재)
                        df_5m_temp = df_5m_temp.sort_values('Date(UTC)', ascending=True).reset_index(drop=True)
                        
                        # df_15m_temp의 15min_group을 유지 (함수 내부에서 재계산 방지)
                        # 함수 내부에서 df_15m_temp의 15min_group을 사용하도록 보장
                        # 임시 컬럼 제거 (df_5m_temp만)
                        if '15min_group' in df_5m_temp.columns:
                            df_5m_temp = df_5m_temp.drop('15min_group', axis=1)
                    else:
                        df_5m_temp = df_binance_ticker_5m.iloc[0:3].copy() if len(df_binance_ticker_5m) >= 3 else df_binance_ticker_5m.copy()
                else:
                    df_5m_temp = df_binance_ticker_5m.iloc[0:3].copy() if len(df_binance_ticker_5m) >= 3 else df_binance_ticker_5m.copy()
                
                # 검증: 해당 15분 그룹에 속하는 5분봉이 3개인지 확인
                if len(df_5m_temp) != 3:
                    print(f"{get_timestamp()} [{stage_prefix}] ⚠️ SB5M 시간 매칭 경고: 15분봉 {target_15min_group}에 해당하는 5분봉이 {len(df_5m_temp)}개 (예상: 3개)")
            except Exception as e:
                # 오류 발생 시 기존 방식 사용 (최신 3개)
                print(f"{get_timestamp()} [{stage_prefix}] ⚠️ SB5M 시간 매칭 실패, 최신 3개 사용: {e}")
                df_5m_temp = df_binance_ticker_5m.iloc[0:3].copy() if len(df_binance_ticker_5m) >= 3 else df_binance_ticker_5m.copy()
            
            # df_15m_temp의 15min_group은 유지 (함수에서 사용)
            df_15m_temp = calculate_sb5m_for_15m(df_15m_temp, df_5m_temp)
            
            # 함수 호출 후 임시 컬럼 제거
            for col in ['15min_group']:
                if col in df_15m_temp.columns:
                    df_15m_temp = df_15m_temp.drop(col, axis=1)
            
            sb5m_value = df_15m_temp.iloc[0].get('SB5M', np.nan)
            # dtype 호환성을 위해 object 타입으로 변환
            if df_binance_ticker_15m['SB5M'].dtype != 'object':
                df_binance_ticker_15m['SB5M'] = df_binance_ticker_15m['SB5M'].astype('object')
            # 빈 문자열은 np.nan으로 변환
            if pd.isna(sb5m_value) or sb5m_value == '':
                sb5m_value = np.nan
            df_binance_ticker_15m.loc[0, 'SB5M'] = sb5m_value
    
    # ORDER 열 계산 (15분봉 시트에만) - LD 계산 후 실행
    if skip_first_row:
        # 1단계: 전체 계산
        df_binance_ticker_15m = calculate_order_column(df_binance_ticker_15m, f"{TICKER}USDT15M")
    else:
        # 2단계: 2행(인덱스 0)만 계산
        if len(df_binance_ticker_15m) > 0:
            # ORDER 컬럼이 없으면 생성 (object 타입으로 명시)
            if 'ORDER' not in df_binance_ticker_15m.columns:
                df_binance_ticker_15m['ORDER'] = ''
                df_binance_ticker_15m['ORDER'] = df_binance_ticker_15m['ORDER'].astype('object')
            
            df_15m_temp = df_binance_ticker_15m.iloc[0:1].copy()
            df_15m_temp = calculate_order_column(df_15m_temp, f"{TICKER}USDT15M")
            order_value = df_15m_temp.iloc[0].get('ORDER', '')
            # dtype 호환성을 위해 object 타입으로 변환
            if df_binance_ticker_15m['ORDER'].dtype != 'object':
                df_binance_ticker_15m['ORDER'] = df_binance_ticker_15m['ORDER'].astype('object')
            df_binance_ticker_15m.loc[0, 'ORDER'] = order_value
    
    # KSC 열 계산 (15분봉 시트에만) - ORDER 계산 후 실행
    if skip_first_row:
        # 1단계: 전체 계산
        df_binance_ticker_15m = calculate_ksc_for_15m(df_binance_ticker_15m)
    else:
        # 2단계: 2행(인덱스 0)만 계산
        df_binance_ticker_15m = calculate_latest_row_only_ksc(df_binance_ticker_15m)
    
    # 1HCLASS, -1HCLASS를 먼저 15분봉으로 복사 (모든 지표가 최신 1H 기준으로 계산되도록 선행)
    if not df_binance_ticker_1h.empty:
        df_binance_ticker_15m = copy_1hclass_to_15m(df_binance_ticker_15m, df_binance_ticker_1h)
        df_binance_ticker_15m = copy_minus_1hclass_to_15m(df_binance_ticker_15m, df_binance_ticker_1h)
        df_binance_ticker_15m = copy_p1h_to_15m_and_set_p(df_binance_ticker_15m, df_binance_ticker_1h)
    
    # StoSP, TP, StoSU, TPC, TPCS, NBS 계산 (PRFT보다 먼저 실행: TP가 있어야 TPOVER 판단 가능)
    if skip_first_row:
        # 1단계: 전체 계산
        df_binance_ticker_15m = calculate_stosp_stosu(df_binance_ticker_15m)
    else:
        # 2단계: 전체 재계산 (StoSP는 누적 계산이므로 전체 재계산 필요)
        df_binance_ticker_15m = calculate_stosp_stosu(df_binance_ticker_15m)
    
    # PRFT 열 계산 (15분봉 시트에만) - TP 계산 후 실행
    if skip_first_row:
        # 1단계: 전체 계산
        df_binance_ticker_15m = calculate_prft_for_15m(df_binance_ticker_15m)
    else:
        # 2단계: 2행(인덱스 0)만 계산
        df_binance_ticker_15m = calculate_latest_row_only_prft(df_binance_ticker_15m)
    
    # PRFT 반영 후 StoSP/TP/NBS 재계산 (TPOVER로 인한 초기화 반영, NBS를 PRFT 이후로 위치)
    if skip_first_row:
        df_binance_ticker_15m = calculate_stosp_stosu(df_binance_ticker_15m)
    else:
        df_binance_ticker_15m = calculate_stosp_stosu(df_binance_ticker_15m)
    
    # 모든 지표 계산 완료 (1단계일 때만 표시)
    if skip_first_row:
        print(f"{get_timestamp()} [{stage_prefix}] ✅ 모든 지표 계산 완료")
    
    # 2단계일 때: 3행 이후(인덱스 1부터)를 previous 값으로 복원 (dateM, LD만 복원)
    if not skip_first_row and not df_prev_15m.empty and len(df_binance_ticker_15m) > 1:
        # 2행(인덱스 0)은 새로 계산된 값 유지, 3행부터(인덱스 1부터)는 previous 값으로 복원
        # dateM, LD 컬럼만 복원 (전체 계산 함수로 인해 변동됨)
        restore_len = min(len(df_binance_ticker_15m) - 1, len(df_prev_15m_backup))
        if restore_len > 0:
            # dateM, LD 컬럼만 복원
            for col in ['dateM', 'LD']:
                if col in df_prev_15m_backup.columns and col in df_binance_ticker_15m.columns:
                    source_values = df_prev_15m_backup[col].values[:restore_len]
                    target_dtype = df_binance_ticker_15m[col].dtype
                    
                    try:
                        if target_dtype == 'object':
                            df_binance_ticker_15m.loc[df_binance_ticker_15m.index[1:1+restore_len], col] = pd.Series(source_values, dtype=object).values
                        elif pd.api.types.is_datetime64_any_dtype(target_dtype):
                            converted = pd.to_datetime(source_values, errors='coerce')
                            df_binance_ticker_15m.loc[df_binance_ticker_15m.index[1:1+restore_len], col] = converted.values
                        elif pd.api.types.is_integer_dtype(target_dtype) or pd.api.types.is_float_dtype(target_dtype):
                            converted_values = pd.to_numeric(source_values, errors='coerce')
                            if converted_values.isna().any() and not pd.api.types.is_float_dtype(target_dtype):
                                df_binance_ticker_15m[col] = df_binance_ticker_15m[col].astype('float64')
                            df_binance_ticker_15m.loc[df_binance_ticker_15m.index[1:1+restore_len], col] = converted_values.values
                        else:
                            df_binance_ticker_15m.loc[df_binance_ticker_15m.index[1:1+restore_len], col] = source_values
                    except Exception as e:
                        df_binance_ticker_15m[col] = df_binance_ticker_15m[col].astype(object)
                        df_binance_ticker_15m.loc[df_binance_ticker_15m.index[1:1+restore_len], col] = pd.Series(source_values, dtype=object).values
    
    # ---- {TICKER}USDT15M 최신행 ORDER 신호 확인 ----
    try:
        if len(df_binance_ticker_15m) == 0:
            print(f"{get_timestamp()} [{stage_prefix}] ⚠️ ORDER 신호 확인 스킵: 15분봉 데이터가 비어있습니다.")
        else:
            latest_row = df_binance_ticker_15m.iloc[0]
            latest_order = str(latest_row.get("ORDER", ""))
            latest_ksc = latest_row.get("KSC", "")
            latest_bomb = latest_row.get("Bomb", "")
            latest_bomb_count = latest_row.get("BombCount", 0)
            latest_prft = latest_row.get("PRFT", "")
            latest_1hmsfast = latest_row.get("1HMSFast", np.nan)
            latest_buyside = latest_row.get("buyside", np.nan)
            latest_1hcl = latest_row.get("1HCL", np.nan)
            latest_minus_1hcl = latest_row.get("-1HCL", np.nan)
            latest_nbs = latest_row.get("NBS", 0)
            
            # spread 계산
            sma25 = latest_row.get("SMA25", np.nan)
            sma100 = latest_row.get("SMA100", np.nan)
            sma200 = latest_row.get("SMA200", np.nan)
            spread = np.nan
            if not pd.isna(sma25) and not pd.isna(sma100) and not pd.isna(sma200):
                sma_vals = [float(sma25), float(sma100), float(sma200)]
                sma_min = min(sma_vals)
                if sma_min > 0:
                    spread = (max(sma_vals) - sma_min) / sma_min
            
            # sprd2threshold 계산
            sprd2_threshold = None
            if 'SPRD2' in df_binance_ticker_15m.columns:
                sort_col = 'Date(UTC)'  # UTC 기준으로 정렬 (바이낸스는 UTC 기준)
                df_sorted = df_binance_ticker_15m.sort_values(sort_col, ascending=False).head(400)
                sprd2_valid = df_sorted['SPRD2'][pd.notna(df_sorted['SPRD2'])]
                if len(sprd2_valid) > 0:
                    sprd2_threshold = sprd2_valid.mean() * 0.3
            
            # 메시지 구성요소 (KSC)
            ksc_str = " | KSC: -"
            if pd.notna(latest_ksc):
                is_bomb = str(latest_bomb).strip() == 'Bomb'
                if is_bomb:
                    ksc_str = f" | KSC: Bomb ({int(latest_bomb_count)})"
                else:
                    try:
                        val = int(float(latest_ksc))
                        ksc_str = f" | KSC: {val}"
                    except: 
                        pass
            
            # 메시지 구성요소 (P: KSC 수열의 p값 = 3 + p1H, 15M 열 p 우선)
            p_str = ""
            try:
                latest_p = latest_row.get("p", np.nan)
                if pd.notna(latest_p):
                    p_val = int(float(latest_p))
                    p_str = f" | P: {p_val}"
                elif pd.notna(latest_1hcl):
                    h1cl_val = int(float(latest_1hcl))
                    p_str = f" | P: {3 + h1cl_val}"
                else:
                    p_str = " | P: 3"
            except:
                p_str = " | P: 3"
            
            # 메시지 구성요소 (PRFT)
            prft_str = " | PRFT: 0"
            if str(latest_prft).strip() == 'PRFT' and pd.notna(latest_buyside):
                prft_mult = 1 + (1 - float(latest_buyside))
                prft_str = f" | PRFT:{prft_mult:.3f}"
            elif str(latest_prft).strip() == 'TPOVER':
                prft_str = " | PRFT:TPOVER"
            elif pd.notna(latest_prft):
                try: 
                    prft_str = f" | PRFT: {int(float(latest_prft))}"
                except: 
                    pass
            
            # 메시지 구성요소 (LS)
            latest_ls = latest_row.get("LS", np.nan)
            if pd.notna(latest_ls):
                try:
                    ls_val = int(float(latest_ls))
                    ls_str = f" | LS:{ls_val}"
                except (TypeError, ValueError):
                    ls_str = " | LS: -"
            else:
                ls_str = " | LS: -"

            # 메시지 구성요소 (기타)
            hms_str = f" | 1HMSF:{float(latest_1hmsfast):.3f}" if pd.notna(latest_1hmsfast) else " | 1HMSF: -"

            # HCL = 1HCL + (-1HCL) 표시용 문자열
            try:
                if pd.notna(latest_1hcl) and pd.notna(latest_minus_1hcl):
                    # 둘 다 값이 있으면 합산
                    h1cl_val = int(float(latest_1hcl))
                    minus_1hcl_val = int(float(latest_minus_1hcl))
                    hcl_sum = h1cl_val + minus_1hcl_val
                    hcl_str = f" | HCL: {hcl_sum}"
                elif pd.isna(latest_1hcl) and pd.isna(latest_minus_1hcl):
                    # 둘 다 NaN이면 0으로 표시
                    hcl_str = " | HCL: 0"
                else:
                    # 둘 중 하나만 NaN이면, 있는 값만 표시
                    if pd.notna(latest_1hcl):
                        h1cl_val = int(float(latest_1hcl))
                        hcl_str = f" | HCL: {h1cl_val}"
                    elif pd.notna(latest_minus_1hcl):
                        minus_1hcl_val = int(float(latest_minus_1hcl))
                        hcl_str = f" | HCL: {minus_1hcl_val}"
                    else:
                        hcl_str = " | HCL: 0"
            except Exception:
                hcl_str = " | HCL: -"

            sprd_str = f" | sprd:{spread*100:.3f}%" if pd.notna(spread) else " | sprd: -"
            if sprd2_threshold is not None:
                th_str = f",sprd2th: {sprd2_threshold*100:.3f}%" 
            else:
                th_str = ", sprd2th: -"
            
            # 직전행 TP 및 TPCS 가져오기
            prev_tp = None
            prev_tpcs = 0.0
            if len(df_binance_ticker_15m) > 1:
                prev_row = df_binance_ticker_15m.iloc[1]
                prev_tp = prev_row.get("TP", np.nan)
                prev_tpcs = prev_row.get("TPCS", 0.0)
            
            # 직전행 TP 안전한 타입 변환 (소수점 둘째자리로 표시)
            try:
                prev_tp_val = float(prev_tp) if pd.notna(prev_tp) else 0.0
            except (TypeError, ValueError):
                prev_tp_val = 0.0
            
            # 최신행 종가 가져오기
            latest_close = latest_row.get("종", None)
            try:
                latest_close_val = float(latest_close) if pd.notna(latest_close) else None
            except (TypeError, ValueError):
                latest_close_val = None
            
            # 티커별 포맷 결정 (엑셀 종가 서식과 동일)
            if 'XRP' in TICKER:
                price_format = '.4f'  # XRP: 소수점 4자리
            elif 'USDT' in TICKER or 'USD' in TICKER or TICKER in ['BTC', 'ETH', 'SOL', 'BNB', 'ADA', 'DOGE', 'DOT', 'LINK', 'LTC', 'MATIC', 'AVAX', 'UNI', 'ATOM', 'ETC', 'XLM', 'ALGO', 'VET', 'ICP', 'FIL', 'TRX', 'EOS', 'AAVE', 'THETA', 'XTZ', 'SAND', 'MANA', 'AXS', 'CHZ', 'ENJ', 'GALA', 'FLOW', 'NEAR', 'APT', 'ARB', 'OP', 'SUI', 'SEI', 'TIA', 'INJ', 'RUNE', 'FET', 'RENDER', 'IMX', 'STRK', 'PIXEL', 'WLD', 'JTO', 'PYTH', 'DYM', 'AI', 'ONDO', 'ALT', 'JUP', 'WIF', 'TNSR', 'SAGA', 'REZ', 'BB', 'NOT', 'IO', 'ZRO', 'ZKSYNC', 'LISTA', 'ZK', 'PEPE', 'FLOKI', 'BONK', 'SHIB', '1000SATS', '1000FLOKI', '1000PEPE', '1000BONK', '1000SHIB']:
                price_format = '.2f'  # USDT/USD: 소수점 2자리
            else:
                price_format = '.0f'  # KRW: 정수
            
            # 종가 포맷 (ORDER 로그는 LS, 1HMSF, 종가 세 개만 표시)
            if latest_close_val is not None:
                close_str = f" | 종가:{latest_close_val:{price_format}}"
            else:
                close_str = " | 종가: -"
            
            # 최종 메시지 조립: LS | 1HMSF | 종가 만 표시
            if latest_order and latest_order.strip() and latest_order.lower() != 'none':
                info_msg = f"{get_timestamp()} [{stage_prefix}] ℹ️ORDER: {TICKER} '{latest_order}'{ls_str}{hms_str}{close_str}"
            else:
                info_msg = f"{get_timestamp()} [{stage_prefix}] ℹ️ORDER: {TICKER} 없음{ls_str}{hms_str}{close_str}"
            
            # 터미널 출력
            print(info_msg)
            
            # Discord로 전송
            send_discord_message(info_msg)
            
    except Exception as e:
        print(f"{get_timestamp()} [{stage_prefix}] ❌ ORDER 신호 확인 실패: {e}")
        import traceback
        traceback.print_exc()
    
    # 모든 지표 계산 완료 (1단계일 때만 표시)
    if skip_first_row:
        print(f"{get_timestamp()} [{stage_prefix}] ✅ 모든 지표 계산 완료")

    # 일봉 컬럼 순서 맞춤
    binance_cols_1d = ["Date(UTC)", "KST", "종", "시", "고", "저", "Vol.", "SMA3", "SMA5", "SMA7", "SMA10", "SMA20", "Max15", "Min15", "하단", "상단", "SFast", "Fast", "Base", "4or1", "buyside", "sellside", "Sell", "Buy", "Samount1D", "Bamount1D"]
    
    # 존재하는 컬럼만 선택
    binance_cols_1d = [col for col in binance_cols_1d if col in df_binance_ticker_1d.columns]
    df_binance_ticker_1d = df_binance_ticker_1d[binance_cols_1d]
    
    # 1분봉 컬럼 순서 맞춤 (Max400, Min400 사용, Source 기준: SMA15, SMA25, SMA35, SMA50, SMA100만)
    binance_cols_1m = ["Date(UTC)", "KST", "종", "시", "고", "저", "Vol.", "SMA15", "SMA25", "SMA35", "SMA50", "SMA100", "Max400", "Min400", "하단", "상단", "SFast", "Fast", "Base", "4or1", "buyside", "sellside", "Sell", "Buy"]
    
    # 존재하는 컬럼만 선택
    if not df_binance_ticker_1m.empty:
        binance_cols_1m = [col for col in binance_cols_1m if col in df_binance_ticker_1m.columns]
        df_binance_ticker_1m = df_binance_ticker_1m[binance_cols_1m]
    
    # 5분봉 컬럼 순서 맞춤 (Max200, Min200 사용, SB1M, 1HMSF 추가)
    binance_cols_5m = ["Date(UTC)", "KST", "종", "시", "고", "저", "Vol.", "SMA3", "SMA5", "SMA7", "SMA10", "SMA20", "Max200", "Min200", "하단", "상단", "SFast", "Fast", "Base", "4or1", "buyside", "sellside", "Sell", "Buy", "SB1M", "1HMSF"]
    
    # 존재하는 컬럼만 선택
    binance_cols_5m = [col for col in binance_cols_5m if col in df_binance_ticker_5m.columns]
    df_binance_ticker_5m = df_binance_ticker_5m[binance_cols_5m]
    
    # 15분봉 컬럼 순서 맞춤 (Max70, Min70 사용, SMA12 추가, SamountW/BamountW, Samount1D/Bamount1D, 최종 Samount/Bamount 추가, dateM, LD, SPRD, SPRD2 추가, SB1M 추가) - Source와 동일한 순서 (김프 제외)
    binance_cols_15m = ["Date(UTC)", "KST", "종", "시", "고", "저", "Vol.", "SMA3", "SMA5", "SMA7", "SMA10", "SMA12", "SMAF", "SMA20", "SMA25", "SMA100", "SMA200", "SMA400", "SMA800", "Max70", "Min70", "하단", "상단", "SFast", "Fast", "Base", "4or1", "buyside", "sellside", "Sell", "Buy", "SB1M", "SB5M", "SB1H", "SB1D", "ORDER", "1HMSFast", "1HCL", "-1HCL", "p", "KSC", "Bomb", "PRFT", "StoSP", "TP", "StoSU", "TPC", "TPCS", "NBS", "LS", "SamountW", "BamountW", "Samount1D", "Bamount1D", "Samount", "Bamount", "dateM", "LD", "SPRD", "SPRD2"]
    
    # 누락된 열들을 기본값으로 추가 (Source와 동일한 구조 유지, SMA400/SMA800 추가)
    if 'SMA400' not in df_binance_ticker_15m.columns:
        df_binance_ticker_15m['SMA400'] = np.nan
    if 'SMA800' not in df_binance_ticker_15m.columns:
        df_binance_ticker_15m['SMA800'] = np.nan
    if '1HCL' not in df_binance_ticker_15m.columns:
        df_binance_ticker_15m['1HCL'] = np.nan
    if '-1HCL' not in df_binance_ticker_15m.columns:
        df_binance_ticker_15m['-1HCL'] = np.nan
    if 'p' not in df_binance_ticker_15m.columns:
        df_binance_ticker_15m['p'] = np.nan
    if 'StoSP' not in df_binance_ticker_15m.columns:
        df_binance_ticker_15m['StoSP'] = np.nan
    if 'TP' not in df_binance_ticker_15m.columns:
        df_binance_ticker_15m['TP'] = np.nan
    if 'StoSU' not in df_binance_ticker_15m.columns:
        df_binance_ticker_15m['StoSU'] = np.nan
    if 'TPC' not in df_binance_ticker_15m.columns:
        df_binance_ticker_15m['TPC'] = 0
    if 'TPCS' not in df_binance_ticker_15m.columns:
        df_binance_ticker_15m['TPCS'] = 0
    if 'NBS' not in df_binance_ticker_15m.columns:
        df_binance_ticker_15m['NBS'] = 0
    if 'LS' not in df_binance_ticker_15m.columns:
        df_binance_ticker_15m['LS'] = ''  # 헤더만, 내용 채우지 않음
    if 'SPRD' not in df_binance_ticker_15m.columns:
        df_binance_ticker_15m['SPRD'] = np.nan
    # bomb → Bomb로 변경 (대소문자 통일)
    if 'bomb' in df_binance_ticker_15m.columns and 'Bomb' not in df_binance_ticker_15m.columns:
        df_binance_ticker_15m['Bomb'] = df_binance_ticker_15m['bomb']
        df_binance_ticker_15m = df_binance_ticker_15m.drop(columns=['bomb'], errors='ignore')
    elif 'Bomb' not in df_binance_ticker_15m.columns:
        df_binance_ticker_15m['Bomb'] = ''
    
    # 존재하는 컬럼만 선택 (Source 순서 유지)
    binance_cols_15m = [col for col in binance_cols_15m if col in df_binance_ticker_15m.columns]
    df_binance_ticker_15m = df_binance_ticker_15m[binance_cols_15m]
    
    
    # 1시간봉 컬럼 순서 맞춤 (Source 기준: SMA25, SMA100, SMA200, SMA400, SMA800, Max200, Min200, 1HCLASS, -1HCLASS)
    binance_cols_1h = ["Date(UTC)", "KST", "종", "시", "고", "저", "Vol.", "SMA25", "SMA100", "SMA200", "SMA400", "SMA800", "Max200", "Min200", "하단", "상단", "SFast", "Fast", "Base", "1HMSFast", "4or1", "buyside", "sellside", "Sell", "Buy", "1HCLASS", "-1HCLASS", "p1H"]
    
    # 누락된 열 추가 (p1H: 내용 채우지 않음)
    if 'p1H' not in df_binance_ticker_1h.columns:
        df_binance_ticker_1h['p1H'] = np.nan
    # 존재하는 컬럼만 선택
    binance_cols_1h = [col for col in binance_cols_1h if col in df_binance_ticker_1h.columns]
    df_binance_ticker_1h = df_binance_ticker_1h[binance_cols_1h]
    
    # 1H4x 시트 컬럼 순서 맞춤 (Source 기준: 1HMSFast 위치를 Base 다음으로)
    binance_cols_1h4x = ["Date(UTC)", "KST", "종", "시", "고", "저", "Vol.", "SMA12", "SMA20", "SMA28", "SMA40", "SMA80", "SMA100", "SMA200", "Max200", "Min200", "하단", "상단", "SFast", "Fast", "Base", "1HMSFast", "4or1", "buyside", "sellside", "Sell", "Buy"]
    
    # 존재하는 컬럼만 선택
    binance_cols_1h4x = [col for col in binance_cols_1h4x if col in df_binance_ticker_1h4x.columns]
    df_binance_ticker_1h4x = df_binance_ticker_1h4x[binance_cols_1h4x]
    
    # 주봉 컬럼 순서 맞춤 (Max25, Min25 사용, 김프 제외, SamountW/BamountW 추가)
    binance_cols_weekly = ["Date(UTC)", "KST", "종", "시", "고", "저", "Vol.", "SMA3", "SMA5", "SMA7", "SMA10", "SMA20", "Max25", "Min25", "하단", "상단", "SFast", "Fast", "Base", "4or1", "buyside", "sellside", "Sell", "Buy", "SamountW", "BamountW"]
    
    # 존재하는 컬럼만 선택
    binance_cols_weekly = [col for col in binance_cols_weekly if col in df_binance_ticker_weekly.columns]
    df_binance_ticker_weekly = df_binance_ticker_weekly[binance_cols_weekly]

    # 엑셀 저장 전 데이터 개수 제한 (CANDLE_COUNT 사용)
    print(f"{get_timestamp()} [{stage_prefix}] 📊 엑셀 저장 전 데이터 제한 중...")
    df_binance_ticker_5m = df_binance_ticker_5m.iloc[:CANDLE_COUNT['5m']].reset_index(drop=True)  # 5분봉: 2400개
    df_binance_ticker_15m = df_binance_ticker_15m.iloc[:CANDLE_COUNT['15m']].reset_index(drop=True)  # 15분봉: 800개
    df_binance_ticker_1h = df_binance_ticker_1h.iloc[:CANDLE_COUNT['1h']].reset_index(drop=True)  # 1시간봉: Source 기준 1600개 (2400개 수집 후 과거 800개 제거)
    df_binance_ticker_1h4x = df_binance_ticker_1h4x.iloc[:400].reset_index(drop=True)  # 1H4x 시트: 400개
    print(f"{get_timestamp()} [{stage_prefix}]    → 5분봉: {len(df_binance_ticker_5m)}개, 15분봉: {len(df_binance_ticker_15m)}개, 1시간봉: {len(df_binance_ticker_1h)}개, 1H4x: {len(df_binance_ticker_1h4x)}개")
    
    # ⚠️NBS=1 발생 시 지정가 매도 주문 실행 (엑셀 저장 전, DataFrame에서 직접 확인)
    # 선물 스크립트: NBS 분할 매도 비활성화 (스팟 전용 기능)
    if False and 'NBS' in df_binance_ticker_15m.columns and len(df_binance_ticker_15m) > 1:
        try:
            # 최신 행(idx=0)의 NBS 확인 (안전한 타입 변환)
            _raw_nbs = df_binance_ticker_15m.iloc[0].get('NBS', 0)
            try:
                nbs_value = int(_raw_nbs) if pd.notna(_raw_nbs) else 0
            except (TypeError, ValueError):
                try:
                    nbs_value = int(float(_raw_nbs)) if pd.notna(_raw_nbs) else 0
                except (TypeError, ValueError):
                    nbs_value = 0
            
            if nbs_value == 1:
                # 직전행(idx=1)의 TPCS와 TP 확인 (안전한 타입 변환)
                _raw_tpcs = df_binance_ticker_15m.iloc[1].get('TPCS', 0)
                try:
                    prev_tpcs = float(_raw_tpcs) if pd.notna(_raw_tpcs) else 0.0
                except (TypeError, ValueError):
                    prev_tpcs = 0.0
                
                _raw_tp = df_binance_ticker_15m.iloc[1].get('TP', 0)
                try:
                    prev_tp = float(_raw_tp) if pd.notna(_raw_tp) else 0.0
                except (TypeError, ValueError):
                    prev_tp = 0.0
                
                # TPCS가 0보다 크고 TP가 유효한 값이면 지정가 매도 주문 실행
                # NBS=1: USDT 제외 다른 티커들은 분할 매도, USDT는 기존 단일 매도
                if prev_tpcs is not None and not pd.isna(prev_tpcs) and prev_tpcs > 0:
                    if prev_tp is not None and not pd.isna(prev_tp) and prev_tp > 0:
                        # 보유 수량 확인
                        ticker_balance = binance_get_account_balance(TICKER)
                        available_balance = ticker_balance.get('free_precise', 0.0)
                        
                        # TPCS는 Unit 단위이므로 코인 수량으로 변환 필요
                        # 변환 공식: TPCS (Unit) × 1UNIT금액 ÷ TP(목표가격) = 코인 수량
                        unit_amount = ROTATION_TRADING_UNITS.get(TICKER, 8)
                        tpcs_unit = float(prev_tpcs)
                        tp_price = float(prev_tp)  # prev_tp = TP = 목표가격
                        
                        # TPCS Unit을 금액으로 변환 후 TP 가격으로 나누어 코인 수량 계산
                        target_amount = tpcs_unit * unit_amount  # TPCS Unit × 1UNIT금액 = 매도 금액
                        target_volume = target_amount / tp_price if tp_price > 0 else 0  # 매도 금액 ÷ TP(목표가격) = 코인 수량
                        
                        symbol = f"{TICKER}USDT"
                        
                        # 바이낸스는 모든 티커에 분할 매도 (3단계) 적용
                        # 기준 수량 결정: available과 target_volume 중 작은 값
                        base_volume = min(available_balance, target_volume) if target_volume > 0 else available_balance
                        
                        if base_volume > 0:
                            # 분할 매도 수량 계산 (30%, 40%, 30%)
                            sell_volume_1 = base_volume * 0.3  # 1차: 30%
                            sell_volume_2 = base_volume * 0.4  # 2차: 40%
                            sell_volume_3 = base_volume * 0.3  # 3차: 30%
                            
                            # 분할 매도 가격 계산 (TP/1.01에 각각 0.45%, 0.50%, 0.55% 추가)
                            avg_price = tp_price / 1.01  # 평단가 역산
                            sell_price_1 = avg_price * (1 + 0.0045)  # 1차: +0.45%
                            sell_price_2 = avg_price * (1 + 0.005)   # 2차: +0.50%
                            sell_price_3 = avg_price * (1 + 0.0055)  # 3차: +0.55%
                            
                            # 틱 사이즈로 조정
                            sell_price_1 = adjust_price_to_tick_binance(symbol, sell_price_1)
                            sell_price_2 = adjust_price_to_tick_binance(symbol, sell_price_2)
                            sell_price_3 = adjust_price_to_tick_binance(symbol, sell_price_3)
                            
                            # NBS 신호 감지 메시지 출력
                            nbs_msg = f"{get_timestamp()} [{stage_prefix}] 🔔NBS=1 신호 감지! 직전행 TPCS={tpcs_unit:.2f} UNIT(={target_amount:.2f} USDT), TP={tp_price:.2f} USDT, 기준수량={base_volume:.8f} {TICKER}, 보유수량={available_balance:.8f} {TICKER} → 분할 매도 실행 (3단계)"
                            print(nbs_msg)
                            send_discord_message(nbs_msg)
                            
                            # 보유 수량 경고 메시지
                            if available_balance < target_volume:
                                warning_msg = f"{get_timestamp()} [{stage_prefix}] ⚠️보유 수량({available_balance:.8f} {TICKER})이 계산된 매도량({target_volume:.8f} {TICKER}, TPCS={tpcs_unit:.2f} UNIT)보다 작아 보유 수량 기준으로 분할 매도합니다"
                                print(warning_msg)
                                send_discord_message(warning_msg)
                            
                            # 1차 매도: 가격 TP/1.01*(1+0.45%), 수량 30%
                            if sell_volume_1 > 0:
                                split_msg_1 = f"{get_timestamp()} [{stage_prefix}] 📤[1/3] 분할 매도 주문: 가격={sell_price_1:.6f} USDT, 수량={sell_volume_1:.8f} {TICKER} (기준수량의 30%)"
                                print(split_msg_1)
                                send_discord_message(split_msg_1)
                                binance_limit_sell(symbol, sell_price_1, sell_volume_1, stage_prefix)
                            
                            # 2차 매도: 가격 TP/1.01*(1+0.50%), 수량 40%
                            if sell_volume_2 > 0:
                                split_msg_2 = f"{get_timestamp()} [{stage_prefix}] 📤[2/3] 분할 매도 주문: 가격={sell_price_2:.6f} USDT, 수량={sell_volume_2:.8f} {TICKER} (기준수량의 40%)"
                                print(split_msg_2)
                                send_discord_message(split_msg_2)
                                binance_limit_sell(symbol, sell_price_2, sell_volume_2, stage_prefix)
                            
                            # 3차 매도: 가격 TP/1.01*(1+0.55%), 수량 30%
                            if sell_volume_3 > 0:
                                split_msg_3 = f"{get_timestamp()} [{stage_prefix}] 📤[3/3] 분할 매도 주문: 가격={sell_price_3:.6f} USDT, 수량={sell_volume_3:.8f} {TICKER} (기준수량의 30%)"
                                print(split_msg_3)
                                send_discord_message(split_msg_3)
                                binance_limit_sell(symbol, sell_price_3, sell_volume_3, stage_prefix)
                        else:
                            no_balance_msg = f"{get_timestamp()} [{stage_prefix}] ⚠️NBS=1 신호 감지했으나 보유 수량이 0입니다 (보유수량: {available_balance:.8f} {TICKER})"
                            print(no_balance_msg)
                            send_discord_message(no_balance_msg)
                    else:
                        invalid_tp_msg = f"{get_timestamp()} [{stage_prefix}] ⚠️NBS=1 신호 감지했으나 직전행 TP가 유효하지 않음 (TP={prev_tp})"
                        print(invalid_tp_msg)
                        send_discord_message(invalid_tp_msg)
                else:
                    invalid_tpcs_msg = f"{get_timestamp()} [{stage_prefix}] ℹ️NBS=1 신호 감지했으나 (TPCS={prev_tpcs})"
                    print(invalid_tpcs_msg)
                    send_discord_message(invalid_tpcs_msg)
        except Exception as nbs_error:
            error_msg = f"{get_timestamp()} [{stage_prefix}] ⚠️NBS=1 신호 확인 중 오류 발생: {nbs_error}"
            print(error_msg)
            send_discord_message(error_msg)
            import traceback
            traceback_str = traceback.format_exc()
            print(f"{get_timestamp()} [{stage_prefix}] 상세 오류:\n{traceback_str}")
            send_discord_message(f"{error_msg}\n상세 오류:\n{traceback_str}")
    
    # 저장 및 열 너비 자동 조정 (UTC 기준)
    print(f"{get_timestamp()} [{stage_prefix}] 💾 엑셀 파일 저장 중...")
    
    # KSC는 숫자만 저장하므로 변환 불필요
    # bomb 열은 "bomb" 문자열로 저장됨
    
    # 날짜 컬럼을 문자열로 변환 (엑셀 날짜 자동 변환 방지, UTC 기준)
    df_binance_ticker_1m = _force_date_text(df_binance_ticker_1m) if not df_binance_ticker_1m.empty else df_binance_ticker_1m
    df_binance_ticker_5m = _force_date_text(df_binance_ticker_5m)
    df_binance_ticker_15m = _force_date_text(df_binance_ticker_15m)
    df_binance_ticker_1h = _force_date_text(df_binance_ticker_1h)
    df_binance_ticker_1h4x = _force_date_text(df_binance_ticker_1h4x)
    df_binance_ticker_1d = _force_date_text(df_binance_ticker_1d)
    df_binance_ticker_weekly = _force_date_text(df_binance_ticker_weekly)
    
    # ExcelWriter로 파일 생성 (총 7개 시트)
    writer = None
    try:
        writer = pd.ExcelWriter(save_path, engine="openpyxl")
        
        # 시트 순서: {TICKER}USDT15M, {TICKER}USDT5M, {TICKER}USDT1M, {TICKER}USDT1H, {TICKER}USDT1H4x, {TICKER}USDT1D, {TICKER}USDTW (바이낸스 전용)
        df_binance_ticker_15m.to_excel(writer, index=False, sheet_name=f"{TICKER}USDT15M")
        df_binance_ticker_5m.to_excel(writer, index=False, sheet_name=f"{TICKER}USDT5M")
        # 1분봉 시트는 항상 생성 (데이터가 없어도 빈 시트 생성)
        if not df_binance_ticker_1m.empty:
            # 1분봉 최종 개수 제한 (12400개 수집 → 400개 제거 → 12000개)
            df_binance_ticker_1m_final = df_binance_ticker_1m.iloc[:CANDLE_COUNT['1m']].copy() if len(df_binance_ticker_1m) > CANDLE_COUNT['1m'] else df_binance_ticker_1m.copy()
            df_binance_ticker_1m_final.to_excel(writer, index=False, sheet_name=f"{TICKER}USDT1M")
        else:
            # 1분봉 데이터가 없어도 빈 시트 생성
            df_binance_ticker_1m_final = pd.DataFrame()
            df_binance_ticker_1m_final.to_excel(writer, index=False, sheet_name=f"{TICKER}USDT1M")
        df_binance_ticker_1h.to_excel(writer, index=False, sheet_name=f"{TICKER}USDT1H")
        df_binance_ticker_1h4x.to_excel(writer, index=False, sheet_name=f"{TICKER}USDT1H4x")
        df_binance_ticker_1d.to_excel(writer, index=False, sheet_name=f"{TICKER}USDT1D")
        df_binance_ticker_weekly.to_excel(writer, index=False, sheet_name=f"{TICKER}USDTW")
        
        # 워크북과 워크시트 가져오기
        workbook = writer.book
        
        # 모든 시트의 열 너비 자동 조정 및 숫자 포맷팅
        print(f"{get_timestamp()} [{stage_prefix}] 🎨 엑셀 포맷팅 중...")
        sheet_names = [f"{TICKER}USDT15M", f"{TICKER}USDT5M", f"{TICKER}USDT1M", f"{TICKER}USDT1H", f"{TICKER}USDT1H4x", f"{TICKER}USDT1D", f"{TICKER}USDTW"]
        
        for sheet_name in sheet_names:
            worksheet = writer.sheets[sheet_name]
            
            # 해당 시트의 DataFrame 가져오기
            df_sheet = None
            if sheet_name == f'{TICKER}USDT5M':  # 5분봉 시트 (바이낸스)
                df_sheet = df_binance_ticker_5m
            elif sheet_name == f'{TICKER}USDT15M':  # 15분봉 시트 (바이낸스)
                df_sheet = df_binance_ticker_15m
            elif sheet_name == f'{TICKER}USDT1M':  # 1분봉 시트 (바이낸스)
                df_sheet = df_binance_ticker_1m_final if not df_binance_ticker_1m.empty else pd.DataFrame()
            elif sheet_name == f'{TICKER}USDT1H':  # 1시간봉 시트 (바이낸스)
                df_sheet = df_binance_ticker_1h
            elif sheet_name == f'{TICKER}USDT1H4x':  # 1H4x 시트 (바이낸스)
                df_sheet = df_binance_ticker_1h4x
            elif sheet_name == f'{TICKER}USDTW':  # 주봉 시트 (바이낸스)
                df_sheet = df_binance_ticker_weekly
            elif sheet_name == f'{TICKER}USDT1D':  # 일봉 시트 (바이낸스)
                df_sheet = df_binance_ticker_1d
            
            # df_sheet가 None인 경우 건너뛰기
            if df_sheet is None:
                print(f"{get_timestamp()} [{stage_prefix}] ⚠️ 시트 '{sheet_name}'에 대한 DataFrame을 찾을 수 없습니다.")
                continue
            
            # 시트별 컬럼명 정의 (Max20/Max200, Min20/Min200, Max70/Min70 구분)
            if sheet_name == f'{TICKER}USDT5M':  # 5분봉 시트 (바이낸스, SB1M, 1HMSF 포함)
                column_names = [
                    'Date(UTC)', 'KST', '종', '시', '고', '저', 'Vol.', 'SMA3', 'SMA5', 'SMA7', 'SMA10', 'SMA20',
                    'Max200', 'Min200', '하단', '상단', 'SFast', 'Fast', 'Base', '4or1',  # 5분봉용
                    'buyside', 'sellside', 'Sell', 'Buy', 'SB1M', '1HMSF'  # SB1M, 1HMSF (저장 컬럼과 동일)
                ]
            elif sheet_name == f'{TICKER}USDT1M':  # 1분봉 시트 (바이낸스, Source 기준: SMA15, SMA25, SMA35, SMA50, SMA100만)
                column_names = [
                    'Date(UTC)', 'KST', '종', '시', '고', '저', 'Vol.', 'SMA15', 'SMA25', 'SMA35', 'SMA50', 'SMA100', 'Max400', 'Min400', '하단', '상단', 'SFast', 'Fast', 'Base', '4or1',  # 1분봉용 (Max400/Min400 사용)
                    'buyside', 'sellside', 'Sell', 'Buy'
                ]
            elif sheet_name == f'{TICKER}USDT15M':  # 15분봉 시트 (바이낸스, SMA400/SMA800 포함, 저장 컬럼과 동일)
                column_names = [
                    'Date(UTC)', 'KST', '종', '시', '고', '저', 'Vol.', 'SMA3', 'SMA5', 'SMA7', 'SMA10', 'SMA12', 'SMAF', 'SMA20',
                    'SMA25', 'SMA100', 'SMA200', 'SMA400', 'SMA800', 'Max70', 'Min70', '하단', '상단', 'SFast', 'Fast', 'Base', '4or1',  # 15분봉용 (SMA400/SMA800 추가)
                    'buyside', 'sellside', 'Sell', 'Buy', 'SB1M', 'SB5M', 'SB1H', 'SB1D', 'ORDER',
                    '1HMSFast', '1HCL', '-1HCL', 'p', 'KSC', 'Bomb', 'PRFT', 'StoSP', 'TP', 'StoSU', 'TPC', 'TPCS', 'NBS', 'LS',
                    'SamountW', 'BamountW', 'Samount1D', 'Bamount1D', 'Samount', 'Bamount', 'dateM', 'LD', 'SPRD', 'SPRD2'
                ]
            elif sheet_name == f'{TICKER}USDT1H':  # 1시간봉 시트 (바이낸스, Source 기준: SMA25, SMA100, SMA200, SMA400, SMA800, Max200, Min200, 1HCLASS, -1HCLASS)
                column_names = [
                    'Date(UTC)', 'KST', '종', '시', '고', '저', 'Vol.', 'SMA25', 'SMA100', 'SMA200', 'SMA400', 'SMA800', 'Max200', 'Min200', '하단', '상단', 'SFast', 'Fast', 'Base', '1HMSFast', '4or1',  # 1시간봉용
                    'buyside', 'sellside', 'Sell', 'Buy', '1HCLASS', '-1HCLASS', 'p1H'
                ]
            elif sheet_name == f'{TICKER}USDT1H4x':  # 1H4x 시트 (바이낸스, Source 기준: 1HMSFast 위치를 Base 다음으로)
                column_names = [
                    'Date(UTC)', 'KST', '종', '시', '고', '저', 'Vol.', 'SMA12', 'SMA20', 'SMA28', 'SMA40', 'SMA80',
                    'SMA100', 'SMA200', 'Max200', 'Min200', '하단', '상단', 'SFast', 'Fast', 'Base', '1HMSFast', '4or1',  # 1H4x 시트용
                    'buyside', 'sellside', 'Sell', 'Buy'
                ]
            elif sheet_name in [f'{TICKER}USDTW']:  # 주봉 시트 (SamountW/BamountW 추가)
                column_names = [
                    'Date(UTC)', 'KST', '종', '시', '고', '저', 'Vol.', 'SMA3', 'SMA5', 'SMA7', 'SMA10', 'SMA20',
                    'Max25', 'Min25', '하단', '상단', 'SFast', 'Fast', 'Base', '4or1',  # 주봉용
                    'buyside', 'sellside', 'Sell', 'Buy', 'SamountW', 'BamountW'
                ]
            else:  # 일봉 시트 ({TICKER}USDT1D, {TICKER}KRW1D)
                column_names = [
                    'Date(UTC)', 'KST', '종', '시', '고', '저', 'Vol.', 'SMA3', 'SMA5', 'SMA7', 'SMA10', 'SMA20',
                    'Max15', 'Min15', '하단', '상단', 'SFast', 'Fast', 'Base', '4or1',  # 일봉용
                    'buyside', 'sellside', 'Sell', 'Buy', 'Samount1D', 'Bamount1D'
                ]
            
            # Excel 컬럼 문자 계산 함수 (26개 이상일 때 AA, AB... 형식)
            def get_column_letter(col_idx):
                if col_idx < 26:
                    return chr(65 + col_idx)  # 0→A, 1→B, 2→C... 25→Z
                else:
                    # 26→AA, 27→AB, 28→AC...
                    first_letter = chr(65 + (col_idx // 26) - 1)  # 26→0→A, 27→1→A, 28→1→A
                    second_letter = chr(65 + (col_idx % 26))      # 26→0→A, 27→1→B, 28→2→C
                    return first_letter + second_letter
            
            # 완벽한 서식 복원 (복사본 파일과 동일)
            # 숫자 포맷팅 먼저 적용
            for row in range(2, worksheet.max_row + 1):  # 헤더 제외하고 데이터 행부터
                # 가격 컬럼들 (종, 시, 고, 저) - 시트별로 다른 정밀도
                for col in ['C', 'D', 'E', 'F']:  # 종, 시, 고, 저 컬럼
                    try:
                        cell = worksheet[f'{col}{row}']
                        if cell.value is not None:
                            # 심볼별 정밀도 설정
                            if 'XRP' in sheet_name and 'USDT' in sheet_name:
                                cell.number_format = '#,##0.0000'  # XRP: 소수점 4자리
                            elif 'USDT' in sheet_name or 'USD' in sheet_name:
                                cell.number_format = '#,##0.00'    # BTC, ETH, SOL, BNB: 소수점 2자리
                            
                            else:
                                cell.number_format = '#,##0'       # KRW 시트: 정수
                    except:
                        pass
                
                # 거래량 컬럼 (Vol.) - 시트별로 다른 포맷팅
                try:
                    cell = worksheet[f'G{row}']  # Vol. 컬럼
                    if cell.value is not None:
                        if 'XRP' in sheet_name and 'USDT' in sheet_name:
                            cell.number_format = '#,##0.0000'
                        elif 'USDT' in sheet_name or 'USD' in sheet_name:
                            cell.number_format = '#,##0.00'
                        else:
                            cell.number_format = '#,##0'
                except:
                    pass
                
                # SMA 컬럼들 (3, 5, 7, 10, 12, 20, 25, 28, 40, 15, 35, 50, 80, 100, 200, 400, 800일) - 시트별로 다른 포맷팅 (1H4x 시트는 SMA12, SMA28, SMA40, SMA80 포함, 1M 시트는 SMA15/SMA25/SMA35/SMA50/SMA100 포함, 1H 시트는 SMA400, SMA800 포함)
                for sma_col_name in ['SMA3', 'SMA5', 'SMA7', 'SMA10', 'SMA12', 'SMAF', 'SMA20', 'SMA25', 'SMA28', 'SMA40', 'SMA15', 'SMA35', 'SMA50', 'SMA80', 'SMA100', 'SMA200', 'SMA400', 'SMA800']:
                    try:
                        sma_col_idx = None
                        if df_sheet is not None and sma_col_name in df_sheet.columns:
                            sma_col_idx = list(df_sheet.columns).index(sma_col_name)
                        else:
                            for i, col_name in enumerate(column_names):
                                if col_name == sma_col_name:
                                    sma_col_idx = i
                                    break
                        
                        if sma_col_idx is not None:
                            sma_col_letter = get_column_letter(sma_col_idx)
                            cell = worksheet[f'{sma_col_letter}{row}']
                            if cell.value is not None:
                                if 'XRP' in sheet_name and 'USDT' in sheet_name:
                                    cell.number_format = '#,##0.0000'
                                elif 'USDT' in sheet_name or 'USD' in sheet_name:
                                    cell.number_format = '#,##0.00'
                                else:
                                    cell.number_format = '#,##0'
                    except:
                        pass
                
                # Max20/Max200/Max100/Max50/Max25/Max400, Min20/Min200/Min100/Min50/Min25/Min400 컬럼들 - 시트별로 다른 포맷팅 (1M 시트는 Max400, Min400 포함)
                for maxmin_col_name in ['Max15', 'Max25', 'Max50', 'Max70', 'Max100', 'Max200', 'Max400', 'Min15', 'Min25', 'Min50', 'Min70', 'Min100', 'Min200', 'Min400']:
                    try:
                        maxmin_col_idx = None
                        if df_sheet is not None and maxmin_col_name in df_sheet.columns:
                            maxmin_col_idx = list(df_sheet.columns).index(maxmin_col_name)
                        else:
                            for i, col_name in enumerate(column_names):
                                if col_name == maxmin_col_name:
                                    maxmin_col_idx = i
                                    break
                        
                        if maxmin_col_idx is not None:
                            maxmin_col_letter = get_column_letter(maxmin_col_idx)
                            cell = worksheet[f'{maxmin_col_letter}{row}']
                            if cell.value is not None:
                                if 'XRP' in sheet_name and 'USDT' in sheet_name:
                                    cell.number_format = '#,##0.0000'
                                elif 'USDT' in sheet_name or 'USD' in sheet_name:
                                    cell.number_format = '#,##0.00'
                                elif sheet_name == 'USDKRW':
                                    cell.number_format = '#,##0.00'
                                else:
                                    cell.number_format = '#,##0'
                    except:
                        pass
                
                # 하단, 상단 컬럼들 - 백분율로 표시 (공통 규칙)
                for col_name in ['하단', '상단']:
                    try:
                        col_idx = None
                        if df_sheet is not None and col_name in df_sheet.columns:
                            col_idx = list(df_sheet.columns).index(col_name)
                        else:
                            for i, cn in enumerate(column_names):
                                if cn == col_name:
                                    col_idx = i
                                    break
                        
                        if col_idx is not None:
                            col_letter = get_column_letter(col_idx)
                            cell = worksheet[f'{col_letter}{row}']
                            if cell.value is not None:
                                cell.number_format = '0.00%'
                    except:
                        pass
                
                # SFast 컬럼 - 소수점 셋째자리까지 (공통 규칙)
                try:
                    sfast_col_idx = None
                    if df_sheet is not None and 'SFast' in df_sheet.columns:
                        sfast_col_idx = list(df_sheet.columns).index('SFast')
                    else:
                        for i, col_name in enumerate(column_names):
                            if col_name == 'SFast':
                                sfast_col_idx = i
                                break
                    
                    if sfast_col_idx is not None:
                        sfast_col_letter = get_column_letter(sfast_col_idx)
                        cell = worksheet[f'{sfast_col_letter}{row}']
                        if cell.value is not None:
                            cell.number_format = '#,##0.000'
                except:
                    pass
                
                # Fast 컬럼 - 소수점 셋째자리까지 (공통 규칙)
                try:
                    fast_col_idx = None
                    if df_sheet is not None and 'Fast' in df_sheet.columns:
                        fast_col_idx = list(df_sheet.columns).index('Fast')
                    else:
                        for i, col_name in enumerate(column_names):
                            if col_name == 'Fast':
                                fast_col_idx = i
                                break
                    
                    if fast_col_idx is not None:
                        fast_col_letter = get_column_letter(fast_col_idx)
                        cell = worksheet[f'{fast_col_letter}{row}']
                        if cell.value is not None:
                            cell.number_format = '#,##0.000'
                except:
                    pass
                
                # Base 컬럼 - 소수점 셋째자리까지 (공통 규칙)
                try:
                    base_col_idx = None
                    if df_sheet is not None and 'Base' in df_sheet.columns:
                        base_col_idx = list(df_sheet.columns).index('Base')
                    else:
                        for i, col_name in enumerate(column_names):
                            if col_name == 'Base':
                                base_col_idx = i
                                break
                    
                    if base_col_idx is not None:
                        base_col_letter = get_column_letter(base_col_idx)
                        cell = worksheet[f'{base_col_letter}{row}']
                        if cell.value is not None:
                            cell.number_format = '#,##0.000'
                except:
                    pass
                
                # 1HMSFast 컬럼 - Base와 동일한 포맷 (소수점 셋째자리까지)
                try:
                    hmfast_col_idx = None
                    if df_sheet is not None and '1HMSFast' in df_sheet.columns:
                        hmfast_col_idx = list(df_sheet.columns).index('1HMSFast')
                    else:
                        for i, col_name in enumerate(column_names):
                            if col_name == '1HMSFast':
                                hmfast_col_idx = i
                                break
                    
                    if hmfast_col_idx is not None:
                        hmfast_col_letter = get_column_letter(hmfast_col_idx)
                        cell = worksheet[f'{hmfast_col_letter}{row}']
                        if cell.value is not None:
                            cell.number_format = '#,##0.000'
                except:
                    pass
                
                # 1HMSF 컬럼 - 소수점 셋째자리까지 (5M 시트용)
                try:
                    hmsf_col_idx = None
                    if df_sheet is not None and '1HMSF' in df_sheet.columns:
                        hmsf_col_idx = list(df_sheet.columns).index('1HMSF')
                    else:
                        for i, col_name in enumerate(column_names):
                            if col_name == '1HMSF':
                                hmsf_col_idx = i
                                break
                    
                    if hmsf_col_idx is not None:
                        hmsf_col_letter = get_column_letter(hmsf_col_idx)
                        cell = worksheet[f'{hmsf_col_letter}{row}']
                        if cell.value is not None:
                            cell.number_format = '#,##0.000'  # 숫자로 인식되면서 표시만 3자리로 함
                except:
                    pass
                
                # SPRD 컬럼 - 백분율, 소수점 셋째자리까지 (15분봉 시트에만)
                if sheet_name == f'{TICKER}USDT15M':
                    try:
                        sprd_col_idx = None
                        if df_sheet is not None and 'SPRD' in df_sheet.columns:
                            sprd_col_idx = list(df_sheet.columns).index('SPRD')
                        else:
                            for i, col_name in enumerate(column_names):
                                if col_name == 'SPRD':
                                    sprd_col_idx = i
                                    break
                        
                        if sprd_col_idx is not None:
                            sprd_col_letter = get_column_letter(sprd_col_idx)
                            cell = worksheet[f'{sprd_col_letter}{row}']
                            if cell.value is not None:
                                cell.number_format = '0.000%'
                    except Exception as e:
                        pass
                
                # SPRD2 컬럼 - 백분율, 소수점 셋째자리까지 (15분봉 시트에만)
                try:
                    # SPRD2 컬럼의 위치를 동적으로 찾기 (df_sheet에서 직접 찾기)
                    sprd2_col_idx = None
                    if df_sheet is not None and 'SPRD2' in df_sheet.columns:
                        sprd2_col_idx = list(df_sheet.columns).index('SPRD2')
                    else:
                        # df_sheet가 없거나 SPRD2가 없으면 column_names에서 찾기
                        for i, col_name in enumerate(column_names):
                            if col_name == 'SPRD2':
                                sprd2_col_idx = i
                                break
                    
                    if sprd2_col_idx is not None:
                        # Excel 컬럼 문자 계산 (26개 이상일 때 AA, AB... 형식)
                        sprd2_col_letter = get_column_letter(sprd2_col_idx)
                        
                        cell = worksheet[f'{sprd2_col_letter}{row}']
                        if cell.value is not None:
                            # SPRD2는 백분율로 표시, 소수점 셋째자리까지
                            cell.number_format = '0.000%'
                except Exception as e:
                    pass
                
                # 15분봉 시트 전용: StoSP, TP, StoSU, TPC, TPCS, NBS, 1HCL, -1HCL, p 포맷팅
                if sheet_name == f'{TICKER}USDT15M':
                    # StoSP, TP 컬럼 - 종가와 동일한 포맷 (티커별 포맷 통일)
                    for target_col in ['StoSP', 'TP']:
                        try:
                            target_col_idx = None
                            if df_sheet is not None and target_col in df_sheet.columns:
                                target_col_idx = list(df_sheet.columns).index(target_col)
                            else:
                                for i, col_name in enumerate(column_names):
                                    if col_name == target_col:
                                        target_col_idx = i
                                        break
                            
                            if target_col_idx is not None:
                                target_col_letter = get_column_letter(target_col_idx)
                                cell = worksheet[f'{target_col_letter}{row}']
                                if cell.value is not None:
                                    # 종가 열과 동일한 포맷 적용 (9240-9257줄 로직과 일치)
                                    if 'XRP' in sheet_name and 'USDT' in sheet_name:
                                        cell.number_format = '#,##0.0000'  # XRP: 소수점 4자리
                                    elif 'USDT' in sheet_name or 'USD' in sheet_name:
                                        cell.number_format = '#,##0.00'    # BTC, ETH, SOL: 소수점 2자리
                                    elif sheet_name == 'USDKRW':
                                        cell.number_format = '#,##0.00'    # 달러환율: 소수점 2자리
                                    else:
                                        cell.number_format = '#,##0'       # KRW 시트: 정수
                        except Exception as e:
                            pass
                    
                    # StoSU 컬럼 - 소수점 포맷
                    try:
                        stosu_col_idx = None
                        if df_sheet is not None and 'StoSU' in df_sheet.columns:
                            stosu_col_idx = list(df_sheet.columns).index('StoSU')
                        else:
                            for i, col_name in enumerate(column_names):
                                if col_name == 'StoSU':
                                    stosu_col_idx = i
                                    break
                        
                        if stosu_col_idx is not None:
                            stosu_col_letter = get_column_letter(stosu_col_idx)
                            cell = worksheet[f'{stosu_col_letter}{row}']
                            if cell.value is not None:
                                cell.number_format = '#,##0.00'
                    except Exception as e:
                        pass
                    
                    # TPC 컬럼 - 정수 포맷
                    try:
                        tpc_col_idx = None
                        if df_sheet is not None and 'TPC' in df_sheet.columns:
                            tpc_col_idx = list(df_sheet.columns).index('TPC')
                        else:
                            for i, col_name in enumerate(column_names):
                                if col_name == 'TPC':
                                    tpc_col_idx = i
                                    break
                        
                        if tpc_col_idx is not None:
                            tpc_col_letter = get_column_letter(tpc_col_idx)
                            cell = worksheet[f'{tpc_col_letter}{row}']
                            if cell.value is not None:
                                cell.number_format = '#,##0'
                    except Exception as e:
                        pass
                    
                    # TPCS 컬럼 - 소수점 포맷
                    try:
                        tpcs_col_idx = None
                        if df_sheet is not None and 'TPCS' in df_sheet.columns:
                            tpcs_col_idx = list(df_sheet.columns).index('TPCS')
                        else:
                            for i, col_name in enumerate(column_names):
                                if col_name == 'TPCS':
                                    tpcs_col_idx = i
                                    break
                        
                        if tpcs_col_idx is not None:
                            tpcs_col_letter = get_column_letter(tpcs_col_idx)
                            cell = worksheet[f'{tpcs_col_letter}{row}']
                            if cell.value is not None:
                                cell.number_format = '#,##0.00'
                    except Exception as e:
                        pass
                    
                    # NBS 컬럼 - 정수 포맷
                    try:
                        nbs_col_idx = None
                        if df_sheet is not None and 'NBS' in df_sheet.columns:
                            nbs_col_idx = list(df_sheet.columns).index('NBS')
                        else:
                            for i, col_name in enumerate(column_names):
                                if col_name == 'NBS':
                                    nbs_col_idx = i
                                    break
                        
                        if nbs_col_idx is not None:
                            nbs_col_letter = get_column_letter(nbs_col_idx)
                            cell = worksheet[f'{nbs_col_letter}{row}']
                            if cell.value is not None:
                                cell.number_format = '#,##0'
                    except Exception as e:
                        pass
                    
                    # 1HCL, -1HCL, p 컬럼 - 정수 포맷
                    for target_col in ['1HCL', '-1HCL', 'p']:
                        try:
                            target_col_idx = None
                            if df_sheet is not None and target_col in df_sheet.columns:
                                target_col_idx = list(df_sheet.columns).index(target_col)
                            else:
                                for i, col_name in enumerate(column_names):
                                    if col_name == target_col:
                                        target_col_idx = i
                                        break
                            
                            if target_col_idx is not None:
                                target_col_letter = get_column_letter(target_col_idx)
                                cell = worksheet[f'{target_col_letter}{row}']
                                if cell.value is not None:
                                    cell.number_format = '#,##0'
                        except Exception as e:
                            pass
                
                # 1H 시트 전용: 1HCLASS, -1HCLASS, p1H 포맷팅
                if sheet_name == f'{TICKER}USDT1H':
                    for target_col in ['1HCLASS', '-1HCLASS', 'p1H']:
                        try:
                            target_col_idx = None
                            if df_sheet is not None and target_col in df_sheet.columns:
                                target_col_idx = list(df_sheet.columns).index(target_col)
                            else:
                                for i, col_name in enumerate(column_names):
                                    if col_name == target_col:
                                        target_col_idx = i
                                        break
                            
                            if target_col_idx is not None:
                                target_col_letter = get_column_letter(target_col_idx)
                                cell = worksheet[f'{target_col_letter}{row}']
                                if cell.value is not None:
                                    cell.number_format = '#,##0'  # 정수 포맷
                        except Exception as e:
                            pass
                
                # 4or1 컬럼 - 소수점 둘째자리까지 (공통 규칙)
                try:
                    fouror1_col_idx = None
                    if df_sheet is not None and '4or1' in df_sheet.columns:
                        fouror1_col_idx = list(df_sheet.columns).index('4or1')
                    else:
                        for i, col_name in enumerate(column_names):
                            if col_name == '4or1':
                                fouror1_col_idx = i
                                break
                    
                    if fouror1_col_idx is not None:
                        fouror1_col_letter = get_column_letter(fouror1_col_idx)
                        cell = worksheet[f'{fouror1_col_letter}{row}']
                        if cell.value is not None:
                            cell.number_format = '#,##0.00'
                except:
                    pass
                
                # buyside 컬럼 - 소수점 넷째자리까지 (공통 규칙)
                try:
                    buyside_col_idx = None
                    if df_sheet is not None and 'buyside' in df_sheet.columns:
                        buyside_col_idx = list(df_sheet.columns).index('buyside')
                    else:
                        for i, col_name in enumerate(column_names):
                            if col_name == 'buyside':
                                buyside_col_idx = i
                                break
                    
                    if buyside_col_idx is not None:
                        buyside_col_letter = get_column_letter(buyside_col_idx)
                        cell = worksheet[f'{buyside_col_letter}{row}']
                        if cell.value is not None:
                            cell.number_format = '#,##0.0000'
                except:
                    pass
                
                # sellside 컬럼 - 소수점 넷째자리까지 (공통 규칙)
                try:
                    sellside_col_idx = None
                    if df_sheet is not None and 'sellside' in df_sheet.columns:
                        sellside_col_idx = list(df_sheet.columns).index('sellside')
                    else:
                        for i, col_name in enumerate(column_names):
                            if col_name == 'sellside':
                                sellside_col_idx = i
                                break
                    
                    if sellside_col_idx is not None:
                        sellside_col_letter = get_column_letter(sellside_col_idx)
                        cell = worksheet[f'{sellside_col_letter}{row}']
                        if cell.value is not None:
                            cell.number_format = '#,##0.0000'
                except:
                    pass
                
                # 김프 컬럼 제거됨
                
                # SB5M 컬럼 - 텍스트 형식 (15분봉 시트에만, 공통 규칙)
                try:
                    # SB5M 컬럼의 위치를 실제 DataFrame 컬럼에서 찾기
                    sb5m_col_idx = None
                    if df_sheet is not None and 'SB5M' in df_sheet.columns:
                        sb5m_col_idx = list(df_sheet.columns).index('SB5M')
                    else:
                        # 폴백: column_names에서 찾기
                        for i, col_name in enumerate(column_names):
                            if col_name == 'SB5M':
                                sb5m_col_idx = i
                                break
                    
                    if sb5m_col_idx is not None:
                        # Excel 컬럼 문자 계산 (26개 이상일 때 AA, AB... 형식)
                        sb5m_col_letter = get_column_letter(sb5m_col_idx)
                        
                        cell = worksheet[f'{sb5m_col_letter}{row}']
                        if cell.value is not None:
                            # SB5M은 텍스트 값이므로 특별한 포맷팅 없음 (기본 텍스트, 공통 규칙)
                            pass
                except Exception as e:
                    pass
                
                # ORDER 컬럼 - 텍스트 형식 (15분봉 시트에만, 공통 규칙)
                try:
                    # ORDER 컬럼의 위치를 실제 DataFrame 컬럼에서 찾기
                    order_col_idx = None
                    if df_sheet is not None and 'ORDER' in df_sheet.columns:
                        order_col_idx = list(df_sheet.columns).index('ORDER')
                    else:
                        # 폴백: column_names에서 찾기
                        for i, col_name in enumerate(column_names):
                            if col_name == 'ORDER':
                                order_col_idx = i
                                break
                    
                    if order_col_idx is not None:
                        # Excel 컬럼 문자 계산 (26개 이상일 때 AA, AB... 형식)
                        order_col_letter = get_column_letter(order_col_idx)
                        
                        cell = worksheet[f'{order_col_letter}{row}']
                        if cell.value is not None:
                            # ORDER는 텍스트 값이므로 특별한 포맷팅 없음 (기본 텍스트, 공통 규칙)
                            pass
                except Exception as e:
                    pass
                
                # dateM 컬럼 - 정수 형식 (15분봉 시트에만, 공통 규칙)
                try:
                    # dateM 컬럼의 위치를 실제 DataFrame 컬럼에서 찾기
                    datem_col_idx = None
                    if df_sheet is not None and 'dateM' in df_sheet.columns:
                        datem_col_idx = list(df_sheet.columns).index('dateM')
                    else:
                        # 폴백: column_names에서 찾기
                        for i, col_name in enumerate(column_names):
                            if col_name == 'dateM':
                                datem_col_idx = i
                                break
                    
                    if datem_col_idx is not None:
                        # Excel 컬럼 문자 계산 (26개 이상일 때 AA, AB... 형식)
                        datem_col_letter = get_column_letter(datem_col_idx)
                        
                        cell = worksheet[f'{datem_col_letter}{row}']
                        if cell.value is not None:
                            # dateM은 정수로 표시 (캔들 개수, 공통 규칙)
                            cell.number_format = '0'
                except Exception as e:
                    pass
                
                # KSC 컬럼 - 정수 형식 (15분봉 시트에만, 공통 규칙)
                try:
                    # KSC 컬럼의 위치를 실제 DataFrame 컬럼에서 찾기
                    ksc_col_idx = None
                    if df_sheet is not None and 'KSC' in df_sheet.columns:
                        ksc_col_idx = list(df_sheet.columns).index('KSC')
                    else:
                        # 폴백: column_names에서 찾기
                        for i, col_name in enumerate(column_names):
                            if col_name == 'KSC':
                                ksc_col_idx = i
                                break
                    
                    if ksc_col_idx is not None:
                        # Excel 컬럼 문자 계산 (26개 이상일 때 AA, AB... 형식)
                        ksc_col_letter = get_column_letter(ksc_col_idx)
                        
                        cell = worksheet[f'{ksc_col_letter}{row}']
                        if cell.value is not None:
                            # KSC는 정수로 표시 (kill 카운트) 또는 "Bomb" 문자열
                            if str(cell.value).strip() == 'Bomb':
                                # "Bomb" 문자열인 경우 포맷팅 없음 (텍스트)
                                pass
                            else:
                                # 숫자인 경우 정수 형식
                                cell.number_format = '0'
                except Exception as e:
                    pass
                
                # PRFT 컬럼 - 정수 형식 (15분봉 시트에만, 공통 규칙)
                try:
                    # PRFT 컬럼의 위치를 실제 DataFrame 컬럼에서 찾기
                    prft_col_idx = None
                    if df_sheet is not None and 'PRFT' in df_sheet.columns:
                        prft_col_idx = list(df_sheet.columns).index('PRFT')
                    else:
                        # 폴백: column_names에서 찾기
                        for i, col_name in enumerate(column_names):
                            if col_name == 'PRFT':
                                prft_col_idx = i
                                break
                    
                    if prft_col_idx is not None:
                        # Excel 컬럼 문자 계산 (26개 이상일 때 AA, AB... 형식)
                        prft_col_letter = get_column_letter(prft_col_idx)
                        
                        cell = worksheet[f'{prft_col_letter}{row}']
                        if cell.value is not None:
                            # PRFT는 정수로 표시 (prft 카운트) 또는 "PRFT" 문자열
                            if str(cell.value).strip() == 'PRFT':
                                # "PRFT" 문자열인 경우 포맷팅 없음 (텍스트)
                                pass
                            else:
                                # 숫자인 경우 정수 형식
                                cell.number_format = '0'
                except Exception as e:
                    pass
                
                # LD 컬럼 - 소수점 셋째자리까지 (15분봉 시트에만, 공통 규칙)
                try:
                    # LD 컬럼의 위치를 실제 DataFrame 컬럼에서 찾기
                    ld_col_idx = None
                    if df_sheet is not None and 'LD' in df_sheet.columns:
                        ld_col_idx = list(df_sheet.columns).index('LD')
                    else:
                        # 폴백: column_names에서 찾기
                        for i, col_name in enumerate(column_names):
                            if col_name == 'LD':
                                ld_col_idx = i
                                break
                    
                    if ld_col_idx is not None:
                        # Excel 컬럼 문자 계산 (26개 이상일 때 AA, AB... 형식)
                        ld_col_letter = get_column_letter(ld_col_idx)
                        
                        cell = worksheet[f'{ld_col_letter}{row}']
                        if cell.value is not None:
                            # LD는 소수점 3자리까지 표시 (콤마 제거, 공통 규칙)
                            cell.number_format = '0.000'
                except Exception as e:
                    pass
                
                # 텍스트 컬럼들 (Sell, Buy, SB1H, SB1D) - 포맷팅 없음 (공통 규칙)
                try:
                    for text_col_name in ['Sell', 'Buy', 'SB1H', 'SB1D']:
                        text_col_idx = None
                        if df_sheet is not None and text_col_name in df_sheet.columns:
                            text_col_idx = list(df_sheet.columns).index(text_col_name)
                        else:
                            # 폴백: column_names에서 찾기
                            for i, col_name in enumerate(column_names):
                                if col_name == text_col_name:
                                    text_col_idx = i
                                    break
                        
                        if text_col_idx is not None:
                            text_col_letter = get_column_letter(text_col_idx)
                            cell = worksheet[f'{text_col_letter}{row}']
                            if cell.value is not None:
                                # 텍스트 컬럼은 포맷팅 없음 (기본 텍스트, 공통 규칙)
                                pass
                except Exception as e:
                    pass
                
                # Samount, Bamount 컬럼 - 티커별 USDT 정밀도 적용
                try:
                    # 시트 이름에서 티커 추출 (예: "BTC 15M" → "BTC")
                    sheet_ticker = sheet_name.split()[0] if ' ' in sheet_name else 'XRP'
                    symbol = f"{sheet_ticker}USDT"
                    usdt_precision = SYMBOL_USDT_PRECISION.get(symbol, 5)
                    
                    # Samount/Bamount 컬럼 찾기
                    for col_name_target in ['Samount', 'Bamount', 'SamountW', 'BamountW', 'Samount1D', 'Bamount1D']:
                        col_idx_target = None
                        if df_sheet is not None and col_name_target in df_sheet.columns:
                            col_idx_target = list(df_sheet.columns).index(col_name_target)
                        else:
                            # 폴백: column_names에서 찾기
                            for i, col_name in enumerate(column_names):
                                if col_name == col_name_target:
                                    col_idx_target = i
                                    break
                        
                        if col_idx_target is not None:
                            col_letter = get_column_letter(col_idx_target)
                            cell = worksheet[f'{col_letter}{row}']
                            if cell.value is not None:
                                # Samount/Bamount는 4자리로 통일 표시
                                cell.number_format = '0.0000'
                except Exception as e:
                    pass
                
                # dateM, KSC, PRFT, LD 포맷 다시 적용 (Samount/Bamount 이후에 덮어씌워지지 않도록)
                try:
                    # dateM 컬럼 포맷 재적용
                    if df_sheet is not None and 'dateM' in df_sheet.columns:
                        datem_col_idx = list(df_sheet.columns).index('dateM')
                        datem_col_letter = get_column_letter(datem_col_idx)
                        cell = worksheet[f'{datem_col_letter}{row}']
                        if cell.value is not None:
                            cell.number_format = '0'  # 정수 형식
                    
                    # KSC 컬럼 포맷 재적용
                    if df_sheet is not None and 'KSC' in df_sheet.columns:
                        ksc_col_idx = list(df_sheet.columns).index('KSC')
                        ksc_col_letter = get_column_letter(ksc_col_idx)
                        cell = worksheet[f'{ksc_col_letter}{row}']
                        if cell.value is not None:
                            # "Bomb" 문자열인 경우 포맷팅 없음 (텍스트)
                            if str(cell.value).strip() == 'Bomb':
                                pass
                            else:
                                cell.number_format = '0'  # 정수 형식
                    
                    # PRFT 컬럼 포맷 재적용
                    if df_sheet is not None and 'PRFT' in df_sheet.columns:
                        prft_col_idx = list(df_sheet.columns).index('PRFT')
                        prft_col_letter = get_column_letter(prft_col_idx)
                        cell = worksheet[f'{prft_col_letter}{row}']
                        if cell.value is not None:
                            # "PRFT" 문자열인 경우 포맷팅 없음 (텍스트)
                            if str(cell.value).strip() == 'PRFT':
                                pass
                            else:
                                cell.number_format = '0'  # 정수 형식
                    
                    # LD 컬럼 포맷 재적용
                    if df_sheet is not None and 'LD' in df_sheet.columns:
                        ld_col_idx = list(df_sheet.columns).index('LD')
                        ld_col_letter = get_column_letter(ld_col_idx)
                        cell = worksheet[f'{ld_col_letter}{row}']
                        if cell.value is not None:
                            cell.number_format = '0.000'  # 소수점 3자리
                except Exception as e:
                    pass
                
            # 서식 적용 후 실제 표시 길이를 읽어서 열너비 계산
            # 실제 DataFrame의 컬럼 순서를 사용 (column_names가 아닌 df_sheet.columns)
            actual_columns = list(df_sheet.columns)
            for col_idx, col_name in enumerate(actual_columns):
                if col_idx >= len(actual_columns):
                    continue
                
                # 헤더 길이
                header_length = len(str(col_name))
                
                # 서식이 적용된 실제 표시 길이 계산 (1행 헤더, 2~10행 데이터에서 읽기)
                max_data_length = 0
                max_check_row = min(10, worksheet.max_row)  # 최대 10행까지 확인
                if worksheet.max_row >= 2:  # 최소 2행(헤더+데이터1) 이상 존재하는 경우
                    try:
                        # 1행(헤더), 2~10행(데이터)의 실제 셀 값을 읽어서 표시 길이 계산
                        for check_row in range(1, max_check_row + 1):
                            cell = worksheet.cell(row=check_row, column=col_idx + 1)
                            if cell.value is not None:
                                # 각 행의 표시 길이를 계산하여 최대값 사용
                                row_length = 0
                                
                                # 1행(헤더)인 경우 단순 문자열 길이
                                if check_row == 1:
                                    row_length = len(str(cell.value))
                                else:
                                    # 2행, 3행(데이터)인 경우 서식 적용된 길이 계산
                                    if col_name in ['하단', '상단']:
                                        # 백분율 형식: 0.1234 → "12.34%"
                                        try:
                                            numeric_value = float(cell.value)
                                            row_length = len(f"{numeric_value:.2%}")
                                        except (ValueError, TypeError):
                                            row_length = len(str(cell.value))
                                    elif col_name in ['SFast', 'Fast', 'Base', '1HMSFast', '1HMSF']:
                                        # 소수점 3자리 형식: 1234.5678 → "1,234.568"
                                        try:
                                            numeric_value = float(cell.value)
                                            row_length = len(f"{numeric_value:,.3f}")
                                        except (ValueError, TypeError):
                                            row_length = len(str(cell.value))
                                    elif col_name in ['buyside', 'sellside']:
                                        # 소수점 4자리 형식: 1234.5678 → "1,234.5678"
                                        try:
                                            numeric_value = float(cell.value)
                                            row_length = len(f"{numeric_value:,.4f}")
                                        except (ValueError, TypeError):
                                            row_length = len(str(cell.value))
                                    
                                    elif col_name == 'ORDER':
                                        # ORDER는 텍스트 값: 'Sell5', 'Sell10', 'Buy5', '' (빈값)
                                        try:
                                            order_value = str(cell.value)
                                            if order_value.startswith('Sell') or order_value.startswith('Buy'):
                                                row_length = len(order_value) + 1  # 여유분 1글자
                                            else:
                                                row_length = 4  # 기본값 (빈값이거나 기타)
                                        except (ValueError, TypeError):
                                            row_length = 4  # 기본값
                                    elif col_name == 'dateM':
                                        # dateM은 정수 형식: 123 → "123"
                                        try:
                                            numeric_value = float(cell.value)
                                            row_length = len(f"{numeric_value:.0f}")  # 정수: "123"
                                        except (ValueError, TypeError):
                                            row_length = len(str(cell.value))
                                    elif col_name == 'KSC':
                                        # KSC는 정수 형식 또는 "Bomb" 문자열
                                        try:
                                            # "Bomb" 문자열인 경우
                                            if str(cell.value).strip() == 'Bomb':
                                                row_length = len('Bomb')
                                            else:
                                                numeric_value = float(cell.value)
                                                row_length = len(f"{numeric_value:.0f}")  # 정수: "123"
                                        except (ValueError, TypeError):
                                            row_length = len(str(cell.value))
                                    elif col_name in ['1HCL', '-1HCL', 'p']:
                                        # 15M 시트: 1HCL, -1HCL, p - 정수 포맷 (#,##0)
                                        try:
                                            numeric_value = float(cell.value)
                                            row_length = len(f"{int(numeric_value):,}") if abs(numeric_value) >= 1000 else len(f"{int(numeric_value)}")
                                        except (ValueError, TypeError):
                                            row_length = len(str(cell.value))
                                    elif col_name in ['1HCLASS', '-1HCLASS', 'p1H']:
                                        # 1H 시트: 1HCLASS, -1HCLASS, p1H - 정수 포맷 (#,##0), 값 범위 0~3 또는 -3~0
                                        try:
                                            numeric_value = float(cell.value)
                                            row_length = len(f"{int(numeric_value)}")  # 1자리
                                        except (ValueError, TypeError):
                                            row_length = len(str(cell.value))
                                    elif col_name == 'PRFT':
                                        # PRFT는 정수 형식 또는 "PRFT" 문자열
                                        try:
                                            # "PRFT" 문자열인 경우
                                            if str(cell.value).strip() == 'PRFT':
                                                row_length = len('PRFT')
                                            else:
                                                numeric_value = float(cell.value)
                                                row_length = len(f"{numeric_value:.0f}")  # 정수: "123"
                                        except (ValueError, TypeError):
                                            row_length = len(str(cell.value))
                                    elif col_name == 'LD':
                                        # LD는 소수점 3자리 형식(콤마 제거): 1234.5678 → "1234.568"
                                        try:
                                            numeric_value = float(cell.value)
                                            row_length = len(f"{numeric_value:.3f}")  # 콤마 없음: "1234.568"
                                        except (ValueError, TypeError):
                                            row_length = len(str(cell.value))
                                    elif col_name == 'SPRD':
                                        # SPRD는 백분율 형식 (소수점 셋째자리): 0.0039 → "0.390%"
                                        # 포맷: 0.000% (예: "0.390%", "12.340%")
                                        try:
                                            numeric_value = float(cell.value)
                                            formatted_value = f"{numeric_value:.3%}"
                                            row_length = max(len(formatted_value), 7)  # 최소 7자리 (12.340% 기준)
                                        except (ValueError, TypeError):
                                            row_length = max(len(str(cell.value)), 7)
                                    elif col_name == 'SPRD2':
                                        # SPRD2는 백분율 형식 (소수점 셋째자리): 0.0039 → "0.390%"
                                        # 최소 너비는 "0.390%" (6자리) 또는 "12.340%" (7자리) 기준
                                        try:
                                            numeric_value = float(cell.value)
                                            formatted_value = f"{numeric_value:.3%}"
                                            row_length = max(len(formatted_value), 7)  # 최소 7자리 (12.340% 기준)
                                        except (ValueError, TypeError):
                                            row_length = max(len(str(cell.value)), 7)
                                    elif col_name == 'StoSU':
                                        # StoSU는 소수점 둘째자리 형식: 1234.56 → "1,234.56"
                                        try:
                                            numeric_value = float(cell.value)
                                            row_length = len(f"{numeric_value:,.2f}")
                                        except (ValueError, TypeError):
                                            row_length = len(str(cell.value))
                                    elif col_name in ['Samount', 'Bamount', 'SamountW', 'BamountW', 'Samount1D', 'Bamount1D']:
                                        # Samount/Bamount는 4자리로 통일 표시
                                        try:
                                            numeric_value = float(cell.value)
                                            row_length = len(f"{numeric_value:.4f}")  # 모든 티커: "1.7234"
                                        except (ValueError, TypeError):
                                            row_length = len(str(cell.value))
                                    elif col_name == '4or1':
                                        # 소수점 둘째자리 형식: 1234.56 → "1,234.56"
                                        try:
                                            numeric_value = float(cell.value)
                                            row_length = len(f"{numeric_value:,.2f}")
                                        except (ValueError, TypeError):
                                            row_length = len(str(cell.value))
                                    elif col_name in ['종', '시', '고', '저', 'SMA3', 'SMA5', 'SMA7', 'SMA10', 'SMA12', 'SMAF', 'SMA20', 'SMA25', 'SMA28', 'SMA40', 'SMA15', 'SMA35', 'SMA50', 'SMA80', 'SMA100', 'SMA200', 'SMA400', 'SMA800']:  # 1H4x 시트용 SMA12, SMA28, SMA40, SMA80 추가, 15M SMAF 추가, 1분봉용 SMA15/SMA25/SMA35/SMA50/SMA100 추가
                                        # 가격/지표 형식: 심볼별 다른 정밀도
                                        try:
                                            numeric_value = float(cell.value)
                                            if 'XRP' in sheet_name and 'USDT' in sheet_name:
                                                row_length = len(f"{numeric_value:,.4f}")  # XRP: "2.4406"
                                            elif 'USDT' in sheet_name or 'USD' in sheet_name:
                                                row_length = len(f"{numeric_value:,.2f}")  # BTC, ETH, SOL, BNB: "67,450.32"
                                            elif sheet_name == 'USDKRW':
                                                row_length = len(f"{numeric_value:,.2f}")  # 달러환율: "1,380.45"
                                            else:
                                                row_length = len(f"{numeric_value:,.0f}")  # KRW 시트: "3,250,000"
                                        except (ValueError, TypeError):
                                            row_length = len(str(cell.value))
                                    elif col_name in ['Max15', 'Min15', 'Max25', 'Min25', 'Max50', 'Min50', 'Max70', 'Min70', 'Max100', 'Min100', 'Max200', 'Min200', 'Max400', 'Min400']:
                                        # Max/Min 형식: 심볼별 다른 정밀도 (모든 Max/Min 변형 포함, 1분봉용 Max400/Min400 추가)
                                        try:
                                            numeric_value = float(cell.value)
                                            if 'XRP' in sheet_name and 'USDT' in sheet_name:
                                                row_length = len(f"{numeric_value:,.4f}")
                                            elif 'USDT' in sheet_name or 'USD' in sheet_name:
                                                row_length = len(f"{numeric_value:,.2f}")
                                            elif sheet_name == 'USDKRW':
                                                row_length = len(f"{numeric_value:,.2f}")
                                            else:
                                                row_length = len(f"{numeric_value:,.0f}")
                                        except (ValueError, TypeError):
                                            row_length = len(str(cell.value))
                                    elif col_name == 'Vol.':
                                        # 거래량 형식: 심볼별 다른 정밀도
                                        try:
                                            numeric_value = float(cell.value)
                                            if 'XRP' in sheet_name and 'USDT' in sheet_name:
                                                row_length = len(f"{numeric_value:,.4f}")
                                            elif 'USDT' in sheet_name or 'USD' in sheet_name:
                                                row_length = len(f"{numeric_value:,.2f}")
                                            elif sheet_name == 'USDKRW':
                                                row_length = len(f"{numeric_value:,.2f}")
                                            else:
                                                row_length = len(f"{numeric_value:,.0f}")
                                        except (ValueError, TypeError):
                                            row_length = len(str(cell.value))
                                    else:
                                        # 텍스트 형식: 그대로 문자열 길이
                                        row_length = len(str(cell.value))
                                
                                # 각 행의 길이를 비교하여 최대값 업데이트
                                max_data_length = max(max_data_length, row_length)
                    except Exception as e:
                        max_data_length = len(str(col_name))  # 폴백: 헤더 길이
                
                # 최대 길이 계산 (헤더와 데이터 중 큰 값)
                max_length = max(header_length, max_data_length)
                
                # ChatGPT 제안: 정확한 글자수 기반 열너비 설정
                import math
                
                def _trunc(x):
                    return math.trunc(x)
                
                def width_to_pixels(width, mdw):
                    return _trunc(((256*width + _trunc(128/mdw)) / 256.0) * mdw)
                
                def pixels_to_display_chars(pixels, mdw):
                    return _trunc(((pixels - 5) / mdw) * 100 + 0.5) / 100.0
                
                def display_from_width(width, mdw):
                    return pixels_to_display_chars(width_to_pixels(width, mdw), mdw)
                
                def width_for_display_chars(target_chars, mdw=8, lo=0.0, hi=100.0):
                    step = 1.0 / 256.0
                    n = int((hi - lo) / step) + 1
                    for i in range(n):
                        w = lo + i * step
                        if abs(display_from_width(w, mdw) - target_chars) < 1e-9:
                            return w
                    return None
                
                def set_colwidth_by_chars(ws, col_letter, chars, mdw=8):
                    w = width_for_display_chars(chars, mdw=mdw)
                    if w is None:
                        raise ValueError(f"목표 {chars}글자에 정확히 맞는 width를 찾지 못했습니다.")
                    ws.column_dimensions[col_letter].width = w
                    return w
                
                # 목표 글자수에 맞는 정확한 width 설정
                target_chars = max_length  # 여유분 없음
                
                col_letter = get_column_letter(col_idx)
                
                try:
                    exact_width = set_colwidth_by_chars(worksheet, col_letter, target_chars, mdw=8)
                except ValueError as e:
                    # 폴백: 기본 방식
                    adjusted_width = max_length
                    worksheet.column_dimensions[col_letter].width = adjusted_width
            
            # 15분봉 시트 전용: StoSP, TP 열 너비 조정 (종가 열과 동일한 너비)
            if sheet_name == f'{TICKER}USDT15M':
                for target_col in ['StoSP', 'TP']:
                    if target_col in column_names:
                        try:
                            target_idx = column_names.index(target_col)
                            target_letter = get_column_letter(target_idx)
                            
                            # 종가 열의 인덱스 찾기
                            종_col_idx = None
                            for i, cn in enumerate(column_names):
                                if cn == '종':
                                    종_col_idx = i
                                    break
                            
                            if 종_col_idx is not None:
                                종_col_letter = get_column_letter(종_col_idx)
                                # 종가 열의 너비를 먼저 계산 (티커별 포맷 적용, 종가 열의 실제 포맷과 동일)
                                if df_sheet is not None and len(df_sheet) > 0 and '종' in df_sheet.columns:
                                    종_value = df_sheet.iloc[0]['종']
                                    if not pd.isna(종_value):
                                        # 종가 열의 실제 포맷과 동일하게 계산 (9953-9966 라인 로직과 일치)
                                        if 'XRP' in sheet_name and 'USDT' in sheet_name:
                                            종_formatted = f"{종_value:,.4f}"  # XRP: "2.4406"
                                        elif 'USDT' in sheet_name or 'USD' in sheet_name:
                                            종_formatted = f"{종_value:,.2f}"  # BTC, ETH, SOL, BNB: "67,450.32"
                                        elif sheet_name == 'USDKRW':
                                            종_formatted = f"{종_value:,.2f}"  # 달러환율: "1,380.45"
                                        else:
                                            종_formatted = f"{종_value:,.0f}"  # KRW 시트: "3,250,000"
                                        종_max_length = max(len('종'), len(종_formatted))
                                    else:
                                        종_max_length = len('종')
                                else:
                                    종_max_length = len('종')
                                
                                try:
                                    # 종가 열의 너비를 먼저 설정
                                    set_colwidth_by_chars(worksheet, 종_col_letter, 종_max_length, mdw=8)
                                    # StoSP/TP도 동일한 너비로 설정
                                    set_colwidth_by_chars(worksheet, target_letter, 종_max_length, mdw=8)
                                except ValueError:
                                    # 폴백: 종가 열의 너비를 직접 읽어서 적용
                                    if 종_col_letter in worksheet.column_dimensions and worksheet.column_dimensions[종_col_letter].width:
                                        worksheet.column_dimensions[target_letter].width = worksheet.column_dimensions[종_col_letter].width
                                    else:
                                        worksheet.column_dimensions[target_letter].width = 종_max_length
                        except Exception:
                            pass
                
                # TPCS 열 너비를 StoSU 열 너비와 동일하게 설정
                if 'StoSU' in column_names and 'TPCS' in column_names:
                    try:
                        stosu_idx = column_names.index('StoSU')
                        tpcs_idx = column_names.index('TPCS')
                        stosu_letter = get_column_letter(stosu_idx)
                        tpcs_letter = get_column_letter(tpcs_idx)
                        
                        # StoSU 열의 너비 계산 (소수점 2자리 포맷: #,##0.00)
                        if df_sheet is not None and len(df_sheet) > 0 and 'StoSU' in df_sheet.columns:
                            stosu_value = df_sheet.iloc[0]['StoSU']
                            if not pd.isna(stosu_value):
                                stosu_formatted = f"{stosu_value:,.2f}"  # "1,234.56"
                                stosu_max_length = max(len('StoSU'), len(stosu_formatted))
                            else:
                                stosu_max_length = len('StoSU')
                        else:
                            stosu_max_length = len('StoSU')
                        
                        try:
                            # StoSU 열의 너비를 먼저 설정
                            set_colwidth_by_chars(worksheet, stosu_letter, stosu_max_length, mdw=8)
                            # TPCS도 동일한 너비로 설정
                            set_colwidth_by_chars(worksheet, tpcs_letter, stosu_max_length, mdw=8)
                        except ValueError:
                            # 폴백: StoSU 열의 너비를 직접 읽어서 적용
                            if stosu_letter in worksheet.column_dimensions and worksheet.column_dimensions[stosu_letter].width:
                                worksheet.column_dimensions[tpcs_letter].width = worksheet.column_dimensions[stosu_letter].width
                            else:
                                worksheet.column_dimensions[tpcs_letter].width = stosu_max_length
                    except Exception:
                        pass
            
            # 헤더 볼드처리 제거 (openpyxl 방식) - 모든 컬럼에 적용
            from openpyxl.styles import Font
            header_font = Font(bold=False)
            
            # 워크시트의 모든 컬럼에 대해 볼드 제거 적용
            for col in range(1, worksheet.max_column + 1):
                try:
                    cell = worksheet.cell(row=1, column=col)
                    if cell.value is not None:  # 헤더가 존재하는 경우만
                        cell.font = header_font
                except Exception as e:
                    continue
            
            # 틀고정 설정 (각 시트별 D2 기준)
            try:
                worksheet.freeze_panes = 'D2'
            except Exception as e:
                print(f"{get_timestamp()} [{stage_prefix}] ⚠️ 시트 '{sheet_name}' 틀고정 설정 실패: {e}")

        # 파일 저장 전에 writer를 명시적으로 닫기
        writer.close()
        writer = None
        
        print(f"{get_timestamp()} [{stage_prefix}] ✅ 저장 완료: {os.path.basename(save_path)}")
        print(f"{get_timestamp()} [{stage_prefix}] 📊 5개 시트 생성 완료: {TICKER}USDT15M, {TICKER}USDT5M, {TICKER}USDT1H, {TICKER}USDT1D, {TICKER}USDTW")
        print(f"{get_timestamp()} [{stage_prefix}] 📊 주력 거래: 바이낸스 USDT (시트 1~5)")
        # 실제 저장된 데이터 개수 계산 (1단계: 수집 후 미완성 제거, 2단계: previous + 새 데이터)
        if skip_first_row:
            # 1단계: 수집 후 미완성 1개 제거된 개수
            actual_5m = minute5_count - 1 if minute5_count > 1 else minute5_count
            actual_15m = minute15_count - 1 if minute15_count > 1 else minute15_count
            actual_1h = hour1_count
            actual_1d = daily_count
        else:
            # 2단계: previous + 새 데이터 (실제 저장된 개수)
            actual_5m = len(df_binance_ticker_5m)
            actual_15m = len(df_binance_ticker_15m)
            actual_1h = len(df_binance_ticker_1h)
            actual_1d = len(df_binance_ticker_1d)
        print(f"{get_timestamp()} [{stage_prefix}] 📊 일봉 {actual_1d}개, 5분봉 {actual_5m}개, 15분봉 {actual_15m}개, 1시간봉 {actual_1h}개, 주봉 {len(df_binance_ticker_weekly)}개 데이터 포함")
        print(f"{get_timestamp()} [{stage_prefix}] 📊 열 너비가 자동으로 조정되었습니다.")
        print(f"{get_timestamp()} [{stage_prefix}] 🎉 모든 작업이 완료되었습니다!")
        
        # 엑셀 파일 저장 완료 후 메모리 정리
        # 폴링 주문 실행 (2단계에서만 실행)
        if not skip_first_row and polling_start_time and ENABLE_TRADING:
            try:
                if len(df_binance_ticker_15m) == 0:
                    print(f"{get_timestamp()} [{stage_prefix}] ⚠️ 주문 실행 스킵: 15분봉 데이터가 비어있습니다.")
                else:
                    latest_order = df_binance_ticker_15m.iloc[0].get("ORDER", "")
                    latest_ksc = df_binance_ticker_15m.iloc[0].get("KSC", 0)
                
                    # LS 시그널(1 또는 -1) 시 선물 전략 실행 — LS 판정된 종가(K)로 스마트 주문 엔진 사용
                    if ENABLE_FUTURES_LS_STRATEGY:
                        latest_ls_raw = df_binance_ticker_15m.iloc[0].get("LS", "")
                        latest_ls = None
                        try:
                            if latest_ls_raw not in ("", None) and pd.notna(latest_ls_raw):
                                v = int(float(latest_ls_raw))
                                if v in (1, -1):
                                    latest_ls = v
                        except (TypeError, ValueError):
                            pass
                        if latest_ls in (1, -1) and TICKER in ROTATION_TICKERS:
                            row0 = df_binance_ticker_15m.iloc[0]
                            K_close = row0.get("종", None)
                            try:
                                K_val = float(K_close) if K_close is not None and pd.notna(K_close) else None
                            except (TypeError, ValueError):
                                K_val = None
                            execute_futures_strategy(latest_ls, f"{TICKER}USDT", stage_prefix=stage_prefix, K=K_val)
                
                if latest_order:
                    print(f"{get_timestamp()} [{stage_prefix}] 🚨 ORDER 신호 감지: {TICKER} {latest_order}")
                    # Samount, Bamount 값 추출
                    latest_samount = df_binance_ticker_15m.iloc[0].get("Samount", 0)
                    latest_bamount = df_binance_ticker_15m.iloc[0].get("Bamount", 0)
                    
                    # NaN 체크
                    if pd.isna(latest_samount):
                        latest_samount = 0
                    if pd.isna(latest_bamount):
                        latest_bamount = 0
                    
                    # ============================================================
                    # KSC 주문량 적용 로직 (Buy 신호에만 적용)
                    # 다른 시트(업비트 등)에도 동일하게 적용 가능
                    # ============================================================
                    # 
                    # [1단계] Multiplier 계산 (수열 규칙)
                    #   - KSC가 3의 배수일 때만 3, 나머지는 0
                    #   - 패턴: 0, 0, 3, 0, 0, 3, 0, 0, 3, ...
                    #   - 함수: calculate_ksc_multiplier(ksc_value, ksc_stack)
                    # 
                    # [2단계] B값 계산 (Bomb 발생 시에만)
                    #   - multiplier == 0: B = ((ksc_stack - 1) % 3) + 1 (1,2,3 반복)
                    #   - multiplier != 0: B = 0
                    #   - 함수: calculate_bomb_b_value(multiplier, ksc_stack)
                    # 
                    # [3단계] Z값 계산
                    #   - Z = multiplier + if(bomb발생, B값, 0)
                    #   - Bomb 미발생: Z = multiplier
                    #   - Bomb 발생: Z = multiplier + B값
                    # 
                    # [4단계] 주문량 계산
                    #   - base_amount = TRADING_UNIT + bamount
                    #   - 주문량 = base_amount × Z
                    #   - 예시: TRADING_UNIT=7, bamount=0, Z=3
                    #     → 주문량 = (7 + 0) × 3 = 21 USDT
                    # ============================================================
                    
                    Z = 1  # 기본값 (multiplier가 없을 때)
                    ksc_numeric = 0  # KSC 값을 숫자로 변환 (조건문 밖에서 초기화)
                    if latest_order in ['Buy5', 'Buy10']:
                        # KSC 값을 숫자로 변환
                        if isinstance(latest_ksc, (int, float)):
                            ksc_numeric = int(latest_ksc)
                        elif isinstance(latest_ksc, str):
                            try:
                                ksc_numeric = int(float(latest_ksc))
                            except:
                                ksc_numeric = 0
                        
                        # KSC stack 값 확인 (숫자만)
                        latest_ksc_stack = df_binance_ticker_15m.iloc[0].get("KSC stack", 0)
                        if pd.isna(latest_ksc_stack):
                            latest_ksc_stack = 0
                        latest_ksc_stack = int(latest_ksc_stack)
                        
                        # Bomb 열 확인
                        latest_bomb = df_binance_ticker_15m.iloc[0].get("Bomb", "")
                        is_bomb = (isinstance(latest_bomb, str) and latest_bomb.strip() == "Bomb")
                        
                        # KSC 스택이 쌓이는 상황인지 확인
                        # - KSC 스택이 쌓이는 상황: KSC > 0 또는 Bomb 발생
                        # - KSC 스택이 쌓이는 상황이 아님: KSC = 0이고 Bomb 아님
                        is_ksc_stack_building = (ksc_numeric > 0 or latest_ksc_stack > 0 or is_bomb)
                        
                        if is_ksc_stack_building:
                            # KSC 스택이 쌓이는 상황: Z값으로 주문 전송 컨트롤 필요
                            # p값: 15M 열 p(= 3+p1H) 우선, 없으면 3 + 1HCL
                            latest_p = df_binance_ticker_15m.iloc[0].get("p", np.nan)
                            if pd.notna(latest_p):
                                try:
                                    p_value = int(float(latest_p))
                                except (TypeError, ValueError):
                                    latest_1hcl = df_binance_ticker_15m.iloc[0].get("1HCL", np.nan)
                                    p_value = 3 + (int(float(latest_1hcl)) if pd.notna(latest_1hcl) else 0)
                            else:
                                latest_1hcl = df_binance_ticker_15m.iloc[0].get("1HCL", np.nan)
                                p_value = 3 + (int(float(latest_1hcl)) if pd.notna(latest_1hcl) else 0)
                            
                            # bomb 발생 시 KSC stack이 0이면 KSC 값을 사용
                            if is_bomb and latest_ksc_stack == 0 and ksc_numeric > 0:
                                latest_ksc_stack = ksc_numeric
                                print(f"{get_timestamp()} [{stage_prefix}] ⚠️ Bomb 발생 시 KSC stack이 0, KSC 값({ksc_numeric})을 사용하여 재계산...")
                            
                            # 1. multiplier 계산 (수열 규칙)
                            # Bomb 발생 시: KSC stack 값을 사용하여 multiplier 계산
                            # 일반 경우: KSC 값을 사용하여 multiplier 계산
                            if is_bomb:
                                # Bomb 발생 시 KSC stack 값을 사용
                                multiplier = calculate_ksc_multiplier(latest_ksc_stack, latest_ksc_stack, p_value)
                            else:
                                # 일반 경우 KSC 값을 사용
                                multiplier = calculate_ksc_multiplier(ksc_numeric, latest_ksc_stack, p_value)
                            
                            # 2. B 값 계산 (Bomb 발생 시)
                            B_value = 0
                            if is_bomb:
                                B_value = calculate_bomb_b_value(multiplier, latest_ksc_stack, p_value)
                                bomb_msg = f"{get_timestamp()} [{stage_prefix}] 💣 Bomb 감지: KSC={latest_ksc}, bomb={latest_bomb}, KSC stack={latest_ksc_stack}, multiplier={multiplier}, B={B_value}, p={p_value}"
                                print(bomb_msg)
                                send_discord_message(bomb_msg)
                            
                            # 3. Z = multiplier(수열) + if(bomb발생, B값, 0)
                            # Z값 계산식: Z = multiplier + if(Bomb 발생, B값, 0)
                            # - Bomb 미발생: Z = multiplier
                            # - Bomb 발생: Z = multiplier + B값
                            # - multiplier와 B값은 독립적으로 계산됨
                            # - Bomb 발생 시 multiplier가 0이어도 B값으로 Z > 0이 되면 주문 가능
                            Z = multiplier + B_value
                            
                            # KSC = 1인 경우: p의 배수가 아니어도 주문 전송 (Z = 0이면 Z = 1로 강제 설정)
                            if ksc_numeric == 1 and Z == 0:
                                Z = 1
                                print(f"{get_timestamp()} [{stage_prefix}] ℹ️ KSC=1: p의 배수가 아니어도 주문 전송 (Z=1로 설정)")
                            
                            if Z > 1:
                                print(f"{get_timestamp()} [{stage_prefix}] 📊 KSC multiplier: {multiplier}, B값: {B_value}, 최종 Z: {Z}")
                            
                            # 선물 스크립트: KSC/p 배수 차단 없음 — Z=0이면 Z=1로 두고 주문 진행
                            if ksc_numeric > 1 and Z == 0 and not is_bomb:
                                Z = 1
                        else:
                            # KSC 스택이 쌓이는 상황이 아님: Z값으로 컨트롤 필요 없음
                            # Z = 1로 유지하여 base_amount (1unit + bamount)로 주문 전송
                            print(f"{get_timestamp()} [{stage_prefix}] ℹ️ KSC=0 : Z=1 base_amount(1unit+bamount)로 주문")
                    
                    # ORDER 신호가 Sell5/Sell10이면 KSC 초기화 (이미 KSC 계산 로직에서 처리됨)
                    elif latest_order in ['Sell5', 'Sell10']:
                        # KSC는 이미 계산 로직에서 0으로 초기화됨
                        Z = 1  # Sell 신호는 multiplier 적용 안 함
                    
                    # PRFT가 되는 타이밍의 ORDER 신호에 multiplier 계산 (Sell 신호에만 적용)
                    # prft multiplier = 1 + (1 - buyside) = 2 - buyside
                    latest_prft = df_binance_ticker_15m.iloc[0].get("PRFT", 0)
                    latest_buyside = df_binance_ticker_15m.iloc[0].get("buyside", np.nan)
                    
                    # buyside_val 초기화 (조건문 밖에서 먼저 정의)
                    buyside_val = float(latest_buyside) if not pd.isna(latest_buyside) else None
                    
                    prft_multiplier = 1
                    if latest_order in ['Sell5', 'Sell10'] and latest_prft == 'PRFT':
                        # PRFT multiplier = 1 + (1 - buyside) = 2 - buyside
                        if buyside_val is not None:
                            prft_multiplier = 1 + (1 - buyside_val)  # = 2 - buyside
                            print(f"{get_timestamp()} [{stage_prefix}] 💰 PRFT 감지: buyside={buyside_val:.4f}, multiplier={prft_multiplier:.4f} (1 + (1 - {buyside_val:.4f}))")
                    
                    # 직전행 TP 가져오기 (수수료 조건 체크용)
                    prev_tp_for_trade = None
                    if len(df_binance_ticker_15m) > 1:
                        prev_row_for_trade = df_binance_ticker_15m.iloc[1]
                        prev_tp_raw = prev_row_for_trade.get("TP", np.nan)
                        if pd.notna(prev_tp_raw):
                            try:
                                prev_tp_for_trade = float(prev_tp_raw)
                                if prev_tp_for_trade <= 0:
                                    prev_tp_for_trade = None
                            except (TypeError, ValueError):
                                prev_tp_for_trade = None
                    
                    # 기타 필요한 값들 가져오기
                    latest_hmsfast = df_binance_ticker_15m.iloc[0].get("1HMSFast", np.nan)
                    hmsfast_val = float(latest_hmsfast) if pd.notna(latest_hmsfast) else None
                    
                    latest_decision_price = df_binance_ticker_15m.iloc[0].get("종", np.nan)
                    decision_price_val = float(latest_decision_price) if pd.notna(latest_decision_price) else None
                    
                    # TPC 값 가져오기 (TPOVER 매도 시 사용)
                    latest_tpc = df_binance_ticker_15m.iloc[0].get("TPC", 0)
                    tpc_value_for_trade = 0.0
                    try:
                        tpc_value_for_trade = float(latest_tpc) if pd.notna(latest_tpc) else 0.0
                    except (TypeError, ValueError):
                        tpc_value_for_trade = 0.0
                    
                    # StoSU 값 가져오기 (TPOVER 매도 시 사용)
                    latest_stosu = df_binance_ticker_15m.iloc[0].get("StoSU", 0.0)
                    stosu_value_for_trade = 0.0
                    try:
                        stosu_value_for_trade = float(latest_stosu) if pd.notna(latest_stosu) else 0.0
                    except (TypeError, ValueError):
                        stosu_value_for_trade = 0.0
                    
                    # 4. 주문량 = (1유닛 + bamount) × Z
                    # p값: 15M 열 p(= 3+p1H) 우선, 없으면 1HCL로 3+1HCL 계산
                    latest_p_for_trade = df_binance_ticker_15m.iloc[0].get("p", np.nan)
                    p_value_for_trade = int(float(latest_p_for_trade)) if pd.notna(latest_p_for_trade) else None
                    latest_1hcl_for_trade = df_binance_ticker_15m.iloc[0].get("1HCL", np.nan)
                    h1cl_for_trade = int(float(latest_1hcl_for_trade)) if pd.notna(latest_1hcl_for_trade) else None
                    if ENABLE_SPOT_TRADING:
                        trade_on_order_signal(latest_order, symbol=f"{TICKER}USDT", samount=float(latest_samount), bamount=float(latest_bamount), bomb_multiplier=Z, prft_value=latest_prft, ksc_numeric=ksc_numeric, prft_multiplier=prft_multiplier, hmsfast=hmsfast_val, buyside=buyside_val, tpc_value=tpc_value_for_trade, stosu=stosu_value_for_trade, decision_price=decision_price_val, prev_tp=prev_tp_for_trade, h1cl=h1cl_for_trade, p_value=p_value_for_trade, stage_prefix=stage_prefix)
                    else:
                        print(f"{get_timestamp()} [{stage_prefix}] ℹ️ 스팟 주문 비활성화 - ORDER 신호만 감지 ({TICKER} {latest_order})")
                else:
                    print(f"{get_timestamp()} [{stage_prefix}] ℹ️ ORDER 신호: {TICKER} 없음")
            except Exception as e:
                print(f"{get_timestamp()} [{stage_prefix}] ❌ 주문 실행 중 오류: {e}")
        elif polling_start_time and not ENABLE_TRADING:
            print(f"{get_timestamp()} [{stage_prefix}] ℹ️ 주문전송 비활성화 - 엑셀만 생성")
        
        # df_15m을 반환하여 폴링에서 사용 가능하게 함 (메모리 정리 전에 반환)
        # result_df 초기화 (예외 발생 시에도 안전하게 처리)
        try:
            result_df = df_binance_ticker_15m.copy()
        except (NameError, AttributeError):
            # df_binance_ticker_15m이 정의되지 않았거나 예외 발생 시 빈 DataFrame 반환
            result_df = pd.DataFrame()
        
        print(f"{get_timestamp()} [{stage_prefix}] 🧹 메모리 정리 중...")
        del df_binance_ticker_1d, df_binance_ticker_15m
        del df_binance_ticker_1h, df_binance_ticker_weekly
        collected = gc.collect()
        print(f"{get_timestamp()} [{stage_prefix}] ✅ 메모리 정리 완료 ({collected}개 객체 해제)")
        
        return result_df

    except Exception as e:
        print(f"{get_timestamp()} [{stage_prefix}] ❌ Excel 파일 저장 중 오류 발생: {e}")
        # 오류 발생 시 writer가 열려있다면 닫기
        if writer is not None:
            try:
                writer.close()
            except:
                pass
        
        # 오류 발생 시에도 빈 DataFrame 반환
        try:
            return df_binance_ticker_15m.copy() if 'df_binance_ticker_15m' in locals() else pd.DataFrame()
        except:
            return pd.DataFrame()
    
    finally:
        # 실행 시간 계산 및 출력
        end_time = time.time()
        elapsed_time = end_time - start_time
        stage_prefix = "PREVIOUS" if skip_first_row else "AFTER"
        print(f"{get_timestamp()} [{stage_prefix}] ⏱️ {TICKER} 처리 완료 - 소요시간: {elapsed_time:.2f}초")
def calculate_LD(df_15m: pd.DataFrame) -> pd.DataFrame:
    """
    15분봉 데이터에 LD 열을 추가합니다.
    dateM은 200개 캔들로 계산하고, 상단+하단은 Python에서 200개 캔들로 계산합니다.
    t^2 + (dateM)t - Spread2*Spread1 = 0 의 양의 근을 계산합니다.
    """
    import math
    
    df = df_15m.copy()
    
    # LD 열 초기화
    df['LD'] = 0.0
    
    for i in range(len(df)):
        # 종, 시, 고, 저, sma3, sma5, sma7, sma10, sma20 값들
        values = [
            df.iloc[i]['종'],
            df.iloc[i]['시'], 
            df.iloc[i]['고'],
            df.iloc[i]['저'],
            df.iloc[i]['SMA3'],
            df.iloc[i]['SMA5'],
            df.iloc[i]['SMA7'],
            df.iloc[i]['SMA10'],
            df.iloc[i]['SMA20']
        ]
        
        # Spread2 계산
        max_val = max(values)
        min_val = min(values)
        spread2 = (max_val - min_val) / min_val * 1000
        
        # dateM 계산 (200개 캔들 기준)
        end_idx = min(i + 200, len(df))
        max_value = df.iloc[i:end_idx]['고'].max()
        
        # 현재 시점부터 미래로 순회하면서 Max 값과 같은 고가를 가진 캔들을 찾기
        dateM = 0
        for j in range(i, end_idx):  # 200개 캔들 범위 내에서만 검색
            if df.iloc[i:end_idx].iloc[j-i]['고'] == max_value:
                dateM = j - i + 1  # 캔들 개수 계산
                break
        
        # Spread1 계산 (Python에서 200개 캔들로 상단+하단 계산)
        # 200개 캔들 범위에서 Max200, Min200 계산
        max200 = df.iloc[i:end_idx][["시", "고", "저", "종"]].max().max()
        min200 = df.iloc[i:end_idx][["시", "고", "저", "종"]].min().min()
        
        # 현재 가격
        current_price = df.iloc[i]['종']
        
        # 상단, 하단 계산 (200개 캔들 기준)
        if min200 != 0:
            하단 = abs((current_price - min200) / min200)
        else:
            하단 = 0
            
        if max200 != 0:
            상단 = abs((current_price - max200) / max200)
        else:
            상단 = 0
        
        spread1 = (상단 + 하단) * 1000
        
        # 이차방정식 t^2 + (dateM)t - Spread2*Spread1 = 0 의 계수
        a = 1
        b = dateM
        c = -spread2 * spread1
        
        # 판별식 계산 (항상 양수)
        discriminant = b * b - 4 * a * c
        
        # 양의 근 계산 (t1이 항상 양수)
        t1 = (-b + math.sqrt(discriminant)) / (2 * a)
        
        # 양의 근을 원래 값으로 저장 (셀 서식에서 소수점 3자리 표시)
        df.iloc[i, df.columns.get_loc('LD')] = t1
    
    return df

def calculate_dateM(df_15m: pd.DataFrame) -> pd.DataFrame:
    """
    15분봉 데이터에 dateM 열을 추가합니다.
    Python에서만 200개 캔들 기준으로 Max 값을 찾아서 계산합니다.
    
    Args:
        df_15m: 15분봉 DataFrame (Date(UTC) 컬럼 포함, UTC 기준)
    
    Returns:
        DataFrame: dateM 컬럼이 추가된 15분봉 DataFrame
    
    Note:
        - Date(UTC) 컬럼 기준으로 계산 (UTC 기준)
    """
    df = df_15m.copy()
    
    # dateM 열 초기화
    df['dateM'] = 0
    
    for i in range(len(df)):
        # 현재 캔들부터 200개 캔들 범위에서 Max 값 찾기
        end_idx = min(i + 200, len(df))
        max_value = df.iloc[i:end_idx]['고'].max()
        
        # 현재 시점부터 미래로 순회하면서 Max 값과 같은 고가를 가진 캔들을 찾기
        max_date = None
        for j in range(i, end_idx):  # 200개 캔들 범위 내에서만 검색
            if df.iloc[j]['고'] == max_value:
                max_date = df.iloc[j]['Date(UTC)']
                break
        
        if max_date is not None:
            # 캔들 개수로 계산 (j - i + 1) - 현재부터 Max 날짜까지의 총 캔들 개수
            candle_count = j - i + 1
            
            # 음수가 되지 않도록 처리
            df.iloc[i, df.columns.get_loc('dateM')] = max(0, candle_count)
        else:
            # Max 값과 같은 가격을 가진 캔들이 없는 경우 0
            df.iloc[i, df.columns.get_loc('dateM')] = 0
    
    return df

# (환율 계산 제거)

def analyze_15m_performance(df_15m: pd.DataFrame, ticker: str) -> dict:
    """
    15분봉 데이터에서 ORDER 컬럼을 분석하여 매수/매도 평균가격과 예상수익률을 계산합니다.
    실제 주문량(1unit + Samount/Bamount) × multiplier를 가중치로 사용하여 가중 평균가격을 계산합니다.
    
    Sell: base_amount = unit_amount + Samount, 실제주문량 = base_amount × prft_multiplier
    Buy: base_amount = unit_amount + Bamount, 실제주문량 = base_amount × Z
    #   Z = multiplier + if(bomb발생, B값, 0)
    #   - multiplier: KSC가 3의 배수일 때만 3, 나머지는 0 (수열: 0,0,3,0,0,3,...)
    #   - B값: multiplier == 0이면 ((ksc_stack - 1) % 3) + 1 (1,2,3 반복), 아니면 0
    #   - 예시: unit_amount=7, bamount=0, Z=3 → 실제주문량 = (7+0) × 3 = 21 USDT
    
    평균단가 = 총주문량 / 총코인수량 = Σ(실제주문량) / Σ(실제주문량/종가)
    
    Parameters:
    - df_15m: 15분봉 데이터프레임
    - ticker: 티커명 (BTC, ETH, XRP, SOL, BNB 등)
    
    Returns:
    - dict: {'ticker', 'sell_avg', 'sell_count', 'sell_total_amount', 'sell_unit', 'buy_avg', 'buy_count', 'buy_total_amount', 'buy_unit', 'return_pct'}
          - sell_total_amount, buy_total_amount: 실제 주문량 합계 (USDT)
          - sell_unit, buy_unit: 총주문량 / unit_amount
    """
    try:
        # 필수 컬럼 확인
        required_cols = ['ORDER', '종', 'Samount', 'Bamount']
        if not all(col in df_15m.columns for col in required_cols):
            unit_amount = ROTATION_TRADING_UNITS.get(ticker, 5)
            return {
                'ticker': ticker,
                'sell_avg': None,
                'sell_count': 0,
                'sell_total_amount': 0,
                'sell_unit': 0,
                'buy_avg': None,
                'buy_count': 0,
                'buy_total_amount': 0,
                'buy_unit': 0,
                'return_pct': None
            }
        
        # 티커별 1unit 금액 (USDT)
        unit_amount = ROTATION_TRADING_UNITS.get(ticker, 5)
        
        # ORDER 값을 소문자로 변환 (대소문자 구분 없이 처리)
        df_15m['ORDER_lower'] = df_15m['ORDER'].astype(str).str.lower()
        
        # sell5 또는 sell10인 행의 거래량 가중 평균가격 계산 (PRFT multiplier 반영)
        sell_mask = df_15m['ORDER_lower'].isin(['sell5', 'sell10'])
        sell_columns = ['종', 'Samount', 'ORDER', 'PRFT']
        if 'buyside' in df_15m.columns:
            sell_columns.append('buyside')
        sell_data = df_15m.loc[sell_mask, sell_columns].copy()
        
        # NaN 제거 및 유효한 데이터만 사용
        sell_data = sell_data.dropna(subset=['종', 'Samount'])
        
        if len(sell_data) > 0:
            sell_총주문량 = 0
            sell_총코인수량 = 0
            sell_실제주문횟수 = 0
            
            for idx, row in sell_data.iterrows():
                base_amount = unit_amount + row['Samount']
                
                # PRFT multiplier 계산: prft_multiplier = 1 + (1 - buyside) = 2 - buyside
                prft_multiplier = 1
                if pd.notna(row['PRFT']) and str(row['PRFT']).strip() == 'PRFT':
                    # PRFT = 'PRFT'이고 buyside를 사용하여 multiplier 계산
                    if 'buyside' in row.index and pd.notna(row['buyside']):
                        buyside_val = float(row['buyside'])
                        prft_multiplier = 1 + (1 - buyside_val)  # = 2 - buyside
                    # buyside가 없으면 기본값 1 사용
                
                실제주문량 = base_amount * prft_multiplier
                
                # H 팩터 적용 (일반 매도)
                hcl_val = 0.0
                if '1HCL' in row.index and pd.notna(row.get('1HCL')):
                    try:
                        hcl_val += float(row['1HCL'])
                    except (TypeError, ValueError):
                        pass
                if '-1HCL' in row.index and pd.notna(row.get('-1HCL')):
                    try:
                        hcl_val += float(row['-1HCL'])
                    except (TypeError, ValueError):
                        pass
                h_factor = _calc_h_factor(hcl_val)
                실제주문량 *= h_factor
                
                # 각 sell 신호마다 실제주문량을 종가로 나눈 코인 수량 계산
                코인수량 = 실제주문량 / row['종'] if row['종'] > 0 else 0
                
                sell_총주문량 += 실제주문량
                sell_총코인수량 += 코인수량
                sell_실제주문횟수 += 1  # 각 sell 신호마다 1회 카운트
            
            # 거래량 가중 평균가격 계산 (각 행의 가중치 = 실제주문량)
            # sell_avg = 총주문량 / 총코인수량 = Σ(실제주문량) / Σ(실제주문량/종가)
            sell_avg = sell_총주문량 / sell_총코인수량 if sell_총코인수량 > 0 else None
            sell_total_amount = sell_총주문량
            sell_count = sell_실제주문횟수
            sell_unit = sell_total_amount / unit_amount if unit_amount > 0 else 0
        else:
            sell_avg = None
            sell_total_amount = 0
            sell_count = 0
            sell_unit = 0
        
        # buy5 또는 buy10인 행의 거래량 가중 평균가격 계산 (KSC multiplier + Bomb B 값 반영)
        buy_mask = df_15m['ORDER_lower'].isin(['buy5', 'buy10'])
        
        # 필요한 컬럼 확인
        buy_columns = ['종', 'ORDER']
        if 'Bamount' in df_15m.columns:
            buy_columns.append('Bamount')
        if 'KSC' in df_15m.columns:
            buy_columns.append('KSC')
        if 'KSC stack' in df_15m.columns:
            buy_columns.append('KSC stack')
        if 'bomb' in df_15m.columns:
            buy_columns.append('bomb')
        
        buy_data = df_15m.loc[buy_mask, buy_columns].copy()
        
        # Bamount 컬럼이 없으면 0으로 채우기
        if 'Bamount' not in buy_data.columns:
            buy_data['Bamount'] = 0
        # KSC stack 컬럼이 없으면 0으로 채우기
        if 'KSC stack' not in buy_data.columns:
            buy_data['KSC stack'] = 0
        
        # NaN 제거 및 유효한 데이터만 사용
        buy_data = buy_data.dropna(subset=['종'])
        
        if len(buy_data) > 0:
            buy_총주문량 = 0
            buy_총코인수량 = 0
            buy_실제주문횟수 = 0
            
            for idx, row in buy_data.iterrows():
                base_amount = unit_amount + row['Bamount']
                
                # ============================================================
                # KSC 스택에 따른 Z값 및 주문량 계산 (다른 시트에도 적용 가능)
                # ============================================================
                # [1단계] Multiplier: KSC가 3의 배수일 때만 3, 나머지는 0 (수열: 0,0,3,0,0,3,...)
                # [2단계] B값: multiplier == 0이면 ((ksc_stack - 1) % 3) + 1 (1,2,3 반복), 아니면 0
                # [3단계] Z = multiplier + if(bomb발생, B값, 0)
                # [4단계] 주문량 = (1유닛 + bamount) × Z
                # ============================================================
                ksc_value = row.get('KSC', 0)
                bomb_value = row.get('Bomb', "")
                ksc_stack = row.get('KSC stack', 0)
                
                # 1. multiplier 계산 (수열 규칙: KSC가 3의 배수일 때만 값이 있음)
                # KSC는 숫자만 저장
                ksc_stack_val = 0
                if pd.notna(ksc_stack):
                    ksc_stack_val = int(ksc_stack)
                
                # 2. Bomb 발생 시 B 값 계산 (multiplier와 독립적으로 계산)
                # Bomb 열 확인
                is_bomb = (isinstance(bomb_value, str) and bomb_value.strip() == "Bomb")
                
                # multiplier 계산: Bomb 발생 시 KSC stack 값을 사용, 일반 경우 KSC 값을 사용
                if is_bomb:
                    # Bomb 발생 시 KSC stack 값을 사용하여 multiplier 계산
                    ksc_multiplier = calculate_ksc_multiplier(ksc_stack_val, ksc_stack_val)
                else:
                    # 일반 경우 KSC 값을 사용하여 multiplier 계산
                    ksc_multiplier = calculate_ksc_multiplier(ksc_value, ksc_stack_val)
                
                bomb_b_value = 0
                if is_bomb:
                    bomb_b_value = calculate_bomb_b_value(ksc_multiplier, ksc_stack_val)
                
                # 3. Z = multiplier(수열) + if(bomb발생, B값, 0)
                z = ksc_multiplier + bomb_b_value
                
                # [4단계] 주문량 계산: (1유닛 + bamount) × Z
                # - base_amount = unit_amount + bamount
                # - 실제주문량 = base_amount × Z
                # - 예시: unit_amount=7, bamount=0, Z=3
                #   → 실제주문량 = (7 + 0) × 3 = 21 USDT
                실제주문량 = base_amount * z
                
                # H 팩터 적용 (Buy5/Buy10)
                hcl_val = 0.0
                if '1HCL' in row.index and pd.notna(row.get('1HCL')):
                    try:
                        hcl_val += float(row['1HCL'])
                    except (TypeError, ValueError):
                        pass
                if '-1HCL' in row.index and pd.notna(row.get('-1HCL')):
                    try:
                        hcl_val += float(row['-1HCL'])
                    except (TypeError, ValueError):
                        pass
                risk_factor = get_buy_risk_factor(str(row.get('ORDER', '')), None, hcl_val)
                실제주문량 *= risk_factor
                
                # 각 buy 신호마다 실제주문량을 종가로 나눈 코인 수량 계산
                코인수량 = 실제주문량 / row['종'] if row['종'] > 0 else 0
                
                buy_총주문량 += 실제주문량
                buy_총코인수량 += 코인수량
                buy_실제주문횟수 += 1  # 각 buy 신호마다 1회 카운트
            
            # 거래량 가중 평균가격 계산 (각 행의 가중치 = 실제주문량)
            # buy_avg = 총주문량 / 총코인수량 = Σ(실제주문량) / Σ(실제주문량/종가)
            buy_avg = buy_총주문량 / buy_총코인수량 if buy_총코인수량 > 0 else None
            buy_total_amount = buy_총주문량
            buy_count = buy_실제주문횟수
            buy_unit = buy_total_amount / unit_amount if unit_amount > 0 else 0
        else:
            buy_avg = None
            buy_total_amount = 0
            buy_count = 0
            buy_unit = 0
        
        # 예상수익률 계산
        if sell_avg is not None and buy_avg is not None and buy_avg != 0:
            expected_return = (sell_avg - buy_avg) / buy_avg * 100
        else:
            expected_return = None
        
        return {
            'ticker': ticker,
            'sell_avg': sell_avg,
            'sell_count': sell_count,
            'sell_total_amount': sell_total_amount,
            'sell_unit': sell_unit,
            'buy_avg': buy_avg,
            'buy_count': buy_count,
            'buy_total_amount': buy_total_amount,
            'buy_unit': buy_unit,
            'return_pct': expected_return
        }
        
    except Exception as e:
        print(f"{get_timestamp()} ❌ [{ticker}] 15M 분석 중 오류 발생: {e}")
        import traceback
        traceback.print_exc()
        unit_amount = ROTATION_TRADING_UNITS.get(ticker, 5)
        return {
            'ticker': ticker,
            'sell_avg': None,
            'sell_count': 0,
            'sell_total_amount': 0,
            'sell_unit': 0,
            'buy_avg': None,
            'buy_count': 0,
            'buy_total_amount': 0,
            'buy_unit': 0,
            'return_pct': None
        }

# 4day 분석 시 티커별 LS·1HMSF·종가 표시용 (run_rotation_sequence에서 설정)
_4day_ticker_snapshots = {}

def analyze_15m_trading_performance():
    """
    선물 계정·포지션 요약을 출력합니다. (Binance USDT-M 선물)
    - 헤더: 선물 잔고, 미실현손익, 사용마진
    - 티커별: 포지션 있으면 LONG/SHORT, 수량, 진입가, 마크가, 미실현손익 / 없으면 포지션 없음
    - 티커별: LS | 1HMSF | 종가 (스냅샷 있을 때만)
    """
    global _4day_ticker_snapshots
    separator = "=" * 100
    discord_msg_buffer = []
    
    # ---------- 헤더 형식 (사용자 정의, $=USDT 기준) ----------
    #   [TotalUSDT]$=[totalWalletBalance]$+B[BNB USDT환산]$+[totalUnrealizedProfit]$([pct]%) |A: [availableBalance]$,L: [포지션증거금]$
    #   예: 3003.05$=2988.99$+B14$+0.06$(+0.00%) |A: 2982.79$,L: 6.26$
    # ----------
    acc = get_futures_account()
    total_wallet = float(acc.get("totalWalletBalance", 0) or 0)
    total_unrealized = float(acc.get("totalUnrealizedProfit", 0) or 0)
    available = float(acc.get("availableBalance", 0) or 0)
    total_margin = float(acc.get("totalPositionInitialMargin", 0) or 0)
    margin_balance = float(acc.get("totalMarginBalance", 0) or 0)
    total_notional = 0.0
    for _t in ROTATION_TICKERS:
        for pos in get_futures_position_risk(f"{_t}USDT"):
            amt = float(pos.get("positionAmt", 0) or 0)
            if amt == 0:
                continue
            mark = float(pos.get("markPrice", 0) or 0)
            total_notional += abs(amt) * mark
    fee_rate_bnb = BINANCE_FUTURES_MAKER_FEE * BINANCE_BNB_FEE_DISCOUNT  # 지정가(Maker) + BNB할인
    fee_estimate = total_notional * fee_rate_bnb * 2
    unrealized_pct = (total_unrealized / total_wallet * 100) if total_wallet and total_wallet != 0 else 0.0
    bnb_wallet = 0.0
    for a in acc.get("assets", []):
        if a.get("asset") == "BNB":
            bnb_wallet = float(a.get("walletBalance", 0) or 0)
            break
    bnb_usdt_header = 0.0
    if "BNB" in _4day_ticker_snapshots:
        close_val = _4day_ticker_snapshots["BNB"].get("종")
        if close_val is not None and pd.notna(close_val) and str(close_val).strip() != "":
            bnb_usdt_header = bnb_wallet * float(close_val)
    if bnb_wallet and bnb_usdt_header == 0.0:
        try:
            bnb_price = binance_ticker_price("BNBUSDT")
            if bnb_price and float(bnb_price) > 0:
                bnb_usdt_header = bnb_wallet * float(bnb_price)
        except Exception:
            pass
    # TotalUSDT = USDT Wallet + BNB(USDT 환산) + 미실현 PnL
    total_usdt_with_bnb = total_wallet + bnb_usdt_header + total_unrealized
    header_content = (
        f" {total_usdt_with_bnb:.2f}$"
        f"={total_wallet:.2f}$+B{bnb_usdt_header:.2f}${total_unrealized:+.2f}$({unrealized_pct:+.2f}%)"
        f" |A: {available:.2f}$,L: {total_margin:.2f}$"
    )
    header_msg = get_timestamp() + header_content
    print(f"\n{separator}")
    print(header_msg)
    print(separator)
    discord_msg_buffer.append(get_timestamp(include_ms=False) + header_content)
    
    for ticker in ROTATION_TICKERS:
        try:
            symbol = f"{ticker}USDT"
            positions = get_futures_position_risk(symbol)
            ticker_display = f"{ticker:<3}"
            # 윗줄: LS | 1HMSF | 종가
            if ticker in _4day_ticker_snapshots:
                s = _4day_ticker_snapshots[ticker]
                ls_val = s.get("LS")
                try:
                    v = int(float(ls_val)) if ls_val is not None and ls_val != "" and not (hasattr(ls_val, "__float__") and pd.isna(ls_val)) else None
                    ls_str = str(v) if v in (1, -1) else "-"
                except (TypeError, ValueError):
                    ls_str = "-"
                hmsf_val = s.get("1HMSFast")
                hmsf_str = f"{float(hmsf_val):.3f}" if hmsf_val is not None and pd.notna(hmsf_val) and str(hmsf_val).strip() != "" else "-"
                close_val = s.get("종")
                price_prec = 4 if ticker == "XRP" else 2
                close_str = f"{float(close_val):.{price_prec}f}" if close_val is not None and pd.notna(close_val) and str(close_val).strip() != "" else "-"
                smaf_val = s.get("SMAF")
                smaf_str = f"{float(smaf_val):.{price_prec}f}" if smaf_val is not None and pd.notna(smaf_val) and str(smaf_val).strip() != "" else "-"
                sma100_val = s.get("SMA100")
                sma100_str = f"{float(sma100_val):.{price_prec}f}" if sma100_val is not None and pd.notna(sma100_val) and str(sma100_val).strip() != "" else "-"
                extra = f" | LS: {ls_str} | 1HMSF: {hmsf_str} | 종가: {close_str} | SMAF: {smaf_str} | SMA100: {sma100_str}"
                print(f"{get_timestamp()} [{ticker_display}]{extra}")
                discord_msg_buffer.append(f"[{ticker_display}]{extra}")
            has_position = False
            for pos in positions:
                amt = float(pos.get("positionAmt", 0) or 0)
                if amt == 0:
                    continue
                has_position = True
                entry = float(pos.get("entryPrice", 0) or 0)
                mark = float(pos.get("markPrice", 0) or 0)
                upnl = float(pos.get("unRealizedProfit", pos.get("unrealizedProfit", 0)) or 0)
                side = "LONG" if amt > 0 else "SHORT"
                qty = abs(amt)
                if ticker == "BTC":
                    entry_str = f"{entry/1000:.2f}K"
                    mark_str = f"{mark/1000:.2f}K"
                    pf = ".2f"
                elif ticker == "BNB":
                    entry_str = f"{entry:.2f}"
                    mark_str = f"{mark:.2f}"
                    pf = ".2f"
                elif ticker == "XRP":
                    entry_str = f"{entry:.4f}"
                    mark_str = f"{mark:.4f}"
                    pf = ".4f"
                elif ticker in ["ETH", "SOL"]:
                    entry_str = f"{entry:.2f}"
                    mark_str = f"{mark:.2f}"
                    pf = ".2f"
                else:
                    entry_str = f"{entry:.3f}"
                    mark_str = f"{mark:.3f}"
                    pf = ".3f"
                notional = qty * mark
                upnl_pct = (upnl / notional * 100) if notional and notional != 0 else 0.0
                side_short = "L" if amt > 0 else "S"
                # P 표시: 전 티커 0.5%, 1%, 1.5%
                tp1 = adjust_price_to_tick_futures(symbol, entry * (1.005 if amt > 0 else 0.995))
                tp2 = adjust_price_to_tick_futures(symbol, entry * (1.01 if amt > 0 else 0.99))
                tp3 = adjust_price_to_tick_futures(symbol, entry * (1.015 if amt > 0 else 0.985))
                # L/BE 표시
                # - 기본: SL 1.2% (롱=진입가 하락 1.2%, 숏=진입가 상승 1.2%)
                # - BE 이동 후: L 대신 BE로 표시. SL은 Algo Order이므로 openAlgoOrders에서 triggerPrice 사용
                l_label = "L"
                sl_display = None
                try:
                    open_algo = get_futures_open_algo_orders(symbol)
                except Exception:
                    open_algo = []
                sl_orders = [
                    o
                    for o in open_algo
                    if (o.get("orderType") or "").upper() == "STOP_MARKET"
                    and (o.get("reduceOnly") in (True, "true", "TRUE") or str(o.get("reduceOnly", "")).lower() == "true")
                ]
                if sl_orders and entry > 0:
                    first_sl_stop = float(sl_orders[0].get("triggerPrice", 0) or 0)
                    if first_sl_stop > 0:
                        tol_pct = 0.001  # 0.1% 이내면 동일 가격으로 간주
                        if amt > 0:
                            # 롱: 원래 SL = 진입가 * (1 - SL%), BE = 진입가 * (1 + BE_OFFSET)
                            orig_sl = adjust_price_to_tick_futures(symbol, entry * (1.0 - FUTURES_SL_PERCENT))
                            be_price = adjust_price_to_tick_futures(symbol, entry * (1.0 + FUTURES_BE_OFFSET_PERCENT))
                        else:
                            # 숏: 원래 SL = 진입가 * (1 + SL%), BE = 진입가 * (1 - BE_OFFSET)
                            orig_sl = adjust_price_to_tick_futures(symbol, entry * (1.0 + FUTURES_SL_PERCENT))
                            be_price = adjust_price_to_tick_futures(symbol, entry * (1.0 - FUTURES_BE_OFFSET_PERCENT))
                        if abs(first_sl_stop - be_price) / entry <= tol_pct:
                            # BE 상태
                            l_label = "BE"
                            sl_display = be_price
                        elif abs(first_sl_stop - orig_sl) / entry <= tol_pct:
                            # 원래 SL 상태
                            l_label = "L"
                            sl_display = orig_sl
                        else:
                            # 수동 조정 등: 실제 stopPrice를 그대로 표시
                            sl_display = first_sl_stop
                if sl_display is None:
                    # SL 주문이 없으면 이론상 SL 기준으로 표시
                    if amt > 0:
                        sl_display = adjust_price_to_tick_futures(symbol, entry * (1.0 - FUTURES_SL_PERCENT))
                    else:
                        sl_display = adjust_price_to_tick_futures(symbol, entry * (1.0 + FUTURES_SL_PERCENT))
                sl_str = f"{sl_display:{pf}}"
                line = (
                    f"[{ticker_display}] {side_short} {round(notional)}$ @{entry:{pf}} |M{mark_str} |{upnl:>+5.2f}$({upnl_pct:>+5.2f}%) |P{tp1:{pf}}-{tp2:{pf}}-{tp3:{pf}} |{l_label}{sl_str}"
                )
                print(f"{get_timestamp()} {line}")
                discord_msg_buffer.append(line)
            
            if not has_position:
                line = f"[{ticker_display}] 포지션 없음"
                print(f"{get_timestamp()} {line}")
                discord_msg_buffer.append(line)
        except Exception as e:
            print(f"{get_timestamp()} [오류] {ticker} 선물 분석 중 오류: {e}")
    
    print(separator)
    if discord_msg_buffer:
        send_discord_message("\n".join(discord_msg_buffer))

def analyze_15m_trading_performance_simple(df_15m, ticker, unit_amount=7500):
    """
    간소화 버전: DataFrame을 직접 받아서 분석
    
    Parameters:
    - df_15m: 15분봉 데이터프레임
    - ticker: 티커명 (BTC, ETH, XRP, SOL, BNB 등)
    - unit_amount: 거래 단위 금액 (기본값: 7500 USDT)
    
    Returns:
    - dict: {'sell_avg', 'buy_avg', 'expected_return', 'sell_count', 'buy_count'}
    """
    try:
        # 필수 컬럼 확인
        required_cols = ['ORDER', '종', 'Samount', 'Bamount']
        if not all(col in df_15m.columns for col in required_cols):
            return {
                'sell_avg': None,
                'buy_avg': None,
                'expected_return': None,
                'sell_count': 0,
                'buy_count': 0
            }
        
        # Sell 평균
        sell_data = df_15m[df_15m['ORDER'].str.lower().isin(['sell5', 'sell10'])]
        sell_avg = (sell_data['종'] * (unit_amount + sell_data['Samount'])).sum() / \
                   (unit_amount + sell_data['Samount']).sum() if len(sell_data) > 0 and (unit_amount + sell_data['Samount']).sum() > 0 else None
        
        # Buy 평균
        buy_data = df_15m[df_15m['ORDER'].str.lower().isin(['buy5', 'buy10'])]
        buy_avg = (buy_data['종'] * (unit_amount + buy_data['Bamount'])).sum() / \
                  (unit_amount + buy_data['Bamount']).sum() if len(buy_data) > 0 and (unit_amount + buy_data['Bamount']).sum() > 0 else None
        
        # 예상수익률
        expected_return = ((sell_avg - buy_avg) / buy_avg * 100) if (sell_avg and buy_avg) else None
        
        return {
            'sell_avg': sell_avg,
            'buy_avg': buy_avg,
            'expected_return': expected_return,
            'sell_count': len(sell_data),
            'buy_count': len(buy_data)
        }
    except Exception as e:
        print(f"{get_timestamp()} ❌ [{ticker}] 15M 간소화 분석 중 오류 발생: {e}")
        import traceback
        traceback.print_exc()
        return {
            'sell_avg': None,
            'buy_avg': None,
            'expected_return': None,
            'sell_count': 0,
            'buy_count': 0
        }

def run_rotation_sequence(polling_start_time=None, skip_first_row=False, target_tickers=None):
    """로테이션 시퀀스 실행: BTC → ETH → XRP → SOL → BNB 순서로 실행
    
    Args:
        polling_start_time: 실행 시작 시간 (UTC 기준)
        skip_first_row: True면 1단계(previous 파일 생성), False면 2단계(after 파일 생성 및 주문 전송)
        target_tickers: 처리할 티커 리스트 (None이면 ROTATION_TICKERS 전체 처리)
    
    Note:
        - 모든 시간 처리는 UTC 기준으로 수행
    """
    global _4day_ticker_snapshots
    # 로테이션 시퀀스 시작 시간 기록 (UTC 기준)
    rotation_start_time = dt.datetime.now(tz.UTC)
    stage_prefix = "PREVIOUS" if skip_first_row else "AFTER"
    stage_label = "1단계 (previous)" if skip_first_row else "2단계 (after)"
    
    # 처리할 티커 결정
    tickers_to_process = target_tickers if target_tickers is not None else ROTATION_TICKERS
    print(f"{get_timestamp()} [{stage_prefix}] 🔄 로테이션 시퀀스 시작 [{stage_label}]: {tickers_to_process}")
    
    # 1·2단계 공통: 시행 시작 시 5개 티커 선물 exchangeInfo 조회 후 터미널·디스코드 출력
    print_futures_exchange_info_summary()
    
    # 2단계 실행 시: 모든 티커의 최신 캔들을 한 번에 수집 (pre_fetched_data)
    pre_fetched_data = {}
    if not skip_first_row and polling_start_time is not None:
        # endTime 계산: 실행 시각에 상관없이 항상 다음 15분 정각 - 1ms로 고정 (업비트 방식)
        current_time_utc = dt.datetime.now(tz.UTC)
        current_minute = current_time_utc.minute
        
        # 1. 실행 시점에 상관없이 '다음 15분 정각'을 타겟으로 설정하여 기준점 고정
        target_dt = current_time_utc
        if 0 <= current_minute < 15:
            target_dt = target_dt.replace(minute=15, second=1, microsecond=0)
        elif 15 <= current_minute < 30:
            target_dt = target_dt.replace(minute=30, second=1, microsecond=0)
        elif 30 <= current_minute < 45:
            target_dt = target_dt.replace(minute=45, second=1, microsecond=0)
        elif 45 <= current_minute < 60:
            target_dt = target_dt.replace(minute=0, second=1, microsecond=0) + dt.timedelta(hours=1)
        
        # 2. 이 타겟 시간에서 1ms를 빼서 '이전 구간의 완성봉'만 가져오도록 고정
        fixed_end_time_ms = int(target_dt.timestamp() * 1000) - 1
        print(f"{get_timestamp()} [AFTER] 🔒 타임라인 동기화 기준점: {target_dt.strftime('%H:%M:%S')} (UTC) - 1ms (1ms 차단 적용)")
        
        # 현재 시간 정보 (UTC 기준)
        current_time_utc = dt.datetime.now(tz.UTC)
        current_hour = current_time_utc.hour
        current_minute = current_time_utc.minute
        current_second = current_time_utc.second
        
        # 1시간봉 조회 개수 결정: XX시 00분 1초에만 2개, 나머지는 1개 (UTC 기준)
        is_00min_01sec = (current_minute == 0 and current_second == 1)
        hour1_count = 2 if is_00min_01sec else 1
        
        # 일봉 조회 개수 결정: 00시 00분 1초에만 2개, 나머지는 1개 (UTC 기준)
        is_00hour_00min_01sec = (current_hour == 0 and current_minute == 0 and current_second == 1)
        daily_count = 2 if is_00hour_00min_01sec else 1
        
        # include_today 설정 (기본값 True)
        include_today_rotation = True
        
        for ticker in tickers_to_process:
            binance_symbol_ticker = f"{ticker}USDT"
            try:
                # 1분봉 수집 (2단계: 16개)
                df_1m = fetch_binance_minutes1(binance_symbol_ticker, 16, include_today=include_today_rotation, fixed_end_time_ms=fixed_end_time_ms, stage_prefix=stage_prefix)
                df_5m = fetch_binance_minutes5(binance_symbol_ticker, 4, include_today=include_today_rotation, fixed_end_time_ms=fixed_end_time_ms)
                df_15m = fetch_binance_minutes15(binance_symbol_ticker, 2, include_today=include_today_rotation, fixed_end_time_ms=fixed_end_time_ms)
                
                # 일봉 수집 (00시 00분 1초에만 2개, 나머지는 1개)
                if is_00hour_00min_01sec:
                    df_1d = fetch_binance_daily(binance_symbol_ticker, 2, include_today=include_today_rotation, fixed_end_time_ms=fixed_end_time_ms)
                else:
                    df_1d = fetch_binance_daily(binance_symbol_ticker, 1, include_today=include_today_rotation, fixed_end_time_ms=fixed_end_time_ms)
                
                # 1시간봉 수집 (XX시 00분 1초에만 2개, 나머지는 1개)
                if is_00min_01sec:
                    df_1h = fetch_binance_hours1(binance_symbol_ticker, 2, include_today=include_today_rotation, fixed_end_time_ms=fixed_end_time_ms)
                else:
                    df_1h = fetch_binance_hours1(binance_symbol_ticker, 1, include_today=include_today_rotation, fixed_end_time_ms=fixed_end_time_ms)
                
                # 수집 직후 개수 저장 (미완성 캔들 제거 전)
                count_1m_before = len(df_1m)
                count_5m_before = len(df_5m)
                count_15m_before = len(df_15m)
                
                print(f"{get_timestamp()} [{stage_prefix}] 📥 {ticker} 캔들 수집 완료 (1M:{count_1m_before}개, 5M:{count_5m_before}개, 15M:{count_15m_before}개, 1H:{len(df_1h)}개, 1D:{len(df_1d)}개)")
                
                # UTC 시간 파싱 함수 (모든 티커에서 공통 사용)
                def extract_datetime_from_utc(utc_str):
                    try:
                        if isinstance(utc_str, str):
                            if ',' in utc_str:
                                date_part, time_part = utc_str.split(',', 1)
                                utc_str = date_part.strip() + ' ' + time_part.strip()
                            if '/' in utc_str:
                                parts = utc_str.split()
                                if len(parts) >= 2:
                                    time_str = parts[1]
                                    if time_str.count(':') == 2:
                                        dt_obj = dt.datetime.strptime(parts[0] + ' ' + time_str, "%y/%m/%d %H:%M:%S")
                                    elif time_str.count(':') == 1:
                                        dt_obj = dt.datetime.strptime(parts[0] + ' ' + time_str, "%y/%m/%d %H:%M")
                                    else:
                                        return None
                                else:
                                    return None
                            else:
                                parts = utc_str.split()
                                if len(parts) >= 2:
                                    time_str = parts[1]
                                    if time_str.count(':') == 2:
                                        dt_obj = dt.datetime.strptime(parts[0] + ' ' + time_str, "%Y-%m-%d %H:%M:%S")
                                    elif time_str.count(':') == 1:
                                        dt_obj = dt.datetime.strptime(parts[0] + ' ' + time_str, "%Y-%m-%d %H:%M")
                                    else:
                                        return None
                                else:
                                    return None
                            return dt_obj
                    except:
                        return None
                    return None
                
                # 5분봉 미완성 캔들 제거 (UTC 기준)
                # ⚠️중요: 30분 1초 실행 시 → 5분봉 03:25~03:30, 03:20~03:25, 03:15~03:20, 03:10~03:15, 03:05~03:10, 03:00~03:05
                # 미완성 03:25~03:30 제거 → 나머지만 남음
                if len(df_5m) > 0 and 'Date(UTC)' in df_5m.columns:
                    try:
                        # 첫 번째 행(최신 캔들)의 시간 확인
                        first_date_val = df_5m.iloc[0]['Date(UTC)']
                        # Timestamp 객체인 경우 그대로 사용, 문자열인 경우 파싱
                        if isinstance(first_date_val, pd.Timestamp):
                            first_date = first_date_val
                        elif pd.api.types.is_datetime64_any_dtype(pd.Series([first_date_val])):
                            first_date = pd.to_datetime(first_date_val)
                        else:
                            first_date = pd.to_datetime(first_date_val, errors='coerce')
                        
                        if pd.notna(first_date):
                            # 현재 실행 시간 확인
                            current_time_utc = dt.datetime.now(tz.UTC)
                            current_minute = current_time_utc.minute
                            current_hour = current_time_utc.hour
                            current_date = current_time_utc.date()
                            
                            # ⚠️중요: 30분 1초 실행 시
                            # API 반환 (최신순): 
                            #   - 첫 번째 행: 03:30 (03:30~03:35 구간, 1초짜리 미완성)
                            #   - 두 번째 행: 03:25 (03:25~03:30 구간, 5분짜리 완성)
                            # → 03:30 미완성 제거 → 03:25만 남음
                            
                            # 5분봉의 시작 시간 계산 (5분 단위로 내림)
                            first_start_minute = (first_date.minute // 5) * 5
                            current_5min_start = (current_minute // 5) * 5
                            
                            # 첫 번째 행(최신)이 현재 5분 구간이면 미완성 → 제거
                            if (first_date.hour == current_hour and 
                                first_start_minute == current_5min_start and
                                first_date.date() == current_date):
                                df_5m = df_5m.iloc[1:].reset_index(drop=True)
                                first_date_str = str(first_date_val) if not isinstance(first_date_val, pd.Timestamp) else first_date_val.strftime('%y/%m/%d,%H:%M')
                                print(f"{get_timestamp()} [{stage_prefix}] ✅ {ticker} 5분봉 미완성 캔들 제거: {first_date_str}")
                    except Exception as e:
                        print(f"{get_timestamp()} [{stage_prefix}] ⚠️ {ticker} 5분봉 미완성 캔들 제거 실패: {e}")
                
                # 1분봉 미완성 캔들 제거 (UTC 기준, main 함수와 동일한 로직)
                # ⚠️중요: 30분 1초 실행 시 → 1분봉 03:29, 03:28, ..., 03:00
                # 미완성 03:30 제거 → 03:29, 03:28, ..., 03:00만 남음
                if len(df_1m) > 0 and 'Date(UTC)' in df_1m.columns:
                    try:
                        latest_date_val = df_1m.iloc[0]['Date(UTC)']
                        # Timestamp 객체인 경우 그대로 사용, 문자열인 경우 파싱
                        if isinstance(latest_date_val, pd.Timestamp):
                            latest_date = latest_date_val
                        elif pd.api.types.is_datetime64_any_dtype(pd.Series([latest_date_val])):
                            latest_date = pd.to_datetime(latest_date_val)
                        else:
                            latest_date = pd.to_datetime(latest_date_val, errors='coerce')
                        
                        if pd.notna(latest_date):
                            # 현재 실행 시간의 '분' 확인
                            current_time_utc = dt.datetime.now(tz.UTC)
                            current_minute = current_time_utc.minute
                            current_hour = current_time_utc.hour
                            current_date = current_time_utc.date()
                            
                            # 최신 캔들의 '분'이 현재 시간의 '분'과 같고, 날짜/시간도 같으면 미완성 -> 제거
                            if (latest_date.minute == current_minute and 
                                latest_date.hour == current_hour and 
                                latest_date.date() == current_date):
                                df_1m = df_1m.iloc[1:].reset_index(drop=True)
                                latest_date_str = str(latest_date_val) if not isinstance(latest_date_val, pd.Timestamp) else latest_date_val.strftime('%y/%m/%d,%H:%M')
                                print(f"{get_timestamp()} [{stage_prefix}] ✅ {ticker} 1분봉 미완성 캔들 제거: {latest_date_str}")
                    except Exception as e:
                        print(f"{get_timestamp()} [{stage_prefix}] ⚠️ {ticker} 1분봉 미완성 캔들 제거 실패: {e}")
                
                # 15분봉 미완성 캔들 제거 (UTC 기준)
                # ⚠️중요: 30분 1초 실행 시 → 15분봉 03:15~03:30 (미완성), 03:00~03:15 (완성)
                # 미완성 03:15~03:30 제거 → 03:00~03:15만 남음 (엑셀 표시: 03:00)
                if len(df_15m) > 0 and 'Date(UTC)' in df_15m.columns:
                    try:
                        # 첫 번째 행(최신 캔들)의 시간 확인
                        first_date_val = df_15m.iloc[0]['Date(UTC)']
                        # Timestamp 객체인 경우 그대로 사용, 문자열인 경우 파싱
                        if isinstance(first_date_val, pd.Timestamp):
                            first_date = first_date_val
                        elif pd.api.types.is_datetime64_any_dtype(pd.Series([first_date_val])):
                            first_date = pd.to_datetime(first_date_val)
                        else:
                            first_date = pd.to_datetime(first_date_val, errors='coerce')
                        
                        if pd.notna(first_date):
                            # 현재 실행 시간 확인
                            current_time_utc = dt.datetime.now(tz.UTC)
                            current_minute = current_time_utc.minute
                            current_hour = current_time_utc.hour
                            current_date = current_time_utc.date()
                            
                            # 15분봉의 시작 시간 계산 (15분 단위로 내림)
                            first_start_minute = (first_date.minute // 15) * 15
                            
                            # ⚠️중요: 30분 1초 실행 시
                            # API 반환 (최신순): 
                            #   - 첫 번째 행: 03:30 (03:30~03:45 구간, 1초짜리 미완성)
                            #   - 두 번째 행: 03:15 (03:15~03:30 구간, 15분짜리 완성)
                            # → 03:30 미완성 제거 → 03:15만 남음
                            current_15min_start = (current_minute // 15) * 15
                            
                            # 첫 번째 행(최신)이 현재 15분 구간이면 미완성 → 제거
                            if (first_date.hour == current_hour and 
                                first_start_minute == current_15min_start and
                                first_date.date() == current_date):
                                df_15m = df_15m.iloc[1:].reset_index(drop=True)
                                first_date_str = str(first_date_val) if not isinstance(first_date_val, pd.Timestamp) else first_date_val.strftime('%y/%m/%d,%H:%M')
                                print(f"{get_timestamp()} [{stage_prefix}] ✅ {ticker} 15분봉 미완성 캔들 제거: {first_date_str}")
                    except Exception as e:
                        print(f"{get_timestamp()} [{stage_prefix}] ⚠️ {ticker} 15분봉 미완성 캔들 제거 실패: {e}")
                
                # 1시간봉 미완성 캔들 제거 (UTC 기준)
                # ⚠️중요: xx시 15/30/45/00분 1초 실행 시
                # API 반환: 05:00 (미완성), 04:00 (완성)
                # → 05:00 미완성 제거 → 04:00만 남음
                if len(df_1h) > 0 and 'Date(UTC)' in df_1h.columns:
                    try:
                        # 첫 번째 행(최신 캔들)의 시간 확인
                        first_date_val = df_1h.iloc[0]['Date(UTC)']
                        # Timestamp 객체인 경우 그대로 사용, 문자열인 경우 파싱
                        if isinstance(first_date_val, pd.Timestamp):
                            first_date = first_date_val
                        elif pd.api.types.is_datetime64_any_dtype(pd.Series([first_date_val])):
                            first_date = pd.to_datetime(first_date_val)
                        else:
                            first_date = pd.to_datetime(first_date_val, errors='coerce')
                        
                        if pd.notna(first_date):
                            # 현재 실행 시간 확인
                            current_time_utc = dt.datetime.now(tz.UTC)
                            current_hour = current_time_utc.hour
                            current_date = current_time_utc.date()
                            
                            # 첫 번째 행(최신)이 현재 시간이면 미완성 → 제거
                            # 1시간봉은 항상 00분으로 표시됨
                            if (first_date.hour == current_hour and
                                first_date.minute == 0 and
                                first_date.date() == current_date):
                                df_1h = df_1h.iloc[1:].reset_index(drop=True)
                                first_date_str = str(first_date_val) if not isinstance(first_date_val, pd.Timestamp) else first_date_val.strftime('%y/%m/%d,%H:%M')
                                print(f"{get_timestamp()} [{stage_prefix}] ✅ {ticker} 1시간봉 미완성 캔들 제거: {first_date_str}")
                    except Exception as e:
                        print(f"{get_timestamp()} [{stage_prefix}] ⚠️ {ticker} 1시간봉 미완성 캔들 제거 실패: {e}")
                
                pre_fetched_data[ticker] = {
                    '1m': df_1m,
                    '5m': df_5m,
                    '15m': df_15m,
                    '1d': df_1d,
                    '1h': df_1h
                }
            except Exception as e:
                print(f"{get_timestamp()} ⚠️ {ticker} 수집 실패: {e}")
                pre_fetched_data[ticker] = {
                    '1m': pd.DataFrame(),
                    '5m': pd.DataFrame(),
                    '15m': pd.DataFrame(),
                    '1d': pd.DataFrame(),
                    '1h': pd.DataFrame()
                }
    
    # 각 티커의 15M 데이터를 저장할 리스트
    ticker_15m_data = []
    
    for i, ticker in enumerate(tickers_to_process):
        print(f"{get_timestamp()} [{stage_prefix}] 📊 [{i+1}/{len(tickers_to_process)}] {ticker} 처리 시작...")
        
        # 전역 변수 업데이트
        global TICKER, TRADING_UNIT
        TICKER = ticker
        TRADING_UNIT = ROTATION_TRADING_UNITS[ticker]
        
        print(f"{get_timestamp()} [{stage_prefix}] ⚙️ 설정: 티커={TICKER}, 거래단위={TRADING_UNIT} USDT")
        
        try:
            # main 함수 실행 (2단계 실행 시 미리 받은 데이터 전달)
            pre_fetched = pre_fetched_data.get(ticker) if not skip_first_row else None
            df_15m = main(polling_start_time=polling_start_time, skip_first_row=skip_first_row, pre_fetched_data=pre_fetched)
            
            # main 함수가 None을 반환하면 (previous 파일 없음 등) 스킵
            if df_15m is None:
                print(f"{get_timestamp()} [{stage_prefix}] ⚠️ {TICKER} 처리 스킵됨 (previous 파일 없음)")
                continue
            
            print(f"{get_timestamp()} [{stage_prefix}] ✅ {TICKER} 처리 완료")
            
            # 15M 데이터 저장 (분석용, 2단계에서만)
            if not skip_first_row:
                ticker_15m_data.append({
                    'ticker': ticker,
                    'df_15m': df_15m.copy()
                })
            
        except Exception as e:
            print(f"{get_timestamp()} [{stage_prefix}] ❌ {TICKER} 처리 실패: {e}")
            import traceback
            traceback.print_exc()
        
        # 다음 티커로 넘어가기 전 잠시 대기 (메모리 정리)
        if i < len(ROTATION_TICKERS) - 1:  # 마지막이 아닌 경우에만
            print(f"{get_timestamp()} [{stage_prefix}] ⏳ 다음 티커로 전환 중...")
            time.sleep(1)
            collected = gc.collect()
            print(f"{get_timestamp()} [{stage_prefix}] 🧹 티커 간 메모리 정리: {collected}개 객체 해제")
    
    # 전체 로테이션 시퀀스 실행 시간 계산 및 출력 (UTC 기준)
    rotation_end_time = dt.datetime.now(tz.UTC)
    total_execution_time = rotation_end_time - rotation_start_time
    
    print(f"{get_timestamp()} [{stage_prefix}] 🎉 로테이션 시퀀스 완료!")
    print(f"{get_timestamp()} [{stage_prefix}] ⏱️ 전체 로테이션 소요시간: {total_execution_time.total_seconds():.2f}초")
    
    # 로테이션 종료 후 메모리 정리
    print(f"{get_timestamp()} [{stage_prefix}] 🧹 메모리 정리 중...")
    # 명시적으로 변수 삭제
    try:
        if 'pre_fetched_data' in locals():
            del pre_fetched_data
        if 'df_5m' in locals():
            del df_5m
        if 'df_15m' in locals():
            del df_15m
        if 'df_1h' in locals():
            del df_1h
        if 'df_1d' in locals():
            del df_1d
    except:
        pass
    
    # 체인 상태 확인을 위해 마지막 티커의 df_15m 저장 (2단계에서만, ticker_15m_data 삭제 전에)
    result_df_15m = None
    if not skip_first_row and ticker_15m_data and len(ticker_15m_data) > 0:
        # 마지막 티커의 df_15m 복사본 저장 (참조가 아닌 복사본)
        result_df_15m = ticker_15m_data[-1]['df_15m'].copy()
    
    # 로테이션 완료 후 4day 분석 (2단계에서만, 한 번에 계산)
    if not skip_first_row:
        try:
            performance_results = []
            _4day_ticker_snapshots = {}
            for data in ticker_15m_data:
                result = analyze_15m_performance(data['df_15m'], data['ticker'])
                performance_results.append(result)
                # 티커별 LS·1HMSFast·종가 스냅샷 (4day 분석 표시용)
                df_15m = data['df_15m']
                if df_15m is not None and len(df_15m) > 0:
                    row0 = df_15m.iloc[0]
                    _4day_ticker_snapshots[data['ticker']] = {
                        'LS': row0.get('LS'),
                        '1HMSFast': row0.get('1HMSFast'),
                        '종': row0.get('종'),
                        'SMAF': row0.get('SMAF'),
                        'SMA100': row0.get('SMA100'),
                    }
            # 엑셀 파일 읽기를 위해 잠시 대기 (파일 저장 완료 대기)
            time.sleep(1)
            
        except Exception as e:
            print(f"{get_timestamp()} ❌ 4day 분석 실패: {e}")
            import traceback
            traceback.print_exc()
        finally:
            # 15M 데이터 정리
            del ticker_15m_data
            collected = gc.collect()
            print(f"{get_timestamp()} [{stage_prefix}] ✅ 메모리 정리 완료 ({collected}개 객체 해제)")
        # 1차 익절 시 SL 본절(BE) 이동: 모든 티커에 대해 TP 1/3 체결 여부 확인 후 BE 재설정
        for _t in ROTATION_TICKERS:
            try:
                check_and_move_sl_to_be(f"{_t}USDT", stage_prefix="[BE]")
            except Exception as e:
                print(f"{get_timestamp()} [BE] ⚠️ {_t} BE 확인 중 오류: {e}")
    else:
        # 1단계도 메모리 정리
        collected1 = gc.collect()
        collected2 = gc.collect()  # 순환 참조 해제를 위해 2번 호출
        collected = collected1 + collected2
        print(f"{get_timestamp()} [{stage_prefix}] ✅ 메모리 정리 완료 ({collected}개 객체 해제)")
    
    # 체인 상태 확인을 위해 df_15m 반환 (2단계에서만)
    if not skip_first_row:
        return result_df_15m  # 2단계: df_15m 반환 (None일 수도 있음)
    else:
        return None  # 1단계: None 반환 (체인 상태 확인 불필요)

def run_every_13_55_28_55_43_55_58_55(initial_window_id: str | None = None, skip_initial_execution: bool = False):
    """
    [스마트 하이브리드 스케줄러]
    - 기본 전략: 2단계(After)만 연속 실행하여 효율성 극대화 (싱글 체인)
    - 안전 장치: 만약 2단계를 놓치거나 실패하면, 다음 1단계(Previous)를 자동으로 실행하여 체인 복구
    
    초회 실행 윈도우 규칙:
    - 0분~15분 사이 1단계 실행 → 15분 전 완료 시 15분1초 2단계, 15분 지나면 스킵
    - 15분~30분 사이 1단계 실행 → 30분 전 완료 시 30분1초 2단계, 30분 지나면 스킵
    - 30분~45분 사이 1단계 실행 → 45분 전 완료 시 45분1초 2단계, 45분 지나면 스킵
    - 45분~60분 사이 1단계 실행 → 0분 전 완료 시 0분1초 2단계, 0분 지나면 스킵
    
    Args:
        skip_initial_execution: True이면 초회 실행을 건너뛰고 정상 폴링으로 시작
    
    Note:
        - 모든 시간 처리는 UTC 기준으로 수행
        - 1단계 실행: UTC 기준 7분, 22분, 37분, 52분 (2단계까지 8분 여유)
        - 2단계 실행: UTC 기준 15분1초, 30분1초, 45분1초, 0분1초
    """
    
    # skip_initial_execution이 True이면 이미 초회 실행 완료된 것으로 간주
    last_executed_key = None  # 마지막 실행 키 (중복 실행 방지)
    if skip_initial_execution:
        # 초기화에서 방금 2단계를 실행했으므로, 현재 속한 2단계 윈도우를 이미 실행한 것으로 표시 (중복 실행 방지)
        now_init = dt.datetime.now(tz.UTC)
        cur_min, cur_sec = now_init.minute, now_init.second
        target_s2_min = None
        if (cur_min == 0 and cur_sec >= 1) or cur_min in (1, 2): target_s2_min = 0
        elif (cur_min == 15 and cur_sec >= 1) or cur_min in (16, 17): target_s2_min = 15
        elif (cur_min == 30 and cur_sec >= 1) or cur_min in (31, 32): target_s2_min = 30
        elif (cur_min == 45 and cur_sec >= 1) or cur_min in (46, 47): target_s2_min = 45
        if target_s2_min is not None:
            if target_s2_min == 0:
                t_s2 = (now_init + dt.timedelta(hours=1)).replace(minute=0, second=1, microsecond=0) if now_init.minute >= 45 else now_init.replace(minute=0, second=1, microsecond=0)
            else:
                t_s2 = now_init.replace(minute=target_s2_min, second=1, microsecond=0)
            if now_init >= t_s2 and (now_init - t_s2).total_seconds() < 120:
                last_executed_key = t_s2.strftime("%Y-%m-%d %H:%M:%S")
    
    # 체인 건강 상태 확인 변수 (True: 정상 연결 중, False: 끊김/초기화 필요)
    # skip_initial_execution=True인 경우 = 초기화에서 체인 검증을 통과한 경우
    # 체인 검증을 통과했을 때만 True로 시작 (체인 검증 실패 시 False로 시작하여 복구 모드)
    chain_is_healthy = True if skip_initial_execution else False
    
    print(f"{get_timestamp()} [스케줄러] 🚀 스마트 하이브리드 스케줄러 가동 (상태: {'정상' if chain_is_healthy else '복구 대기중'})")
    
    # print(f"{get_timestamp()} [스케줄러] 🚀 스마트 하이브리드 스케줄러 가동 (상태: {'정상' if chain_is_healthy else '복구 대기중'})")

    while True:
        try:
            # 현재 시간 (UTC 기준)
            now = dt.datetime.now(tz.UTC)
            current_minute = now.minute
            current_second = now.second
            
            # --- [A] 1단계 (Previous) 실행 로직 (7, 22, 37, 52분) ---
            # 역할: 체인이 끊겼을 때(chain_is_healthy == False)만 실행되는 비상 발전기
            # 시간 마진: 2단계(15, 30, 45, 0분)까지 8분 여유 확보
            target_s1_min = None
            if 7 <= current_minute < 15: target_s1_min = 7
            elif 22 <= current_minute < 30: target_s1_min = 22
            elif 37 <= current_minute < 45: target_s1_min = 37
            elif 52 <= current_minute < 60: target_s1_min = 52
            
            if target_s1_min is not None:
                # 해당 1단계 구간의 시작인지 확인 (구간 내 1회 실행 보장용 키)
                s1_key = f"{now.strftime('%Y-%m-%d %H')}:{target_s1_min:02d}:S1"
                
                # [핵심 로직] 체인이 건강하지 않을 때만 1단계 실행
                if not chain_is_healthy and s1_key != last_executed_key:
                    print(f"\n{get_timestamp()} [PREVIOUS] ⚠️체인 복구 모드 발동: 1단계 실행 ({current_minute}분)")
                    print(f"{get_timestamp()} [PREVIOUS] (직전 2단계를 놓쳤거나 초기화가 필요하여 Previous 파일을 재생성합니다)")
                    
                    # 1단계 실행 (Previous 생성)
                    run_rotation_sequence(polling_start_time=now, skip_first_row=True)
                    
                    last_executed_key = s1_key
                    # 1단계를 수행했으므로, 다음 2단계 실행 자격 획득 (단, 아직 체인이 완성된 건 아님)
                    # 여기서는 chain_is_healthy를 True로 바꾸지 않음 (2단계 성공 시 True로 변경)

            # --- [B] 2단계 (After) 실행 로직 (15, 30, 45, 00분) ---
            target_s2_min = None
            # 현재 시간이 2단계 실행 윈도우 내인지 확인 (정시 1초~+2분 0초)
            if (current_minute == 0 and current_second >= 1) or (current_minute == 1) or (current_minute == 2 and current_second == 0):
                target_s2_min = 0
            elif (current_minute == 15 and current_second >= 1) or (current_minute == 16) or (current_minute == 17 and current_second == 0):
                target_s2_min = 15
            elif (current_minute == 30 and current_second >= 1) or (current_minute == 31) or (current_minute == 32 and current_second == 0):
                target_s2_min = 30
            elif (current_minute == 45 and current_second >= 1) or (current_minute == 46) or (current_minute == 47 and current_second == 0):
                target_s2_min = 45
            
            if target_s2_min is not None:
                # 2단계 타겟 시간 계산
                if target_s2_min == 0:
                    if current_minute >= 45:
                        target_s2_dt = (now + dt.timedelta(hours=1)).replace(minute=0, second=1, microsecond=0)
                    else:
                        target_s2_dt = now.replace(minute=0, second=1, microsecond=0)
                else:
                    target_s2_dt = now.replace(minute=target_s2_min, second=1, microsecond=0)
                
                # 정시 실행 (오차 2분 이내)
                if now >= target_s2_dt and (now - target_s2_dt).total_seconds() < 120:
                    s2_key = target_s2_dt.strftime("%Y-%m-%d %H:%M:%S")
                    
                    if s2_key != last_executed_key:
                        print(f"\n{get_timestamp()} [AFTER] ⏰ 2단계 정시 실행: {current_minute}분 {current_second}초")
                        
                        # 2단계 실행
                        result_df = run_rotation_sequence(polling_start_time=now, skip_first_row=False)
                        last_executed_key = s2_key  # 즉시 갱신하여 동일 윈도우 중복 실행 방지 (이후 예외 발생 시에도 재실행 안 함)
                        
                        # 분석 및 자산기록 (실행 성공 여부와 상관없이 hour=0이면 기록)
                        try:
                            if result_df is not None:
                                analyze_15m_trading_performance()
                            
                            # ========================================================
                            # UTC 0시 자산기록 및 PNLcal.py 실행 (ENABLE_ASSET_RECORDING 시에만, 로그는 LOG_DIR 별도 기록)
                            # ========================================================
                            if ENABLE_ASSET_RECORDING:
                                try:
                                    now_utc_check = dt.datetime.now(tz.UTC)
                                    
                                    # 조건: UTC 시간이 00:00~00:14 범위 (15분 윈도우 확대)
                                    if now_utc_check.hour == 0 and now_utc_check.minute < 15:
                                        # 오늘 기록이 없거나, 혹은 지금이 딱 00시 정각 실행 회차라면 강제 기록
                                        if not check_today_snapshot_exists() or current_minute == 0:
                                            print(f"{get_timestamp()} [자산기록] 📅UTC 00:00 일일 정산 완료. 자산기록을 저장합니다.")
                                            try:
                                                # 1단계: 자산기록 스냅샷 실행
                                                record_total_balance_snapshot(stage_prefix="[자산기록]")
                                                
                                                # 2단계: 자산 기록 완료 디스코드 메시지 전송
                                                snapshot_complete_msg = f"{get_timestamp()} [자산기록] ✅ UTC 00:00 일일 자산기록 완료"
                                                send_discord_message(snapshot_complete_msg)
                                                
                                                # 3단계: 어제(Yesterday) 로그 파일 경로 생성 (로그는 LOG_DIR 별도)
                                                yesterday_utc = now_utc_check - dt.timedelta(days=1)
                                                yesterday_str = yesterday_utc.strftime('%Y%m%d')
                                                
                                                log_dir = os.path.join(script_dir, LOG_DIR)
                                                yesterday_log_file = os.path.join(log_dir, f'BINANCE_FUTURES_log_{yesterday_str}.txt')
                                                
                                                # 4단계: 어제 로그 파일이 존재하면 PNL 분석 실행
                                                if os.path.exists(yesterday_log_file):
                                                    print(f"{get_timestamp()} [자산기록] 📊 {yesterday_str} 일일 정산을 자동 시작합니다.")
                                                    try:
                                                        pnlcal_script = os.path.join(script_dir, 'PNLcal.py')
                                                        if sys.platform == 'win32':
                                                            subprocess.Popen(
                                                                [sys.executable, pnlcal_script, '<자산기록다음시행 바이낸스>'],
                                                                cwd=script_dir,
                                                                stdout=None,
                                                                stderr=None,
                                                                creationflags=subprocess.CREATE_NEW_PROCESS_GROUP
                                                            )
                                                        else:
                                                            subprocess.Popen(
                                                                [sys.executable, pnlcal_script, '<자산기록다음시행 바이낸스>'],
                                                                cwd=script_dir,
                                                                stdout=None,
                                                                stderr=None,
                                                                start_new_session=True
                                                            )
                                                        print(f"{get_timestamp()} [자산기록] ✅ PNL 분석 프로세스 시작: {yesterday_str}")
                                                        pnl_start_msg = f"{get_timestamp()} [자산기록] 📊 {yesterday_str} 일일 PNL 분석 시작"
                                                        send_discord_message(pnl_start_msg)
                                                    except Exception as pnl_error:
                                                        error_msg = f"{get_timestamp()} [자산기록] ⚠️ PNL 자동 정산 중 오류: {pnl_error}"
                                                        print(error_msg)
                                                        send_discord_message(error_msg)
                                                else:
                                                    no_log_msg = f"{get_timestamp()} [자산기록] ℹ️ 어제 로그 파일이 없습니다: {yesterday_str}"
                                                    print(no_log_msg)
                                                    send_discord_message(no_log_msg)
                                                
                                            except Exception as snapshot_error:
                                                error_msg = f"{get_timestamp()} ⚠️자산기록 실행 오류: {snapshot_error}"
                                                print(error_msg)
                                                send_discord_message(error_msg)
                                except Exception as snapshot_error:
                                    print(f"{get_timestamp()} ⚠️자산기록 확인 중 오류 발생: {snapshot_error}")
                            # ========================================================
                        except Exception as e:
                            print(f"{get_timestamp()} [자산기록/분석] ⚠️오류: {e}")
                        
                        # 실행 성공 여부 확인 (결과가 있으면 성공)
                        if result_df is not None:
                            chain_is_healthy = True  # ✅성공! 체인 연결됨
                            print(f"{get_timestamp()} [AFTER] ✅2단계 완료. 체인 상태: 정상 (다음 1단계는 스킵됩니다)")
                        else:
                            chain_is_healthy = False # ❌실패! 체인 끊김
                            print(f"{get_timestamp()} [AFTER] ❌2단계 실패/데이터 없음. 체인 상태: 끊김 (다음 1단계가 실행됩니다)")

            # --- [C] 대기 로직 (0.1초 단위로 나눠서 Ctrl+C 즉시 응답) ---
            for _ in range(10):
                time.sleep(0.1)
                
        except KeyboardInterrupt:
            print(f"{get_timestamp()} [스케줄러] 🛑 사용자에 의해 중단되었습니다.")
            break
        except Exception as e:
            print(f"{get_timestamp()} ❌ 폴링 중 오류: {e}")
            import traceback
            traceback.print_exc()  # 상세 오류 정보 출력
            # 예외 발생 시 체인 상태를 끊김으로 설정 (복구 모드 발동)
            chain_is_healthy = False
            print(f"{get_timestamp()} [스케줄러] ⚠️예외 발생으로 체인 상태를 끊김으로 설정 (다음 1단계에서 복구)")
            # 오류 시 메모리 정리
            gc.collect()
            time.sleep(10)  # 오류 시 10초 대기

if __name__ == "__main__":
    # 폴링·주문·자산기록 상태 (로그는 LOG_DIR에 별도 기록)
    polling_status = "활성화" if ENABLE_POLLING else "비활성화"
    trading_status = "활성화" if ENABLE_TRADING else "비활성화"
    asset_recording_status = "활성화" if ENABLE_ASSET_RECORDING else "비활성화"
    print(f"{get_timestamp()} [활성] 폴링: {polling_status}, 주문 전송: {trading_status}, 자산기록: {asset_recording_status} (로그: {LOG_DIR} 별도)")
    if ENABLE_TRADING:
        spot_status = "활성화" if ENABLE_SPOT_TRADING else "비활성화"
        futures_status = "활성화" if ENABLE_FUTURES_LS_STRATEGY else "비활성화"
        print(f"{get_timestamp()} [활성] 스팟 주문: {spot_status}, 선물(LS) 주문: {futures_status}")
    if not ENABLE_POLLING:
        print(f"{get_timestamp()} [활성] 🔒 폴링 비활성화 → 1회 캔들 수집만 수행 후 종료합니다.")
    print(f"{get_timestamp()} [엑셀] 엑셀 파일 생성 및 자동매매를 수행합니다.")
    print(f"{get_timestamp()} [2단계] 1단계(7분/22분/37분/52분): previous 파일 생성")
    print(f"{get_timestamp()} [2단계] 2단계(15분1초/30분1초/45분1초/0분1초): after 파일 생성 및 주문 전송")
    
    try:
        # 현재 시간
        start_time = dt.datetime.now()
        
        # [스마트 재시작] 최신 캔들 시간을 확인하여 싱글체인 복귀 판단
        # 정확한 캔들 시간(expected_candle_time)과 파일 내 시간을 비교
        has_valid_chain, failed_tickers = check_recent_after_files_exist()
        
        # ==============================================================================
        # [추가] 봇 시작 시 '오늘자 자산기록' 누락 확인 및 즉시 보충 (ENABLE_ASSET_RECORDING 시에만)
        # ==============================================================================
        if ENABLE_ASSET_RECORDING:
            try:
                if not check_today_snapshot_exists():
                    print(f"{get_timestamp()} [자산기록] ⚠️오늘자 자산기록이 없습니다. 보충 기록을 실행합니다")
                    record_total_balance_snapshot(stage_prefix="[자산기록]")
                else:
                    print(f"{get_timestamp()} [자산기록] ✅오늘자 자산기록이 이미 존재합니다.")
            except Exception as e:
                print(f"{get_timestamp()} [초기화] ⚠️자산기록 확인 중 오류 (무시하고 진행): {e}")
        # ==============================================================================
        
        skip_phase1 = False
        
        # 티커별 개별화: 체인 검증을 한 바퀴 돌고, 통과하지 못한 티커들만 곧바로 1단계 실행
        if failed_tickers:
            print(f"{get_timestamp()} [초기화] 🔥 체인 검증 완료. 실패 티커: {', '.join(failed_tickers)} ({len(failed_tickers)}개)")
            print(f"{get_timestamp()} [초기화] 🛠️통과하지 못한 티커들만 곧바로 1단계(Previous) 파일 생성 시작")
            # 실패한 티커만 처리 (통과한 티커는 스킵)
            run_rotation_sequence(polling_start_time=start_time, skip_first_row=True, target_tickers=failed_tickers)
            skip_phase1 = False  # 실패한 티커가 있으면 1단계 실행했으므로 skip_phase1 = False
        elif has_valid_chain:
            print(f"{get_timestamp()} [초기화] 🔗 모든 티커 체인 검증 통과! (캔들 시간 정상)")
            print(f"{get_timestamp()} [초기화] ⏭️1단계(Previous) 생성을 건너뛰고 2단계 대기 모드로 진입합니다.")
            skip_phase1 = True
        else:
            # failed_tickers가 없고 has_valid_chain도 False인 경우 (파일 자체가 없는 경우)
                print(f"{get_timestamp()} [초기화] 🔥 스크립트 시작! (유효 체인 없음 / 오래된 데이터)")
                print(f"{get_timestamp()} [초기화] 🛠️최초 1회 1단계(Previous) 파일 생성 시작")
                run_rotation_sequence(polling_start_time=start_time, skip_first_row=True)
        
        # 1단계 완료(또는 스킵) 시각
        finish_time = dt.datetime.now()
        
        # [2단계 타겟 계산] 현재 시점 기준 다가오는 2단계(After) 타겟 시간 계산
        target_s2 = None
        curr_m = finish_time.minute
        
        if 0 <= curr_m < 15: 
            target_s2 = finish_time.replace(minute=15, second=1, microsecond=0)
        elif 15 <= curr_m < 30: 
            target_s2 = finish_time.replace(minute=30, second=1, microsecond=0)
        elif 30 <= curr_m < 45: 
            target_s2 = finish_time.replace(minute=45, second=1, microsecond=0)
        else:
            # 45분 이상이면 다음 시간 00분
            target_s2 = (finish_time + dt.timedelta(hours=1)).replace(minute=0, second=1, microsecond=0)
        
        # 래깅(Ragging) 체크:
        # - 2단계 실행 시간은 각 15분 단위 "01초" (00:01/15:01/30:01/45:01)
        # - 1단계를 실제로 실행했는데 늦게 끝나 해당 회차 01초를 놓친 경우에만 Ragging → 복구 모드
        # - 체인 검증 통과 후 1단계 스킵(올라탐)인 경우: Ragging 적용 안 함 → 다음 2단계까지 대기
        #   (45:01에 2단계 정상 종료 → 코드 중단 → 재시작 → 체인 검증 → 올라탐 → 다음 00:01 대기)
        is_ragging = False
        if target_s2 and not skip_phase1:
            rem15 = finish_time.minute % 15
            # 0분대에서는 01초 초과부터 "놓침", 1~2분대는 이미 "놓침"으로 간주
            is_ragging = (rem15 == 0 and finish_time.second > 1) or (rem15 in (1, 2))

        # 판단: 2단계 대기 후 실행할지, 즉시 복구 모드로 진입할지
        # - finish_time < target_s2 이면서, 동시에 ragging이 아닐 때만 대기
        if target_s2 and finish_time < target_s2 and not is_ragging:
            wait_sec = (target_s2 - finish_time).total_seconds()
            
            if skip_phase1:
                print(f"{get_timestamp()} [초기화] ⏳기존 체인에 탑승. 2단계({target_s2.strftime('%H:%M:%S')}) 대기 중 ({wait_sec:.0f}초)")
            else:
                print(f"{get_timestamp()} [초기화] ✅1단계 완료. 2단계({target_s2.strftime('%H:%M:%S')}) 대기 ({wait_sec:.0f}초)")
            
            # 대기 루프
            try:
                while dt.datetime.now() < target_s2:
                    remaining = (target_s2 - dt.datetime.now()).total_seconds()
                    if remaining > 1: time.sleep(0.5)
                    else: time.sleep(0.1)
            except KeyboardInterrupt:
                raise  # 상위로 전파하여 메인 예외 처리에서도 처리되도록
            
            print(f"{get_timestamp()} [초기화] 2단계(After) 실행")
            # 2단계 실행
            result_df = run_rotation_sequence(polling_start_time=dt.datetime.now(), skip_first_row=False)
            
            # 초기화 2단계 실행 결과 확인 (체인 검증 통과 + 2단계 성공 시에만 chain_is_healthy = True)
            if result_df is not None:
                try:
                    analyze_15m_trading_performance()
                except Exception as e:
                    print(f"{get_timestamp()} [초기화] ⚠️분석 오류: {e}")
                    # 분석 오류는 체인 자체에 영향 없음, 정상 진입
                
                # 정상 진입 (체인 건강함: 체인 검증 통과 + 2단계 성공)
                # quote 정밀도 사전 로드 (한 번만 조회)
                try:
                    init_symbol_quote_precisions([f"{t}USDT" for t in ROTATION_TICKERS])
                    # print(f"{get_timestamp()} [초기] 심볼 정밀도 로드 완료: {QUOTE_PRECISION_MAP}")
                except Exception as e:
                    print(f"{get_timestamp()} [경고] 심볼 정밀도 로드 실패: {e}")
                
                # 폴링 실행 (2단계 구조) - 체인 건강하므로 초회 실행 스킵 (폴링 비활성화 시 1회만 수집 후 종료)
                # skip_initial_execution=True = 체인 검증 통과 + 2단계 성공 → chain_is_healthy = True
                if ENABLE_POLLING:
                    run_every_13_55_28_55_43_55_58_55(skip_initial_execution=True)
                else:
                    print(f"{get_timestamp()} [활성] 🔒 1회 캔들 수집 완료. 폴링 비활성화로 종료합니다.")
            else:
                # 초기화 2단계 실패 → 체인 복구 모드로 진입
                print(f"{get_timestamp()} [초기화] ❌2단계 실패/데이터 없음. 체인 복구 모드로 진입합니다.")
                try:
                    init_symbol_quote_precisions([f"{t}USDT" for t in ROTATION_TICKERS])
                except Exception as e:
                    print(f"{get_timestamp()} [경고] 심볼 정밀도 로드 실패: {e}")
                # skip_initial_execution=False = 체인 검증 통과했지만 2단계 실패 → chain_is_healthy = False
                if ENABLE_POLLING:
                    run_every_13_55_28_55_43_55_58_55(skip_initial_execution=False)
                else:
                    print(f"{get_timestamp()} [활성] 🔒 1회 캔들 수집 완료. 폴링 비활성화로 종료합니다.")
            
        else:
            # ragging 또는 시간 초과 시 복구 모드로 진입
            if is_ragging:
                msg = "1단계 지연 종료(Ragging)"
            else:
                msg = "타겟 시간 초과" if skip_phase1 else "1단계 지연 종료"
            if target_s2:
                print(f"{get_timestamp()} [초기화] ⚠️{msg}. (2단계 타임테이블 {target_s2.strftime('%H:%M:%S')} 경과)")
            else:
                print(f"{get_timestamp()} [초기화] ⚠️{msg}. (2단계 타겟 시간 계산 실패)")
            print(f"{get_timestamp()} [초기화] ⏭️정규 스케줄러로 진입하여 '체인 복구 모드(다음 1단계)'부터 시작합니다.")
            
            # quote 정밀도 사전 로드 (한 번만 조회)
            try:
                init_symbol_quote_precisions([f"{t}USDT" for t in ROTATION_TICKERS])
                # print(f"{get_timestamp()} [초기] 심볼 정밀도 로드 완료: {QUOTE_PRECISION_MAP}")
            except Exception as e:
                print(f"{get_timestamp()} [경고] 심볼 정밀도 로드 실패: {e}")
            
            # skip_initial_execution=False로 진입하여 다음 1단계(7분, 22분, 37분, 52분)를 기다리게 함
            # 이렇게 하면 chain_is_healthy = False로 시작하여 다음 1단계에서 체인 복구 실행
            if ENABLE_POLLING:
                run_every_13_55_28_55_43_55_58_55(skip_initial_execution=False)
            else:
                print(f"{get_timestamp()} [활성] 🔒 1회 캔들 수집 완료. 폴링 비활성화로 종료합니다.")
            
    except KeyboardInterrupt:
        print(f"\n{get_timestamp()} 🛑 프로그램이 사용자에 의해 중단되었습니다.")
    except Exception as e:
        print(f"{get_timestamp()} [오류] 실행 오류: {e}")
        import traceback
        traceback.print_exc()
    finally:
        # 프로그램 종료 시 메모리 정리
        gc.collect()
        gc.collect()